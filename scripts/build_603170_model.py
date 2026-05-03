#!/usr/bin/env python3
"""Build a financial valuation model + top-10 shareholder history for 603170.SH (宝立食品).

Uses akshare for all data retrieval. Produces a single Excel file with:
  Sheet 1: Financial Data (annual 2018-2024 + 2025 TTM)
  Sheet 2: ROIC Analysis
  Sheet 3: Valuation Model
  Sheet 4: Top 10 Shareholders (quarterly since IPO 2022-06)
  Sheet 5: Data Correlation (notes)

Note: 宝立食品 IPO'd in June 2022, so public data pre-2022 comes from IPO prospectus
(available via akshare back to 2018 fiscal year). Only 7 full annual years + 1 TTM are available.
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import time

import akshare as ak
import pandas as pd


def _retry(fn, *args, **kwargs):
    last_err = None
    for i in range(4):
        try:
            return fn(*args, **kwargs)
        except Exception as e:  # noqa: BLE001
            last_err = e
            print(f"[retry] {fn.__name__} attempt {i+1}: {e}")
            time.sleep(1.5 * (i + 1))
    raise last_err
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TICKER = "603170"
TICKER_SH = "sh603170"
COMPANY = "宝立食品 Baolige Food"
CCY = "CNY"

# Annual reporting periods available (FY-end). IPO 2022-06.
ANNUAL_DATES = ["20181231", "20191231", "20201231", "20211231",
                "20221231", "20231231", "20241231"]
# Quarterly dates for top-10 shareholders (from IPO onward, past ~3.5 years).
SHAREHOLDER_DATES = [
    "20220630", "20220930", "20221231",
    "20230331", "20230630", "20230930", "20231231",
    "20240331", "20240630", "20240930", "20241231",
    "20250331", "20250630", "20250930",
]

# CN 10yr bond yield reference (year-end approx, pct)
CN_10Y_YIELD = {
    "2018": 3.23, "2019": 3.14, "2020": 3.15, "2021": 2.78,
    "2022": 2.84, "2023": 2.56, "2024": 1.68, "2025": 1.83,
}


# ──────────────────────────────────────────────────────────────
# Data fetch
# ──────────────────────────────────────────────────────────────
def fetch_statements() -> dict:
    print("[fetch] income / balance / cash flow from sina …")
    inc = _retry(ak.stock_financial_report_sina, stock=TICKER_SH, symbol="利润表")
    bs = _retry(ak.stock_financial_report_sina, stock=TICKER_SH, symbol="资产负债表")
    cf = _retry(ak.stock_financial_report_sina, stock=TICKER_SH, symbol="现金流量表")
    return {"inc": inc, "bs": bs, "cf": cf}


def fetch_prices() -> pd.DataFrame:
    print("[fetch] daily prices (sina) …")
    df = _retry(ak.stock_zh_a_daily, symbol=TICKER_SH,
                start_date="20220101", end_date="20251231", adjust="")
    # Normalize to columns [日期, 收盘]
    df = df.rename(columns={"date": "日期", "close": "收盘"})
    df["日期"] = pd.to_datetime(df["日期"])
    return df


def fetch_abstract() -> pd.DataFrame:
    print("[fetch] financial abstract (for ROE / EPS) …")
    return _retry(ak.stock_financial_abstract, symbol=TICKER)


def fetch_shareholders() -> dict:
    out = {}
    for d in SHAREHOLDER_DATES:
        try:
            df = _retry(ak.stock_gdfx_top_10_em, symbol=TICKER_SH, date=d)
            out[d] = df
            print(f"[fetch] top-10 shareholders {d}: {len(df)} rows")
        except Exception as e:
            print(f"[warn] {d}: {e}")
    return out


# ──────────────────────────────────────────────────────────────
# Derivation helpers
# ──────────────────────────────────────────────────────────────
def row_for(df: pd.DataFrame, date: str) -> pd.Series | None:
    m = df[df["报告日"] == date]
    return m.iloc[0] if len(m) else None


def safe(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def scale_m(v):
    """Convert CNY → CNY millions."""
    if v is None:
        return None
    return v / 1_000_000.0


def build_annual_data(statements: dict) -> dict:
    """Return dict[year_label] = {metric: value_in_CNYm}."""
    inc, bs, cf = statements["inc"], statements["bs"], statements["cf"]
    data = {}
    for d in ANNUAL_DATES:
        year = d[:4]
        i = row_for(inc, d)
        b = row_for(bs, d)
        c = row_for(cf, d)
        if i is None or b is None:
            continue
        # Short-term debt: 短期借款 + 一年内到期的非流动负债
        st_debt = (safe(b.get("短期借款")) or 0) + (safe(b.get("一年内到期的非流动负债")) or 0)
        lt_debt = (safe(b.get("长期借款")) or 0) + (safe(b.get("应付债券")) or 0)
        total_debt = st_debt + lt_debt
        # Cash & equivalents
        cash = safe(b.get("货币资金")) or 0
        # Leases
        lt_lease = safe(b.get("租赁负债")) or 0
        st_lease = 0  # not broken out in sina; often bundled in "一年内到期" already
        # Working capital proxies
        ca = safe(b.get("流动资产合计"))
        cl = safe(b.get("流动负债合计"))
        wc = (ca - cl) if ca is not None and cl is not None else None
        nwc = None
        if ca is not None and cl is not None:
            nwc = ca - cash - cl + st_debt  # (CA-Cash) − (CL−STDebt)
        # Equity
        equity_parent = safe(b.get("归属于母公司股东权益合计"))
        mi = safe(b.get("少数股东权益")) or 0
        total_equity = (equity_parent or 0) + mi

        # Cash flow
        ocf = safe(c.get("经营活动产生的现金流量净额")) if c is not None else None
        div_paid = safe(c.get("分配股利、利润或偿付利息所支付的现金")) if c is not None else None
        if div_paid is not None:
            div_paid = -abs(div_paid)  # stored as positive outflow in sina; show negative
        buyback = 0  # not separately shown in sina CF for most periods

        data[year] = {
            "Total Revenues": scale_m(safe(i.get("营业总收入"))),
            "Cost Of Revenues": scale_m(safe(i.get("营业成本"))),
            "EBIT": scale_m(safe(i.get("营业利润"))),
            "Income Tax Expense": scale_m(safe(i.get("所得税费用"))),
            "D&A": None,  # not in sina CF summary; leave blank for manual input
            "Total Assets": scale_m(safe(b.get("资产总计"))),
            "Total Equity (Stockholders)": scale_m(equity_parent),
            "Total Debt": scale_m(total_debt),
            "Cash and Equivalents": scale_m(cash),
            "Minority Interest": scale_m(mi),
            "Working Capital": scale_m(wc),
            "Net Working Capital": scale_m(nwc),
            "Cash from Operations": scale_m(ocf),
            "Dividends Paid": scale_m(div_paid),
            "Share Repurchases": scale_m(buyback),
            # ROIC Analysis inputs
            "Long Term Debt": scale_m(lt_debt),
            "Short Term Debt": scale_m(st_debt),
            "Total Equity (incl MI)": scale_m(total_equity),
            "Current Portion of Leases": scale_m(st_lease),
            "Long Term Leases": scale_m(lt_lease),
        }
    return data


def build_market_data(prices: pd.DataFrame, abstract: pd.DataFrame, annual: dict) -> dict:
    """Compute year-end market cap = price × shares outstanding."""
    out = {}
    # Shares outstanding: use 实收资本(或股本) (股本) from balance sheet (万股 units? actually in yuan — it's 股本*面值)
    # For 603170, 面值=1元, 股本 = 400,010,000 shares post-IPO. Pre-IPO ~340m (per prospectus).
    # Use 实收资本(或股本) /1 (par value) as share count.
    bs = ak.stock_financial_report_sina(stock=TICKER_SH, symbol="资产负债表")
    for y in annual:
        d = f"{y}1231"
        row = row_for(bs, d)
        shares = safe(row.get("实收资本(或股本)")) if row is not None else None
        # Get last trading day ≤ Dec 31 of that year
        mask = prices["日期"] <= pd.Timestamp(f"{y}-12-31")
        px = prices[mask].tail(1)["收盘"].values
        if len(px) and shares:
            out[y] = {"shares_m": shares / 1_000_000, "price": float(px[0]),
                      "market_cap_m": shares * float(px[0]) / 1_000_000}
        else:
            out[y] = {"shares_m": shares / 1_000_000 if shares else None,
                      "price": float(px[0]) if len(px) else None,
                      "market_cap_m": None}
    return out


# ──────────────────────────────────────────────────────────────
# Excel build
# ──────────────────────────────────────────────────────────────
BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="008000")
BOLD = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFF")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="305496")
CENTER = Alignment(horizontal="center")


def write_financial_data_sheet(ws, annual: dict, market: dict, years: list[str]):
    ws.title = "Financial Data"
    ws["A1"] = f"{COMPANY} ({TICKER}.SH) - Financial Data"
    ws["A1"].font = BOLD
    ws["B1"] = f"Units: {CCY}M unless noted"
    ws["A2"] = "Note: Company IPO'd 2022-06. Pre-2022 figures from prospectus (sina/akshare)."
    # Header row 3
    ws["A3"] = f"Metric ({CCY}M)"
    ws["A3"].font = BOLD
    for idx, y in enumerate(years):
        col = get_column_letter(2 + idx)
        ws[f"{col}3"] = int(y)
        ws[f"{col}3"].font = BOLD
        ws[f"{col}3"].alignment = CENTER

    rows = [
        # (label, key_or_formula_template, kind)
        # kind: 'input' (blue), 'formula' (black), 'link' (green), 'section' (gray)
        ("Income Statement", None, "section"),
        ("Total Revenues", "Total Revenues", "input"),
        ("Cost Of Revenues", "Cost Of Revenues", "input"),
        ("Gross Profit", "={c}5-{c}6", "formula"),
        ("EBIT", "EBIT", "input"),
        ("Income Tax Expense", "Income Tax Expense", "input"),
        ("NOPAT", "={c}8-{c}9", "formula"),
        ("D&A", "D&A", "input"),
        ("Gross Profit Margin %", "=IFERROR({c}7/{c}5,\"-\")", "formula_pct"),
        ("EBITDA", "=IF(ISNUMBER({c}11),{c}8+{c}11,\"-\")", "formula"),
        ("Balance Sheet", None, "section"),
        ("Total Assets", "Total Assets", "input"),
        ("Total Equity (Stockholders)", "Total Equity (Stockholders)", "input"),
        ("Total Debt", "Total Debt", "input"),
        ("Cash and Equivalents", "Cash and Equivalents", "input"),
        ("Minority Interest", "Minority Interest", "input"),
        ("Working Capital", "Working Capital", "input"),
        ("Net Working Capital", "Net Working Capital", "input"),
        ("Cash Flow", None, "section"),
        ("Cash from Operations", "Cash from Operations", "input"),
        ("Change in Net Working Capital", "change_nwc", "formula"),  # special: C21-B21
        ("Dividends Paid", "Dividends Paid", "input"),
        ("Share Repurchases", "Share Repurchases", "input"),
        ("Free Cash Flow to Equity", "={c}23+{c}24", "formula"),
        ("Market Data", None, "section"),
        ("Market Capitalisation", "market_cap", "input"),
        ("Enterprise Value", "={c}29+{c}17-{c}18+{c}19", "formula"),
        ("EV/EBITDA", "=IF(OR({c}13=0,NOT(ISNUMBER({c}13))),\"-\",{c}30/{c}13)", "formula_mult"),
        ("ROIC", "='ROIC Analysis'!{c}17", "link_pct"),
    ]

    # Map label→row number
    row_no = 4
    label_to_row = {}
    for label, key, kind in rows:
        ws[f"A{row_no}"] = label
        if kind == "section":
            ws[f"A{row_no}"].fill = GRAY_FILL
            ws[f"A{row_no}"].font = BOLD
            for idx in range(len(years)):
                col = get_column_letter(2 + idx)
                ws[f"{col}{row_no}"].fill = GRAY_FILL
        label_to_row[label] = row_no
        row_no += 1

    # Write data
    for idx, y in enumerate(years):
        col = get_column_letter(2 + idx)
        for label, key, kind in rows:
            if kind == "section":
                continue
            r = label_to_row[label]
            cell = ws[f"{col}{r}"]
            if kind == "input":
                if key == "market_cap":
                    v = market.get(y, {}).get("market_cap_m")
                else:
                    v = annual.get(y, {}).get(key)
                if v is not None:
                    cell.value = round(v, 2)
                cell.font = BLUE
            elif kind in ("formula", "formula_pct", "formula_mult"):
                if key == "change_nwc":
                    if idx == 0:
                        cell.value = None
                    else:
                        prev = get_column_letter(1 + idx)
                        cell.value = f"={col}21-{prev}21"
                else:
                    cell.value = key.format(c=col)
                cell.font = BLACK
                if kind == "formula_pct":
                    cell.number_format = "0.0%"
                elif kind == "formula_mult":
                    cell.number_format = "0.0\"x\""
                else:
                    cell.number_format = "#,##0.0"
            elif kind == "link_pct":
                cell.value = key.format(c=col)
                cell.font = GREEN
                cell.number_format = "0.0%"
            cell.alignment = CENTER

    ws.column_dimensions["A"].width = 34
    for idx in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 13
    return label_to_row


def write_roic_sheet(ws, annual: dict, years: list[str]):
    ws.title = "ROIC Analysis"
    ws["A1"] = f"{COMPANY} ({TICKER}.SH) - ROIC Decomposition"
    ws["A1"].font = BOLD
    ws["A3"] = "ROIC = NOPAT / Avg Invested Capital"
    ws["A3"].font = BOLD
    for idx, y in enumerate(years):
        col = get_column_letter(2 + idx)
        ws[f"{col}3"] = int(y)
        ws[f"{col}3"].font = BOLD
        ws[f"{col}3"].alignment = CENTER

    # rows
    layout = [
        (4, "NOPAT Calculation", "section"),
        (5, "EBIT", "link", "='Financial Data'!{c}8"),
        (6, "Less: Income Tax Expense", "link", "='Financial Data'!{c}9"),
        (7, "NOPAT", "formula", "={c}5-{c}6"),
        (8, "Invested Capital Components", "section", None),
        (9, "Long Term Debt", "input", "Long Term Debt"),
        (10, "Short Term Debt", "input", "Short Term Debt"),
        (11, "Total Equity", "input", "Total Equity (incl MI)"),
        (12, "Current Portion of Leases", "input", "Current Portion of Leases"),
        (13, "Long Term Leases", "input", "Long Term Leases"),
        (14, "Invested Capital", "formula", "={c}9+{c}10+{c}11+{c}12+{c}13"),
        (15, "Avg Invested Capital", "avg_ic", None),
        (17, "ROIC", "roic_calc", None),
        (19, "Year-over-Year Changes", "section", None),
        (20, "NOPAT Growth", "yoy_nopat", None),
        (21, "IC Growth", "yoy_ic", None),
    ]
    for r, label, kind, *rest in layout:
        ws[f"A{r}"] = label
        if kind == "section":
            ws[f"A{r}"].fill = GRAY_FILL
            ws[f"A{r}"].font = BOLD
            for idx in range(len(years)):
                c = get_column_letter(2 + idx)
                ws[f"{c}{r}"].fill = GRAY_FILL

    for idx, y in enumerate(years):
        col = get_column_letter(2 + idx)
        prev = get_column_letter(1 + idx) if idx > 0 else None
        for r, label, kind, *rest in layout:
            if kind == "section":
                continue
            cell = ws[f"{col}{r}"]
            if kind == "link":
                cell.value = rest[0].format(c=col)
                cell.font = GREEN
                cell.number_format = "#,##0.0"
            elif kind == "input":
                v = annual.get(y, {}).get(rest[0])
                if v is not None:
                    cell.value = round(v, 2)
                cell.font = BLUE
                cell.number_format = "#,##0.0"
            elif kind == "formula":
                cell.value = rest[0].format(c=col)
                cell.font = BLACK
                cell.number_format = "#,##0.0"
            elif kind == "avg_ic":
                if prev:
                    cell.value = f"=({prev}14+{col}14)/2"
                    cell.font = BLACK
                    cell.number_format = "#,##0.0"
            elif kind == "roic_calc":
                if prev:
                    cell.value = f"=IF({col}15=0,\"-\",{col}7/{col}15)"
                    cell.font = BLACK
                    cell.number_format = "0.0%"
            elif kind == "yoy_nopat":
                if prev:
                    cell.value = f"=IF({prev}7=0,\"-\",({col}7-{prev}7)/{prev}7)"
                    cell.number_format = "0.0%"
            elif kind == "yoy_ic":
                if prev:
                    cell.value = f"=IF({prev}14=0,\"-\",({col}14-{prev}14)/{prev}14)"
                    cell.number_format = "0.0%"
            cell.alignment = CENTER

    ws.column_dimensions["A"].width = 34
    for idx in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 13


def write_valuation_sheet(ws, years: list[str]):
    ws.title = "Valuation Model"
    ws["A1"] = f"{COMPANY} ({TICKER}.SH) - Valuation Model"
    ws["A1"].font = BOLD
    ws["A2"] = "Based on Dividend Yield + ROIC Multiple Methodology"
    ws["A4"] = f"Valuation Parameters ({CCY}M)"
    ws["A4"].font = BOLD
    for idx, y in enumerate(years):
        col = get_column_letter(2 + idx)
        ws[f"{col}4"] = int(y)
        ws[f"{col}4"].font = BOLD
        ws[f"{col}4"].alignment = CENTER

    # Row layout matching template_spec.md
    rows = [
        (5, "Key Assumptions", "section", None),
        (6, "CN 10yr Bond Yield (%)", "bond", None),
        (7, "ROIC", "link", "='ROIC Analysis'!{c}17", "0.0%"),
        (8, "Valuation Multiple (ROIC×100+5)", "f", "=IF(ISNUMBER({c}7),{c}7*100+5,\"-\")", "#,##0.00"),
        (10, "Input Data (from Financial Data)", "section", None),
        (11, "Dividends Paid", "f", "=ABS('Financial Data'!{c}25)", "#,##0.0"),
        (12, "Share Repurchases", "f", "=ABS('Financial Data'!{c}26)", "#,##0.0"),
        (14, "Operating Cash Flow", "f", "='Financial Data'!{c}23", "#,##0.0"),
        (15, "Change in Net Working Capital", "f", "='Financial Data'!{c}24", "#,##0.0"),
        (16, "adj. OCF", "f", "=IF(ISNUMBER({c}15),{c}14+{c}15,{c}14)", "#,##0.0"),
        (17, "D&A", "f", "='Financial Data'!{c}11", "#,##0.0"),
        (19, "NOPAT", "f", "='Financial Data'!{c}10", "#,##0.0"),
        (21, "NOPAT-Dividend-Repurchase(NDR)", "f", "={c}19-{c}11-{c}12", "#,##0.0"),
        (23, "Change of Debt", "chg_debt", None, "#,##0.0"),
        (24, "Change of Equity", "chg_eq", None, "#,##0.0"),
        (26, "Change of IC", "chg_ic", None, "#,##0.0"),
        (28, "Dividend Yield (on Mkt Cap)", "f", "=IF({c}42=0,\"-\",{c}11/{c}42)", "0.00%"),
        (29, "NDR/IC", "f", "=IF(ISNUMBER('ROIC Analysis'!{c}14),{c}21/'ROIC Analysis'!{c}14,\"-\")", "0.0%"),
        (30, "ROIC", "f", "={c}7", "0.0%"),
        (32, "NOPAT Growth", "nopat_gr", None, "0.0%"),
        (34, "Multiple fr. NOPAT Growth", "mult_gr", None, "#,##0.00"),
        (36, "Valuation fr. NOPAT Growth", "val_gr", None, "#,##0.0"),
        (37, "Gap NOPAT Gr vs. Actual", "gap_gr", None, "0.0%"),
        (39, "Total Debt", "f", "='Financial Data'!{c}17", "#,##0.0"),
        (40, "Cash and Equivalents", "f", "='Financial Data'!{c}18", "#,##0.0"),
        (41, "Minority Interest", "f", "='Financial Data'!{c}19", "#,##0.0"),
        (42, "Market Cap (Actual)", "f", "='Financial Data'!{c}29", "#,##0.0"),
        (44, "Valuation Calculations", "section", None),
        (45, "Mkt Cap 1 (Div Yield)", "f", "=IF({c}6=0,\"-\",{c}11/({c}6/100))", "#,##0.0"),
        (46, "Adjusted FCF (OpCF+ChgNWC-Div)", "f", "={c}16-{c}11", "#,##0.0"),
        (47, "EV (ROIC Multiple)", "f", "=IF(ISNUMBER({c}8),{c}46*{c}8,\"-\")", "#,##0.0"),
        (48, "Mkt Cap 2 (EV-Debt+Cash-MI)", "f", "=IF(ISNUMBER({c}47),{c}47-{c}39+{c}40-{c}41,\"-\")", "#,##0.0"),
        (50, "Total Valuation", "f", "=IF(AND(ISNUMBER({c}45),ISNUMBER({c}48)),{c}45+{c}48,\"-\")", "#,##0.0"),
        (51, "Actual Market Cap", "f", "={c}42", "#,##0.0"),
        (52, "Valuation Premium/(Discount)", "f", "=IF(AND(ISNUMBER({c}50),{c}51<>0),{c}50/{c}51,\"-\")", "0.0\"x\""),
        (54, "Capital Structure Analysis", "section", None),
        (55, "Dividend Yield (on Mkt Cap)", "f", "={c}28", "0.00%"),
        (56, "Div Payout Ratio (Div/OpCF)", "f", "=IF({c}14=0,\"-\",{c}11/{c}14)", "0.0%"),
        (57, "Debt / Equity Ratio", "f", "=IF('ROIC Analysis'!{c}11=0,\"-\",'Financial Data'!{c}17/'ROIC Analysis'!{c}11)", "0.00"),
        (58, "Total Equity", "f", "='ROIC Analysis'!{c}11", "#,##0.0"),
        (60, "Mkt Value/Equity", "f", "=IF({c}58=0,\"-\",{c}42/{c}58)", "0.0\"x\""),
    ]
    for r, label, *_ in rows:
        ws[f"A{r}"] = label
    for r, label, kind, *_ in rows:
        if kind == "section":
            ws[f"A{r}"].fill = GRAY_FILL
            ws[f"A{r}"].font = BOLD
            for idx in range(len(years)):
                c = get_column_letter(2 + idx)
                ws[f"{c}{r}"].fill = GRAY_FILL

    for idx, y in enumerate(years):
        col = get_column_letter(2 + idx)
        prev = get_column_letter(1 + idx) if idx > 0 else None
        nextc = get_column_letter(3 + idx) if idx < len(years) - 1 else None
        for r, label, kind, *rest in rows:
            if kind == "section":
                continue
            cell = ws[f"{col}{r}"]
            fmt = rest[1] if len(rest) > 1 else "#,##0.0"
            if kind == "bond":
                cell.value = CN_10Y_YIELD.get(y)
                cell.font = BLUE
                cell.fill = YELLOW_FILL
                cell.number_format = "0.00"
            elif kind == "link":
                cell.value = rest[0].format(c=col)
                cell.font = GREEN
                cell.number_format = fmt
            elif kind == "f":
                cell.value = rest[0].format(c=col)
                cell.font = BLACK
                cell.number_format = fmt
            elif kind == "chg_debt":
                if nextc:
                    cell.value = f"='Financial Data'!{nextc}17-'Financial Data'!{col}17"
                    cell.number_format = fmt
            elif kind == "chg_eq":
                if nextc:
                    cell.value = f"='ROIC Analysis'!{nextc}11-'ROIC Analysis'!{col}11"
                    cell.number_format = fmt
            elif kind == "chg_ic":
                if nextc:
                    cell.value = f"='ROIC Analysis'!{nextc}14-'ROIC Analysis'!{col}14"
                    cell.number_format = fmt
            elif kind == "nopat_gr":
                if prev:
                    cell.value = f"=IF({prev}19=0,\"-\",({col}19-{prev}19)/{prev}19)"
                    cell.number_format = fmt
            elif kind == "mult_gr":
                if prev:
                    cell.value = f"=IF(ISNUMBER({col}32),MAX(0,{col}32*100+5),\"-\")"
                    cell.number_format = fmt
            elif kind == "val_gr":
                if prev:
                    cell.value = f"=IF(ISNUMBER({col}34),{col}19*{col}34,\"-\")"
                    cell.number_format = fmt
            elif kind == "gap_gr":
                if prev:
                    cell.value = f"=IF(AND(ISNUMBER({col}36),{col}42<>0),{col}36/{col}42-1,\"-\")"
                    cell.number_format = fmt
            cell.alignment = CENTER

    ws.column_dimensions["A"].width = 38
    for idx in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + idx)].width = 13


def write_shareholder_sheet(ws, holders: dict):
    ws.title = "Top 10 Shareholders"
    ws["A1"] = f"{COMPANY} ({TICKER}.SH) - Top 10 Shareholders (Quarterly)"
    ws["A1"].font = BOLD
    ws["A2"] = "Source: akshare stock_gdfx_top_10_em (East Money). Holdings in shares, % of total shares outstanding."
    ws["A2"].font = Font(italic=True, color="808080")

    # Layout: columns per quarter-end, for each shareholder rank 1..10, we show: 股东名称 / 持股数 / 占比
    dates = sorted(holders.keys())
    # Build unified shareholder master list (all unique names ever in top-10)
    all_names = []
    for d in dates:
        for n in holders[d]["股东名称"].tolist():
            if n not in all_names:
                all_names.append(n)

    # Header
    ws["A4"] = "Shareholder"
    ws["A4"].font = BOLD_WHITE
    ws["A4"].fill = HEADER_FILL
    for i, d in enumerate(dates):
        col1 = get_column_letter(2 + i * 2)
        col2 = get_column_letter(3 + i * 2)
        ws[f"{col1}3"] = f"{d[:4]}-{d[4:6]}-{d[6:]}"
        ws[f"{col1}3"].font = BOLD
        ws[f"{col1}3"].alignment = CENTER
        ws.merge_cells(f"{col1}3:{col2}3")
        ws[f"{col1}4"] = "持股数 (shares)"
        ws[f"{col2}4"] = "占比 (%)"
        ws[f"{col1}4"].font = BOLD_WHITE
        ws[f"{col1}4"].fill = HEADER_FILL
        ws[f"{col2}4"].font = BOLD_WHITE
        ws[f"{col2}4"].fill = HEADER_FILL
        ws[f"{col1}4"].alignment = CENTER
        ws[f"{col2}4"].alignment = CENTER

    # Rows
    for r_idx, name in enumerate(all_names, start=5):
        ws[f"A{r_idx}"] = name
        for i, d in enumerate(dates):
            col1 = get_column_letter(2 + i * 2)
            col2 = get_column_letter(3 + i * 2)
            df = holders[d]
            match = df[df["股东名称"] == name]
            if len(match):
                row = match.iloc[0]
                ws[f"{col1}{r_idx}"] = int(row["持股数"])
                ws[f"{col1}{r_idx}"].number_format = "#,##0"
                ws[f"{col2}{r_idx}"] = float(row["占总股本持股比例"]) / 100
                ws[f"{col2}{r_idx}"].number_format = "0.00%"
            ws[f"{col1}{r_idx}"].alignment = CENTER
            ws[f"{col2}{r_idx}"].alignment = CENTER

    # Summary section: count of quarters each holder appears
    summary_row = 5 + len(all_names) + 2
    ws[f"A{summary_row}"] = "Summary"
    ws[f"A{summary_row}"].font = BOLD
    ws[f"A{summary_row}"].fill = GRAY_FILL
    ws[f"A{summary_row+1}"] = "Total periods covered"
    ws[f"B{summary_row+1}"] = len(dates)
    ws[f"A{summary_row+2}"] = "Unique top-10 shareholders"
    ws[f"B{summary_row+2}"] = len(all_names)
    ws[f"A{summary_row+3}"] = "Note"
    ws[f"B{summary_row+3}"] = "Empty cell = shareholder was not in top-10 that quarter (may still hold shares)."
    ws[f"B{summary_row+3}"].font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 42
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width = 15
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = 10
    ws.freeze_panes = "B5"


def write_notes_sheet(ws, years: list[str]):
    ws.title = "Data Correlation"
    ws["A1"] = f"{COMPANY} ({TICKER}.SH) - Methodology & Data Sources"
    ws["A1"].font = BOLD
    notes = [
        "",
        "1. Data sources",
        "   • Income / Balance / Cash-Flow statements: akshare.stock_financial_report_sina (source: Sina Finance, 定期报告).",
        "   • Market prices: akshare.stock_zh_a_hist (daily close, 不复权).",
        "   • Top-10 shareholders: akshare.stock_gdfx_top_10_em (source: East Money).",
        "   • CN 10yr bond yield: PBoC / ChinaBond year-end approximations.",
        "",
        "2. Coverage caveats",
        "   • 603170.SH IPO'd 2022-06-23. Pre-2022 statements are from IPO prospectus (restated).",
        "   • Only 7 annual fiscal years (2018-2024) are available. User requested 10 yr — earlier periods not disclosed.",
        "   • D&A is not broken out in Sina CF summary; left blank for manual entry (enter to drive EBITDA).",
        "",
        "3. Valuation methodology (Dual Capitalization)",
        "   Mkt_Cap_1 = Dividends / Bond_Yield",
        "   Adjusted_FCF = OCF + ΔNWC − Dividends",
        "   EV = Adjusted_FCF × (ROIC × 100 + 5)",
        "   Mkt_Cap_2 = EV − Debt + Cash − MI",
        "   Total_Valuation = Mkt_Cap_1 + Mkt_Cap_2",
        "   Premium/(Discount) = Total_Valuation / Actual_Market_Cap",
        "",
        "4. Formula conventions",
        "   • BLUE  text = hardcoded inputs (edit these to change assumptions)",
        "   • BLACK text = formulas",
        "   • GREEN text = cross-sheet references",
        "   • YELLOW fill = key assumption to review (e.g., bond yield)",
        "",
        "5. Cross-checks",
        "   • Financial Data!ROIC (row 32) = ROIC Analysis!row 17",
        "   • Valuation Model row 42 = Financial Data!row 29 = price × shares",
        "",
        f"Built on {dt.date.today().isoformat()} via scripts/build_603170_model.py",
    ]
    for i, n in enumerate(notes, start=2):
        ws[f"A{i}"] = n
    ws.column_dimensions["A"].width = 110


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────
def main():
    out_path = Path.home() / "Desktop" / "BUI_Investment_Management" / "603170.SH"
    out_path.mkdir(parents=True, exist_ok=True)
    fname = out_path / f"603170.SH.Valuation.v{dt.date.today():%y%m%d}.xlsx"

    statements = fetch_statements()
    prices = fetch_prices()
    abstract = fetch_abstract()
    holders = fetch_shareholders()

    annual = build_annual_data(statements)
    market = build_market_data(prices, abstract, annual)
    years = sorted(annual.keys())
    print(f"[build] years covered: {years}")

    wb = Workbook()
    ws1 = wb.active
    write_financial_data_sheet(ws1, annual, market, years)
    write_roic_sheet(wb.create_sheet(), annual, years)
    write_valuation_sheet(wb.create_sheet(), years)
    write_shareholder_sheet(wb.create_sheet(), holders)
    write_notes_sheet(wb.create_sheet(), years)

    wb.save(fname)
    print(f"[done] wrote {fname}")
    print(f"        size = {fname.stat().st_size/1024:.1f} KB")


if __name__ == "__main__":
    main()
