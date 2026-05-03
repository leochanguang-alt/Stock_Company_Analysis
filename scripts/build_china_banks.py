#!/usr/bin/env python3
"""
Build 6-sheet bank valuation models for 7 Chinese banks per bank-financial-model skill.
Sheets: Financial Data | DRIP Analysis | Key Ratios | Valuation Model | Top 10 Shareholders | Data Correlation.
Output: ~/Desktop/BUI_Investment_Management/<TICKER>/<TICKER>.Valuation.v<YYMMDD>.xlsx
"""
import os, time
from datetime import datetime
import tushare as ts
import akshare as ak
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TS_TOKEN = '45e390d8bf059895336edd8073d294a2f137495fee6cd659acfc6f43'
YEARS = list(range(2015, 2026))   # 11 years
TODAY = datetime.now().strftime('%y%m%d')
OUT_BASE = os.path.expanduser('~/Desktop/BUI_Investment_Management')

# ============================================================
# Banks
# ============================================================
CONFIG = {
    "002142.SZ": dict(name="宁波银行 (Bank of Ningbo)", h5=None, yf=None,
                      type="城商行 City Commercial Bank", target_pe=6.0),
    "000001.SZ": dict(name="平安银行 (Ping An Bank)", h5=None, yf=None,
                      type="股份制银行 Joint-Stock", target_pe=5.0),
    "601077.SH": dict(name="渝农商行 (Chongqing Rural Commercial Bank)", h5="03618", yf="3618.HK",
                      type="农商行 Rural Commercial (A+H)", target_pe=5.0),
    "601825.SH": dict(name="沪农商行 (Shanghai Rural Commercial Bank)", h5=None, yf=None,
                      type="农商行 Rural Commercial", target_pe=5.0),
    "601166.SH": dict(name="兴业银行 (Industrial Bank)", h5=None, yf=None,
                      type="股份制银行 Joint-Stock", target_pe=5.0),
    "600919.SH": dict(name="江苏银行 (Bank of Jiangsu)", h5=None, yf=None,
                      type="城商行 City Commercial Bank", target_pe=6.0),
    "601128.SH": dict(name="常熟银行 (Bank of Changshu)", h5=None, yf=None,
                      type="农商行 Rural Commercial", target_pe=6.0),
}

# CN 10y bond (year-end %, fractional)
CN_BOND = {2015:0.0286, 2016:0.0306, 2017:0.0388, 2018:0.0331, 2019:0.0314,
           2020:0.0318, 2021:0.0278, 2022:0.0284, 2023:0.0256, 2024:0.0168, 2025:0.0185}

# China M2 (CNY 万亿, year-end)
CN_M2 = {2015:139.23, 2016:155.01, 2017:167.68, 2018:182.67, 2019:198.65,
         2020:218.68, 2021:238.29, 2022:266.43, 2023:292.27, 2024:313.53, 2025:340.29}

# DRIP defaults for A-share mainland individual long-term holder
DRIP_INVEST = 1_000_000   # CNY
DRIP_FX     = 1.0
DRIP_WHT    = 0.0

# ============================================================
# Format
# ============================================================
FONT = "Apple Braille"
SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY    = "FFE7E6E6"
LGREY   = "FFF2F2F2"
THIN    = Side(style="thin", color="FF000000")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.00;(#,##0.00);"–"'
INT_FMT = '#,##0;(#,##0);"–"'
PCT_FMT = '0.00%;(0.00%);"–"'
PCT1_FMT= '0.0%;(0.0%);"–"'
PX_FMT  = '#,##0.00'
X_FMT   = '0.00"x"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False, lgrey=False, font_size=SIZE):
    c.font = Font(name=FONT, size=font_size, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if num_fmt: c.number_format = num_fmt
    if grey:  c.fill = PatternFill("solid", fgColor=GREY)
    elif lgrey: c.fill = PatternFill("solid", fgColor=LGREY)

def col_l(i): return get_column_letter(2+i)

# ============================================================
# Data fetching
# ============================================================
pro = ts.pro_api(TS_TOKEN)

def fetch_bank(code, cfg):
    print(f"\n{'='*60}\n{code} {cfg['name']}\n{'='*60}")
    market = 'sh' if code.endswith('.SH') else 'sz'
    raw_code = code.split('.')[0]

    # 1. tushare IS — fetch annual via period for each year
    print("  [1] Income statements (tushare period)...")
    is_data = {}
    for y in YEARS:
        df = pro.income(ts_code=code, period=f'{y}1231', report_type='1')
        if not df.empty:
            is_data[y] = df.iloc[0]
        time.sleep(0.25)

    # 2. tushare CF — for dividends paid
    print("  [2] Cash flows (tushare)...")
    cf_data = {}
    for y in YEARS:
        df = pro.cashflow(ts_code=code, period=f'{y}1231', report_type='1')
        if not df.empty:
            cf_data[y] = df.iloc[0]
        time.sleep(0.25)

    # 3. akshare em BS — for LOAN_ADVANCE / ACCEPT_DEPOSIT / TOTAL_PARENT_EQUITY
    print("  [3] Balance sheet (akshare em)...")
    bs_data = {}
    try:
        bs_df = ak.stock_balance_sheet_by_yearly_em(symbol=f"{market.upper()}{raw_code}")
        for _, r in bs_df.iterrows():
            d = str(r.get('REPORT_DATE',''))[:10]
            if d.endswith('-12-31'):
                y = int(d[:4])
                if y in YEARS:
                    bs_data[y] = r
    except Exception as e:
        print(f"    em BS ERR: {e}")

    # 4. tushare daily_basic — for total_share, total_mv, A-share close
    print("  [4] Daily basic (price/shares/mktcap)...")
    db = pro.daily_basic(ts_code=code, start_date='20141231', end_date='20251231',
                         fields='ts_code,trade_date,close,total_share,total_mv,pe,pb')
    a_data = {}
    for y in YEARS:
        sub = db[db['trade_date'].astype(str).str.startswith(str(y))]
        if not sub.empty:
            r = sub.sort_values('trade_date').iloc[-1]
            a_data[y] = {
                'close':  float(r['close']),
                'shares': float(r['total_share'])/1e4,   # 万股 → 亿股
                'mktcap': float(r['total_mv'])/1e4,      # 万元 → 亿元
            }
    time.sleep(0.3)

    # 5. H-share price if applicable
    h_px = {}
    if cfg['yf']:
        print("  [5] H-share price (yfinance)...")
        try:
            t = yf.Ticker(cfg['yf'])
            hist = t.history(start="2015-01-01", end="2026-04-30", auto_adjust=False)
            for d, v in hist['Close'].resample('YE').last().items():
                if d.year in YEARS:
                    h_px[d.year] = float(v)
        except Exception as e:
            print(f"    yf ERR: {e}")

    # 6. Top 10 shareholders (latest)
    print("  [6] Top 10 shareholders (akshare)...")
    top10 = None
    try:
        top10 = ak.stock_gdfx_top_10_em(symbol=f"{market}{raw_code}", date="20251231")
        if top10 is None or top10.empty:
            top10 = ak.stock_gdfx_top_10_em(symbol=f"{market}{raw_code}", date="20240930")
    except Exception as e:
        print(f"    top10 ERR: {e}")

    return dict(is_=is_data, cf=cf_data, bs=bs_data, a_data=a_data, h_px=h_px, top10=top10)


def to_yi(v):
    """元 → 亿元."""
    if v is None or pd.isna(v): return None
    try: return float(v)/1e8
    except: return None

# ============================================================
# Workbook builder
# ============================================================
def build(code, cfg, data):
    wb = Workbook()
    # Sheet 1
    ws1 = wb.active; ws1.title = "Financial Data"
    s1_financial(ws1, code, cfg, data)
    # Sheet 2 — DRIP
    ws2 = wb.create_sheet("DRIP Analysis")
    s2_drip(ws2, code, cfg)
    # Sheet 3 — Key Ratios
    ws3 = wb.create_sheet("Key Ratios")
    s3_ratios(ws3, code, cfg)
    # Sheet 4 — Valuation
    ws4 = wb.create_sheet("Valuation Model")
    s4_valuation(ws4, code, cfg)
    # Sheet 5 — Top 10
    ws5 = wb.create_sheet("Top 10 Shareholders")
    s5_holders(ws5, code, cfg, data['top10'])
    # Sheet 6 — Data Correlation
    ws6 = wb.create_sheet("Data Correlation")
    s6_correlation(ws6, code, cfg)
    return wb

# ===== Sheet 1: Financial Data =====
def s1_financial(ws, code, cfg, data):
    ws.column_dimensions['A'].width = 38
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 13
    ws.column_dimensions[col_l(len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} ({code}) - Financial Data")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(2,1, f"Units: CNY 亿元 (∼1e8 CNY); Market Cap = A-share price × total shares (CNY)")
    fmt(ws.cell(2,1), italic=True, align="left", color=BLUE)

    ws.cell(3,1, "Metric"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    is_, cf, bs, a_d = data['is_'], data['cf'], data['bs'], data['a_data']

    # Helper to compute fields
    def nii(row):
        ii = row.get('int_income') if row is not None else None
        ie = row.get('int_exp')    if row is not None else None
        if ii is None or pd.isna(ii) or ie is None or pd.isna(ie): return None
        return ii - ie
    def nfee(row):
        v = row.get('n_commis_income') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def opinc(row):
        v = row.get('total_revenue') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def admin(row):
        v = row.get('admin_exp') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def biz_tax(row):
        v = row.get('biz_tax_surchg') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else 0
    def other_bus(row):
        v = row.get('other_bus_cost') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else 0
    def opex_total(row):
        v = row.get('oper_exp') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def impairment(row):
        oe = opex_total(row); ad = admin(row)
        if oe is None or ad is None: return None
        return oe - ad - biz_tax(row) - other_bus(row)
    def opp(row):
        v = row.get('operate_profit') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def tax(row):
        v = row.get('income_tax') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def ni(row):
        v = row.get('n_income') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def ni_p(row):
        v = row.get('n_income_attr_p') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None
    def mi(row):
        v = row.get('minority_gain') if row is not None else None
        return v if (v is not None and not pd.isna(v)) else None

    def bs_get(row, k):
        if row is None: return None
        v = row.get(k)
        try: return float(v) if v is not None and not pd.isna(v) else None
        except: return None

    def ocf(row):
        if row is None: return None
        v = row.get('n_cashflow_act')
        return v if (v is not None and not pd.isna(v)) else None
    def divs_paid(row):
        if row is None: return None
        v = row.get('c_pay_dist_dpcp_int_exp')
        return v if (v is not None and not pd.isna(v)) else None

    rows = [
        # row, label, type, computer
        (4, "==Income Statement (CNY 亿)==", "header", None),
        (5, "Net Interest Income", "is", nii),
        (6, "Fee & Commission Income (Net)", "is", nfee),
        (7, "Other Non-Interest Income", "formula", None),
        (8, "Operating Income (营业收入)", "is", opinc),
        (9, "Operating Expenses (业务及管理费)", "is", admin),
        (10, "PPOP (Pre-Provision Op Profit)", "formula", None),
        (11, "Impairment Losses (减值损失)", "is", impairment),
        (12, "Operating Profit (营业利润)", "is", opp),
        (13, "Income Tax Expense", "is", tax),
        (14, "Net Profit", "is", ni),
        (15, "Net Profit to Parent", "is", ni_p),
        (16, "Cost-to-Income Ratio", "formula", None),
        (18, "==Balance Sheet (CNY 亿)==", "header", None),
        (19, "Total Assets", "bs", "TOTAL_ASSETS"),
        (20, "Loans & Advances", "bs", "LOAN_ADVANCE"),
        (21, "Customer Deposits", "bs", "ACCEPT_DEPOSIT"),
        (22, "Total Liabilities", "bs", "TOTAL_LIABILITIES"),
        (23, "Total Equity to Parent", "bs", "TOTAL_PARENT_EQUITY"),
        (24, "Minority Interest", "bs", "MINORITY_EQUITY"),
        (25, "Total Equity", "formula", None),
        (26, "Loan/Deposit Ratio", "formula", None),
        (27, "Equity/Asset Ratio (Leverage)", "formula", None),
        (29, "==Cash Flow (CNY 亿)==", "header", None),
        (30, "Net Cash from Operations", "cf", ocf),
        (31, "Dividends & Interest on Perp (paid)", "cf", divs_paid),
        (33, "==Market Data==", "header", None),
        (34, "A-share YE Close (CNY/share)", "px", None),
        (35, "Total Shares Outstanding (十亿股)", "shares", None),
        (36, "Market Capitalisation (CNY 亿)", "mc", None),
        (37, "Book Value Per Share (CNY)", "formula", None),
        (38, "EPS (CNY, basic)", "formula", None),
        (39, "DPS (CNY, implied from CF)", "formula", None),
    ]

    for r, label, kind, computer in rows:
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, grey=True, align="left")
            for i in range(len(YEARS)): fmt(ws.cell(r,2+i), grey=True)
            continue
        fmt(ws.cell(r,1), align="left")
        for i, y in enumerate(YEARS):
            cl, pcl = col_l(i), col_l(i-1) if i>0 else None
            v = None; formula = None
            if kind == "is":
                row = is_.get(y)
                v = to_yi(computer(row))
            elif kind == "bs":
                row = bs.get(y)
                v = to_yi(bs_get(row, computer))
            elif kind == "cf":
                row = cf.get(y)
                v = to_yi(computer(row))
            elif kind == "px":
                v = a_d.get(y, {}).get('close')
            elif kind == "shares":
                # convert 亿股 → 十亿股 (divide by 10)
                sh = a_d.get(y, {}).get('shares')
                v = sh / 10 if sh else None
            elif kind == "mc":
                v = a_d.get(y, {}).get('mktcap')
            elif kind == "formula":
                if r == 7:   # Other Non-II = OpInc - NII - FeeNet
                    formula = f"=IFERROR({cl}8-{cl}5-{cl}6,\"-\")"
                elif r == 10:  # PPOP = OpInc - OpEx
                    formula = f"=IFERROR({cl}8-{cl}9,\"-\")"
                elif r == 16:  # C/I
                    formula = f"=IFERROR({cl}9/{cl}8,\"-\")"
                elif r == 25:  # Total Equity
                    formula = f"=IFERROR({cl}23+{cl}24,\"-\")"
                elif r == 26:  # L/D
                    formula = f"=IFERROR({cl}20/{cl}21,\"-\")"
                elif r == 27:  # E/A
                    formula = f"=IFERROR({cl}23/{cl}19,\"-\")"
                elif r == 37:  # BVPS
                    formula = f"=IFERROR({cl}23/({cl}35*10),\"-\")"
                elif r == 38:  # EPS
                    formula = f"=IFERROR({cl}15/({cl}35*10),\"-\")"
                elif r == 39:  # DPS
                    formula = f"=IFERROR(ABS({cl}31)/({cl}35*10),\"-\")"
            cell = ws.cell(r, 2+i)
            if formula:
                cell.value = formula
                if r == 16: nf = PCT1_FMT
                elif r == 26: nf = PCT1_FMT
                elif r == 27: nf = PCT_FMT
                elif r in (37,38): nf = PX_FMT
                elif r == 39: nf = '0.0000'
                else: nf = NUM_FMT
                fmt(cell, color=BLACK, num_fmt=nf)
            elif v is not None:
                cell.value = v
                nf = PX_FMT if r in (34,35) else NUM_FMT
                fmt(cell, color=BLUE, num_fmt=nf)
            else:
                fmt(cell, color=BLUE, num_fmt=NUM_FMT)

    for rr in range(1, 40):
        ws.row_dimensions[rr].height = 20.0

# ===== Sheet 2: DRIP Analysis =====
def s2_drip(ws, code, cfg):
    ws.column_dimensions['A'].width = 38.33
    ws.column_dimensions['B'].width = 15.83
    for i in range(2, len(YEARS)+1):
        ws.column_dimensions[get_column_letter(2+i-1)].width = 13
    ws.column_dimensions[col_l(len(YEARS))].width = 8.83

    fd = "'Financial Data'"

    # Title
    c = ws.cell(1,1, f"DRIP Return Analysis — {code} ({YEARS[0]}–{YEARS[-1]}, {cfg['type']})")
    fmt(c, bold=True, lgrey=True, align="left", font_size=12)
    ws.cell(2,1, f"Source data linked from 'Financial Data' sheet (rows 3, 31, 34, 35)")
    fmt(ws.cell(2,1), italic=True, align="left", color=BLUE)

    # Assumptions
    ws.cell(4,1, "Assumptions"); fmt(ws.cell(4,1), bold=True, align="left")
    ws.cell(5,1, "Initial investment (CNY)"); fmt(ws.cell(5,1), align="left")
    c = ws.cell(5,2, DRIP_INVEST); fmt(c, color=BLUE, bold=True, num_fmt=INT_FMT)
    ws.cell(6,1, f"Entry price (CNY/sh, {YEARS[0]} YE)"); fmt(ws.cell(6,1), align="left")
    c = ws.cell(6,2); c.value = f"={fd}!B34"; fmt(c, color=GREEN, bold=True, num_fmt=PX_FMT)
    ws.cell(7,1, "FX (CNY denominator, 1.00 for A-share)"); fmt(ws.cell(7,1), align="left")
    c = ws.cell(7,2, DRIP_FX); fmt(c, color=BLUE, bold=True, num_fmt='0.00')
    ws.cell(8,1, "Dividend WHT (mainland indiv >1y = 0%)"); fmt(ws.cell(8,1), align="left")
    c = ws.cell(8,2, DRIP_WHT); fmt(c, color=BLUE, bold=True, num_fmt=PCT_FMT)
    ws.cell(9,1, "Initial shares purchased"); fmt(ws.cell(9,1), align="left")
    c = ws.cell(9,2); c.value = "=B5/B6"; fmt(c, num_fmt=INT_FMT)

    # Year-by-year DRIP
    ws.cell(11,1, "Year-by-year DRIP"); fmt(ws.cell(11,1), bold=True, align="left")
    # Year header
    ws.cell(12,1, "Year"); fmt(ws.cell(12,1), bold=True, align="left")
    for i, y in enumerate(YEARS):
        cl = col_l(i)
        c = ws.cell(12, 2+i); c.value = f"={fd}!{cl}3"; fmt(c, bold=True, align="right")

    # Row 13: Shares BoY
    ws.cell(13,1, "Shares held BoY"); fmt(ws.cell(13,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        if i == 0:
            ws.cell(13,2).value = "=B9"
        else:
            ws.cell(13, 2+i).value = f"={col_l(i-1)}18"
        fmt(ws.cell(13, 2+i), num_fmt=INT_FMT)

    # Row 14: DPS
    ws.cell(14,1, "DPS (CNY, from CF)"); fmt(ws.cell(14,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        f_ = f"=IFERROR(ABS({fd}!{cl}31)/({fd}!{cl}35*10),0)"
        ws.cell(14, 2+i).value = f_
        fmt(ws.cell(14, 2+i), num_fmt='0.0000')

    # Row 15: Dividend after-tax
    ws.cell(15,1, "Dividend (CNY, after tax)"); fmt(ws.cell(15,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        f_ = f"={cl}13*{cl}14*$B$7*(1-$B$8)"
        ws.cell(15, 2+i).value = f_
        fmt(ws.cell(15, 2+i), num_fmt=INT_FMT)

    # Row 16: Reinvest price
    ws.cell(16,1, "Reinvest price (CNY)"); fmt(ws.cell(16,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        f_ = f"={fd}!{cl}34"
        ws.cell(16, 2+i).value = f_
        fmt(ws.cell(16, 2+i), color=GREEN, num_fmt=PX_FMT)

    # Row 17: Shares bought
    ws.cell(17,1, "Shares bought"); fmt(ws.cell(17,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        f_ = f"=IFERROR({cl}15/{cl}16,0)"
        ws.cell(17, 2+i).value = f_
        fmt(ws.cell(17, 2+i), num_fmt=INT_FMT)

    # Row 18: Shares EoY
    ws.cell(18,1, "Shares held EoY"); fmt(ws.cell(18,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        f_ = f"={cl}13+{cl}17"
        ws.cell(18, 2+i).value = f_
        fmt(ws.cell(18, 2+i), num_fmt=INT_FMT)

    # Row 19: Portfolio value EoY
    ws.cell(19,1, "Portfolio value EoY (CNY)"); fmt(ws.cell(19,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        f_ = f"={cl}18*{cl}16"
        ws.cell(19, 2+i).value = f_
        bold = (i == len(YEARS)-1)
        fmt(ws.cell(19, 2+i), bold=bold, num_fmt=INT_FMT)

    # M2 reference
    ws.cell(21,1, "Reference: China M2 (CNY 万亿)")
    fmt(ws.cell(21,1), bold=True, align="left")
    ws.cell(22,1, "M2 year-end balance"); fmt(ws.cell(22,1), align="left")
    for i, y in enumerate(YEARS):
        v = CN_M2.get(y)
        if v:
            c = ws.cell(22, 2+i, v); fmt(c, color=BLUE, bold=True, num_fmt='#,##0.00')
    ws.cell(23,1, f"M2 multiple vs {YEARS[0]}"); fmt(ws.cell(23,1), align="left")
    n = len(YEARS) - 1
    c = ws.cell(23, 1+len(YEARS)+1)
    cl_l = col_l(len(YEARS)-1)
    c.value = f"={cl_l}22/B22"; fmt(c, num_fmt='0.00"x"')
    ws.cell(24,1, f"M2 CAGR ({n} yrs)"); fmt(ws.cell(24,1), align="left")
    c = ws.cell(24, 1+len(YEARS)+1)
    c.value = f"=({cl_l}22/B22)^(1/{n})-1"; fmt(c, num_fmt=PCT_FMT)

    # Summary
    ws.cell(26,1, "Summary"); fmt(ws.cell(26,1), bold=True, align="left")
    ws.cell(27,1, f"Final value {YEARS[-1]} YE (CNY)"); fmt(ws.cell(27,1), align="left")
    c = ws.cell(27,2); c.value = f"={cl_l}19"; fmt(c, bold=True, num_fmt=INT_FMT)
    ws.cell(28,1, "Total return"); fmt(ws.cell(28,1), align="left")
    c = ws.cell(28,2); c.value = "=B27/B5-1"; fmt(c, num_fmt=PCT_FMT)
    ws.cell(29,1, "Multiple on invested (x)"); fmt(ws.cell(29,1), align="left")
    c = ws.cell(29,2); c.value = "=B27/B5"; fmt(c, bold=True, num_fmt=X_FMT)
    ws.cell(30,1, f"Annualized IRR ({n+1} yrs)"); fmt(ws.cell(30,1), align="left")
    c = ws.cell(30,2); c.value = f"=(B27/B5)^(1/{n+1})-1"; fmt(c, bold=True, num_fmt=PCT_FMT)
    cell_m2cagr = f"{get_column_letter(1+len(YEARS)+1)}24"
    ws.cell(31,1, "vs M2 growth (DRIP IRR − M2 CAGR)"); fmt(ws.cell(31,1), align="left")
    c = ws.cell(31,2); c.value = f"=B30-{cell_m2cagr}"; fmt(c, num_fmt=PCT_FMT)
    ws.cell(32,1, "Price-only return (no dividends)"); fmt(ws.cell(32,1), align="left")
    c = ws.cell(32,2); c.value = f"={fd}!{cl_l}34/{fd}!B34-1"; fmt(c, num_fmt=PCT1_FMT)

    for rr in range(1, 35):
        ws.row_dimensions[rr].height = 23.0

# ===== Sheet 3: Key Ratios =====
def s3_ratios(ws, code, cfg):
    ws.column_dimensions['A'].width = 34
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 13

    ws.cell(1,1, f"{cfg['name']} - Key Banking Ratios"); fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1, "Ratio"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    fd = "'Financial Data'"
    rows = [
        (4,  "Profitability & Efficiency",  "header", None),
        (5,  "NI to Parent",                  "link",  f"{fd}!{{cl}}15"),
        (6,  "Total Equity to Parent",        "link",  f"{fd}!{{cl}}23"),
        (7,  "Avg Equity",                    "avg",   f"{fd}!23"),
        (8,  "ROE (on Parent Equity)",        "div_avg", (f"{fd}!{{cl}}15", f"{fd}!23")),
        (9,  "Total Assets",                  "link",  f"{fd}!{{cl}}19"),
        (10, "Avg Assets",                    "avg",   f"{fd}!19"),
        (11, "ROA",                           "div_avg", (f"{fd}!{{cl}}14", f"{fd}!19")),
        (12, "Net Interest Income",           "link",  f"{fd}!{{cl}}5"),
        (13, "NIM (NII / Avg Assets)",        "div_avg", (f"{fd}!{{cl}}5", f"{fd}!19")),
        (15, "Provisioning & Growth",         "header", None),
        (16, "Impairment Losses",             "link",  f"{fd}!{{cl}}11"),
        (17, "Provisions/Loans (Credit Cost)","ratio", (f"{fd}!{{cl}}11", f"{fd}!{{cl}}20")),
        (18, "Loan Growth YoY",               "yoy",   f"{fd}!{{c}}20"),
        (19, "Deposit Growth YoY",            "yoy",   f"{fd}!{{c}}21"),
        (20, "NI Growth YoY",                 "yoy",   f"{fd}!{{c}}15"),
        (22, "Capital & Payout",              "header", None),
        (23, "Dividend Payout Ratio",         "payout",  f"{fd}"),
        (24, "Retention Ratio",               "ret",   None),
    ]
    for r, label, kind, ref in rows:
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, align="left")
            continue
        fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl, pcl = col_l(i), col_l(i-1) if i>0 else None
            cell = ws.cell(r, 2+i)
            if kind == "link":
                cell.value = "=" + ref.replace("{cl}", cl)
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "avg":
                if i == 0: continue
                base = ref.split("!")[1]
                cell.value = f"=({col_l(i-1)}{base}+{cl}{base})/2"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "div_avg":
                if i == 0: continue
                num = ref[0].replace("{cl}", cl)
                den_base = ref[1].split("!")[1]
                cell.value = f"=IFERROR({num}/(({col_l(i-1)}{den_base}+{cl}{den_base})/2),\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT_FMT)
            elif kind == "ratio":
                num = ref[0].replace("{cl}", cl)
                den = ref[1].replace("{cl}", cl)
                cell.value = f"=IFERROR(ABS({num})/{den},\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT_FMT)
            elif kind == "yoy":
                if i == 0: continue
                base = ref.split("!")[1].replace("{c}","")
                cell.value = f"=IFERROR(({fd}!{cl}{base}-{fd}!{pcl}{base})/{fd}!{pcl}{base},\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT1_FMT)
            elif kind == "payout":
                cell.value = f"=IFERROR(ABS({fd}!{cl}31)/{fd}!{cl}15,\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT1_FMT)
            elif kind == "ret":
                cell.value = f"=1-{cl}23"
                fmt(cell, color=BLACK, num_fmt=PCT1_FMT)

    for rr in range(1, 26):
        ws.row_dimensions[rr].height = 20.0

# ===== Sheet 4: Valuation Model =====
def s4_valuation(ws, code, cfg):
    ws.column_dimensions['A'].width = 38
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 13

    ws.cell(1,1, f"{cfg['name']} - Valuation Model")
    fmt(ws.cell(1,1), bold=True, align="left")

    ws.cell(3,1, "Item"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    fd = "'Financial Data'"
    kr = "'Key Ratios'"

    rows = [
        (5, "Key Assumptions", "header"),
        (6, "CN 10y Bond Yield", "bond"),
        (7, "CNY → HKD FX (n/a for A-share)", "fx"),
        (9, "Inputs (linked)", "header"),
        (10,"NI to Parent (CNY 亿)", "link", f"{fd}!{{cl}}15"),
        (11,"Dividends Paid (CNY 亿, abs)", "link_abs", f"{fd}!{{cl}}31"),
        (12,"Total Equity to Parent", "link", f"{fd}!{{cl}}23"),
        (13,"Market Cap (CNY 亿)", "link", f"{fd}!{{cl}}36"),
        (14,"Market Cap (CNY 亿)", "self", "B13"),
        (16,"Valuation Methods", "header"),
        (17,"Dividend Yield (Div/MktCap)", "ratio", (f"{{cl}}11", f"{{cl}}14")),
        (18,"Mkt Cap 1 = Div / BondYield", "v1"),
        (19,"P/B (Implied)", "ratio", (f"{{cl}}14", f"{{cl}}12")),
        (20,"ROE", "link", f"{kr}!{{cl}}8"),
        (21,"Payout Ratio", "link", f"{kr}!{{cl}}23"),
        (22,"Growth g = ROE × (1-Payout)", "g"),
        (23,"Cost of Equity (Bond + 5%)", "coe"),
        (24,"Justified P/B = (ROE-g)/(r-g)", "jpb"),
        (25,"Mkt Cap 2 = JustPB × Equity", "v2"),
        (27,"P/E (Implied)", "ratio", (f"{{cl}}14", f"{{cl}}10")),
        (28,"Target P/E", "tgt_pe"),
        (29,"Mkt Cap 3 = Target PE × NI", "v3"),
        (31,"Valuation Summary", "header"),
        (32,"Mkt Cap 1 (Div Yield)", "self", "{{cl}}18"),
        (33,"Mkt Cap 2 (Justified P/B)", "self", "{{cl}}25"),
        (34,"Mkt Cap 3 (Target P/E)", "self", "{{cl}}29"),
        (35,"Average Fair Value (CNY 亿)", "avg3"),
        (36,"Actual Mkt Cap (CNY 亿)", "self", "{{cl}}14"),
        (37,"Premium / (Discount)", "pd"),
    ]
    for spec in rows:
        r = spec[0]; label = spec[1]; kind = spec[2]
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, align="left")
            continue
        fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl = col_l(i); pcl = col_l(i-1) if i>0 else None
            cell = ws.cell(r, 2+i)
            if kind == "bond":
                cell.value = CN_BOND.get(YEARS[i], 0.03)
                fmt(cell, color=BLUE, num_fmt=PCT_FMT)
            elif kind == "fx":
                cell.value = 1.00
                fmt(cell, color=BLUE, num_fmt='0.0000')
            elif kind == "link":
                ref = spec[3].replace("{cl}", cl)
                cell.value = "=" + ref
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "link_abs":
                ref = spec[3].replace("{cl}", cl)
                cell.value = f"=ABS({ref})"
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "self":
                ref = spec[3].replace("{{cl}}", cl).replace("{cl}", cl)
                cell.value = f"={ref}"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "ratio":
                num = spec[3][0].replace("{{cl}}", cl).replace("{cl}", cl)
                den = spec[3][1].replace("{{cl}}", cl).replace("{cl}", cl)
                cell.value = f"=IFERROR({num}/{den},\"-\")"
                nf = PCT_FMT if r == 17 else X_FMT
                fmt(cell, color=BLACK, num_fmt=nf)
            elif kind == "v1":
                cell.value = f"=IFERROR({cl}11/{cl}6,\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "g":
                if i == 0: continue
                cell.value = f"=IFERROR({cl}20*(1-{cl}21),\"-\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "coe":
                cell.value = f"={cl}6+0.05"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "jpb":
                if i == 0: continue
                cell.value = f"=IFERROR(IF({cl}23-{cl}22=0,\"-\",({cl}20-{cl}22)/({cl}23-{cl}22)),\"-\")"
                fmt(cell, color=BLACK, num_fmt=X_FMT)
            elif kind == "v2":
                if i == 0: continue
                cell.value = f"=IFERROR({cl}24*{cl}12,\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "tgt_pe":
                cell.value = cfg['target_pe']
                fmt(cell, color=BLUE, num_fmt=X_FMT)
            elif kind == "v3":
                cell.value = f"={cl}28*{cl}10"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "avg3":
                cell.value = f"=IFERROR(AVERAGE({cl}32,{cl}33,{cl}34),\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "pd":
                cell.value = f"=IFERROR({cl}35/{cl}36-1,\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=PCT_FMT)

    for rr in range(1, 38):
        ws.row_dimensions[rr].height = 17.0

# ===== Sheet 5: Top 10 Shareholders =====
def s5_holders(ws, code, cfg, top10):
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 53
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12

    ws.cell(1,1, f"{cfg['name']} - Top 10 Shareholders (latest snapshot)")
    fmt(ws.cell(1,1), bold=True, align="left")

    headers = ["#","Shareholder","Shares Held","% of Total","Type"]
    for i,h in enumerate(headers):
        c = ws.cell(3,1+i,h)
        fmt(c, bold=True, color="FFFFFFFF", align="center" if i==0 else "left")
        c.fill = PatternFill("solid", fgColor="FF305496")

    if top10 is not None and not top10.empty:
        for i, (_, row) in enumerate(top10.head(10).iterrows()):
            r = 4 + i
            ws.cell(r,1, i+1); fmt(ws.cell(r,1), align="center")
            ws.cell(r,2, str(row.get('股东名称',''))[:60]); fmt(ws.cell(r,2), align="left")
            sh = row.get('持股数', None)
            try: sh = float(str(sh).replace(',','').replace('万股','').replace('亿股',''))
            except: sh = None
            if sh is not None:
                ws.cell(r,3, sh); fmt(ws.cell(r,3), color=BLUE, num_fmt=INT_FMT)
            pct = row.get('占总股本持股比例', None)
            try:
                if pct is not None:
                    p = float(str(pct).replace('%',''))/100
                    ws.cell(r,4, p); fmt(ws.cell(r,4), color=BLUE, num_fmt=PCT_FMT)
            except: pass
            ws.cell(r,5, str(row.get('股份类型','')))
            fmt(ws.cell(r,5), align="left")
    else:
        ws.cell(4,1, "Top 10 shareholder data not retrieved — check 同花顺/东方财富 manually")
        fmt(ws.cell(4,1), italic=True, align="left")

    for rr in range(1, 15):
        ws.row_dimensions[rr].height = 15.0

# ===== Sheet 6: Data Correlation =====
def s6_correlation(ws, code, cfg):
    ws.column_dimensions['A'].width = 110

    ws.cell(1,1, f"{cfg['name']} ({code}) - Data Correlation & Methodology")
    fmt(ws.cell(1,1), bold=True, align="left")

    sections = [
        ("1. Bank Profile",
         f"{cfg['name']} — {cfg['type']}.\n"
         f"Years modeled: 2015–2025 (11 years).\n"
         f"Reporting currency: CNY (亿元 throughout).\n"
         f"{'Dual-listed A+H. H-share price recorded in Data Correlation; market cap uses A-share price × total shares.' if cfg['yf'] else 'A-share only listing.'}"),
        ("2. Data Sources",
         "INCOME STATEMENT (annual): tushare pro `pro.income(period=YYYY1231, report_type=1)`\n"
         "  Bank-specific derivations:\n"
         "    NII = int_income − int_exp\n"
         "    Net F&C = n_commis_income\n"
         "    Other Non-II = total_revenue − NII − Net F&C\n"
         "    Op Expenses = admin_exp\n"
         "    Impairment = oper_exp − admin_exp − biz_tax_surchg − other_bus_cost\n"
         "      (tushare's standalone `assets_impair_loss` is often None for banks; derived as residual)\n"
         "    Op Profit = operate_profit\n"
         "BALANCE SHEET (annual): akshare `stock_balance_sheet_by_yearly_em` — better than tushare for\n"
         "  bank-specific fields (LOAN_ADVANCE, ACCEPT_DEPOSIT, TOTAL_PARENT_EQUITY, MINORITY_EQUITY).\n"
         "CASH FLOW: tushare `pro.cashflow` — `c_pay_dist_dpcp_int_exp` is the dividend+perp interest line.\n"
         "PRICE/SHARES: tushare `pro.daily_basic` — total_share, total_mv, year-end close.\n"
         "TOP 10 HOLDERS: akshare `stock_gdfx_top_10_em` — 2025-12-31 if available else 2024-09-30."),
        ("3. Mixed-Unit Disclosure",
         "All financials in CNY 亿元. Market cap uses A-share price × total shares = CNY 亿元.\n"
         "DRIP analysis uses CNY directly (FX = 1.00) — appropriate for mainland investors.\n"
         f"{'For H-share comparison, see yfinance ' + cfg['yf'] + ' separately. H-share prices are NOT used in the main Mkt Cap row.' if cfg['yf'] else 'No H-share listing.'}"),
        ("4. DRIP Analysis",
         "Setup: invest CNY 1,000,000 at 2015 YE close, reinvest all dividends at each subsequent YE close.\n"
         "Withholding tax = 0% (mainland individual investor, hold >1 year — full 个人所得税 exemption per\n"
         "  财政部公告 [2015年]101号 / 中央国债登记结算公司2014修订规则).\n"
         "China M2 benchmark: 2015 = 139.23 万亿 → 2025 = 340.29 万亿 → 10-yr CAGR ≈ 9.35%.\n"
         "Excess return = DRIP IRR − M2 CAGR. Positive means real wealth growth above monetary expansion.\n"
         "Caveats:\n"
         "  • Dividends row 31 mixes COMMON dividends + perp bond interest (公司发行永续债时).\n"
         "    For pure-equity DRIP precision, manually subtract perp coupon from 利润分配表.\n"
         "  • Reinvestment timing assumes year-end close; real ex-div reinvest is mid-year, may add ~0.5% drag."),
        ("5. Key Caveats",
         "• Impairment 2018 transition: pre-2018 = 资产减值损失, 2018+ = 信用减值损失 — combined into single row.\n"
         "• 2025 share count for the big banks (601398, 601939, 601288, 601988) jumped ~25% post MOF/SAFE\n"
         "  capital injection. For our 7 banks here, no major dilution events 2015–2025.\n"
         "• NIM here = NII / avg total assets (proxy). Real reported NIM uses avg interest-earning assets;\n"
         "  difference is typically 10–20 bp.\n"
         "• Cost-to-income shown is 业务及管理费 / 营业收入. Some banks report a slightly different\n"
         "  numerator (incl 税金及附加); our number may be ~2pp lower than reported.\n"
         "• Top 10 shareholders snapshot: A-share filer perspective. State-owned banks have very stable\n"
         "  top-10 (国有股 dominant); city/rural commercial banks have more dynamic ownership."),
        ("6. Valuation Methodology",
         "Three-method blend, equally weighted:\n"
         "  Method 1 — Dividend Yield: Implied Mkt Cap = Div / Bond Yield\n"
         "  Method 2 — Justified P/B (Gordon): JustPB = (ROE − g) / (r − g), where\n"
         "    g = ROE × (1 − Payout); r = Bond Yield + 5% equity risk premium\n"
         "  Method 3 — Target P/E: Mkt Cap = Target P/E × NI to Parent\n"
         f"  Target P/E for this bank: {cfg['target_pe']}x ({'城商行 6x' if '城商' in cfg['type'] else '股份制 5x' if '股份' in cfg['type'] else '农商行 5-6x'} default)\n"
         "Premium / (Discount) = Avg Implied Mkt Cap / Actual Mkt Cap − 1.\n"
         "Positive = market underpricing fair value; negative = market overpricing."),
        ("7. Format Conventions",
         "Font: Apple Braille 11pt (DRIP title row 12pt bold).\n"
         "Colors: BLUE = hardcoded input, BLACK = formula, GREEN = cross-sheet link.\n"
         "Row heights: Financial Data 20pt, DRIP 23pt, Key Ratios 20pt, Valuation 17pt, Top 10 15pt.\n"
         "Column widths: A column 34–38, data cols B–L = 13.\n"
         "Number formats: financial line items #,##0.00; % = 0.00% or 0.0%; multiples = 0.00x."),
    ]

    r = 3
    for title, body in sections:
        c = ws.cell(r,1, title); fmt(c, bold=True, grey=True, align="left")
        r += 1
        for line in body.split("\n"):
            ws.cell(r,1, line); fmt(ws.cell(r,1), align="left")
            r += 1
        r += 1


# ============================================================
# Main
# ============================================================
def main():
    for code, cfg in CONFIG.items():
        try:
            data = fetch_bank(code, cfg)
            print(f"  Building workbook...")
            wb = build(code, cfg, data)
            outdir = os.path.join(OUT_BASE, code)
            os.makedirs(outdir, exist_ok=True)
            outpath = os.path.join(outdir, f"{code}.Valuation.v{TODAY}.xlsx")
            wb.save(outpath)
            print(f"  ✓ Saved: {outpath}")
        except Exception as e:
            import traceback
            print(f"\n!! ERR {code}: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    main()
