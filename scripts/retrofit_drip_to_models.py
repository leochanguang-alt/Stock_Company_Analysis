#!/usr/bin/env python3
"""
Retrofit unified-format DRIP Analysis sheet onto existing generic financial models.
Targets: DDOG, NUS, ILMN, PHG, JNJ, TMO.

For each: open the most recent v260*.xlsx, insert 'DRIP Analysis' as Sheet 2,
rebuild from data already in 'Financial Data' sheet. Save as v260501.xlsx.
"""
import os
from datetime import datetime
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TODAY = datetime.now().strftime('%y%m%d')
OUT_BASE = os.path.expanduser('~/Desktop/BUI_Investment_Management')

# ============================================================
# Per-ticker config: source filename, FD row layout, DRIP setup
# ============================================================
# All 4 healthcare + DDOG + NUS share the same FD row layout (built by
# build_DDOG_model.py / build_NUS_model.py / build_healthcare_models.py):
#   Row 5: Revenue
#   Row 17: Net Income
#   Row 18: NI to Parent
#   Row 26: Total Assets
#   Row 30: Equity
#   Row 32: OCF
#   Row 35: Capex (negative)
#   Row 40: Dividends (negative)
#   Row 42: YE Price
#   Row 43: Shares (M)
#   Row 44: Market Cap
# Note: NUS model has D&A inserted at row 40, Dividends at row 41 → different.
# Note: DDOG has SBC at 39, no Dividends row.
# Let me handle each individually with explicit row mapping.

CONFIG = {
    "DDOG": dict(
        src="BUI-公司研究/DDOG.O/DDOG.O.Valuation.v260429.xlsx",
        ccy="USD", local_label="USD",
        entry_year=2020,    # IPO Sept 2019, first full year 2020
        wht=0.15,           # foreign treaty
        # FD row indexes (per build_DDOG_model.py, validated)
        fd_rows=dict(rev=5, ni_p=18, eq=31, mc=43, px=41, shares=42, divs=None),
        m2_label="US M2 (USD T)",
        m2={2020:19.10, 2021:21.61, 2022:21.39, 2023:20.89, 2024:21.45, 2025:22.20},
        # Hardcoded YE prices and shares for DRIP Market Data section
        ye_px={2020:198.40, 2021:178.11, 2022:73.50, 2023:121.38, 2024:142.89, 2025:135.99},
        shares_m={2020:305, 2021:313, 2022:318, 2023:325, 2024:340, 2025:348},
        holder="Foreign Treaty (US-China 15% WHT)",
        no_div=True,
    ),
    "NUS.N": dict(
        src="BUI-公司研究/NUS.N/NUS.N.Valuation.v260429.xlsx",
        ccy="USD", local_label="USD",
        entry_year=2015,
        wht=0.15,
        # NUS has D&A at row 40, Divs at row 41, Market at 42-45
        fd_rows=dict(rev=5, ni_p=18, eq=31, mc=45, px=43, shares=44, divs=41),
        m2_label="US M2 (USD T)",
        m2={2015:12.34, 2016:13.21, 2017:13.85, 2018:14.38, 2019:15.32,
            2020:19.10, 2021:21.61, 2022:21.39, 2023:20.89, 2024:21.45, 2025:22.20},
        ye_px={2015:37.89, 2016:47.78, 2017:68.23, 2018:61.33, 2019:40.98,
               2020:54.63, 2021:50.75, 2022:42.16, 2023:19.42, 2024:6.89, 2025:9.62},
        shares_m={2015:58.6, 2016:55.7, 2017:54.6, 2018:56.0, 2019:55.4,
                  2020:51.0, 2021:50.5, 2022:50.0, 2023:49.7, 2024:49.7, 2025:50.7},
        holder="Foreign Treaty (US-China 15% WHT)",
    ),
    "ILMN": dict(
        src="ILMN/ILMN.Valuation.v260501.xlsx",
        ccy="USD", local_label="USD",
        entry_year=2016,    # SEC data starts 2016 for ILMN
        wht=0.15,
        fd_rows=dict(rev=5, ni_p=18, eq=30, mc=44, px=42, shares=43, divs=None),
        m2_label="US M2 (USD T)",
        m2={2016:13.21, 2017:13.85, 2018:14.38, 2019:15.32,
            2020:19.10, 2021:21.61, 2022:21.39, 2023:20.89, 2024:21.45, 2025:22.20},
        # YE prices via yfinance — fetch live
        holder="Foreign Treaty (US-China 15% WHT)",
        no_div=True,
    ),
    "PHG": dict(
        src="PHG/PHG.Valuation.v260501.xlsx",
        ccy="EUR", local_label="EUR",
        entry_year=2015,
        wht=0.15,
        fd_rows=dict(rev=5, ni_p=18, eq=30, mc=44, px=42, shares=43, divs=40),
        m2_label="Eurozone M2 (EUR T)",
        m2={2015:9.95, 2016:10.45, 2017:10.95, 2018:11.40, 2019:12.05,
            2020:13.55, 2021:14.65, 2022:15.45, 2023:15.85, 2024:16.20, 2025:16.50},
        holder="Foreign Treaty (15% WHT, varies by jurisdiction)",
    ),
    "JNJ": dict(
        src="JNJ/JNJ.Valuation.v260501.xlsx",
        ccy="USD", local_label="USD",
        entry_year=2016,    # SEC data starts 2016 for JNJ
        wht=0.15,
        fd_rows=dict(rev=5, ni_p=18, eq=30, mc=44, px=42, shares=43, divs=40),
        m2_label="US M2 (USD T)",
        m2={2016:13.21, 2017:13.85, 2018:14.38, 2019:15.32,
            2020:19.10, 2021:21.61, 2022:21.39, 2023:20.89, 2024:21.45, 2025:22.20},
        holder="Foreign Treaty (US-China 15% WHT)",
    ),
    "TMO": dict(
        src="TMO/TMO.Valuation.v260501.xlsx",
        ccy="USD", local_label="USD",
        entry_year=2015,
        wht=0.15,
        fd_rows=dict(rev=5, ni_p=18, eq=30, mc=44, px=42, shares=43, divs=40),
        m2_label="US M2 (USD T)",
        m2={2015:12.34, 2016:13.21, 2017:13.85, 2018:14.38, 2019:15.32,
            2020:19.10, 2021:21.61, 2022:21.39, 2023:20.89, 2024:21.45, 2025:22.20},
        holder="Foreign Treaty (US-China 15% WHT)",
    ),
}

# ============================================================
# Format
# ============================================================
FONT = "Apple Braille"; SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
LGREY = "FFF2F2F2"
THIN = Side(style="thin", color="FF000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = '#,##0.0'
INT = '#,##0'
PCT = '0.00%'
PCT1 = '0.0%'
PX = '#,##0.00'
X = '0.00"x"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", lgrey=False, font_size=SIZE):
    c.font = Font(name=FONT, size=font_size, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if num_fmt: c.number_format = num_fmt
    if lgrey: c.fill = PatternFill("solid", fgColor=LGREY)

def col_l(i): return get_column_letter(2+i)

# ============================================================
# Build DRIP sheet (generic)
# ============================================================
def build_drip(wb, ticker, cfg, fd_data):
    """Insert DRIP Analysis sheet at position 1 (after Financial Data)."""
    fd_idx = wb.sheetnames.index('Financial Data')
    ws = wb.create_sheet('DRIP Analysis', fd_idx + 1)

    rows_fd = cfg['fd_rows']
    YEARS = sorted(fd_data['years'])   # all years from FD with data
    entry = cfg['entry_year']
    drip_years = [y for y in YEARS if y >= entry]
    n = len(drip_years)
    last_col = col_l(n - 1)
    cagr_n = max(n - 1, 1)

    # Year offset: drip_years[0] maps to FD col B + (entry - YEARS[0])
    year_offset = entry - YEARS[0]
    def fdc(i): return get_column_letter(2 + year_offset + i)

    # Column widths
    ws.column_dimensions['A'].width = 38.33
    ws.column_dimensions['B'].width = 15.83
    for i in range(1, n):
        ws.column_dimensions[col_l(i)].width = 13
    ws.column_dimensions[col_l(n)].width = 8.83

    fd = "'Financial Data'"

    # Title
    title = f"DRIP Return Analysis — {ticker} ({entry}–{drip_years[-1]}, {cfg['holder']})"
    c = ws.cell(1, 1, title)
    c.font = Font(name=FONT, size=12, color=BLACK, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor=LGREY)
    for j in range(2, 2+n+1):
        ws.cell(1, j).fill = PatternFill("solid", fgColor=LGREY)

    note = f"Source: 'Financial Data' rows {rows_fd['ni_p']} (NI), {rows_fd['eq']} (Eq), {rows_fd['mc']} (MC)"
    if rows_fd.get('divs'): note += f", {rows_fd['divs']} (Div)"
    note += f". {'Non-dividend payer — DRIP reduces to price-only.' if cfg.get('no_div') else ''}"
    ws.cell(2, 1, note); fmt(ws.cell(2,1), italic=True, color=BLUE, align="left")

    # ---- Assumptions ----
    ws.cell(4,1, "Assumptions"); fmt(ws.cell(4,1), bold=True, align="left")
    ws.cell(5,1, f"Initial investment ({cfg['ccy']})"); fmt(ws.cell(5,1), align="left")
    fmt(ws.cell(5,2, 1_000_000), color=BLUE, bold=True, num_fmt=INT)

    ws.cell(6,1, f"Entry price ({cfg['ccy']}/sh, {entry} YE)"); fmt(ws.cell(6,1), align="left")
    c = ws.cell(6,2); c.value = "=B39"
    fmt(c, color=GREEN, bold=True, num_fmt=PX)

    ws.cell(7,1, f"FX ({cfg['ccy']} denominator)"); fmt(ws.cell(7,1), align="left")
    fmt(ws.cell(7,2, 1.00), color=BLUE, bold=True, num_fmt='0.00')

    ws.cell(8,1, f"Dividend WHT ({cfg['holder'].split('(')[0].strip()})"); fmt(ws.cell(8,1), align="left")
    fmt(ws.cell(8,2, cfg['wht']), color=BLUE, bold=True, num_fmt=PCT)

    ws.cell(9,1, "Initial shares purchased"); fmt(ws.cell(9,1), align="left")
    c = ws.cell(9,2); c.value = "=B5/B6"; fmt(c, num_fmt=INT)

    # ---- Year-by-year DRIP ----
    ws.cell(11,1, "Year-by-year DRIP"); fmt(ws.cell(11,1), bold=True, align="left")

    ws.cell(12,1, "Year"); fmt(ws.cell(12,1), bold=True, align="left")
    for i in range(n):
        c = ws.cell(12, 2+i); c.value = f"={fd}!{fdc(i)}3"; fmt(c, bold=True, align="right")

    # Row 13: Shares BoY
    ws.cell(13,1, "Shares held BoY"); fmt(ws.cell(13,1), align="left")
    for i in range(n):
        if i == 0: ws.cell(13,2).value = "=B9"
        else: ws.cell(13, 2+i).value = f"={col_l(i-1)}18"
        fmt(ws.cell(13, 2+i), num_fmt=INT)

    # Row 14: DPS
    ws.cell(14,1, f"DPS ({cfg['ccy']}, from CF)"); fmt(ws.cell(14,1), align="left")
    for i in range(n):
        if rows_fd.get('divs'):
            f_ = f"=IFERROR(ABS({fd}!{fdc(i)}{rows_fd['divs']})*$B$7/{fdc(i)}40,0)"
        else:
            f_ = "=0"
        ws.cell(14, 2+i).value = f_
        fmt(ws.cell(14, 2+i), num_fmt='0.0000')

    # Row 15: After-tax dividend
    ws.cell(15,1, f"Dividend ({cfg['ccy']}, after WHT)"); fmt(ws.cell(15,1), align="left")
    for i in range(n):
        ws.cell(15, 2+i).value = f"={col_l(i)}13*{col_l(i)}14*(1-$B$8)"
        fmt(ws.cell(15, 2+i), num_fmt=INT)

    # Row 16: Reinvest price
    ws.cell(16,1, f"Reinvest price ({cfg['ccy']})"); fmt(ws.cell(16,1), align="left")
    for i in range(n):
        ws.cell(16, 2+i).value = f"={col_l(i)}39"
        fmt(ws.cell(16, 2+i), color=GREEN, num_fmt=PX)

    # Row 17: Shares bought
    ws.cell(17,1, "Shares bought"); fmt(ws.cell(17,1), align="left")
    for i in range(n):
        ws.cell(17, 2+i).value = f"=IFERROR({col_l(i)}15/{col_l(i)}16,0)"
        fmt(ws.cell(17, 2+i), num_fmt=INT)

    # Row 18: Shares EoY
    ws.cell(18,1, "Shares held EoY"); fmt(ws.cell(18,1), align="left")
    for i in range(n):
        ws.cell(18, 2+i).value = f"={col_l(i)}13+{col_l(i)}17"
        fmt(ws.cell(18, 2+i), num_fmt=INT)

    # Row 19: Portfolio value
    ws.cell(19,1, f"Portfolio value EoY ({cfg['ccy']})"); fmt(ws.cell(19,1), align="left")
    for i in range(n):
        bold = (i == n-1)
        ws.cell(19, 2+i).value = f"={col_l(i)}18*{col_l(i)}16"
        fmt(ws.cell(19, 2+i), bold=bold, num_fmt=INT)

    # Row 20: P/B at YE
    ws.cell(20,1, "P/B at Year-End (= MC / Equity)"); fmt(ws.cell(20,1), align="left")
    for i in range(n):
        ws.cell(20, 2+i).value = f"=IFERROR({fd}!{fdc(i)}{rows_fd['mc']}/{fd}!{fdc(i)}{rows_fd['eq']},\"-\")"
        fmt(ws.cell(20, 2+i), color=GREEN, num_fmt=X)

    # Row 21: Dividend Yield
    ws.cell(21,1, "Dividend Yield (Div / MktCap)"); fmt(ws.cell(21,1), align="left")
    for i in range(n):
        if rows_fd.get('divs'):
            f_ = f"=IFERROR(ABS({fd}!{fdc(i)}{rows_fd['divs']})*$B$7/{fd}!{fdc(i)}{rows_fd['mc']},0)"
        else:
            f_ = "=0"
        ws.cell(21, 2+i).value = f_
        fmt(ws.cell(21, 2+i), num_fmt=PCT)

    # ---- M2 reference ----
    ws.cell(23,1, f"Reference: {cfg['m2_label']}"); fmt(ws.cell(23,1), bold=True, align="left")
    ws.cell(24,1, "M2 year-end balance"); fmt(ws.cell(24,1), align="left")
    for i, y in enumerate(drip_years):
        v = cfg['m2'].get(y)
        if v: fmt(ws.cell(24, 2+i, v), color=BLUE, bold=True, num_fmt='#,##0.00')

    ws.cell(25,1, f"M2 multiple vs {entry}"); fmt(ws.cell(25,1), align="left")
    c = ws.cell(25, 2+n); c.value = f"={last_col}24/B24"; fmt(c, num_fmt=X)
    ws.cell(26,1, f"M2 CAGR ({cagr_n} yrs)"); fmt(ws.cell(26,1), align="left")
    c = ws.cell(26, 2+n); c.value = f"=({last_col}24/B24)^(1/{cagr_n})-1"; fmt(c, num_fmt=PCT)

    # ---- Summary ----
    ws.cell(28,1, "Summary"); fmt(ws.cell(28,1), bold=True, align="left")
    ws.cell(29,1, f"Final value {drip_years[-1]} YE ({cfg['ccy']})"); fmt(ws.cell(29,1), align="left")
    c = ws.cell(29,2); c.value = f"={last_col}19"; fmt(c, bold=True, num_fmt=INT)
    ws.cell(30,1, "Total return"); fmt(ws.cell(30,1), align="left")
    c = ws.cell(30,2); c.value = "=B29/B5-1"; fmt(c, num_fmt=PCT)
    ws.cell(31,1, "Multiple on invested (x)"); fmt(ws.cell(31,1), align="left")
    c = ws.cell(31,2); c.value = "=B29/B5"; fmt(c, bold=True, num_fmt=X)
    ws.cell(32,1, f"Annualized IRR ({n} yrs)"); fmt(ws.cell(32,1), align="left")
    c = ws.cell(32,2); c.value = f"=(B29/B5)^(1/{n})-1"; fmt(c, bold=True, num_fmt=PCT)
    cell_m2cagr = f"{get_column_letter(2+n)}26"
    ws.cell(33,1, "vs M2 growth (DRIP IRR − M2 CAGR)"); fmt(ws.cell(33,1), align="left")
    c = ws.cell(33,2); c.value = f"=B32-{cell_m2cagr}"; fmt(c, num_fmt=PCT)
    ws.cell(34,1, "Price-only return (no dividends)"); fmt(ws.cell(34,1), align="left")
    c = ws.cell(34,2); c.value = f"={last_col}39/B39-1"; fmt(c, num_fmt=PCT1)
    ws.cell(35,1, "Avg P/B over period"); fmt(ws.cell(35,1), align="left")
    c = ws.cell(35,2); c.value = f"=AVERAGE({col_l(0)}20:{last_col}20)"; fmt(c, num_fmt=X)

    # ---- Market Data Reference (BLUE editable) ----
    ws.cell(37,1, "Market Data Reference (BLUE = editable inputs)")
    fmt(ws.cell(37,1), bold=True, align="left")
    ws.cell(38,1, "Year"); fmt(ws.cell(38,1), bold=True, align="left")
    for i, y in enumerate(drip_years):
        c = ws.cell(38, 2+i, y); fmt(c, bold=True, num_fmt="0", align="center")

    ws.cell(39,1, f"YE Close ({cfg['ccy']}/share)"); fmt(ws.cell(39,1), align="left")
    for i, y in enumerate(drip_years):
        # Try cfg ye_px first, fall back to FD link
        v = cfg.get('ye_px', {}).get(y)
        if v is not None:
            fmt(ws.cell(39, 2+i, v), color=BLUE, num_fmt=PX)
        else:
            # Link to FD price row
            ws.cell(39, 2+i).value = f"={fd}!{fdc(i)}{rows_fd['px']}"
            fmt(ws.cell(39, 2+i), color=GREEN, num_fmt=PX)

    ws.cell(40,1, "Total Shares (M)"); fmt(ws.cell(40,1), align="left")
    for i, y in enumerate(drip_years):
        v = cfg.get('shares_m', {}).get(y)
        if v is not None:
            fmt(ws.cell(40, 2+i, v), color=BLUE, num_fmt=INT)
        else:
            ws.cell(40, 2+i).value = f"={fd}!{fdc(i)}{rows_fd['shares']}"
            fmt(ws.cell(40, 2+i), color=GREEN, num_fmt=INT)

    # Row heights 23pt
    for rr in range(1, 41):
        ws.row_dimensions[rr].height = 23.0

    return ws


def get_fd_years(wb, fd_rows):
    """Inspect Financial Data sheet to find which year columns have data."""
    fd = wb['Financial Data']
    years = []
    # Year header row 3
    for col_idx in range(2, fd.max_column+1):
        v = fd.cell(3, col_idx).value
        if isinstance(v, int) and 2010 <= v <= 2030:
            # Check if revenue row has data for this year
            rev = fd.cell(fd_rows['rev'], col_idx).value
            if rev is not None:
                years.append(v)
    return years


def main():
    for ticker, cfg in CONFIG.items():
        try:
            src_path = os.path.join(OUT_BASE, cfg['src'])
            print(f"\n{'='*60}\nProcessing {ticker}\n  src: {src_path}\n{'='*60}")
            if not os.path.exists(src_path):
                print(f"  ✗ Source not found, skip")
                continue
            wb = load_workbook(src_path)
            years = get_fd_years(wb, cfg['fd_rows'])
            print(f"  Years available in FD: {years}")
            fd_data = dict(years=years)
            # Skip if DRIP already exists (idempotent)
            if 'DRIP Analysis' in wb.sheetnames:
                print(f"  'DRIP Analysis' already exists — removing for clean re-add")
                del wb['DRIP Analysis']
            build_drip(wb, ticker, cfg, fd_data)
            # Save with today's version
            dst_path = src_path.rsplit('.v', 1)[0] + f'.v{TODAY}.xlsx'
            wb.save(dst_path)
            print(f"  ✓ Saved: {dst_path}")
            print(f"  Sheets: {wb.sheetnames}")
        except Exception as e:
            import traceback
            print(f"\n!! ERR {ticker}: {e}")
            traceback.print_exc()


if __name__ == "__main__":
    main()
