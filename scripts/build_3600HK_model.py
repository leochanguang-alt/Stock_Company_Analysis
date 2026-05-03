#!/usr/bin/env python3
"""
Financial Valuation Model Generator — 3600.HK 现代牙科 (Modern Dental Group)
HK-listed dental prosthetics manufacturer. Reports in HKD.
Data source: akshare stock_financial_hk_report_em + analysis_indicator.
Units: HKD 亿元 (÷1e8 from raw HKD)
"""

import akshare as ak
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import time
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ─── Config ───────────────────────────────────────────────────────────────────
TS_CODE    = '3600.HK'
HK_CODE5   = '03600'           # 5-digit HK code for akshare
COMPANY    = '现代牙科'
CCY        = 'HKD'
UNIT_LABEL = 'HKD 亿元'
UNIT_DIV   = 1e8               # raw HKD → 亿 HKD

YEARS = list(range(2017, 2026))   # 9 annual years: 2017-2025

# CN 10yr sovereign bond yield (China + global dental prosthetics exposure)
CN_BOND = {
    2017: 0.0390, 2018: 0.0323, 2019: 0.0314, 2020: 0.0315, 2021: 0.0278,
    2022: 0.0284, 2023: 0.0256, 2024: 0.0168, 2025: 0.0183,
}

OUTPUT_DIR  = os.path.expanduser(f'~/Desktop/BUI_Investment_Management/{TS_CODE}/')
os.makedirs(OUTPUT_DIR, exist_ok=True)
TODAY_STR   = datetime.now().strftime('%y%m%d')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'{TS_CODE}.Valuation.v{TODAY_STR}.xlsx')

# ─── Style helpers ────────────────────────────────────────────────────────────
def _font(color, bold=False):
    return Font(name='Apple Braille', size=11, color=color, bold=bold)

F_BLUE  = _font('FF0000FF')
F_BLACK = _font('FF000000')
F_GREEN = _font('FF008000')
F_BOLD  = Font(name='Apple Braille', size=11, bold=True)
F_HDR   = Font(name='Apple Braille', size=11, bold=True, color='FFFFFFFF')

SOLID      = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
GRAY_FILL  = PatternFill(fill_type='solid', fgColor='FFD9D9D9')
HDR_FILL   = PatternFill(fill_type='solid', fgColor='FF305496')

R_ALIGN  = Alignment(horizontal='right')
C_ALIGN  = Alignment(horizontal='center')

def _w(ws, r, c, val, font=None, fill=None, align=None, nf=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:  cell.font  = font
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    if nf:    cell.number_format = nf
    return cell

def w_blue(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, F_BLUE, SOLID, R_ALIGN, nf)

def w_blk(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, F_BLACK, SOLID, R_ALIGN, nf)

def w_grn(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, F_GREEN, SOLID, R_ALIGN, nf)

def w_lbl(ws, r, c, val, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _font('FF000000', bold=bold)
    cell.fill = SOLID
    return cell

def w_sec(ws, r, c, val):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name='Apple Braille', size=11, bold=True)
    cell.fill = GRAY_FILL
    return cell

def set_heights(ws, h, r1=1, r2=None):
    r2 = r2 or ws.max_row
    for r in range(r1, r2+1):
        ws.row_dimensions[r].height = h

def na(v):
    if v is None: return None
    try:
        f = float(v)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None

def to_yi(v):
    """Divide raw HKD by 1e8 to get HKD 亿."""
    v2 = na(v)
    return round(v2 / UNIT_DIV, 4) if v2 is not None else None


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — FETCH HK FINANCIAL DATA
# ═══════════════════════════════════════════════════════════════════════════════

print("=" * 60)
print(f"Fetching {TS_CODE} ({COMPANY}) annual data...")
print("=" * 60)

print("  [1/4] Income statement (年度)...")
df_is = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="利润表", indicator="年度")
time.sleep(0.5)

print("  [2/4] Balance sheet (年度)...")
df_bs = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="资产负债表", indicator="年度")
time.sleep(0.5)

print("  [3/4] Cash flow (年度)...")
df_cf = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="现金流量表", indicator="年度")
time.sleep(0.5)

print("  [4/4] Analysis indicators (年度)...")
df_ind = ak.stock_financial_hk_analysis_indicator_em(symbol=HK_CODE5, indicator="年度")
time.sleep(0.5)

# Price history — try akshare first, fallback to yfinance
print("  [+] Historical price...")
df_px = pd.DataFrame()
for attempt in range(3):
    try:
        df_px = ak.stock_hk_hist(symbol=HK_CODE5, period="daily",
                                 start_date="20160101", end_date="20261231", adjust="")
        print(f"    ✓ akshare: {len(df_px)} rows")
        break
    except Exception as e:
        print(f"    akshare attempt {attempt+1}/3 failed: {type(e).__name__}")
        time.sleep(2)

if df_px is None or df_px.empty:
    print("    → Fallback to yfinance...")
    try:
        import yfinance as yf
        t = yf.Ticker("3600.HK")
        hist = t.history(start="2016-01-01", end="2026-12-31")
        if not hist.empty:
            df_px = pd.DataFrame({
                '日期': hist.index.tz_localize(None) if hist.index.tz else hist.index,
                '收盘': hist['Close'].values,
            })
            df_px['日期'] = pd.to_datetime(df_px['日期'])
            print(f"    ✓ yfinance: {len(df_px)} rows")
    except Exception as e:
        print(f"    yfinance failed: {e}")
time.sleep(0.5)


def pivot_hk(df_long, years):
    """Pivot long-format HK report into dict: {year: {item_name: amount}}."""
    out = {yr: {} for yr in years}
    for _, row in df_long.iterrows():
        rd = str(row.get('REPORT_DATE', ''))[:10]
        if len(rd) < 4: continue
        yr = int(rd[:4])
        if yr not in years: continue
        item = str(row.get('STD_ITEM_NAME', ''))
        amt  = na(row.get('AMOUNT'))
        if amt is not None and item:
            out[yr][item] = amt
    return out

is_d  = pivot_hk(df_is, YEARS)
bs_d  = pivot_hk(df_bs, YEARS)
cf_d  = pivot_hk(df_cf, YEARS)

# Analysis indicator: each row is a year
ind_d = {}
for _, row in df_ind.iterrows():
    rd = str(row.get('REPORT_DATE', ''))[:10]
    if len(rd) < 4: continue
    yr = int(rd[:4])
    if yr in YEARS:
        ind_d[yr] = row


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — PROCESS INTO MODEL DATA
# ═══════════════════════════════════════════════════════════════════════════════

def pick(d, keys):
    """Return first non-null value for any key in keys list."""
    for k in keys:
        if k in d:
            v = na(d[k])
            if v is not None: return v
    return None

# Year-end prices
px_by_year = {}
if df_px is not None and not df_px.empty:
    df_px['日期'] = pd.to_datetime(df_px['日期'])
    df_px['年'] = df_px['日期'].dt.year
    for yr in YEARS:
        sub = df_px[df_px['年'] == yr]
        if not sub.empty:
            last_row = sub.sort_values('日期').iloc[-1]
            px_by_year[yr] = float(last_row['收盘'])

data = {}
for yr in YEARS:
    isd = is_d.get(yr, {})
    bsd = bs_d.get(yr, {})
    cfd = cf_d.get(yr, {})
    ind = ind_d.get(yr, {})

    revenue   = to_yi(pick(isd, ['营业额', '营运收入']))
    gross_p   = to_yi(pick(isd, ['毛利']))
    cogs      = to_yi(pick(isd, ['营运支出']))
    ebit      = to_yi(pick(isd, ['经营溢利']))
    tax       = to_yi(pick(isd, ['税项']))
    net_inc   = to_yi(pick(isd, ['除税后溢利']))
    ni_parent = to_yi(pick(isd, ['股东应占溢利']))

    tot_assets = to_yi(pick(bsd, ['总资产']))
    eq_excl_mi = to_yi(pick(bsd, ['股东权益']))
    eq_incl_mi = to_yi(pick(bsd, ['总权益', '净资产']))
    cash_eq    = to_yi(pick(bsd, ['现金及等价物']))
    st_deposit = to_yi(pick(bsd, ['短期存款']))
    mi         = to_yi(pick(bsd, ['少数股东权益']))
    cur_assets = to_yi(pick(bsd, ['流动资产合计']))
    cur_liab   = to_yi(pick(bsd, ['流动负债合计']))
    st_debt    = to_yi(pick(bsd, ['短期贷款']))
    lt_debt    = to_yi(pick(bsd, ['长期贷款']))
    st_lease   = to_yi(pick(bsd, ['融资租赁负债(流动)']))
    lt_lease   = to_yi(pick(bsd, ['融资租赁负债(非流动)']))

    ocf        = to_yi(pick(cfd, ['经营业务现金净额']))
    div_paid   = to_yi(pick(cfd, ['已付股息(融资)']))
    capex      = to_yi(pick(cfd, ['购建固定资产']))

    # Treat cash + ST deposits as "cash and equivalents" (wider def common for HK)
    cash_tot = None
    if cash_eq is not None or st_deposit is not None:
        cash_tot = round((cash_eq or 0) + (st_deposit or 0), 4)

    # Total Debt
    debt_parts = [x for x in [st_debt, lt_debt] if x is not None]
    tot_debt = round(sum(debt_parts), 4) if debt_parts else 0.0

    # Working Capital
    wc = round(cur_assets - cur_liab, 4) if (cur_assets is not None and cur_liab is not None) else None

    # Net Working Capital = (CA - Cash) - (CL - ST Debt)
    nwc = None
    if cur_assets is not None and cur_liab is not None:
        ca_adj = cur_assets - (cash_tot or 0)
        cl_adj = cur_liab   - (st_debt or 0)
        nwc = round(ca_adj - cl_adj, 4)

    # D&A — not broken out in HK CF; estimate from industry typical or skip
    # HK CF doesn't break out D&A separately; we leave None and note it
    dna = None

    # Share count (reverse-engineer from HOLDER_PROFIT / BASIC_EPS)
    shares = None
    hp = na(ind.get('HOLDER_PROFIT')) if len(ind) else None
    eps = na(ind.get('BASIC_EPS')) if len(ind) else None
    if hp and eps and eps != 0:
        shares = hp / eps  # in shares
        shares = round(shares / 1e8, 4)  # in 亿股

    # Market cap
    mktcap = None
    price = px_by_year.get(yr)
    if price and shares:
        mktcap = round(price * shares, 4)   # HKD 亿

    # Dividends paid — expressed negative (outflow)
    divs_paid = -round(abs(div_paid), 4) if div_paid is not None else None

    data[yr] = dict(
        revenue=revenue, cogs=cogs, gross_p=gross_p, ebit=ebit, tax=tax,
        net_inc=net_inc, ni_parent=ni_parent, dna=dna,
        tot_assets=tot_assets, eq_excl_mi=eq_excl_mi, eq_incl_mi=eq_incl_mi,
        cash_eq=cash_tot, mi=mi, cur_assets=cur_assets, cur_liab=cur_liab,
        st_debt=st_debt, lt_debt=lt_debt, st_lease=st_lease, lt_lease=lt_lease,
        tot_debt=tot_debt, wc=wc, nwc=nwc, shares=shares, price=price,
        ocf=ocf, capex=capex, divs_paid=divs_paid,
        mktcap=mktcap,
    )

print("\nAnnual data summary:")
for yr in YEARS:
    d = data[yr]
    print(f"  {yr}: Rev={d['revenue']} EBIT={d['ebit']} NI={d['ni_parent']} "
          f"Shares={d['shares']} Price={d['price']} MktCap={d['mktcap']}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — INTERIM (SEMI-ANNUAL) DATA
# ═══════════════════════════════════════════════════════════════════════════════

print("\nFetching interim (中报) data for H1/H2 analysis...")
try:
    df_is_int  = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="利润表", indicator="报告期")
    time.sleep(0.5)
    df_bs_int  = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="资产负债表", indicator="报告期")
    time.sleep(0.5)
    df_cf_int  = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="现金流量表", indicator="报告期")
    time.sleep(0.5)
    df_ind_int = ak.stock_financial_hk_analysis_indicator_em(symbol=HK_CODE5, indicator="报告期")
    time.sleep(0.5)
    print(f"  IS interim rows: {len(df_is_int)}  BS: {len(df_bs_int)}  CF: {len(df_cf_int)}  Ind: {len(df_ind_int)}")
except Exception as e:
    print(f"  Interim fetch error: {e}")
    df_is_int = df_bs_int = df_cf_int = pd.DataFrame()
    df_ind_int = pd.DataFrame()

# Collect all unique period dates (YYYY-MM-DD) from IS
periods = []
if not df_is_int.empty:
    for rd in df_is_int['REPORT_DATE'].unique():
        rd_str = str(rd)[:10]
        if rd_str and rd_str not in periods:
            periods.append(rd_str)
periods = sorted(periods, reverse=True)   # newest first
print(f"  Interim periods: {len(periods)}")

def get_by_period(df, period_date, item_name):
    if df.empty: return None
    sub = df[(df['REPORT_DATE'].astype(str).str[:10] == period_date) &
             (df['STD_ITEM_NAME'] == item_name)]
    if sub.empty: return None
    return na(sub.iloc[0]['AMOUNT'])

# Build per-period rows
period_rows = []
for pd_str in periods:
    rev  = to_yi(get_by_period(df_is_int, pd_str, '营业额'))
    cogs = to_yi(get_by_period(df_is_int, pd_str, '营运支出'))
    ebit = to_yi(get_by_period(df_is_int, pd_str, '经营溢利'))
    ni_p = to_yi(get_by_period(df_is_int, pd_str, '股东应占溢利'))
    ocf  = to_yi(get_by_period(df_cf_int, pd_str, '经营业务现金净额'))

    tot_a = to_yi(get_by_period(df_bs_int, pd_str, '总资产'))
    eq_e  = to_yi(get_by_period(df_bs_int, pd_str, '股东权益'))
    eq_i  = to_yi(get_by_period(df_bs_int, pd_str, '总权益'))
    cash  = to_yi(get_by_period(df_bs_int, pd_str, '现金及等价物'))
    stdep = to_yi(get_by_period(df_bs_int, pd_str, '短期存款'))
    cash_tot = round((cash or 0) + (stdep or 0), 4) if (cash or stdep) else None
    mi    = to_yi(get_by_period(df_bs_int, pd_str, '少数股东权益'))
    std   = to_yi(get_by_period(df_bs_int, pd_str, '短期贷款'))
    ltd   = to_yi(get_by_period(df_bs_int, pd_str, '长期贷款'))
    tot_debt = round((std or 0) + (ltd or 0), 4)

    # Period-end price (pick last trading day on/before period date)
    p_dt = pd.Timestamp(pd_str)
    px_val = None
    if df_px is not None and not df_px.empty:
        sub = df_px[df_px['日期'] <= p_dt]
        if not sub.empty:
            px_val = float(sub.sort_values('日期').iloc[-1]['收盘'])

    # Shares from indicator row
    shares = None
    ind_row = df_ind_int[df_ind_int['REPORT_DATE'].astype(str).str[:10] == pd_str]
    if not ind_row.empty:
        hp  = na(ind_row.iloc[0].get('HOLDER_PROFIT'))
        eps = na(ind_row.iloc[0].get('BASIC_EPS'))
        if hp and eps and eps != 0:
            shares = round(hp / eps / 1e8, 4)

    mkt = None
    if px_val and shares:
        mkt = round(px_val * shares, 4)

    period_rows.append(dict(
        period=pd_str, rev=rev, cogs=cogs, ebit=ebit, ni_p=ni_p, ocf=ocf,
        tot_a=tot_a, eq_e=eq_e, eq_i=eq_i, cash=cash_tot, mi=mi,
        tot_debt=tot_debt, price=px_val, shares=shares, mktcap=mkt,
    ))


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — BUILD EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*60); print("Building Excel workbook..."); print("="*60)

wb = Workbook()
ws_names = [
    'Financial Data', 'ROIC Analysis', 'Valuation Model',
    'Semi-Annual Analysis', 'Top 10 Shareholders',
    'Shareholder Analytics', 'Data Correlation'
]
sheets = {}
for i, name in enumerate(ws_names):
    if i == 0:
        ws = wb.active; ws.title = name
    else:
        ws = wb.create_sheet(name)
    sheets[name] = ws

def yr_col(yr): return 2 + YEARS.index(yr)
def yr_cl(yr):  return get_column_letter(yr_col(yr))

# ═══ SHEET 1: Financial Data ═══════════════════════════════════════════════
print("  [1/7] Financial Data...")
ws1 = sheets['Financial Data']

_w(ws1, 1, 1, f'{COMPANY} ({TS_CODE}) - Financial Data', F_BOLD, SOLID)
_w(ws1, 1, 2, f'Units: {UNIT_LABEL} unless noted', _font('FF000000'), SOLID)

_w(ws1, 3, 1, f'Metric ({UNIT_LABEL})', F_BOLD, SOLID)
for yr in YEARS:
    _w(ws1, 3, yr_col(yr), yr, F_BOLD, SOLID, C_ALIGN, 'General')

# Income Statement
w_sec(ws1, 4, 1, 'Income Statement')
ROWS_IS = [
    (5,  'Total Revenues',     'revenue'),
    (6,  'Cost Of Revenues',   'cogs'),
    (8,  'EBIT',               'ebit'),
    (9,  'Income Tax Expense', 'tax'),
    (11, 'D&A',                'dna'),
]
for row, lbl, key in ROWS_IS:
    w_lbl(ws1, row, 1, lbl)
    for yr in YEARS:
        v = data[yr].get(key)
        if v is not None:
            w_blue(ws1, row, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 7, 1, 'Gross Profit')
w_lbl(ws1, 10, 1, 'NOPAT')
w_lbl(ws1, 12, 1, 'Gross Profit Margin %')
w_lbl(ws1, 13, 1, 'EBITDA')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws1, 7, c, f'={cl}5-{cl}6', '#,##0.0')
    w_blk(ws1, 10, c, f'={cl}8-{cl}9', '#,##0.0')
    w_blk(ws1, 12, c, f'=IF({cl}5=0,"-",{cl}7/{cl}5)', '0.0%')
    # EBITDA: since D&A often missing, formula handles empty D&A
    w_blk(ws1, 13, c, f'=IF(ISNUMBER({cl}11),{cl}8+{cl}11,{cl}8)', '#,##0.0')

# Balance Sheet
w_sec(ws1, 14, 1, 'Balance Sheet')
ROWS_BS = [
    (15, 'Total Assets',                    'tot_assets'),
    (16, 'Total Equity (Stockholders)',      'eq_excl_mi'),
    (17, 'Total Debt',                       'tot_debt'),
    (18, 'Cash and Equivalents',             'cash_eq'),
    (19, 'Minority Interest',                'mi'),
    (20, 'Working Capital',                  'wc'),
    (21, 'Net Working Capital',              'nwc'),
]
for row, lbl, key in ROWS_BS:
    w_lbl(ws1, row, 1, lbl)
    for yr in YEARS:
        v = data[yr].get(key)
        if v is not None:
            w_blue(ws1, row, yr_col(yr), v, '#,##0.0')

# Cash Flow
w_sec(ws1, 22, 1, 'Cash Flow')
w_lbl(ws1, 23, 1, 'Cash from Operations')
for yr in YEARS:
    v = data[yr].get('ocf')
    if v is not None: w_blue(ws1, 23, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 24, 1, 'Change in Net Working Capital')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c, cl, pcl = yr_col(yr), yr_cl(yr), yr_cl(YEARS[i-1])
    w_blk(ws1, 24, c, f'={cl}21-{pcl}21', '#,##0.0')

w_lbl(ws1, 25, 1, 'Dividends Paid')
for yr in YEARS:
    v = data[yr].get('divs_paid')
    if v is not None: w_blue(ws1, 25, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 26, 1, 'Share Repurchases')
for yr in YEARS:
    w_blue(ws1, 26, yr_col(yr), 0.0, '#,##0.0')

w_lbl(ws1, 27, 1, 'Free Cash Flow to Equity')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws1, 27, c, f'={cl}23+{cl}24', '#,##0.0')

# Market Data
w_sec(ws1, 28, 1, 'Market Data')
w_lbl(ws1, 29, 1, 'Market Capitalisation')
for yr in YEARS:
    v = data[yr].get('mktcap')
    if v is not None: w_blue(ws1, 29, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 30, 1, 'Enterprise Value')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws1, 30, c, f'={cl}29+{cl}17-{cl}18+{cl}19', '#,##0.0')

w_lbl(ws1, 31, 1, 'EV/EBITDA')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws1, 31, c, f'=IF({cl}13=0,"-",{cl}30/{cl}13)', '0.0"x"')

w_lbl(ws1, 32, 1, 'ROIC')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_grn(ws1, 32, c, f"='ROIC Analysis'!{cl}17", '0.0%')

ws1.column_dimensions['A'].width = 34
for yr in YEARS:
    ws1.column_dimensions[yr_cl(yr)].width = 13
ws1.column_dimensions[get_column_letter(yr_col(YEARS[-1]) + 1)].width = 8.83
set_heights(ws1, 20)


# ═══ SHEET 2: ROIC Analysis ════════════════════════════════════════════════
print("  [2/7] ROIC Analysis...")
ws2 = sheets['ROIC Analysis']

_w(ws2, 1, 1, f'{COMPANY} ({TS_CODE}) - ROIC Decomposition', F_BOLD, SOLID)
_w(ws2, 3, 1, 'ROIC = NOPAT / Avg Invested Capital', F_BOLD, SOLID)
for yr in YEARS:
    _w(ws2, 3, yr_col(yr), yr, F_BOLD, SOLID, C_ALIGN, 'General')

w_sec(ws2, 4, 1, 'NOPAT Calculation')
w_lbl(ws2, 5, 1, 'EBIT')
w_lbl(ws2, 6, 1, 'Less: Income Tax Expense')
w_lbl(ws2, 7, 1, 'NOPAT')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_grn(ws2, 5, c, f"='Financial Data'!{cl}8", '#,##0.0')
    w_grn(ws2, 6, c, f"='Financial Data'!{cl}9", '#,##0.0')
    w_blk(ws2, 7, c, f'={cl}5-{cl}6', '#,##0.0')

w_sec(ws2, 8, 1, 'Invested Capital Components')
ROWS_IC = [
    (9,  'Long Term Debt',             'lt_debt'),
    (10, 'Short Term Debt',            'st_debt'),
    (11, 'Total Equity (incl. MI)',    'eq_incl_mi'),
    (12, 'Current Portion of Leases',  'st_lease'),
    (13, 'Long Term Leases',           'lt_lease'),
]
for row, lbl, key in ROWS_IC:
    w_lbl(ws2, row, 1, lbl)
    for yr in YEARS:
        v = data[yr].get(key)
        w_blue(ws2, row, yr_col(yr), v if v is not None else 0.0, '#,##0.0')

w_lbl(ws2, 14, 1, 'Invested Capital')
w_lbl(ws2, 15, 1, 'Avg Invested Capital')
w_lbl(ws2, 17, 1, 'ROIC', bold=True)
for i, yr in enumerate(YEARS):
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws2, 14, c, f'={cl}9+{cl}10+{cl}11+{cl}12+{cl}13', '#,##0.0')
    if i > 0:
        pcl = yr_cl(YEARS[i-1])
        w_blk(ws2, 15, c, f'=({cl}14+{pcl}14)/2', '#,##0.0')
        w_blk(ws2, 17, c, f'=IF({cl}15=0,"-",{cl}7/{cl}15)', '0.0%')

w_sec(ws2, 19, 1, 'Year-over-Year Changes')
w_lbl(ws2, 20, 1, 'NOPAT Growth')
w_lbl(ws2, 21, 1, 'IC Growth')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c, cl, pcl = yr_col(yr), yr_cl(yr), yr_cl(YEARS[i-1])
    w_blk(ws2, 20, c, f'=IF({pcl}7=0,"-",({cl}7-{pcl}7)/{pcl}7)', '0.0%')
    w_blk(ws2, 21, c, f'=IF({pcl}14=0,"-",({cl}14-{pcl}14)/{pcl}14)', '0.0%')

ws2.column_dimensions['A'].width = 34
for yr in YEARS: ws2.column_dimensions[yr_cl(yr)].width = 13
ws2.column_dimensions[get_column_letter(yr_col(YEARS[-1]) + 1)].width = 8.83
set_heights(ws2, 20)


# ═══ SHEET 3: Valuation Model ══════════════════════════════════════════════
print("  [3/7] Valuation Model...")
ws3 = sheets['Valuation Model']

_w(ws3, 1, 1, f'{COMPANY} ({TS_CODE}) - Valuation Model', F_BOLD, SOLID)
_w(ws3, 2, 1, 'Based on Dividend Yield + ROIC Multiple Methodology', _font('FF000000'), SOLID)
_w(ws3, 4, 1, f'Valuation Parameters ({UNIT_LABEL})', F_BOLD, SOLID)
for yr in YEARS:
    _w(ws3, 4, yr_col(yr), yr, F_BOLD, SOLID, C_ALIGN, 'General')

w_sec(ws3, 5, 1, 'Key Assumptions')
w_lbl(ws3, 6, 1, 'China 10yr Bond Yield (%)')
for yr in YEARS:
    w_blue(ws3, 6, yr_col(yr), CN_BOND[yr], '0.00%')
w_lbl(ws3, 7, 1, 'ROIC')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_grn(ws3, 7, c, f"='ROIC Analysis'!{cl}17", '0.0%')
w_lbl(ws3, 8, 1, 'Valuation Multiple (ROIC×100+5)')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 8, c, f'=IF(ISNUMBER({cl}7),{cl}7*100+5,"-")', '0.0')

w_sec(ws3, 10, 1, 'Input Data (from Financial Data)')
LINKS_VM = [
    (11, 'Dividends Paid',                "=ABS('Financial Data'!{}25)"),
    (12, 'Share Repurchases',             "=ABS('Financial Data'!{}26)"),
    (14, 'Operating Cash Flow',           "='Financial Data'!{}23"),
    (15, 'Change in Net Working Capital', "='Financial Data'!{}24"),
    (17, 'D&A',                           "='Financial Data'!{}11"),
    (19, 'NOPAT',                         "='Financial Data'!{}10"),
    (39, 'Total Debt',                    "='Financial Data'!{}17"),
    (40, 'Cash and Equivalents',          "='Financial Data'!{}18"),
    (41, 'Minority Interest',             "='Financial Data'!{}19"),
    (42, 'Market Cap (Actual)',           "='Financial Data'!{}29"),
]
for row, lbl, tmpl in LINKS_VM:
    w_lbl(ws3, row, 1, lbl)
    for yr in YEARS:
        c, cl = yr_col(yr), yr_cl(yr)
        w_grn(ws3, row, c, tmpl.format(cl), '#,##0.0')

w_lbl(ws3, 16, 1, 'adj. OCF')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 16, c, f'=IF(ISNUMBER({cl}15),{cl}14+{cl}15,{cl}14)', '#,##0.0')

w_lbl(ws3, 21, 1, 'NOPAT-Dividend-Repurchase (NDR)')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 21, c, f'={cl}19-{cl}11-{cl}12', '#,##0.0')

w_lbl(ws3, 23, 1, 'Change of Debt')
w_lbl(ws3, 24, 1, 'Change of Equity')
w_lbl(ws3, 26, 1, 'Change of IC')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c, cl, pcl = yr_col(yr), yr_cl(yr), yr_cl(YEARS[i-1])
    w_blk(ws3, 23, c, f"='Financial Data'!{cl}17-'Financial Data'!{pcl}17", '#,##0.0')
    w_blk(ws3, 24, c, f"='ROIC Analysis'!{cl}11-'ROIC Analysis'!{pcl}11", '#,##0.0')
    w_blk(ws3, 26, c, f"='ROIC Analysis'!{cl}14-'ROIC Analysis'!{pcl}14", '#,##0.0')

w_lbl(ws3, 28, 1, 'Dividend Yield (on Mkt Cap)')
w_lbl(ws3, 29, 1, 'NDR/IC')
w_lbl(ws3, 30, 1, 'ROIC')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 28, c, f'=IF({cl}42=0,"-",{cl}11/{cl}42)', '0.0%')
    w_blk(ws3, 29, c, f"=IF(ISNUMBER('ROIC Analysis'!{cl}14),{cl}21/'ROIC Analysis'!{cl}14,\"-\")", '0.0%')
    w_blk(ws3, 30, c, f'={cl}7', '0.0%')

w_lbl(ws3, 32, 1, 'NOPAT Growth')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c, cl, pcl = yr_col(yr), yr_cl(yr), yr_cl(YEARS[i-1])
    w_blk(ws3, 32, c, f'=IF({pcl}19=0,"-",({cl}19-{pcl}19)/{pcl}19)', '0.0%')

w_lbl(ws3, 34, 1, 'Multiple fr. NOPAT Growth')
w_lbl(ws3, 36, 1, 'Valuation fr. NOPAT Growth')
w_lbl(ws3, 37, 1, 'Gap NOPAT Gr vs. Actual')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 34, c, f'=IF(ISNUMBER({cl}32),MAX(0,{cl}32*100+5),"-")', '0.0')
    w_blk(ws3, 36, c, f'=IF(ISNUMBER({cl}34),{cl}19*{cl}34,"-")', '#,##0.0')
    w_blk(ws3, 37, c, f'=IF(AND(ISNUMBER({cl}36),{cl}42<>0),{cl}36/{cl}42-1,"-")', '0.0%')

w_sec(ws3, 44, 1, 'Valuation Calculations')
w_lbl(ws3, 45, 1, 'Mkt Cap 1 (Div Yield)', bold=True)
w_lbl(ws3, 46, 1, 'Adjusted FCF (adj OCF - Div)')
w_lbl(ws3, 47, 1, 'EV (ROIC Multiple)')
w_lbl(ws3, 48, 1, 'Mkt Cap 2 (EV-Debt+Cash-MI)', bold=True)
w_lbl(ws3, 50, 1, 'Total Valuation', bold=True)
w_lbl(ws3, 51, 1, 'Actual Market Cap')
w_lbl(ws3, 52, 1, 'Valuation Premium/(Discount)', bold=True)
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 45, c, f'=IF({cl}6=0,"-",{cl}11/{cl}6)', '#,##0.0')
    w_blk(ws3, 46, c, f'={cl}16-{cl}11', '#,##0.0')
    w_blk(ws3, 47, c, f'=IF(ISNUMBER({cl}8),{cl}46*{cl}8,"-")', '#,##0.0')
    w_blk(ws3, 48, c, f'=IF(ISNUMBER({cl}47),{cl}47-{cl}39+{cl}40-{cl}41,"-")', '#,##0.0')
    w_blk(ws3, 50, c, f'=IF(AND(ISNUMBER({cl}45),ISNUMBER({cl}48)),{cl}45+{cl}48,"-")', '#,##0.0')
    w_blk(ws3, 51, c, f'={cl}42', '#,##0.0')
    w_blk(ws3, 52, c, f'=IF(AND(ISNUMBER({cl}50),{cl}51<>0),{cl}50/{cl}51,"-")', '0.0"x"')

w_sec(ws3, 54, 1, 'Capital Structure Analysis')
w_lbl(ws3, 55, 1, 'Dividend Yield (on Mkt Cap)')
w_lbl(ws3, 56, 1, 'Div Payout Ratio (Div/OCF)')
w_lbl(ws3, 57, 1, 'Debt / Equity Ratio')
w_lbl(ws3, 58, 1, 'Total Equity (incl. MI)')
w_lbl(ws3, 60, 1, 'Mkt Value / Equity')
for yr in YEARS:
    c, cl = yr_col(yr), yr_cl(yr)
    w_blk(ws3, 55, c, f'={cl}28', '0.0%')
    w_blk(ws3, 56, c, f'=IF({cl}14=0,"-",{cl}11/{cl}14)', '0.0%')
    w_blk(ws3, 57, c, f"=IF('ROIC Analysis'!{cl}11=0,\"-\",'Financial Data'!{cl}17/'ROIC Analysis'!{cl}11)", '0.0"x"')
    w_grn(ws3, 58, c, f"='ROIC Analysis'!{cl}11", '#,##0.0')
    w_blk(ws3, 60, c, f'=IF({cl}58=0,"-",{cl}42/{cl}58)', '0.0"x"')

ws3.column_dimensions['A'].width = 38
for yr in YEARS: ws3.column_dimensions[yr_cl(yr)].width = 13
ws3.column_dimensions[get_column_letter(yr_col(YEARS[-1]) + 1)].width = 8.83
set_heights(ws3, 17)


# ═══ SHEET 4: Semi-Annual Analysis ══════════════════════════════════════════
print("  [4/7] Semi-Annual Analysis...")
ws4 = sheets['Semi-Annual Analysis']

_w(ws4, 1, 1, f'{COMPANY} ({TS_CODE}) - Semi-Annual Analysis', F_BOLD, SOLID)
_w(ws4, 2, 1, 'HK-listed companies report semi-annually (H1 interim + annual). Flow items = H1 or full-year cumulative. Units: HKD 亿',
   _font('FF000000'), SOLID)

for j, row_d in enumerate(period_rows):
    c = j + 2
    _w(ws4, 3, c, row_d['period'], F_BOLD, SOLID, C_ALIGN, 'General')

w_sec(ws4, 4, 1, 'Income Statement (cumulative)')
w_lbl(ws4, 5, 1, 'Revenue')
w_lbl(ws4, 6, 1, 'COGS')
w_lbl(ws4, 7, 1, 'Gross Profit')
w_lbl(ws4, 8, 1, 'Gross Margin %')
w_lbl(ws4, 9, 1, 'EBIT (Operating Profit)')
w_lbl(ws4, 10, 1, 'EBIT Margin %')
w_lbl(ws4, 11, 1, 'NI to Parent')
w_lbl(ws4, 12, 1, 'Net Margin %')

w_sec(ws4, 13, 1, 'Cash Flow')
w_lbl(ws4, 14, 1, 'Operating Cash Flow')

w_sec(ws4, 15, 1, 'Balance Sheet (period-end)')
w_lbl(ws4, 16, 1, 'Total Assets')
w_lbl(ws4, 17, 1, 'Total Equity (Parent)')
w_lbl(ws4, 18, 1, 'Total Equity (incl MI)')
w_lbl(ws4, 19, 1, 'Total Debt')
w_lbl(ws4, 20, 1, 'Cash & Equivalents')
w_lbl(ws4, 21, 1, 'Minority Interest')

w_sec(ws4, 22, 1, 'Market Data')
w_lbl(ws4, 23, 1, 'Period-end Close (HKD)')
w_lbl(ws4, 24, 1, 'Shares (亿股)')
w_lbl(ws4, 25, 1, 'Market Cap (HKD 亿)')
w_lbl(ws4, 26, 1, 'Enterprise Value')

w_sec(ws4, 27, 1, 'Valuation Multiples')
w_lbl(ws4, 28, 1, 'P/E (cumulative × 2 if H1)')
w_lbl(ws4, 29, 1, 'P/B')
w_lbl(ws4, 30, 1, 'ROE (cumulative)')

for j, r in enumerate(period_rows):
    c  = j + 2
    cl = get_column_letter(c)

    vals = [
        (5,  r['rev']),     (6, r['cogs']),   (9, r['ebit']),
        (11, r['ni_p']),    (14,r['ocf']),    (16,r['tot_a']),
        (17, r['eq_e']),    (18,r['eq_i']),   (19,r['tot_debt']),
        (20, r['cash']),    (21,r['mi']),     (23,r['price']),
        (24, r['shares']),  (25,r['mktcap']),
    ]
    for row, v in vals:
        if v is not None:
            nf = '0.00' if row == 23 else '#,##0.0'
            w_blue(ws4, row, c, v, nf)

    # Formulas
    w_blk(ws4, 7,  c, f'={cl}5-{cl}6', '#,##0.0')
    w_blk(ws4, 8,  c, f'=IFERROR({cl}7/{cl}5,"-")', '0.0%')
    w_blk(ws4, 10, c, f'=IFERROR({cl}9/{cl}5,"-")', '0.0%')
    w_blk(ws4, 12, c, f'=IFERROR({cl}11/{cl}5,"-")', '0.0%')
    w_blk(ws4, 26, c, f'={cl}25+{cl}19-{cl}20+{cl}21', '#,##0.0')
    w_blk(ws4, 28, c, f'=IFERROR({cl}25/{cl}11,"-")', '0.0"x"')
    w_blk(ws4, 29, c, f'=IFERROR({cl}25/{cl}17,"-")', '0.0"x"')
    w_blk(ws4, 30, c, f'=IFERROR({cl}11/{cl}17,"-")', '0.0%')

ws4.column_dimensions['A'].width = 34
for j in range(len(period_rows)):
    ws4.column_dimensions[get_column_letter(j+2)].width = 13
ws4.column_dimensions[get_column_letter(len(period_rows)+2)].width = 8.83
set_heights(ws4, 17)


# ═══ SHEET 5: Top 10 Shareholders (note sheet) ══════════════════════════════
print("  [5/7] Top 10 Shareholders (limited data note)...")
ws5 = sheets['Top 10 Shareholders']

_w(ws5, 1, 1, f'{COMPANY} ({TS_CODE}) - Top 10 Shareholders', F_BOLD, SOLID)
notes = [
    (2, 'HK-listed stocks: granular quarterly top-10 shareholder data is not freely available via akshare/tushare.'),
    (3, 'For detailed ownership information, refer to:'),
    (4, '  1) HKEXnews — Disclosure of Interest filings (>5% shareholders)'),
    (5, '     https://www.hkexnews.hk/sdw/search/searchsdw.aspx'),
    (6, '  2) Annual/Interim reports — Top shareholder section (Angelalign 2024 AR)'),
    (7, '  3) CCASS Shareholder Analysis on HKEX'),
    (8, ''),
    (9, 'Known key shareholders of 3600.HK 现代牙科 (per latest disclosures):'),
    (10, '  • Founder / management holding vehicles'),
    (11, '  • International institutional investors'),
    (12, '  • Public float'),
]
for r, text in notes:
    _w(ws5, r, 1, text, _font('FF000000'), SOLID)

ws5.column_dimensions['A'].width = 100


# ═══ SHEET 6: Shareholder Analytics (note sheet) ═════════════════════════════
print("  [6/7] Shareholder Analytics (limited data note)...")
ws6 = sheets['Shareholder Analytics']

_w(ws6, 1, 1, f'{COMPANY} ({TS_CODE}) - Shareholder Analytics', F_BOLD, SOLID)
notes6 = [
    (2, 'HK-listed stocks: shareholder count time series data is not available via the free A-share APIs used in this model.'),
    (3, '(akshare.stock_zh_a_gdhs_detail_em covers mainland listings only.)'),
    (4, ''),
    (5, 'Market data (derived):'),
    (6, f'  Year-end close price (HKD):'),
]
for r, text in notes6:
    _w(ws6, r, 1, text, _font('FF000000'), SOLID)

# Add year-end price table
_w(ws6, 7, 1, 'Year', F_BOLD, SOLID)
_w(ws6, 7, 2, 'Close (HKD)', F_BOLD, SOLID, C_ALIGN)
_w(ws6, 7, 3, 'Shares (亿)', F_BOLD, SOLID, C_ALIGN)
_w(ws6, 7, 4, 'Mkt Cap (HKD 亿)', F_BOLD, SOLID, C_ALIGN)
for i, yr in enumerate(YEARS):
    r = 8 + i
    _w(ws6, r, 1, yr, _font('FF000000'), SOLID)
    d = data[yr]
    if d.get('price') is not None:   w_blue(ws6, r, 2, d['price'], '0.00')
    if d.get('shares') is not None:  w_blue(ws6, r, 3, d['shares'], '#,##0.0')
    if d.get('mktcap') is not None:  w_blue(ws6, r, 4, d['mktcap'], '#,##0.0')

ws6.column_dimensions['A'].width = 14
for c in range(2, 5):
    ws6.column_dimensions[get_column_letter(c)].width = 18
set_heights(ws6, 17)


# ═══ SHEET 7: Data Correlation ════════════════════════════════════════════
print("  [7/7] Data Correlation...")
ws7 = sheets['Data Correlation']

doc = [
    (1,  f'{COMPANY} ({TS_CODE}) - Data Correlation & Methodology', True),
    (2,  '', False),
    (3,  '1. Company Overview', True),
    (4,  '现代牙科 (Modern Dental Group) — global manufacturer of custom dental prosthetic devices.', False),
    (5,  'IPO: December 2015 on HKEX main board. Ticker: 3600.HK.', False),
    (6,  'Reporting currency: HKD. Revenue denominated primarily in EUR/USD/HKD (EU + N. America).', False),
    (7,  'Largest customer base in Europe & North America; production hubs in China.', False),
    (8,  '', False),
    (9,  '2. Data Sources', True),
    (10, 'Primary: akshare', False),
    (11, '  stock_financial_hk_report_em(stock="06699", symbol=..., indicator="年度") — IS/BS/CF annual', False),
    (12, '  stock_financial_hk_report_em(... indicator="报告期") — interim/semi-annual', False),
    (13, '  stock_financial_hk_analysis_indicator_em — ROE/ROIC/EPS/margins', False),
    (14, '  stock_hk_hist(symbol="06699") — daily prices from IPO (2021-06-16) to present', False),
    (15, '', False),
    (16, '3. Key Computations & Assumptions', True),
    (17, 'Units: all financials in HKD 亿元 (raw HKD ÷ 1e8).', False),
    (18, 'Shares Outstanding: back-solved from HOLDER_PROFIT / BASIC_EPS (亿股).', False),
    (19, 'Market Cap: year-end close price × shares. 2025 uses latest available price.', False),
    (20, 'Cash & Equivalents: 现金及等价物 + 短期存款 (short-term bank deposits — treated as cash).', False),
    (21, 'Total Debt: 短期贷款 + 长期贷款 (finance leases excluded from "Total Debt" but included in Invested Capital).', False),
    (22, 'D&A: HK cash-flow layout does NOT break out D&A; row left blank — EBITDA formula falls back to EBIT.', False),
    (23, 'Dividends Paid: 已付股息(融资) from financing section.', False),
    (24, '', False),
    (25, '4. Valuation Methodology', True),
    (26, 'Leg 1 (Dividend Yield): Mkt_Cap_1 = Dividends_Paid / Sovereign_Bond_Yield', False),
    (27, 'Leg 2 (ROIC Multiple): Multiple = ROIC×100+5; EV = Adj_FCF × Multiple', False),
    (28, 'Adj FCF = adj. OCF - Dividends. Mkt_Cap_2 = EV - Debt + Cash - MI', False),
    (29, 'Total Valuation = Mkt_Cap_1 + Mkt_Cap_2', False),
    (30, 'Bond yield: CN 10yr used (underlying business is in China; HKD reporting is currency-translation only).', False),
    (31, '', False),
    (32, '5. Data Caveats', True),
    (33, '• 9 years of history (2017-2025). Detailed 2025 statements may still be pending release.', False),
    (34, '• Business has meaningful FX exposure (EUR, USD, HKD, CNY) — reported in HKD.', False),
    (35, '• Quarterly TTM analysis NOT applicable — HK requires only semi-annual (H1) + annual disclosure.', False),
    (36, '• Top-10 shareholder quarterly grid NOT available via free HK APIs; see HKEX Disclosure of Interest for >5% holders.', False),
    (37, '• Shareholder-count time series not available for HK listings.', False),
    (38, '', False),
    (39, '6. Bond Yield Reference (CN 10yr year-end)', True),
    (40, '2017=3.90% 2018=3.23% 2019=3.14% 2020=3.15% 2021=2.78% 2022=2.84% 2023=2.56% 2024=1.68% 2025=1.83%', False),
    (41, '', False),
    (42, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', False),
]
for r, text, bold in doc:
    _w(ws7, r, 1, text, Font(name='Apple Braille', size=11, bold=bold), SOLID)
ws7.column_dimensions['A'].width = 110


# ═══════════════════════════════════════════════════════════════════════════
print(f"\nSaving → {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)
print("="*60)
print(f"✓ Model saved: {OUTPUT_FILE}")
print("="*60)
print("\nSummary:")
print(f"  Sheet 1: Financial Data — {len(YEARS)} years ({YEARS[0]}-{YEARS[-1]})")
print(f"  Sheet 2: ROIC Analysis")
print(f"  Sheet 3: Valuation Model")
print(f"  Sheet 4: Semi-Annual Analysis — {len(period_rows)} periods")
print(f"  Sheet 5: Top 10 Shareholders — reference notes (limited HK data)")
print(f"  Sheet 6: Shareholder Analytics — reference notes + price table")
print(f"  Sheet 7: Data Correlation — methodology & caveats")
