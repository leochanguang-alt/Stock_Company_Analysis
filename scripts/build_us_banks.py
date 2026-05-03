#!/usr/bin/env python3
"""
Build 6-sheet bank valuation models for 8 US banks per bank-financial-model skill.
Sheets: Financial Data | DRIP Analysis | Key Ratios | Valuation Model | Top Holders | Data Correlation.
Currency: USD millions. Shares: millions. Years: 2015-2025.
Output: ~/Desktop/BUI_Investment_Management/<TICKER>/<TICKER>.Valuation.v<YYMMDD>.xlsx
"""
import os, time
from datetime import datetime, date
import requests
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

YEARS = list(range(2015, 2026))
TODAY = datetime.now().strftime('%y%m%d')
OUT_BASE = os.path.expanduser('~/Desktop/BUI_Investment_Management')
SEC_HDR = {'User-Agent':'BUI Research bui@example.com'}

# ============================================================
# Banks
# ============================================================
CONFIG = {
    "JPM":  dict(cik="0000019617", name="JPMorgan Chase & Co.",   type="Money-Center / G-SIB",  target_pe=11.0),
    "BAC":  dict(cik="0000070858", name="Bank of America Corp.",  type="Money-Center / G-SIB",  target_pe=11.0),
    "WFC":  dict(cik="0000072971", name="Wells Fargo & Co.",      type="Money-Center / G-SIB",  target_pe=10.0),
    "C":    dict(cik="0000831001", name="Citigroup Inc.",         type="Money-Center / G-SIB",  target_pe=9.0),
    "USB":  dict(cik="0000036104", name="U.S. Bancorp",           type="Super-Regional",         target_pe=10.0),
    "STT":  dict(cik="0000093751", name="State Street Corp.",     type="Custodial / Trust Bank", target_pe=13.0),
    "COF":  dict(cik="0000927628", name="Capital One Financial",  type="Card-focused / Bank",    target_pe=9.0),
    "EWBC": dict(cik="0001069157", name="East West Bancorp",      type="Regional (China-US niche)", target_pe=10.0),
}

# US 10y Treasury (year-end)
US_10Y = {2015:0.0227, 2016:0.0244, 2017:0.0241, 2018:0.0269, 2019:0.0192,
          2020:0.0091, 2021:0.0151, 2022:0.0388, 2023:0.0388, 2024:0.0458, 2025:0.0428}

# US M2 (USD T, year-end)
US_M2 = {2015:12.34, 2016:13.21, 2017:13.85, 2018:14.38, 2019:15.32,
         2020:19.10, 2021:21.61, 2022:21.39, 2023:20.89, 2024:21.45, 2025:22.20}

# DRIP defaults
DRIP_INVEST = 100_000   # USD
DRIP_FX     = 1.0
DRIP_WHT    = 0.15      # 15% US-China treaty rate for foreign investors (default)

# ============================================================
# Format
# ============================================================
FONT = "Apple Braille"
SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY    = "FFE7E6E6"
LGREY   = "FFF2F2F2"
THIN    = Side(style="thin", color="FF000000")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.0;(#,##0.0);"–"'
INT_FMT = '#,##0;(#,##0);"–"'
PCT_FMT = '0.00%;(0.00%);"–"'
PCT1_FMT= '0.0%;(0.0%);"–"'
PX_FMT  = '#,##0.00'
X_FMT   = '0.00"x"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False, lgrey=False, font_size=SIZE):
    c.font = Font(name=FONT, size=font_size, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if num_fmt: c.number_format = num_fmt
    if grey:  c.fill = PatternFill("solid", fgColor=GREY)
    elif lgrey: c.fill = PatternFill("solid", fgColor=LGREY)

def col_l(i): return get_column_letter(2+i)

# ============================================================
# SEC EDGAR fetch
# ============================================================
def fetch_bank(ticker, cfg):
    print(f"\n{'='*60}\n{ticker} {cfg['name']}\n{'='*60}")
    print("  [1] SEC EDGAR XBRL...")
    r = requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cfg['cik']}.json",
                     headers=SEC_HDR, timeout=30)
    j = r.json()
    facts = j.get('facts', {}).get('us-gaap', {})

    def fy_annual(tags, units='USD'):
        """Try multiple tags in order. Return {year: value}."""
        if isinstance(tags, str): tags = [tags]
        for tag in tags:
            if tag not in facts: continue
            arr = facts[tag].get('units', {}).get(units, [])
            out = {}
            for u in arr:
                if u.get('form') not in ('10-K','10-K/A'): continue
                end = u.get('end','')
                if not end.endswith('-12-31'): continue
                start = u.get('start')
                if start and units == 'USD':
                    d1, d2 = date.fromisoformat(start), date.fromisoformat(end)
                    if (d2-d1).days < 350: continue
                yr = int(end[:4])
                if yr not in out or u.get('filed','') > out[yr][1]:
                    out[yr] = (u['val'], u.get('filed',''))
            if out: return {k: v[0] for k, v in out.items()}
        return {}

    def fy_instant(tags, units='USD'):
        if isinstance(tags, str): tags = [tags]
        for tag in tags:
            if tag not in facts: continue
            arr = facts[tag].get('units', {}).get(units, [])
            out = {}
            for u in arr:
                if u.get('form') not in ('10-K','10-K/A'): continue
                end = u.get('end','')
                if not end.endswith('-12-31'): continue
                if u.get('start'): continue
                yr = int(end[:4])
                if yr not in out or u.get('filed','') > out[yr][1]:
                    out[yr] = (u['val'], u.get('filed',''))
            if out: return {k: v[0] for k, v in out.items()}
        return {}

    # Income statement (bank-specific XBRL tags)
    nii = fy_annual(['InterestIncomeExpenseNet', 'InterestIncomeExpenseAfterProvisionForLoanLoss'])
    int_inc = fy_annual(['InterestAndDividendIncomeOperating', 'InterestIncomeOperating'])
    int_exp = fy_annual(['InterestExpense'])
    if not nii:
        # Compute as int_inc - int_exp
        nii = {y: int_inc.get(y,0) - int_exp.get(y,0) for y in YEARS
               if y in int_inc and y in int_exp}
    nonint_inc = fy_annual(['NoninterestIncome'])
    nonint_exp = fy_annual(['NoninterestExpense'])
    provision = fy_annual(['ProvisionForLoanLeaseAndOtherLosses', 'ProvisionForCreditLosses', 'ProvisionForLoanAndLeaseLosses'])
    pretax = fy_annual(['IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest',
                        'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments'])
    tax = fy_annual(['IncomeTaxExpenseBenefit'])
    ni  = fy_annual(['NetIncomeLoss'])
    ni_p = fy_annual(['NetIncomeLossAvailableToCommonStockholdersBasic', 'NetIncomeLoss'])
    fees = fy_annual(['FeesAndCommissionsAdministrativeAndOther', 'NoninterestIncome'])
    revenue_total = fy_annual(['Revenues'])

    # Balance sheet
    cash    = fy_instant(['CashAndCashEquivalentsAtCarryingValue'])
    loans   = fy_instant(['LoansAndLeasesReceivableNetReportedAmount', 'LoansAndLeasesReceivableNetOfDeferredIncome',
                          'NotesReceivableNet', 'FinancingReceivableNetOfAllowance'])
    deposits= fy_instant(['Deposits'])
    assets  = fy_instant(['Assets'])
    liab    = fy_instant(['Liabilities'])
    equity  = fy_instant(['StockholdersEquity', 'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest'])
    eq_par  = fy_instant(['StockholdersEquity'])
    minor   = fy_instant(['MinorityInterest'])

    # Cash flow
    ocf     = fy_annual(['NetCashProvidedByUsedInOperatingActivities'])
    divs    = fy_annual(['PaymentsOfDividendsCommonStock', 'PaymentsOfDividends'])

    # Prices via yfinance
    print("  [2] Prices and shares (yfinance)...")
    yt = yf.Ticker(ticker)
    hist = yt.history(start="2014-01-01", end="2026-04-30", auto_adjust=False)
    ye_px = {d.year: float(v) for d, v in hist['Close'].resample('YE').last().items() if d.year in YEARS}
    yinfo = {}
    try:
        yinfo = yt.info
    except: pass

    # Top holders
    holders = None
    try:
        holders = yt.institutional_holders
    except: pass

    return dict(
        nii=nii, int_inc=int_inc, int_exp=int_exp,
        nonint_inc=nonint_inc, nonint_exp=nonint_exp,
        provision=provision, pretax=pretax, tax=tax, ni=ni, ni_p=ni_p,
        revenue_total=revenue_total,
        cash=cash, loans=loans, deposits=deposits, assets=assets,
        liab=liab, equity=equity, eq_par=eq_par, minor=minor,
        ocf=ocf, divs=divs, ye_px=ye_px, yinfo=yinfo, holders=holders,
    )

# ============================================================
# Convert raw → millions
# ============================================================
def to_M(v):
    if v is None or pd.isna(v): return None
    return float(v) / 1e6

# Historical share counts (M shares) — embedded since SEC EDGAR shares data is patchy
SHARES_M = {
    "JPM":  {2015:3663, 2016:3559, 2017:3425, 2018:3275, 2019:3084,
             2020:3043, 2021:2944, 2022:2925, 2023:2864, 2024:2805, 2025:2820},
    "BAC":  {2015:10401, 2016:9979, 2017:10288, 2018:9669, 2019:8836,
             2020:8651, 2021:8081, 2022:7997, 2023:7870, 2024:7660, 2025:7600},
    "WFC":  {2015:5092, 2016:5018, 2017:4892, 2018:4560, 2019:4147,
             2020:4123, 2021:3792, 2022:3743, 2023:3617, 2024:3322, 2025:3260},
    "C":    {2015:2953, 2016:2772, 2017:2570, 2018:2369, 2019:2179,
             2020:2082, 2021:1956, 2022:1936, 2023:1907, 2024:1881, 2025:1860},
    "USB":  {2015:1745, 2016:1716, 2017:1677, 2018:1611, 2019:1551,
             2020:1517, 2021:1487, 2022:1499, 2023:1525, 2024:1559, 2025:1565},
    "STT":  {2015:402,  2016:392,  2017:368,  2018:367,  2019:362,
             2020:354,  2021:348,  2022:354,  2023:308,  2024:294,  2025:283},
    "COF":  {2015:534,  2016:485,  2017:484,  2018:480,  2019:466,
             2020:457,  2021:441,  2022:381,  2023:381,  2024:382,  2025:455},
    "EWBC": {2015:144,  2016:144,  2017:144,  2018:144,  2019:142,
             2020:141,  2021:142,  2022:141,  2023:141,  2024:139,  2025:140},
}

# ============================================================
# Workbook builder
# ============================================================
def build(ticker, cfg, data):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Financial Data"
    s1_financial(ws1, ticker, cfg, data)
    ws2 = wb.create_sheet("DRIP Analysis")
    s2_drip(ws2, ticker, cfg)
    ws3 = wb.create_sheet("Key Ratios")
    s3_ratios(ws3, ticker, cfg)
    ws4 = wb.create_sheet("Valuation Model")
    s4_valuation(ws4, ticker, cfg)
    ws5 = wb.create_sheet("Top Holders")
    s5_holders(ws5, ticker, cfg, data['holders'])
    ws6 = wb.create_sheet("Data Correlation")
    s6_correlation(ws6, ticker, cfg)
    return wb

# ===== Sheet 1 =====
def s1_financial(ws, ticker, cfg, d):
    ws.column_dimensions['A'].width = 38
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 13
    ws.column_dimensions[col_l(len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} ({ticker}) - Financial Data")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(2,1, f"Units: USD millions; Shares: millions; Market Cap = price × shares (USD M)")
    fmt(ws.cell(2,1), italic=True, align="left", color=BLUE)

    ws.cell(3,1, "Metric"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    rows = [
        (4,  "==Income Statement (USD M)==", "header", None),
        (5,  "Net Interest Income", "data", d['nii']),
        (6,  "Noninterest Income (Fees etc.)", "data", d['nonint_inc']),
        (7,  "Total Revenue (NII + Non-II)", "formula", None),
        (8,  "Operating Income (= Total Revenue)", "formula_alt", None),
        (9,  "Noninterest Expense", "data", d['nonint_exp']),
        (10, "PPOP (Pre-Provision Op Profit)", "formula", None),
        (11, "Provision for Credit Losses", "data", d['provision']),
        (12, "Pretax Income", "data", d['pretax']),
        (13, "Income Tax Expense", "data", d['tax']),
        (14, "Net Income", "data", d['ni']),
        (15, "NI to Common Stockholders", "data", d['ni_p']),
        (16, "Cost-to-Income Ratio", "formula", None),
        (18, "==Balance Sheet (USD M)==", "header", None),
        (19, "Total Assets", "data", d['assets']),
        (20, "Loans & Leases (Net)", "data", d['loans']),
        (21, "Total Deposits", "data", d['deposits']),
        (22, "Total Liabilities", "data", d['liab']),
        (23, "Stockholders' Equity (Common)", "data", d['eq_par']),
        (24, "Minority Interest", "data", d['minor']),
        (25, "Total Equity (incl. MI)", "formula", None),
        (26, "Loan/Deposit Ratio", "formula", None),
        (27, "Equity/Asset Ratio (Leverage)", "formula", None),
        (29, "==Cash Flow (USD M)==", "header", None),
        (30, "Net Cash from Operations", "data", d['ocf']),
        (31, "Dividends Paid (Common)", "data_neg", d['divs']),  # convention: negative
        (33, "==Market Data==", "header", None),
        (34, "Year-End Close (USD/share)", "px", None),
        (35, "Total Shares Outstanding (M)", "shares", None),
        (36, "Market Cap (USD M)", "mc", None),
        (37, "Book Value Per Share (USD)", "formula", None),
        (38, "EPS (USD, basic)", "formula", None),
        (39, "DPS (USD, implied from CF)", "formula", None),
    ]
    for r, label, kind, src in rows:
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, grey=True, align="left")
            for i in range(len(YEARS)): fmt(ws.cell(r,2+i), grey=True)
            continue
        fmt(ws.cell(r,1), align="left")
        for i, y in enumerate(YEARS):
            cl = col_l(i)
            cell = ws.cell(r, 2+i)
            if kind == "data":
                v = to_M(src.get(y)) if src else None
                if v is not None:
                    cell.value = v
                    fmt(cell, color=BLUE, num_fmt=NUM_FMT)
                else:
                    fmt(cell, color=BLUE, num_fmt=NUM_FMT)
            elif kind == "data_neg":
                v = to_M(src.get(y)) if src else None
                if v is not None:
                    cell.value = -abs(v)   # convention: negative cash outflow
                    fmt(cell, color=BLUE, num_fmt=NUM_FMT)
            elif kind == "formula":
                if r == 7:   # Total Rev = NII + Non-II
                    cell.value = f"=IFERROR({cl}5+{cl}6,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=NUM_FMT)
                elif r == 10:  # PPOP = TotRev - NoninterestExp
                    cell.value = f"=IFERROR({cl}7-{cl}9,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=NUM_FMT)
                elif r == 16:  # C/I
                    cell.value = f"=IFERROR({cl}9/{cl}7,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=PCT1_FMT)
                elif r == 25:
                    cell.value = f"=IFERROR({cl}23+{cl}24,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=NUM_FMT)
                elif r == 26:
                    cell.value = f"=IFERROR({cl}20/{cl}21,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=PCT1_FMT)
                elif r == 27:
                    cell.value = f"=IFERROR({cl}23/{cl}19,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=PCT_FMT)
                elif r == 37:
                    cell.value = f"=IFERROR({cl}23/{cl}35,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=PX_FMT)
                elif r == 38:
                    cell.value = f"=IFERROR({cl}15/{cl}35,\"-\")"
                    fmt(cell, color=BLACK, num_fmt=PX_FMT)
                elif r == 39:
                    cell.value = f"=IFERROR(ABS({cl}31)/{cl}35,\"-\")"
                    fmt(cell, color=BLACK, num_fmt='0.0000')
            elif kind == "formula_alt":
                if r == 8:   # Operating Income = Total Revenue
                    cell.value = f"={cl}7"
                    fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "px":
                v = d['ye_px'].get(y)
                if v is not None:
                    cell.value = v; fmt(cell, color=BLUE, num_fmt=PX_FMT)
                else:
                    fmt(cell, color=BLUE, num_fmt=PX_FMT)
            elif kind == "shares":
                v = SHARES_M[ticker].get(y)
                if v is not None:
                    cell.value = v; fmt(cell, color=BLUE, num_fmt=INT_FMT)
            elif kind == "mc":
                cell.value = f"={cl}34*{cl}35"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT, bold=True)

    for rr in range(1, 40):
        ws.row_dimensions[rr].height = 20.0

# ===== Sheet 2 DRIP =====
def s2_drip(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 38.33
    ws.column_dimensions['B'].width = 15.83
    for i in range(2, len(YEARS)+1):
        ws.column_dimensions[get_column_letter(2+i-1)].width = 13
    ws.column_dimensions[col_l(len(YEARS))].width = 8.83

    fd = "'Financial Data'"

    c = ws.cell(1,1, f"DRIP Return Analysis — {ticker} ({YEARS[0]}–{YEARS[-1]}, Foreign treaty 15% WHT)")
    fmt(c, bold=True, lgrey=True, align="left", font_size=12)
    ws.cell(2,1, f"Source data linked from 'Financial Data' sheet (rows 3, 31, 34, 35)")
    fmt(ws.cell(2,1), italic=True, align="left", color=BLUE)

    ws.cell(4,1, "Assumptions"); fmt(ws.cell(4,1), bold=True, align="left")
    ws.cell(5,1, "Initial investment (USD)"); fmt(ws.cell(5,1), align="left")
    fmt(ws.cell(5,2, DRIP_INVEST), color=BLUE, bold=True, num_fmt=INT_FMT)
    ws.cell(6,1, f"Entry price (USD/sh, {YEARS[0]} YE)"); fmt(ws.cell(6,1), align="left")
    c = ws.cell(6,2); c.value = f"={fd}!B34"; fmt(c, color=GREEN, bold=True, num_fmt=PX_FMT)
    ws.cell(7,1, "FX (USD denominator, 1.00)"); fmt(ws.cell(7,1), align="left")
    fmt(ws.cell(7,2, DRIP_FX), color=BLUE, bold=True, num_fmt='0.00')
    ws.cell(8,1, "Dividend WHT (foreign treaty)"); fmt(ws.cell(8,1), align="left")
    fmt(ws.cell(8,2, DRIP_WHT), color=BLUE, bold=True, num_fmt=PCT_FMT)
    ws.cell(9,1, "Initial shares purchased"); fmt(ws.cell(9,1), align="left")
    c = ws.cell(9,2); c.value = "=B5/B6"; fmt(c, num_fmt=INT_FMT)

    ws.cell(11,1, "Year-by-year DRIP"); fmt(ws.cell(11,1), bold=True, align="left")
    ws.cell(12,1, "Year"); fmt(ws.cell(12,1), bold=True, align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        c = ws.cell(12, 2+i); c.value = f"={fd}!{cl}3"; fmt(c, bold=True, align="right")

    ws.cell(13,1, "Shares held BoY"); fmt(ws.cell(13,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        if i == 0:
            ws.cell(13,2).value = "=B9"
        else:
            ws.cell(13, 2+i).value = f"={col_l(i-1)}18"
        fmt(ws.cell(13, 2+i), num_fmt=INT_FMT)

    ws.cell(14,1, "DPS (USD, from CF)"); fmt(ws.cell(14,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        ws.cell(14, 2+i).value = f"=IFERROR(ABS({fd}!{cl}31)/{fd}!{cl}35,0)"
        fmt(ws.cell(14, 2+i), num_fmt='0.0000')

    ws.cell(15,1, "Dividend (USD, after tax)"); fmt(ws.cell(15,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        ws.cell(15, 2+i).value = f"={cl}13*{cl}14*$B$7*(1-$B$8)"
        fmt(ws.cell(15, 2+i), num_fmt=INT_FMT)

    ws.cell(16,1, "Reinvest price (USD)"); fmt(ws.cell(16,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        ws.cell(16, 2+i).value = f"={fd}!{cl}34"
        fmt(ws.cell(16, 2+i), color=GREEN, num_fmt=PX_FMT)

    ws.cell(17,1, "Shares bought"); fmt(ws.cell(17,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        ws.cell(17, 2+i).value = f"=IFERROR({cl}15/{cl}16,0)"
        fmt(ws.cell(17, 2+i), num_fmt=INT_FMT)

    ws.cell(18,1, "Shares held EoY"); fmt(ws.cell(18,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        ws.cell(18, 2+i).value = f"={cl}13+{cl}17"
        fmt(ws.cell(18, 2+i), num_fmt=INT_FMT)

    ws.cell(19,1, "Portfolio value EoY (USD)"); fmt(ws.cell(19,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        ws.cell(19, 2+i).value = f"={cl}18*{cl}16"
        bold = (i == len(YEARS)-1)
        fmt(ws.cell(19, 2+i), bold=bold, num_fmt=INT_FMT)

    # M2 reference (US)
    ws.cell(21,1, "Reference: US M2 (USD T)"); fmt(ws.cell(21,1), bold=True, align="left")
    ws.cell(22,1, "M2 year-end balance"); fmt(ws.cell(22,1), align="left")
    for i, y in enumerate(YEARS):
        v = US_M2.get(y)
        if v: fmt(ws.cell(22, 2+i, v), color=BLUE, bold=True, num_fmt='#,##0.00')
    n = len(YEARS) - 1
    cl_l = col_l(len(YEARS)-1)
    ws.cell(23,1, f"M2 multiple vs {YEARS[0]}"); fmt(ws.cell(23,1), align="left")
    c = ws.cell(23, 1+len(YEARS)+1)
    c.value = f"={cl_l}22/B22"; fmt(c, num_fmt='0.00"x"')
    ws.cell(24,1, f"M2 CAGR ({n} yrs)"); fmt(ws.cell(24,1), align="left")
    c = ws.cell(24, 1+len(YEARS)+1)
    c.value = f"=({cl_l}22/B22)^(1/{n})-1"; fmt(c, num_fmt=PCT_FMT)

    ws.cell(26,1, "Summary"); fmt(ws.cell(26,1), bold=True, align="left")
    ws.cell(27,1, f"Final value {YEARS[-1]} YE (USD)"); fmt(ws.cell(27,1), align="left")
    c = ws.cell(27,2); c.value = f"={cl_l}19"; fmt(c, bold=True, num_fmt=INT_FMT)
    ws.cell(28,1, "Total return"); fmt(ws.cell(28,1), align="left")
    c = ws.cell(28,2); c.value = "=B27/B5-1"; fmt(c, num_fmt=PCT_FMT)
    ws.cell(29,1, "Multiple on invested (x)"); fmt(ws.cell(29,1), align="left")
    c = ws.cell(29,2); c.value = "=B27/B5"; fmt(c, bold=True, num_fmt=X_FMT)
    ws.cell(30,1, f"Annualized IRR ({n+1} yrs)"); fmt(ws.cell(30,1), align="left")
    c = ws.cell(30,2); c.value = f"=(B27/B5)^(1/{n+1})-1"; fmt(c, bold=True, num_fmt=PCT_FMT)
    cell_m2cagr = f"{get_column_letter(1+len(YEARS)+1)}24"
    ws.cell(31,1, "vs M2 growth (DRIP IRR − M2 CAGR)"); fmt(ws.cell(31,1), align="left")
    c = ws.cell(31,2); c.value = f"=B30-{cell_m2cagr}"; fmt(c, num_fmt=PCT_FMT)
    ws.cell(32,1, "Price-only return (no dividends)"); fmt(ws.cell(32,1), align="left")
    c = ws.cell(32,2); c.value = f"={fd}!{cl_l}34/{fd}!B34-1"; fmt(c, num_fmt=PCT1_FMT)

    for rr in range(1, 35):
        ws.row_dimensions[rr].height = 23.0

# ===== Sheet 3 Key Ratios =====
def s3_ratios(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 34
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 13

    ws.cell(1,1, f"{cfg['name']} - Key Banking Ratios")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1, "Ratio"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    fd = "'Financial Data'"
    rows = [
        (4,  "Profitability & Efficiency", "header", None),
        (5,  "NI to Common", "link",  f"{fd}!{{cl}}15"),
        (6,  "Stockholders' Equity", "link",  f"{fd}!{{cl}}23"),
        (7,  "Avg Equity", "avg",   f"{fd}!23"),
        (8,  "ROE (on Common Equity)", "div_avg", (f"{fd}!{{cl}}15", f"{fd}!23")),
        (9,  "Total Assets", "link",  f"{fd}!{{cl}}19"),
        (10, "Avg Assets", "avg",   f"{fd}!19"),
        (11, "ROA", "div_avg", (f"{fd}!{{cl}}14", f"{fd}!19")),
        (12, "Net Interest Income", "link",  f"{fd}!{{cl}}5"),
        (13, "NIM (NII / Avg Earning Assets ≈ Avg Assets)", "div_avg", (f"{fd}!{{cl}}5", f"{fd}!19")),
        (15, "Provisioning & Growth", "header", None),
        (16, "Provision for Credit Losses", "link",  f"{fd}!{{cl}}11"),
        (17, "Provision/Loans (Credit Cost)", "ratio", (f"{fd}!{{cl}}11", f"{fd}!{{cl}}20")),
        (18, "Loan Growth YoY", "yoy",   f"{fd}!{{c}}20"),
        (19, "Deposit Growth YoY", "yoy",   f"{fd}!{{c}}21"),
        (20, "NI Growth YoY", "yoy",   f"{fd}!{{c}}15"),
        (22, "Capital & Payout", "header", None),
        (23, "Dividend Payout Ratio", "payout", f"{fd}"),
        (24, "Retention Ratio", "ret", None),
    ]
    for r, label, kind, ref in rows:
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, align="left")
            continue
        fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl, pcl = col_l(i), col_l(i-1) if i>0 else None
            cell = ws.cell(r, 2+i)
            if kind == "link":
                cell.value = "=" + ref.replace("{cl}", cl)
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "avg":
                if i == 0: continue
                base = ref.split("!")[1]
                cell.value = f"=({col_l(i-1)}{base}+{cl}{base})/2"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "div_avg":
                if i == 0: continue
                num = ref[0].replace("{cl}", cl)
                den_base = ref[1].split("!")[1]
                cell.value = f"=IFERROR({num}/(({col_l(i-1)}{den_base}+{cl}{den_base})/2),\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT_FMT)
            elif kind == "ratio":
                num = ref[0].replace("{cl}", cl)
                den = ref[1].replace("{cl}", cl)
                cell.value = f"=IFERROR(ABS({num})/{den},\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT_FMT)
            elif kind == "yoy":
                if i == 0: continue
                base = ref.split("!")[1].replace("{c}","")
                cell.value = f"=IFERROR(({fd}!{cl}{base}-{fd}!{pcl}{base})/{fd}!{pcl}{base},\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT1_FMT)
            elif kind == "payout":
                cell.value = f"=IFERROR(ABS({fd}!{cl}31)/{fd}!{cl}15,\"-\")"
                fmt(cell, color=GREEN, num_fmt=PCT1_FMT)
            elif kind == "ret":
                cell.value = f"=1-{cl}23"
                fmt(cell, color=BLACK, num_fmt=PCT1_FMT)

    for rr in range(1, 26):
        ws.row_dimensions[rr].height = 20.0

# ===== Sheet 4 Valuation =====
def s4_valuation(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 38
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 13

    ws.cell(1,1, f"{cfg['name']} - Valuation Model")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1, "Item"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    fd = "'Financial Data'"
    kr = "'Key Ratios'"

    rows = [
        (5,  "Key Assumptions", "header"),
        (6,  "US 10y Treasury Yield", "bond"),
        (7,  "FX (USD = 1.00)", "fx"),
        (9,  "Inputs (linked)", "header"),
        (10, "NI to Common (USD M)", "link", f"{fd}!{{cl}}15"),
        (11, "Dividends Paid (USD M, abs)", "link_abs", f"{fd}!{{cl}}31"),
        (12, "Stockholders' Equity (USD M)", "link", f"{fd}!{{cl}}23"),
        (13, "Market Cap (USD M)", "link", f"{fd}!{{cl}}36"),
        (14, "Market Cap (USD M)", "self", "B13"),
        (16, "Valuation Methods", "header"),
        (17, "Dividend Yield (Div/MktCap)", "ratio", (f"{{cl}}11", f"{{cl}}14")),
        (18, "Mkt Cap 1 = Div / Bond Yield", "v1"),
        (19, "P/B (Implied)", "ratio", (f"{{cl}}14", f"{{cl}}12")),
        (20, "ROE", "link", f"{kr}!{{cl}}8"),
        (21, "Payout Ratio", "link", f"{kr}!{{cl}}23"),
        (22, "Growth g = ROE × (1-Payout)", "g"),
        (23, "Cost of Equity (Bond + 5%)", "coe"),
        (24, "Justified P/B = (ROE-g)/(r-g)", "jpb"),
        (25, "Mkt Cap 2 = JustPB × Equity", "v2"),
        (27, "P/E (Implied)", "ratio", (f"{{cl}}14", f"{{cl}}10")),
        (28, "Target P/E", "tgt_pe"),
        (29, "Mkt Cap 3 = Target PE × NI", "v3"),
        (31, "Valuation Summary", "header"),
        (32, "Mkt Cap 1 (Div Yield)", "self", "{{cl}}18"),
        (33, "Mkt Cap 2 (Justified P/B)", "self", "{{cl}}25"),
        (34, "Mkt Cap 3 (Target P/E)", "self", "{{cl}}29"),
        (35, "Average Fair Value (USD M)", "avg3"),
        (36, "Actual Mkt Cap (USD M)", "self", "{{cl}}14"),
        (37, "Premium / (Discount)", "pd"),
    ]
    for spec in rows:
        r = spec[0]; label = spec[1]; kind = spec[2]
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, align="left")
            continue
        fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl = col_l(i)
            cell = ws.cell(r, 2+i)
            if kind == "bond":
                cell.value = US_10Y.get(YEARS[i], 0.04)
                fmt(cell, color=BLUE, num_fmt=PCT_FMT)
            elif kind == "fx":
                cell.value = 1.00
                fmt(cell, color=BLUE, num_fmt='0.0000')
            elif kind == "link":
                ref = spec[3].replace("{cl}", cl)
                cell.value = "=" + ref
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "link_abs":
                ref = spec[3].replace("{cl}", cl)
                cell.value = f"=ABS({ref})"
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "self":
                ref = spec[3].replace("{{cl}}", cl).replace("{cl}", cl)
                cell.value = f"={ref}"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "ratio":
                num = spec[3][0].replace("{{cl}}", cl).replace("{cl}", cl)
                den = spec[3][1].replace("{{cl}}", cl).replace("{cl}", cl)
                cell.value = f"=IFERROR({num}/{den},\"-\")"
                nf = PCT_FMT if r == 17 else X_FMT
                fmt(cell, color=BLACK, num_fmt=nf)
            elif kind == "v1":
                cell.value = f"=IFERROR({cl}11/{cl}6,\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "g":
                if i == 0: continue
                cell.value = f"=IFERROR({cl}20*(1-{cl}21),\"-\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "coe":
                cell.value = f"={cl}6+0.05"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "jpb":
                if i == 0: continue
                cell.value = f"=IFERROR(IF({cl}23-{cl}22=0,\"-\",({cl}20-{cl}22)/({cl}23-{cl}22)),\"-\")"
                fmt(cell, color=BLACK, num_fmt=X_FMT)
            elif kind == "v2":
                if i == 0: continue
                cell.value = f"=IFERROR({cl}24*{cl}12,\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "tgt_pe":
                cell.value = cfg['target_pe']
                fmt(cell, color=BLUE, num_fmt=X_FMT)
            elif kind == "v3":
                cell.value = f"={cl}28*{cl}10"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "avg3":
                cell.value = f"=IFERROR(AVERAGE({cl}32,{cl}33,{cl}34),\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "pd":
                cell.value = f"=IFERROR({cl}35/{cl}36-1,\"-\")"
                fmt(cell, bold=True, color=BLACK, num_fmt=PCT_FMT)

    for rr in range(1, 38):
        ws.row_dimensions[rr].height = 17.0

# ===== Sheet 5 Top Holders =====
def s5_holders(ws, ticker, cfg, holders):
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    ws.cell(1,1, f"{cfg['name']} - Top Institutional Holders (yfinance snapshot)")
    fmt(ws.cell(1,1), bold=True, align="left")

    headers = ["#","Holder","Shares","Date Reported","Value (USD)"]
    for i, h in enumerate(headers):
        c = ws.cell(3,1+i,h)
        fmt(c, bold=True, color="FFFFFFFF", align="center" if i==0 else "left")
        c.fill = PatternFill("solid", fgColor="FF305496")

    if holders is not None and not holders.empty:
        for i, (_, row) in enumerate(holders.head(15).iterrows()):
            r = 4 + i
            ws.cell(r,1, i+1); fmt(ws.cell(r,1), align="center")
            ws.cell(r,2, str(row.get('Holder',''))[:60]); fmt(ws.cell(r,2), align="left")
            try:
                sh = float(row.get('Shares', 0)) if pd.notna(row.get('Shares')) else None
                if sh is not None:
                    ws.cell(r,3, sh); fmt(ws.cell(r,3), color=BLUE, num_fmt=INT_FMT)
            except: pass
            ws.cell(r,4, str(row.get('Date Reported','')))
            fmt(ws.cell(r,4), align="center")
            try:
                v = float(row.get('Value', 0)) if pd.notna(row.get('Value')) else None
                if v is not None:
                    ws.cell(r,5, v); fmt(ws.cell(r,5), color=BLUE, num_fmt=INT_FMT)
            except: pass
    else:
        ws.cell(4,1, "Top holders data not retrieved — use 13F filings on SEC EDGAR")
        fmt(ws.cell(4,1), italic=True, align="left")

    for rr in range(1, 20):
        ws.row_dimensions[rr].height = 15.0

# ===== Sheet 6 Data Correlation =====
def s6_correlation(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 110

    ws.cell(1,1, f"{cfg['name']} ({ticker}) - Data Correlation & Methodology")
    fmt(ws.cell(1,1), bold=True, align="left")

    sections = [
        ("1. Bank Profile",
         f"{cfg['name']} — {cfg['type']}.\n"
         f"Years modeled: 2015–2025 (11 years). Reporting currency: USD millions.\n"
         f"NYSE-listed (or NASDAQ for some) US bank holding company.\n"
         f"CIK: {cfg['cik']}"),
        ("2. Data Sources",
         f"INCOME STATEMENT / BALANCE SHEET / CASH FLOW: SEC EDGAR XBRL companyfacts API\n"
         f"  https://data.sec.gov/api/xbrl/companyfacts/CIK{cfg['cik']}.json\n"
         "  10-K annual filings, fiscal year = calendar year.\n"
         "Bank-specific XBRL tags used:\n"
         "  NII: InterestIncomeExpenseNet OR (InterestAndDividendIncomeOperating − InterestExpense)\n"
         "  Noninterest Income: NoninterestIncome\n"
         "  Noninterest Expense: NoninterestExpense\n"
         "  Provision for Credit Losses: ProvisionForCreditLosses (post-CECL 2020) OR\n"
         "    ProvisionForLoanLeaseAndOtherLosses (pre-2020)\n"
         "  Loans: LoansAndLeasesReceivableNetReportedAmount\n"
         "  Deposits: Deposits\n"
         "  Equity: StockholdersEquity\n"
         "  Dividends: PaymentsOfDividendsCommonStock\n"
         "PRICE: yfinance — year-end close (auto_adjust=False, pre-split adjusted).\n"
         "TOP HOLDERS: yfinance institutional_holders snapshot (refreshes ~quarterly)."),
        ("3. Currency & Units",
         "All financials in USD millions (raw USD ÷ 1e6).\n"
         "Shares in millions. BVPS = Equity / Shares = USD/share.\n"
         "Market Cap = Year-End Price × Shares Outstanding (USD M)."),
        ("4. DRIP Analysis",
         f"Setup: invest USD 100,000 at {YEARS[0]} YE close, reinvest all dividends at each subsequent YE close.\n"
         f"Withholding tax = 15% (US-China tax treaty rate for foreign investors; adjust for your tax residency:\n"
         "  - US resident in taxable account: ~0% (qualified dividend tax handled separately)\n"
         "  - US resident in IRA/401k: 0%\n"
         "  - HK/Singapore investor: 30% (no treaty)\n"
         "  - Mainland China resident via QDII: 10–15% depending on fund structure)\n"
         "US M2 benchmark: 2015 = $12.34T → 2025 ≈ $22.20T → 10-yr CAGR ≈ 6.05%.\n"
         "Note: US M2 had unprecedented expansion 2020–2021 (COVID stimulus +$5T). Use IRR vs M2 as a\n"
         "directional benchmark, not a precise comparison.\n"
         "Excess return = DRIP IRR − M2 CAGR. Banks should comfortably beat M2 for shareholder value preservation."),
        ("5. Key Caveats",
         "• CECL transition (2020): Provision for Credit Losses reporting changed under CECL standard;\n"
         "  pre-2020 figures used incurred-loss model, post-2020 use lifetime expected-loss. Discontinuity ~Q1-2020.\n"
         "• 2020 COVID: provisions jumped sharply (e.g., JPM took ~$10B Q1-2020 reserve build).\n"
         "  Most banks released reserves in 2021, creating negative provision artifacts.\n"
         "• 2023 SVB crisis: regional banks (esp WAL, FHN, ZION) had liquidity events. Not modeled here.\n"
         "• Some banks split or reclassify revenue lines; total_revenue tag may be inconsistent year-to-year.\n"
         "  We compute Operating Income = NII + Noninterest Income (more reliable).\n"
         "• Loan growth: post-2020 unusual due to PPP loans (forgiven by 2022) — temporary spike.\n"
         "• Share buybacks: US banks aggressively repurchase. Share count declines steadily for most banks."),
        ("6. Valuation Methodology",
         "Three-method blend, equally weighted:\n"
         "  Method 1 — Dividend Yield: Mkt Cap = Div / Bond Yield\n"
         "  Method 2 — Justified P/B (Gordon): JustPB = (ROE − g) / (r − g), where\n"
         "    g = ROE × (1 − Payout); r = US 10y + 5% equity risk premium\n"
         "  Method 3 — Target P/E: Mkt Cap = Target P/E × NI to Common\n"
         f"  Target P/E for this bank: {cfg['target_pe']}x (calibrated to {cfg['type']} peer group)\n"
         "Premium/(Discount) = Avg Implied / Actual − 1.\n"
         "Caveat: Justified P/B assumes perpetual ROE > r (Gordon constant-growth). For banks with\n"
         "  ROE near or below 10% in some years, JustPB can produce nonsensical results — flag and exclude."),
        ("7. Format Conventions",
         "Font: Apple Braille 11pt (DRIP title row 12pt bold, light grey fill).\n"
         "Colors: BLUE = hardcoded input, BLACK = formula, GREEN = cross-sheet link.\n"
         "Row heights: Financial Data 20pt, DRIP 23pt, Key Ratios 20pt, Valuation 17pt, Top Holders 15pt.\n"
         "Column widths: A column 34–38, data cols B–L = 13."),
    ]
    r = 3
    for title, body in sections:
        c = ws.cell(r,1, title); fmt(c, bold=True, grey=True, align="left")
        r += 1
        for line in body.split("\n"):
            ws.cell(r,1, line); fmt(ws.cell(r,1), align="left")
            r += 1
        r += 1


def main():
    for ticker, cfg in CONFIG.items():
        try:
            data = fetch_bank(ticker, cfg)
            print(f"  Building workbook...")
            wb = build(ticker, cfg, data)
            outdir = os.path.join(OUT_BASE, ticker)
            os.makedirs(outdir, exist_ok=True)
            outpath = os.path.join(outdir, f"{ticker}.Valuation.v{TODAY}.xlsx")
            wb.save(outpath)
            print(f"  ✓ Saved: {outpath}")
        except Exception as e:
            import traceback
            print(f"\n!! ERR {ticker}: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    main()
