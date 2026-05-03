#!/usr/bin/env python3
"""
Bank Financial Valuation Model — 5.HK HSBC Holdings plc
UK-headquartered global universal bank; dual-primary LSE + HKEX listing.
Data source: akshare stock_financial_hk_report_em (returns HKD equivalents).
Units: HKD 亿 throughout (÷1e8 from raw HKD).
"""

import akshare as ak
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os, time
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ─── Config ───────────────────────────────────────────────────────────────────
TS_CODE   = '5.HK'
HK_CODE5  = '00005'
COMPANY   = 'HSBC Holdings plc (汇丰控股)'
CCY       = 'HKD'
UNIT_LABEL = 'HKD 亿'
UNIT_DIV  = 1e8

YEARS = list(range(2017, 2026))   # 9 years

# USD 10yr Treasury yield, year-end (HSBC reports USD originally; HGB/HIBOR closely track USD)
USD_BOND = {
    2017: 0.0240, 2018: 0.0268, 2019: 0.0192, 2020: 0.0091,
    2021: 0.0151, 2022: 0.0388, 2023: 0.0388, 2024: 0.0458, 2025: 0.0425,
}
TARGET_PE = 8.0   # per bank-template default for HK banks

OUTPUT_DIR  = os.path.expanduser(f'~/Desktop/BUI_Investment_Management/{TS_CODE}/')
os.makedirs(OUTPUT_DIR, exist_ok=True)
TODAY = datetime.now().strftime('%y%m%d')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'{TS_CODE}.Valuation.v{TODAY}.xlsx')

# ─── Style helpers ────────────────────────────────────────────────────────────
def _font(color='FF000000', bold=False):
    return Font(name='Apple Braille', size=11, color=color, bold=bold)
F_BOLD   = _font(bold=True)
SOLID    = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
GRAY_FILL= PatternFill(fill_type='solid', fgColor='FFD9D9D9')
HDR_FILL = PatternFill(fill_type='solid', fgColor='FF305496')
YLW_FILL = PatternFill(fill_type='solid', fgColor='FFFFF2CC')
R_ALIGN  = Alignment(horizontal='right')
C_ALIGN  = Alignment(horizontal='center')

def _w(ws, r, c, val, font=None, fill=None, align=None, nf=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:  cell.font = font
    if fill:  cell.fill = fill
    if align: cell.alignment = align
    if nf:    cell.number_format = nf
    return cell

def w_blue(ws, r, c, val, nf='#,##0.0', fill=None):
    return _w(ws, r, c, val, _font('FF0000FF'), fill or SOLID, R_ALIGN, nf)
def w_blk(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, _font('FF000000'), SOLID, R_ALIGN, nf)
def w_grn(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, _font('FF008000'), SOLID, R_ALIGN, nf)
def w_lbl(ws, r, c, val, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _font(bold=bold); cell.fill = SOLID
    return cell
def w_sec(ws, r, c, val):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _font(bold=True); cell.fill = GRAY_FILL
    return cell

def set_heights(ws, h):
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = h

def na(v):
    if v is None: return None
    try:
        f = float(v)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None

def yi(v):
    v2 = na(v)
    return round(v2 / UNIT_DIV, 4) if v2 is not None else None


# ═══════════════════════════════════════════════════════════════════════════════
# FETCH DATA
# ═══════════════════════════════════════════════════════════════════════════════

print("="*60); print(f"Fetching {TS_CODE} ({COMPANY})..."); print("="*60)

print("  [1/5] IS annual...")
df_is  = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="利润表",    indicator="年度"); time.sleep(0.5)
print("  [2/5] BS annual...")
df_bs  = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="资产负债表", indicator="年度"); time.sleep(0.5)
print("  [3/5] CF annual...")
df_cf  = ak.stock_financial_hk_report_em(stock=HK_CODE5, symbol="现金流量表", indicator="年度"); time.sleep(0.5)
print("  [4/5] Indicators...")
df_ind = ak.stock_financial_hk_analysis_indicator_em(symbol=HK_CODE5, indicator="年度"); time.sleep(0.5)

print("  [5/5] Prices (yfinance)...")
df_px = pd.DataFrame()
try:
    import yfinance as yf
    t = yf.Ticker("0005.HK")
    hist = t.history(start="2016-01-01", end="2026-12-31", auto_adjust=False)
    if not hist.empty:
        df_px = pd.DataFrame({
            '日期': hist.index.tz_localize(None) if hist.index.tz else hist.index,
            '收盘': hist['Close'].values,
        })
        df_px['日期'] = pd.to_datetime(df_px['日期'])
        print(f"    ✓ yfinance: {len(df_px)} rows")
except Exception as e:
    print(f"    yfinance error: {e}")

# Pivot long → year-dict
def pivot(df_long, years):
    out = {yr: {} for yr in years}
    for _, row in df_long.iterrows():
        rd = str(row.get('REPORT_DATE', ''))[:10]
        if len(rd) < 4: continue
        yr = int(rd[:4])
        if yr not in years: continue
        item = str(row.get('STD_ITEM_NAME', ''))
        amt  = na(row.get('AMOUNT'))
        if amt is not None and item:
            out[yr][item] = amt
    return out

is_d = pivot(df_is, YEARS)
bs_d = pivot(df_bs, YEARS)
cf_d = pivot(df_cf, YEARS)
ind_d = {int(str(r['REPORT_DATE'])[:4]): r for _, r in df_ind.iterrows()
         if int(str(r['REPORT_DATE'])[:4]) in YEARS}

# Year-end prices
px_by_year = {}
if not df_px.empty:
    df_px['年'] = df_px['日期'].dt.year
    for yr in YEARS:
        sub = df_px[df_px['年'] == yr]
        if not sub.empty:
            px_by_year[yr] = float(sub.sort_values('日期').iloc[-1]['收盘'])

def pick(d, *keys):
    for k in keys:
        if k in d:
            v = na(d[k])
            if v is not None: return v
    return None

# ─── Build yearly data dict ────────────────────────────────────────────────────
data = {}
for yr in YEARS:
    isd = is_d.get(yr, {})
    bsd = bs_d.get(yr, {})
    cfd = cf_d.get(yr, {})
    ind = ind_d.get(yr, {})

    # Income statement
    nii        = yi(pick(isd, '净利息收入'))
    fee_net    = yi(pick(isd, '手续费及佣金净收入'))
    op_income  = yi(pick(isd, '经营收入总额', '营业收入'))
    op_exp     = yi(pick(isd, '经营支出总额', '营业支出'))
    op_profit  = yi(pick(isd, '经营溢利', '营业利润'))
    tax        = yi(pick(isd, '税项', '所得税'))
    net_profit = yi(pick(isd, '除税后溢利', '净利润'))
    ni_parent  = yi(pick(isd, '股东应占溢利', '归属于母公司股东的净利润'))

    # Impairment — absent from standard IS; pulled from CF reconciliation
    impair     = yi(pick(cfd, '加:减值及拨备', '减值损失'))

    # Balance sheet
    tot_assets = yi(pick(bsd, '总资产'))
    loans      = yi(pick(bsd, '客户贷款及其他款项'))
    deposits   = yi(pick(bsd, '客户存款'))
    tot_liab   = yi(pick(bsd, '总负债'))
    eq_parent  = yi(pick(bsd, '归属于母公司股东权益', '股东权益'))
    mi         = yi(pick(bsd, '非控股股东权益', '少数股东权益'))

    # Cash flow
    ocf        = yi(pick(cfd, '经营业务现金净额'))
    div_paid   = yi(pick(cfd, '已付股息(融资)'))
    buybacks   = yi(pick(cfd, '回购股份'))

    # Shares (derive from EPS and NI_parent)
    shares = None
    eps = na(ind.get('BASIC_EPS'))
    hp  = na(ind.get('HOLDER_PROFIT'))
    if eps and hp and eps != 0:
        shares = round(hp / eps / 1e9, 4)   # 十亿股

    # Market cap
    price = px_by_year.get(yr)
    mktcap = None
    if price and shares:
        mktcap = round(price * shares * 10, 4)   # 十亿股 × HKD/share → 十亿 HKD; × 10 → 亿

    # Dividends paid — negative convention
    divs_neg = -round(abs(div_paid), 4) if div_paid is not None else None

    data[yr] = dict(
        nii=nii, fee_net=fee_net, op_income=op_income, op_exp=op_exp,
        op_profit=op_profit, tax=tax, net_profit=net_profit, ni_parent=ni_parent,
        impair=impair,
        tot_assets=tot_assets, loans=loans, deposits=deposits,
        tot_liab=tot_liab, eq_parent=eq_parent, mi=mi,
        ocf=ocf, div_paid=div_paid, divs_neg=divs_neg, buybacks=buybacks,
        price=price, shares=shares, mktcap=mktcap,
    )

print("\nSummary:")
for yr in YEARS:
    d = data[yr]
    print(f"  {yr}: OpInc={d['op_income']} NI={d['ni_parent']} Assets={d['tot_assets']} "
          f"Eq={d['eq_parent']} Px={d['price']} MktCap={d['mktcap']}")


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding workbook...")
wb = Workbook()
sheets = {}
for i, name in enumerate(['Financial Data', 'Key Ratios', 'Valuation Model',
                           'Top 10 Shareholders', 'Data Correlation']):
    if i == 0:
        ws = wb.active; ws.title = name
    else:
        ws = wb.create_sheet(name)
    sheets[name] = ws

def yc(yr): return 2 + YEARS.index(yr)
def ycl(yr): return get_column_letter(yc(yr))


# ═══ SHEET 1 — Financial Data ════════════════════════════════════════════════
ws1 = sheets['Financial Data']
_w(ws1, 1, 1, f'{COMPANY} ({TS_CODE}) - Financial Data', F_BOLD, SOLID)
_w(ws1, 1, 2, f'Units: {UNIT_LABEL}', _font(), SOLID)
_w(ws1, 2, 1, 'Bank-specific template. HSBC reports USD; akshare returns HKD equivalents.', _font(), SOLID)
_w(ws1, 3, 1, f'Metric', F_BOLD, SOLID)
for yr in YEARS:
    _w(ws1, 3, yc(yr), yr, F_BOLD, SOLID, C_ALIGN, 'General')

# Income Statement
w_sec(ws1, 4, 1, 'Income Statement (HKD 亿)')

IS_ROWS_BLUE = [
    (5,  'Net Interest Income',          'nii'),
    (6,  'Fee & Commission Income (net)', 'fee_net'),
    (8,  'Operating Income (经营收入总额)', 'op_income'),
    (9,  'Operating Expenses (经营支出总额)','op_exp'),
    (11, 'Impairment (from CF reconcile)', 'impair'),
    (12, 'Operating Profit (经营溢利)',    'op_profit'),
    (13, 'Income Tax Expense',             'tax'),
    (14, 'Net Profit',                     'net_profit'),
    (15, 'Net Profit to Parent',           'ni_parent'),
]
for row, lbl, key in IS_ROWS_BLUE:
    w_lbl(ws1, row, 1, lbl)
    for yr in YEARS:
        v = data[yr].get(key)
        if v is not None: w_blue(ws1, row, yc(yr), v, '#,##0.0')

w_lbl(ws1, 7, 1, 'Other Non-Interest Income')
w_lbl(ws1, 10, 1, 'PPOP (Pre-Provision Op Profit)')
w_lbl(ws1, 16, 1, 'Cost-to-Income Ratio')
for yr in YEARS:
    c, cl = yc(yr), ycl(yr)
    w_blk(ws1, 7, c, f'={cl}8-{cl}5-{cl}6', '#,##0.0')
    w_blk(ws1, 10, c, f'={cl}8-{cl}9', '#,##0.0')
    w_blk(ws1, 16, c, f'=IFERROR({cl}9/{cl}8,"-")', '0.0%')

# Balance Sheet
w_sec(ws1, 18, 1, 'Balance Sheet (HKD 亿)')
BS_ROWS_BLUE = [
    (19, 'Total Assets',                   'tot_assets'),
    (20, 'Loans & Advances (贷款)',         'loans'),
    (21, 'Customer Deposits (存款)',        'deposits'),
    (22, 'Total Liabilities',              'tot_liab'),
    (23, 'Total Equity to Parent',         'eq_parent'),
    (24, 'Minority Interest',              'mi'),
]
for row, lbl, key in BS_ROWS_BLUE:
    w_lbl(ws1, row, 1, lbl)
    for yr in YEARS:
        v = data[yr].get(key)
        if v is not None: w_blue(ws1, row, yc(yr), v, '#,##0.0')

w_lbl(ws1, 25, 1, 'Total Equity')
w_lbl(ws1, 26, 1, 'Loan/Deposit Ratio')
w_lbl(ws1, 27, 1, 'Equity/Asset Ratio (Leverage)')
for yr in YEARS:
    c, cl = yc(yr), ycl(yr)
    w_blk(ws1, 25, c, f'={cl}23+{cl}24', '#,##0.0')
    w_blk(ws1, 26, c, f'=IFERROR({cl}20/{cl}21,"-")', '0.0%')
    w_blk(ws1, 27, c, f'=IFERROR({cl}23/{cl}19,"-")', '0.00%')

# Cash Flow
w_sec(ws1, 29, 1, 'Cash Flow (HKD 亿)')
w_lbl(ws1, 30, 1, 'Net Cash from Operations')
w_lbl(ws1, 31, 1, 'Dividends & Interest on Perp (paid, neg)')
w_lbl(ws1, 32, 1, 'Share Buybacks (paid, positive)')
for yr in YEARS:
    d = data[yr]
    if d['ocf']      is not None: w_blue(ws1, 30, yc(yr), d['ocf'], '#,##0.0')
    if d['divs_neg'] is not None: w_blue(ws1, 31, yc(yr), d['divs_neg'], '#,##0.0')
    if d['buybacks'] is not None: w_blue(ws1, 32, yc(yr), d['buybacks'], '#,##0.0')

# Market Data
w_sec(ws1, 33, 1, 'Market Data')
w_lbl(ws1, 34, 1, 'HK Year-End Close (HKD/share)')
w_lbl(ws1, 35, 1, 'Total Shares (十亿股)')
w_lbl(ws1, 36, 1, 'Market Capitalisation (HKD 亿)')
w_lbl(ws1, 37, 1, 'Book Value Per Share (HKD)')
w_lbl(ws1, 38, 1, 'EPS (HKD, basic)')
w_lbl(ws1, 39, 1, 'DPS (HKD, implied)')
for yr in YEARS:
    c, cl = yc(yr), ycl(yr)
    d = data[yr]
    if d['price']  is not None: w_blue(ws1, 34, c, d['price'], '0.00')
    if d['shares'] is not None: w_blue(ws1, 35, c, d['shares'], '#,##0.000')
    if d['mktcap'] is not None: w_blue(ws1, 36, c, d['mktcap'], '#,##0.0')
    w_blk(ws1, 37, c, f'=IFERROR({cl}23/({cl}35*10),"-")', '0.00')
    w_blk(ws1, 38, c, f'=IFERROR({cl}15/({cl}35*10),"-")', '0.00')
    w_blk(ws1, 39, c, f'=IFERROR(ABS({cl}31)/({cl}35*10),"-")', '0.00')

ws1.column_dimensions['A'].width = 40
for yr in YEARS: ws1.column_dimensions[ycl(yr)].width = 13
ws1.column_dimensions[get_column_letter(yc(YEARS[-1])+1)].width = 8.83
set_heights(ws1, 20)


# ═══ SHEET 2 — Key Ratios ════════════════════════════════════════════════════
ws2 = sheets['Key Ratios']
_w(ws2, 1, 1, f'{COMPANY} - Key Banking Ratios', F_BOLD, SOLID)
for yr in YEARS:
    _w(ws2, 3, yc(yr), yr, F_BOLD, SOLID, C_ALIGN, 'General')

w_sec(ws2, 3, 1, 'Profitability & Efficiency')
LINKS_KR = [
    (4,  'NI to Parent',              "='Financial Data'!{}15"),
    (5,  'Total Equity to Parent',    "='Financial Data'!{}23"),
    (8,  'Total Assets',              "='Financial Data'!{}19"),
    (11, 'Net Interest Income',       "='Financial Data'!{}5"),
    (16, 'Impairment Losses',         "='Financial Data'!{}11"),
]
for row, lbl, tmpl in LINKS_KR:
    w_lbl(ws2, row, 1, lbl)
    for yr in YEARS:
        c, cl = yc(yr), ycl(yr)
        w_grn(ws2, row, c, tmpl.format(cl), '#,##0.0')

w_lbl(ws2, 6, 1, 'Avg Equity')
w_lbl(ws2, 7, 1, 'ROE (on Parent Equity)', bold=True)
w_lbl(ws2, 9, 1, 'Avg Assets')
w_lbl(ws2, 10, 1, 'ROA')
w_lbl(ws2, 12, 1, 'Avg Interest-Earning Assets (≈Assets)')
w_lbl(ws2, 13, 1, 'NIM (approx NII / AvgAssets)')

for i, yr in enumerate(YEARS):
    c, cl = yc(yr), ycl(yr)
    if i > 0:
        pcl = ycl(YEARS[i-1])
        w_blk(ws2, 6, c, f'=({pcl}5+{cl}5)/2', '#,##0.0')
        w_blk(ws2, 7, c, f'=IFERROR({cl}4/{cl}6,"-")', '0.0%')
        w_blk(ws2, 9, c, f'=({pcl}8+{cl}8)/2', '#,##0.0')
        w_blk(ws2, 10, c, f'=IFERROR({cl}4/{cl}9,"-")', '0.00%')
        w_blk(ws2, 12, c, f'={cl}9', '#,##0.0')
        w_blk(ws2, 13, c, f'=IFERROR({cl}11/{cl}12,"-")', '0.00%')

w_sec(ws2, 15, 1, 'Provisioning & Growth')
w_lbl(ws2, 17, 1, 'Provisions/Loans (Credit Cost)')
w_lbl(ws2, 18, 1, 'Loan Growth YoY')
w_lbl(ws2, 19, 1, 'Deposit Growth YoY')
w_lbl(ws2, 20, 1, 'NI Growth YoY')
for i, yr in enumerate(YEARS):
    c, cl = yc(yr), ycl(yr)
    w_blk(ws2, 17, c, f"=IFERROR({cl}16/'Financial Data'!{cl}20,\"-\")", '0.00%')
    if i > 0:
        pcl = ycl(YEARS[i-1])
        w_blk(ws2, 18, c, f"=IFERROR(('Financial Data'!{cl}20-'Financial Data'!{pcl}20)/'Financial Data'!{pcl}20,\"-\")", '0.0%')
        w_blk(ws2, 19, c, f"=IFERROR(('Financial Data'!{cl}21-'Financial Data'!{pcl}21)/'Financial Data'!{pcl}21,\"-\")", '0.0%')
        w_blk(ws2, 20, c, f'=IFERROR(({cl}4-{pcl}4)/{pcl}4,"-")', '0.0%')

w_sec(ws2, 22, 1, 'Capital & Payout')
w_lbl(ws2, 23, 1, 'Dividend Payout Ratio')
w_lbl(ws2, 24, 1, 'Retention Ratio')
for yr in YEARS:
    c, cl = yc(yr), ycl(yr)
    w_blk(ws2, 23, c, f"=IFERROR(ABS('Financial Data'!{cl}31)/{cl}4,\"-\")", '0.0%')
    w_blk(ws2, 24, c, f'=1-{cl}23', '0.0%')

ws2.column_dimensions['A'].width = 40
for yr in YEARS: ws2.column_dimensions[ycl(yr)].width = 13
ws2.column_dimensions[get_column_letter(yc(YEARS[-1])+1)].width = 8.83
set_heights(ws2, 20)


# ═══ SHEET 3 — Valuation Model ═══════════════════════════════════════════════
ws3 = sheets['Valuation Model']
_w(ws3, 1, 1, f'{COMPANY} - Valuation Model', F_BOLD, SOLID)
_w(ws3, 2, 1, '3 methods: Dividend Yield / Justified P/B / Target P/E. Currency: HKD 亿 (market cap and financials aligned).', _font(), SOLID)

for yr in YEARS:
    _w(ws3, 4, yc(yr), yr, F_BOLD, SOLID, C_ALIGN, 'General')

w_sec(ws3, 5, 1, 'Key Assumptions')
w_lbl(ws3, 6, 1, 'USD 10yr Bond Yield (year-end)')
for yr in YEARS:
    w_blue(ws3, 6, yc(yr), USD_BOND[yr], '0.00%', fill=YLW_FILL)

w_sec(ws3, 9, 1, 'Inputs (linked)')
LINKS_VM = [
    (10, 'NI to Parent (HKD 亿)',           "='Financial Data'!{}15"),
    (11, 'Dividends Paid (abs, HKD 亿)',     "=ABS('Financial Data'!{}31)"),
    (12, 'Total Equity to Parent (HKD 亿)',  "='Financial Data'!{}23"),
    (13, 'Market Cap (HKD 亿)',              "='Financial Data'!{}36"),
    (14, 'Market Cap (HKD 亿, same unit)',   "={}13"),
]
for row, lbl, tmpl in LINKS_VM:
    w_lbl(ws3, row, 1, lbl)
    for yr in YEARS:
        c, cl = yc(yr), ycl(yr)
        if '{}' in tmpl and 'Financial Data' in tmpl:
            w_grn(ws3, row, c, tmpl.format(cl), '#,##0.0')
        else:
            w_blk(ws3, row, c, tmpl.format(cl), '#,##0.0')

w_sec(ws3, 16, 1, 'Valuation Methods')
w_lbl(ws3, 17, 1, 'Dividend Yield (Div/MktCap)')
w_lbl(ws3, 18, 1, 'Mkt Cap 1 = Div / BondYield', bold=True)
w_lbl(ws3, 19, 1, 'P/B (Implied)')
w_lbl(ws3, 20, 1, 'ROE')
w_lbl(ws3, 21, 1, 'Payout Ratio')
w_lbl(ws3, 22, 1, 'Growth g = ROE × (1−Payout)')
w_lbl(ws3, 23, 1, 'Cost of Equity (r = Bond + 5%)')
w_lbl(ws3, 24, 1, 'Justified P/B = (ROE−g)/(r−g)')
w_lbl(ws3, 25, 1, 'Mkt Cap 2 = JustPB × Equity', bold=True)
w_lbl(ws3, 27, 1, 'P/E (Implied)')
w_lbl(ws3, 28, 1, f'Target P/E (manual, default {TARGET_PE:.1f}x)')
w_lbl(ws3, 29, 1, 'Mkt Cap 3 = Target PE × NI', bold=True)

for i, yr in enumerate(YEARS):
    c, cl = yc(yr), ycl(yr)
    w_blk(ws3, 17, c, f'=IFERROR({cl}11/{cl}14,"-")', '0.00%')
    w_blk(ws3, 18, c, f'=IFERROR({cl}11/{cl}6,"-")', '#,##0.0')
    w_blk(ws3, 19, c, f'=IFERROR({cl}14/{cl}12,"-")', '0.0"x"')
    w_grn(ws3, 20, c, f"='Key Ratios'!{cl}7", '0.0%')
    w_grn(ws3, 21, c, f"='Key Ratios'!{cl}23", '0.0%')
    if i > 0:
        w_blk(ws3, 22, c, f'=IFERROR({cl}20*(1-{cl}21),"-")', '0.0%')
        w_blk(ws3, 24, c, f'=IFERROR(IF({cl}23-{cl}22=0,"-",({cl}20-{cl}22)/({cl}23-{cl}22)),"-")', '0.0"x"')
        w_blk(ws3, 25, c, f'=IFERROR({cl}24*{cl}12,"-")', '#,##0.0')
    w_blk(ws3, 23, c, f'={cl}6+0.05', '0.0%')
    w_blk(ws3, 27, c, f'=IFERROR({cl}14/{cl}10,"-")', '0.0"x"')
    w_blue(ws3, 28, c, TARGET_PE, '0.0"x"', fill=YLW_FILL)
    w_blk(ws3, 29, c, f'={cl}28*{cl}10', '#,##0.0')

w_sec(ws3, 31, 1, 'Valuation Summary')
w_lbl(ws3, 32, 1, 'Mkt Cap 1 (Div Yield)')
w_lbl(ws3, 33, 1, 'Mkt Cap 2 (Justified P/B)')
w_lbl(ws3, 34, 1, 'Mkt Cap 3 (Target P/E)')
w_lbl(ws3, 35, 1, 'Average Fair Value (HKD 亿)', bold=True)
w_lbl(ws3, 36, 1, 'Actual Mkt Cap (HKD 亿)')
w_lbl(ws3, 37, 1, 'Premium / (Discount)', bold=True)
for i, yr in enumerate(YEARS):
    c, cl = yc(yr), ycl(yr)
    w_blk(ws3, 32, c, f'={cl}18', '#,##0.0')
    if i > 0:
        w_blk(ws3, 33, c, f'={cl}25', '#,##0.0')
    w_blk(ws3, 34, c, f'={cl}29', '#,##0.0')
    if i > 0:
        w_blk(ws3, 35, c, f'=IFERROR(AVERAGE({cl}32,{cl}33,{cl}34),"-")', '#,##0.0')
    else:
        w_blk(ws3, 35, c, f'=IFERROR(AVERAGE({cl}32,{cl}34),"-")', '#,##0.0')
    w_blk(ws3, 36, c, f'={cl}14', '#,##0.0')
    w_blk(ws3, 37, c, f'=IFERROR({cl}35/{cl}36,"-")', '0.0"x"')

ws3.column_dimensions['A'].width = 42
for yr in YEARS: ws3.column_dimensions[ycl(yr)].width = 13
ws3.column_dimensions[get_column_letter(yc(YEARS[-1])+1)].width = 8.83
set_heights(ws3, 17)


# ═══ SHEET 4 — Top 10 Shareholders (note) ════════════════════════════════════
ws4 = sheets['Top 10 Shareholders']
_w(ws4, 1, 1, f'{COMPANY} - Top 10 Shareholders', F_BOLD, SOLID)
notes = [
    (2, 'HSBC is UK-listed (LSE: HSBA) + HK-listed (5.HK) with primary-primary dual listing.'),
    (3, 'Granular quarterly top-10 shareholder data is not available via free A-share APIs (akshare / tushare).'),
    (4, ''),
    (5, 'For detailed shareholder information, refer to:'),
    (6, '  1) HSBC Annual Report — Substantial Shareholders section (>3% holders per UK/HK rules)'),
    (7, '  2) HKEXnews — CCASS Shareholding Analysis (https://www.hkexnews.hk/sdw/search/searchsdw.aspx)'),
    (8, '  3) UK PLC filings — Companies House / LSE RNS'),
    (9, ''),
    (10, 'Known largest beneficial holders of HSBC (per 2024 Annual Report):'),
    (11, '  • Ping An Insurance Group — largest single shareholder (~8-9%)'),
    (12, '  • BlackRock — institutional ~7%'),
    (13, '  • The Capital Group, Vanguard, Norges Bank — other major institutional holders'),
    (14, '  • Free float — very broadly held; ~150K+ beneficial holders worldwide'),
]
for r, text in notes:
    _w(ws4, r, 1, text, _font(), SOLID)
ws4.column_dimensions['A'].width = 110


# ═══ SHEET 5 — Data Correlation ══════════════════════════════════════════════
ws5 = sheets['Data Correlation']
doc = [
    (1,  f'{COMPANY} ({TS_CODE}) - Data Correlation & Methodology', True),
    (2,  '', False),
    (3,  '1. Company Overview', True),
    (4,  'HSBC Holdings plc — UK-headquartered global universal bank.', False),
    (5,  'Dual-primary listing: LSE (HSBA.L) + HKEX (5.HK). ADR: HSBC (NYSE).', False),
    (6,  'Historical roots: founded 1865 in Hong Kong as The Hongkong and Shanghai Banking Corporation.', False),
    (7,  'Operations span Europe, Asia, Middle East, N.America, L.America; ~$2.8 trillion USD total assets.', False),
    (8,  '', False),
    (9,  '2. Data Sources', True),
    (10, 'Primary: akshare HK financial report endpoints', False),
    (11, '  stock_financial_hk_report_em(stock="00005", symbol=..., indicator="年度")', False),
    (12, '  stock_financial_hk_analysis_indicator_em(symbol="00005", indicator="年度")', False),
    (13, 'Price: yfinance (ticker "0005.HK"), year-end close.', False),
    (14, '', False),
    (15, '3. Currency & Units', True),
    (16, 'HSBC reports originally in USD. akshare returns data with CURRENCY="HKD".', False),
    (17, 'Per akshare metadata, amounts shown are HKD equivalents; market cap (HK price × shares) is also HKD.', False),
    (18, 'Model unit: HKD 亿 (100M HKD). Financials and market cap use the same unit → no mixed-unit issue.', False),
    (19, 'USD base readers: divide any HKD 亿 number by ~7.8 for USD billion rough equivalent.', False),
    (20, '', False),
    (21, '4. Bank-Specific Line-Item Mapping (HK IS format)', True),
    (22, 'Net Interest Income: 净利息收入 (line 35 of HK standard IS)', False),
    (23, 'Fee & Commission Income (net): 手续费及佣金净收入', False),
    (24, 'Operating Income: 经营收入总额 (includes NII + fee + trading + insurance + other)', False),
    (25, 'Operating Expenses: 经营支出总额 (staff + admin + depreciation, etc.)', False),
    (26, 'Operating Profit: 经营溢利', False),
    (27, 'Tax: 税项; Net Profit: 除税后溢利; NI to Parent: 股东应占溢利', False),
    (28, '', False),
    (29, '5. Critical Caveats', True),
    (30, '• Impairment losses are NOT a distinct top-line category in HK IS format — sourced from CF', False),
    (31, '  statement reconciliation ("加:减值及拨备"). Differs from IFRS ECL reporting convention.', False),
    (32, '• Dividends row ("已付股息(融资)") mixes common dividends + perpetual capital coupons.', False),
    (33, '• Share buybacks are significant for HSBC (~HK$110B in 2024) but shown separately from dividends.', False),
    (34, '• NIM here is NII / avg total assets. True NIM uses avg interest-earning assets only', False),
    (35, '  (excludes trading assets, derivatives, etc.). Reported HSBC NIM is typically ~150-170bp,', False),
    (36, '  while this model shows lower numbers because total-asset denominator is much larger.', False),
    (37, '• Operating income includes insurance and trading revenue swings — more volatile than peers.', False),
    (38, '• Share count varies year-to-year due to active buyback programs.', False),
    (39, '', False),
    (40, '6. Valuation Methodology', True),
    (41, 'Method 1 — Dividend Yield Capitalization: Mkt Cap = Dividends / USD 10yr Bond Yield', False),
    (42, 'Method 2 — Justified P/B (Gordon): (ROE − g) / (r − g) × Equity,  g = ROE × (1−Payout), r = Bond+5%', False),
    (43, f'Method 3 — Target P/E: Target × NI, default {TARGET_PE:.1f}x per bank-template for HK banks', False),
    (44, 'Average of the three = Fair Value; Premium/(Discount) = Fair Value / Actual Market Cap', False),
    (45, '', False),
    (46, '7. Bond Yield Reference (USD 10yr Treasury, year-end)', True),
    (47, '2017=2.40% 2018=2.68% 2019=1.92% 2020=0.91% 2021=1.51%', False),
    (48, '2022=3.88% 2023=3.88% 2024=4.58% 2025=4.25%', False),
    (49, '', False),
    (50, '8. Format Conventions', True),
    (51, 'Font: Apple Braille 11pt everywhere.', False),
    (52, 'Colors: BLUE=hardcoded input, BLACK=formula, GREEN=cross-sheet link.', False),
    (53, 'Yellow fill highlights key assumption cells (bond yield, target P/E).', False),
    (54, '', False),
    (55, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', False),
]
for r, text, bold in doc:
    _w(ws5, r, 1, text, Font(name='Apple Braille', size=11, bold=bold), SOLID)
ws5.column_dimensions['A'].width = 110


# ═══ SAVE ════════════════════════════════════════════════════════════════════
print(f"\nSaving → {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)
print("="*60)
print(f"✓ Saved: {OUTPUT_FILE}")
print("="*60)
print("\nSheets:")
print(f"  1. Financial Data ({len(YEARS)} years {YEARS[0]}-{YEARS[-1]})")
print(f"  2. Key Ratios")
print(f"  3. Valuation Model (3-method)")
print(f"  4. Top 10 Shareholders (reference notes)")
print(f"  5. Data Correlation")
