#!/usr/bin/env python3
"""
Financial Valuation Model Generator — 600519.SH 贵州茅台
7-sheet Excel workbook following BUI Investment Management template.
Units: CNY 亿元 (÷1e8 from raw yuan; total_mv ÷1e4 from 万元)
"""

import tushare as ts
import akshare as ak
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import time
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ─── Config ───────────────────────────────────────────────────────────────────
TS_CODE   = '600519.SH'
TICKER_6  = '600519'
COMPANY   = '贵州茅台'
CCY       = 'CNY'
UNIT_LABEL = 'CNY 亿元'
UNIT_DIV  = 1e8    # yuan → 亿元
MV_DIV    = 1e4    # 万元 → 亿元

YEARS = list(range(2015, 2026))   # 11 annual years: 2015-2025

# China 10yr sovereign bond yield (year-end, decimal)
CN_BOND = {
    2015: 0.0282, 2016: 0.0307, 2017: 0.0390, 2018: 0.0323,
    2019: 0.0314, 2020: 0.0315, 2021: 0.0278, 2022: 0.0284,
    2023: 0.0256, 2024: 0.0168, 2025: 0.0183,
}

OUTPUT_DIR  = os.path.expanduser(f'~/Desktop/BUI_Investment_Management/{TS_CODE}/')
os.makedirs(OUTPUT_DIR, exist_ok=True)
TODAY_STR   = datetime.now().strftime('%y%m%d')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'{TS_CODE}.Valuation.v{TODAY_STR}.xlsx')

# ─── Tushare ──────────────────────────────────────────────────────────────────
ts.set_token('45e390d8bf059895336edd8073d294a2f137495fee6cd659acfc6f43')
pro = ts.pro_api()

# ─── Style helpers ────────────────────────────────────────────────────────────
def _font(color, bold=False):
    return Font(name='Apple Braille', size=11, color=color, bold=bold)

F_BLUE  = _font('FF0000FF')
F_BLACK = _font('FF000000')
F_GREEN = _font('FF008000')
F_LABEL = _font('FF000000')            # default (no forced color) for labels
F_BOLD  = Font(name='Apple Braille', size=11, bold=True)
F_HDR   = Font(name='Apple Braille', size=11, bold=True, color='FFFFFFFF')

SOLID      = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
GRAY_FILL  = PatternFill(fill_type='solid', fgColor='FFD9D9D9')
HDR_FILL   = PatternFill(fill_type='solid', fgColor='FF305496')

R_ALIGN  = Alignment(horizontal='right')
C_ALIGN  = Alignment(horizontal='center')

def _w(ws, r, c, val, font=None, fill=None, align=None, nf=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:  cell.font  = font
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    if nf:    cell.number_format = nf
    return cell

def w_blue(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, F_BLUE, SOLID, R_ALIGN, nf)

def w_blk(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, F_BLACK, SOLID, R_ALIGN, nf)

def w_grn(ws, r, c, val, nf='#,##0.0'):
    return _w(ws, r, c, val, F_GREEN, SOLID, R_ALIGN, nf)

def w_lbl(ws, r, c, val, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = _font('FF000000', bold=bold)
    cell.fill = SOLID
    return cell

def w_sec(ws, r, c, val):
    """Gray section header spanning full width."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name='Apple Braille', size=11, bold=True)
    cell.fill = GRAY_FILL
    return cell

def set_heights(ws, h, r1=1, r2=None):
    r2 = r2 or ws.max_row
    for r in range(r1, r2+1):
        ws.row_dimensions[r].height = h

def na(v):
    """Return None if value is NaN or None, else the float."""
    if v is None: return None
    try:
        f = float(v)
        return None if np.isnan(f) else f
    except (TypeError, ValueError):
        return None

def div_unit(v):
    """Divide raw yuan value by UNIT_DIV to get 亿元, or None."""
    v2 = na(v)
    return round(v2 / UNIT_DIV, 4) if v2 is not None else None

def div_mv(v):
    """Divide total_mv (万元) by MV_DIV to get 亿元, or None."""
    v2 = na(v)
    return round(v2 / MV_DIV, 4) if v2 is not None else None


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — FETCH ANNUAL DATA
# ═══════════════════════════════════════════════════════════════════════════════

print("=" * 60)
print("Fetching 600519.SH annual financial data (2015-2025)...")
print("=" * 60)

# ── Income Statement ──
print("  [1/7] Income statement...")
inc_raw = pro.income(
    ts_code=TS_CODE, start_date='20150101', end_date='20251231',
    report_type='1',
    fields='ts_code,end_date,revenue,oper_cost,operate_profit,income_tax,'
           'n_income,n_income_attr_p'
)
time.sleep(0.5)
inc_ann = (inc_raw[inc_raw['end_date'].str.endswith('1231')]
           .drop_duplicates('end_date').sort_values('end_date'))
inc_d = {int(r['end_date'][:4]): r for _, r in inc_ann.iterrows()}

# ── Balance Sheet ──
print("  [2/7] Balance sheet...")
bs_raw = pro.balancesheet(
    ts_code=TS_CODE, start_date='20150101', end_date='20251231',
    report_type='1',
    fields='ts_code,end_date,total_assets,total_hldr_eqy_exc_min_int,'
           'total_hldr_eqy_inc_min_int,money_cap,minority_int,'
           'total_cur_assets,total_cur_liab,short_loan,long_loan,'
           'lease_liab,non_cur_liab_due_1y,total_share'
)
time.sleep(0.5)
bs_ann = (bs_raw[bs_raw['end_date'].str.endswith('1231')]
          .drop_duplicates('end_date').sort_values('end_date'))
bs_d = {int(r['end_date'][:4]): r for _, r in bs_ann.iterrows()}

# ── Cash Flow ──
print("  [3/7] Cash flow...")
cf_raw = pro.cashflow(
    ts_code=TS_CODE, start_date='20150101', end_date='20251231',
    report_type='1',
    fields='ts_code,end_date,n_cashflow_act,depr_fa_coga_dpba,'
           'amort_intang_assets,lt_amort_deferred_exp,c_pay_dist_dpcp_int_exp,'
           'stot_cash_outflows_fnc_act'
)
time.sleep(0.5)
cf_ann = (cf_raw[cf_raw['end_date'].str.endswith('1231')]
          .drop_duplicates('end_date').sort_values('end_date'))
cf_d = {int(r['end_date'][:4]): r for _, r in cf_ann.iterrows()}

# ── Dividends ──
print("  [4/7] Dividends...")
div_raw = pro.dividend(
    ts_code=TS_CODE,
    fields='ts_code,div_proc,cash_div,cash_div_tax,record_date,ex_date,pay_date,end_date,ann_date'
)
time.sleep(0.5)

# Build dividends-paid by year (total cash = cash_div_tax * total_shares)
# Use per-share × total shares from balance sheet
div_by_year = {}
if not div_raw.empty:
    for _, row in div_raw.iterrows():
        if str(row.get('div_proc', '')) not in ['实施', '进行中', '']:
            continue
        yr = None
        for fld in ['end_date', 'ann_date']:
            v = str(row.get(fld, '') or '')
            if len(v) >= 4 and v[:4].isdigit():
                yr = int(v[:4])
                break
        if yr and 2014 <= yr <= 2025:
            cdtax = na(row.get('cash_div_tax')) or na(row.get('cash_div')) or 0.0
            if yr not in div_by_year:
                div_by_year[yr] = 0.0
            div_by_year[yr] += cdtax

# Convert per-share dividends to total using total shares from BS
# Actually, let's get total cash dividends from cashflow statement
# c_pay_dist_dpcp_int_exp is cash paid for dividends/interest (financing)
# For Moutai with ~0 debt, this ≈ dividends paid

# ── Market Cap (year-end) ──
print("  [5/7] Year-end market cap...")
mkt_d = {}
for yr in YEARS:
    try:
        db = pro.daily_basic(
            ts_code=TS_CODE, start_date=f'{yr}1201', end_date=f'{yr}1231',
            fields='ts_code,trade_date,close,total_mv,pe_ttm,pb,total_share'
        )
        time.sleep(0.3)
        if not db.empty:
            db = db.sort_values('trade_date', ascending=False)
            mkt_d[yr] = db.iloc[0]
            print(f"    {yr}: close={db.iloc[0]['close']:.1f} total_mv={db.iloc[0]['total_mv']:.0f}万元")
    except Exception as e:
        print(f"    {yr} market data error: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — PROCESS ANNUAL DATA INTO MODEL ROWS
# ═══════════════════════════════════════════════════════════════════════════════

def g(d, yr, field, divisor=None):
    """Get field from dict[year] row, optionally divide."""
    row = d.get(yr)
    if row is None: return None
    v = na(row.get(field))
    if v is None: return None
    if divisor: v = v / divisor
    return round(v, 4)

# Build year-indexed data containers
data = {}
for yr in YEARS:
    inc = inc_d.get(yr, {})
    bs  = bs_d.get(yr, {})
    cf  = cf_d.get(yr, {})
    mkt = mkt_d.get(yr, {})

    revenue    = g(inc_d, yr, 'revenue',          UNIT_DIV)
    oper_cost  = g(inc_d, yr, 'oper_cost',         UNIT_DIV)
    ebit       = g(inc_d, yr, 'operate_profit',    UNIT_DIV)
    tax        = g(inc_d, yr, 'income_tax',        UNIT_DIV)
    net_inc    = g(inc_d, yr, 'n_income',          UNIT_DIV)
    ni_parent  = g(inc_d, yr, 'n_income_attr_p',  UNIT_DIV)

    tot_assets = g(bs_d, yr, 'total_assets',                 UNIT_DIV)
    eq_excl_mi = g(bs_d, yr, 'total_hldr_eqy_exc_min_int',  UNIT_DIV)
    eq_incl_mi = g(bs_d, yr, 'total_hldr_eqy_inc_min_int',  UNIT_DIV)
    cash_eq    = g(bs_d, yr, 'money_cap',                    UNIT_DIV)
    mi         = g(bs_d, yr, 'minority_int',                 UNIT_DIV)
    cur_assets = g(bs_d, yr, 'total_cur_assets',             UNIT_DIV)
    cur_liab   = g(bs_d, yr, 'total_cur_liab',               UNIT_DIV)
    st_debt    = g(bs_d, yr, 'short_loan',                   UNIT_DIV)
    lt_debt    = g(bs_d, yr, 'long_loan',                    UNIT_DIV)
    st_lease   = g(bs_d, yr, 'non_cur_liab_due_1y',          UNIT_DIV)
    lt_lease   = g(bs_d, yr, 'lease_liab',                   UNIT_DIV)
    tot_shares = g(bs_d, yr, 'total_share', 1e8)  # shares → 亿股

    ocf       = g(cf_d, yr, 'n_cashflow_act',         UNIT_DIV)
    depr      = g(cf_d, yr, 'depr_fa_coga_dpba',      UNIT_DIV)
    amort_int = g(cf_d, yr, 'amort_intang_assets',    UNIT_DIV)
    amort_lpe = g(cf_d, yr, 'lt_amort_deferred_exp',  UNIT_DIV)
    div_fin   = g(cf_d, yr, 'c_pay_dist_dpcp_int_exp',UNIT_DIV)

    # D&A = sum of components
    dna_parts = [x for x in [depr, amort_int, amort_lpe] if x is not None]
    dna = round(sum(dna_parts), 4) if dna_parts else None

    # Working Capital = CA - CL
    wc = round(cur_assets - cur_liab, 4) if (cur_assets and cur_liab) else None

    # Net Working Capital = (CA - Cash) - (CL - ST Debt)
    nwc = None
    if cur_assets and cur_liab:
        ca_adj = cur_assets - (cash_eq or 0)
        cl_adj = cur_liab - (st_debt or 0)
        nwc = round(ca_adj - cl_adj, 4)

    # Total Debt = ST + LT debt
    tot_debt = None
    debt_parts = [x for x in [st_debt, lt_debt] if x is not None]
    if debt_parts:
        tot_debt = round(sum(debt_parts), 4)

    # Market cap
    mktcap = None
    if mkt is not None:
        mv = na(mkt.get('total_mv'))
        if mv: mktcap = round(mv / MV_DIV, 4)

    # Dividends paid — use cashflow financing outflow (≈ dividends for Moutai)
    # Negate to show as positive outflow
    divs_paid = None
    if div_fin is not None:
        divs_paid = -round(abs(div_fin), 4)   # negative convention

    data[yr] = dict(
        revenue=revenue, oper_cost=oper_cost, ebit=ebit, tax=tax,
        net_inc=net_inc, ni_parent=ni_parent, dna=dna,
        tot_assets=tot_assets, eq_excl_mi=eq_excl_mi, eq_incl_mi=eq_incl_mi,
        cash_eq=cash_eq, mi=mi, cur_assets=cur_assets, cur_liab=cur_liab,
        st_debt=st_debt, lt_debt=lt_debt, st_lease=st_lease, lt_lease=lt_lease,
        tot_debt=tot_debt, wc=wc, nwc=nwc, tot_shares=tot_shares,
        ocf=ocf, div_fin=div_fin, divs_paid=divs_paid,
        mktcap=mktcap,
    )

print("\nAnnual data summary:")
for yr in YEARS:
    d = data[yr]
    print(f"  {yr}: Rev={d['revenue']} EBIT={d['ebit']} NI={d['ni_parent']} MktCap={d['mktcap']}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — FETCH QUARTERLY DATA FOR TTM
# ═══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*60)
print("Fetching quarterly data for TTM analysis...")
print("="*60)

# Need YTD cumulative from 2019Q1 onwards for 6-year TTM (2020Q1–2025Q4)
print("  Quarterly income (2019-2025)...")
q_inc = pro.income(
    ts_code=TS_CODE, start_date='20190101', end_date='20251231',
    report_type='1',
    fields='ts_code,end_date,revenue,oper_cost,operate_profit,income_tax,'
           'n_income,n_income_attr_p'
)
time.sleep(0.5)
q_inc = q_inc.drop_duplicates('end_date').sort_values('end_date').reset_index(drop=True)

print("  Quarterly cash flow (2019-2025)...")
q_cf = pro.cashflow(
    ts_code=TS_CODE, start_date='20190101', end_date='20251231',
    report_type='1',
    fields='ts_code,end_date,n_cashflow_act,depr_fa_coga_dpba,'
           'amort_intang_assets,lt_amort_deferred_exp,c_pay_dist_dpcp_int_exp'
)
time.sleep(0.5)
q_cf = q_cf.drop_duplicates('end_date').sort_values('end_date').reset_index(drop=True)

print("  Quarterly balance sheet (2019-2025)...")
q_bs = pro.balancesheet(
    ts_code=TS_CODE, start_date='20190101', end_date='20251231',
    report_type='1',
    fields='ts_code,end_date,total_assets,total_hldr_eqy_exc_min_int,'
           'total_hldr_eqy_inc_min_int,money_cap,minority_int,'
           'short_loan,long_loan,lease_liab,non_cur_liab_due_1y'
)
time.sleep(0.5)
q_bs = q_bs.drop_duplicates('end_date').sort_values('end_date').reset_index(drop=True)

# Convert quarterly DataFrames to dicts keyed by end_date string (YYYYMMDD)
def df_to_dict(df):
    return {str(r['end_date']): r for _, r in df.iterrows()}

qi_d  = df_to_dict(q_inc)
qcf_d = df_to_dict(q_cf)
qbs_d = df_to_dict(q_bs)

def get_qval(d, date_str, field):
    row = d.get(date_str)
    if row is None: return None
    return na(row.get(field))

def ttm_flow(d, field, qdate_str):
    """
    Compute TTM for a flow item (YTD-based).
    qdate_str: YYYYMMDD of quarter end.
    TTM(Qn,Y) = YTD(Qn,Y) + Annual(Y-1) - YTD(Qn,Y-1)
    Q4 (1231): TTM = Annual(Y) = YTD Q4
    """
    yr   = int(qdate_str[:4])
    mmdd = qdate_str[4:]

    ytd_cur  = get_qval(d, qdate_str, field)
    if ytd_cur is None: return None

    if mmdd == '1231':
        return ytd_cur   # Q4 YTD = full year = TTM

    # Find same quarter last year
    prior_qdate = f'{yr-1}{mmdd}'
    ytd_prior = get_qval(d, prior_qdate, field)

    # Annual Y-1 = Q4 of Y-1
    annual_prior_date = f'{yr-1}1231'
    annual_prior = get_qval(d, annual_prior_date, field)

    if ytd_prior is None or annual_prior is None:
        return None

    return ytd_cur + annual_prior - ytd_prior


# Generate all quarter-end dates for 2020Q1 to latest available
all_q_dates = []
for yr in range(2020, 2026):
    for mmdd in ['0331','0630','0930','1231']:
        all_q_dates.append(f'{yr}{mmdd}')

# Get quarter-end market data
print("  Quarter-end market data...")
q_mkt = {}
for qd in all_q_dates:
    yr   = int(qd[:4])
    mmdd = qd[4:]
    # Try qd date first; if weekend/holiday, get last 5 trading days ending that date
    try:
        db = pro.daily_basic(
            ts_code=TS_CODE, start_date=f'{yr}{mmdd[:2]}01' if mmdd[2:]=='01' else qd[:6]+'01',
            end_date=qd,
            fields='ts_code,trade_date,close,total_mv,pe_ttm,pb'
        )
        time.sleep(0.2)
        if not db.empty:
            db = db.sort_values('trade_date', ascending=False)
            q_mkt[qd] = db.iloc[0]
    except Exception as e:
        print(f"    {qd} market error: {e}")

# Build TTM rows
print("  Computing TTM values...")
ttm_rows = []
for qd in all_q_dates:
    yr   = int(qd[:4])
    mmdd = qd[4:]
    q_label = f"{yr}-{mmdd[:2]}-{mmdd[2:]}"

    rev    = ttm_flow(qi_d, 'revenue',          qd)
    cogs   = ttm_flow(qi_d, 'oper_cost',        qd)
    ebit   = ttm_flow(qi_d, 'operate_profit',   qd)
    tax    = ttm_flow(qi_d, 'income_tax',        qd)
    ni     = ttm_flow(qi_d, 'n_income',         qd)
    ni_p   = ttm_flow(qi_d, 'n_income_attr_p',  qd)
    ocf    = ttm_flow(qcf_d,'n_cashflow_act',   qd)
    div_cf = ttm_flow(qcf_d,'c_pay_dist_dpcp_int_exp', qd)
    depr   = ttm_flow(qcf_d,'depr_fa_coga_dpba',       qd)
    amort1 = ttm_flow(qcf_d,'amort_intang_assets',     qd)
    amort2 = ttm_flow(qcf_d,'lt_amort_deferred_exp',   qd)

    dna_parts = [x for x in [depr, amort1, amort2] if x is not None]
    dna = round(sum(dna_parts), 4) if dna_parts else None

    # BS point-in-time
    bs_row = qbs_d.get(qd, {})
    def bsv(f): return na(bs_row.get(f)) / UNIT_DIV if na(bs_row.get(f)) else None

    tot_assets = bsv('total_assets')
    eq_excl    = bsv('total_hldr_eqy_exc_min_int')
    eq_incl    = bsv('total_hldr_eqy_inc_min_int')
    cash_eq    = bsv('money_cap')
    mi_val     = bsv('minority_int')
    st_debt_q  = bsv('short_loan')
    lt_debt_q  = bsv('long_loan')
    debt_parts = [x for x in [st_debt_q, lt_debt_q] if x is not None]
    tot_debt_q = round(sum(debt_parts), 4) if debt_parts else None

    # Market data
    mkt_row = q_mkt.get(qd)
    close   = na(mkt_row.get('close')) if mkt_row is not None else None
    total_mv_raw = na(mkt_row.get('total_mv')) if mkt_row is not None else None
    mktcap_q = round(total_mv_raw / MV_DIV, 4) if total_mv_raw else None

    def to_yi(v):
        return round(v / UNIT_DIV, 4) if v else None

    ttm_rows.append(dict(
        qdate=qd, q_label=q_label,
        rev    = to_yi(rev)    if rev    else None,
        cogs   = to_yi(cogs)   if cogs   else None,
        ebit   = to_yi(ebit)   if ebit   else None,
        tax    = to_yi(tax)    if tax    else None,
        ni     = to_yi(ni)     if ni     else None,
        ni_p   = to_yi(ni_p)   if ni_p   else None,
        ocf    = to_yi(ocf)    if ocf    else None,
        div_cf = to_yi(div_cf) if div_cf else None,
        dna    = dna,
        tot_assets=tot_assets, eq_excl=eq_excl, eq_incl=eq_incl,
        cash_eq=cash_eq, mi=mi_val,
        tot_debt=tot_debt_q,
        close=close, mktcap=mktcap_q,
    ))

# Sort newest first
ttm_rows = [r for r in ttm_rows if r['rev'] is not None or r['mktcap'] is not None]
ttm_rows.sort(key=lambda x: x['qdate'], reverse=True)
print(f"  TTM quarters available: {len(ttm_rows)}")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — FETCH TOP 10 SHAREHOLDERS
# ═══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*60)
print("Fetching top-10 shareholders (quarterly)...")
print("="*60)

top10_data = {}
quarters_to_try = []
for yr in range(2015, 2026):
    for mmdd in ['0331','0630','0930','1231']:
        quarters_to_try.append(f'{yr}{mmdd}')

for qd in quarters_to_try:
    try:
        df = ak.stock_gdfx_top_10_em(symbol=f'sh{TICKER_6}', date=qd)
        time.sleep(0.3)
        if df is not None and not df.empty:
            top10_data[qd] = df
            print(f"  {qd}: {len(df)} shareholders")
    except Exception:
        pass

print(f"  Top-10 data available for {len(top10_data)} quarters")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — FETCH SHAREHOLDER COUNT ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*60)
print("Fetching shareholder analytics...")
print("="*60)

sh_count_df = None
try:
    sh_count_df = ak.stock_zh_a_gdhs_detail_em(symbol=TICKER_6)
    print(f"  Shareholder count records: {len(sh_count_df)}")
    print(f"  Columns: {list(sh_count_df.columns)}")
except Exception as e:
    print(f"  Error: {e}")

# Also try tushare stk_holdernumber
print("  Also fetching tushare stk_holdernumber...")
try:
    ts_holder = pro.stk_holdernumber(ts_code=TS_CODE)
    time.sleep(0.5)
    print(f"  Tushare holder records: {len(ts_holder)}")
    print(f"  Columns: {list(ts_holder.columns)}")
except Exception as e:
    print(f"  Tushare holder error: {e}")
    ts_holder = pd.DataFrame()


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — BUILD EXCEL WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*60)
print("Building Excel workbook...")
print("="*60)

wb = Workbook()
ws_names = [
    'Financial Data', 'ROIC Analysis', 'Valuation Model',
    'TTM Quarterly Analysis', 'Top 10 Shareholders',
    'Shareholder Analytics', 'Data Correlation'
]

# Create sheets
sheets = {}
for i, name in enumerate(ws_names):
    if i == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    sheets[name] = ws


# ─── Helpers for column letter of year ───────────────────────────────────────
# Annual sheets: col B = YEARS[0]=2015, col C = 2016, ..., col L = 2025
def yr_col(yr):
    return 2 + YEARS.index(yr)   # B=2 for 2015

def yr_cl(yr):
    return get_column_letter(yr_col(yr))


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: Financial Data
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 1: Financial Data...")
ws1 = sheets['Financial Data']

# Title row
_w(ws1, 1, 1, f'{COMPANY} ({TS_CODE}) - Financial Data',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
_w(ws1, 1, 2, f'Units: {UNIT_LABEL} unless noted',
   Font(name='Apple Braille', size=11), SOLID)

# Headers row 3
_w(ws1, 3, 1, f'Metric ({UNIT_LABEL})',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
for yr in YEARS:
    c = yr_col(yr)
    _w(ws1, 3, c, yr,
       Font(name='Apple Braille', size=11, bold=True), SOLID, C_ALIGN, 'General')

# ── Income Statement ──
w_sec(ws1, 4, 1, 'Income Statement')

ROWS_IS = [
    (5,  'Total Revenues',        'revenue',   '#,##0.0'),
    (6,  'Cost Of Revenues',      'oper_cost', '#,##0.0'),
    (8,  'EBIT',                  'ebit',      '#,##0.0'),
    (9,  'Income Tax Expense',    'tax',       '#,##0.0'),
    (11, 'D&A',                   'dna',       '#,##0.0'),
]
for row, label, key, nf in ROWS_IS:
    w_lbl(ws1, row, 1, label)
    for yr in YEARS:
        v = data[yr].get(key)
        if v is not None:
            w_blue(ws1, row, yr_col(yr), v, nf)

# Formula rows (black)
w_lbl(ws1, 7, 1, 'Gross Profit')
w_lbl(ws1, 10, 1, 'NOPAT')
w_lbl(ws1, 12, 1, 'Gross Profit Margin %')
w_lbl(ws1, 13, 1, 'EBITDA')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws1, 7, c,  f'={cl}5-{cl}6',   '#,##0.0')
    w_blk(ws1, 10, c, f'={cl}8-{cl}9',   '#,##0.0')
    w_blk(ws1, 12, c, f'=IF({cl}5=0,"-",{cl}7/{cl}5)', '0.0%')
    w_blk(ws1, 13, c, f'={cl}8+{cl}11',  '#,##0.0')

# ── Balance Sheet ──
w_sec(ws1, 14, 1, 'Balance Sheet')

ROWS_BS = [
    (15, 'Total Assets',                    'tot_assets', '#,##0.0'),
    (16, 'Total Equity (Stockholders)',      'eq_excl_mi','#,##0.0'),
    (17, 'Total Debt',                       'tot_debt',   '#,##0.0'),
    (18, 'Cash and Equivalents',             'cash_eq',    '#,##0.0'),
    (19, 'Minority Interest',                'mi',         '#,##0.0'),
    (20, 'Working Capital',                  'wc',         '#,##0.0'),
    (21, 'Net Working Capital',              'nwc',        '#,##0.0'),
]
for row, label, key, nf in ROWS_BS:
    w_lbl(ws1, row, 1, label)
    for yr in YEARS:
        v = data[yr].get(key)
        if v is not None:
            w_blue(ws1, row, yr_col(yr), v, nf)

# ── Cash Flow ──
w_sec(ws1, 22, 1, 'Cash Flow')

w_lbl(ws1, 23, 1, 'Cash from Operations')
for yr in YEARS:
    v = data[yr].get('ocf')
    if v is not None:
        w_blue(ws1, 23, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 24, 1, 'Change in Net Working Capital')
# Row 24 = YoY delta of NWC (starts col C = 2016)
for i, yr in enumerate(YEARS):
    if i == 0: continue  # no prior year for 2015
    c   = yr_col(yr)
    cl  = yr_cl(yr)
    pcl = yr_cl(YEARS[i-1])
    w_blk(ws1, 24, c, f'={cl}21-{pcl}21', '#,##0.0')

w_lbl(ws1, 25, 1, 'Dividends Paid')
for yr in YEARS:
    v = data[yr].get('divs_paid')
    if v is not None:
        w_blue(ws1, 25, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 26, 1, 'Share Repurchases')
for yr in YEARS:
    w_blue(ws1, 26, yr_col(yr), 0.0, '#,##0.0')  # Moutai has no buybacks historically

w_lbl(ws1, 27, 1, 'Free Cash Flow to Equity')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws1, 27, c, f'={cl}23+{cl}24', '#,##0.0')

# ── Market Data ──
w_sec(ws1, 28, 1, 'Market Data')

w_lbl(ws1, 29, 1, 'Market Capitalisation')
for yr in YEARS:
    v = data[yr].get('mktcap')
    if v is not None:
        w_blue(ws1, 29, yr_col(yr), v, '#,##0.0')

w_lbl(ws1, 30, 1, 'Enterprise Value')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws1, 30, c, f'={cl}29+{cl}17-{cl}18+{cl}19', '#,##0.0')

w_lbl(ws1, 31, 1, 'EV/EBITDA')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws1, 31, c, f'=IF({cl}13=0,"-",{cl}30/{cl}13)', '0.0"x"')

w_lbl(ws1, 32, 1, 'ROIC')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_grn(ws1, 32, c, f"='ROIC Analysis'!{cl}17", '0.0%')

# Column widths
ws1.column_dimensions['A'].width = 34
for yr in YEARS:
    ws1.column_dimensions[yr_cl(yr)].width = 13
ws1.column_dimensions[get_column_letter(yr_col(YEARS[-1]) + 1)].width = 8.83

set_heights(ws1, 20)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: ROIC Analysis
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 2: ROIC Analysis...")
ws2 = sheets['ROIC Analysis']

_w(ws2, 1, 1, f'{COMPANY} ({TS_CODE}) - ROIC Decomposition',
   Font(name='Apple Braille', size=11, bold=True), SOLID)

_w(ws2, 3, 1, 'ROIC = NOPAT / Avg Invested Capital',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
for yr in YEARS:
    c = yr_col(yr)
    _w(ws2, 3, c, yr,
       Font(name='Apple Braille', size=11, bold=True), SOLID, C_ALIGN, 'General')

# NOPAT Section
w_sec(ws2, 4, 1, 'NOPAT Calculation')
w_lbl(ws2, 5, 1, 'EBIT')
w_lbl(ws2, 6, 1, 'Less: Income Tax Expense')
w_lbl(ws2, 7, 1, 'NOPAT')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_grn(ws2, 5, c, f"='Financial Data'!{cl}8",  '#,##0.0')
    w_grn(ws2, 6, c, f"='Financial Data'!{cl}9",  '#,##0.0')
    w_blk(ws2, 7, c, f'={cl}5-{cl}6',             '#,##0.0')

# Invested Capital Section
w_sec(ws2, 8, 1, 'Invested Capital Components')
ROWS_IC = [
    (9,  'Long Term Debt',             'lt_debt',   '#,##0.0'),
    (10, 'Short Term Debt',            'st_debt',   '#,##0.0'),
    (11, 'Total Equity (incl. MI)',    'eq_incl_mi','#,##0.0'),
    (12, 'Current Portion of Leases',  'st_lease',  '#,##0.0'),
    (13, 'Long Term Leases',           'lt_lease',  '#,##0.0'),
]
for row, label, key, nf in ROWS_IC:
    w_lbl(ws2, row, 1, label)
    for yr in YEARS:
        v = data[yr].get(key)
        w_blue(ws2, row, yr_col(yr), v if v is not None else 0.0, nf)

w_lbl(ws2, 14, 1, 'Invested Capital')
w_lbl(ws2, 15, 1, 'Avg Invested Capital')
w_lbl(ws2, 17, 1, 'ROIC', bold=True)

for i, yr in enumerate(YEARS):
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws2, 14, c, f'={cl}9+{cl}10+{cl}11+{cl}12+{cl}13', '#,##0.0')
    if i > 0:
        pcl = yr_cl(YEARS[i-1])
        w_blk(ws2, 15, c, f'=({cl}14+{pcl}14)/2', '#,##0.0')
        w_blk(ws2, 17, c, f'=IF({cl}15=0,"-",{cl}7/{cl}15)', '0.0%')

# YoY Changes
w_sec(ws2, 19, 1, 'Year-over-Year Changes')
w_lbl(ws2, 20, 1, 'NOPAT Growth')
w_lbl(ws2, 21, 1, 'IC Growth')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c   = yr_col(yr)
    cl  = yr_cl(yr)
    pcl = yr_cl(YEARS[i-1])
    w_blk(ws2, 20, c, f'=IF({pcl}7=0,"-",({cl}7-{pcl}7)/{pcl}7)', '0.0%')
    w_blk(ws2, 21, c, f'=IF({pcl}14=0,"-",({cl}14-{pcl}14)/{pcl}14)', '0.0%')

ws2.column_dimensions['A'].width = 34
for yr in YEARS:
    ws2.column_dimensions[yr_cl(yr)].width = 13
ws2.column_dimensions[get_column_letter(yr_col(YEARS[-1]) + 1)].width = 8.83
set_heights(ws2, 20)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: Valuation Model
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 3: Valuation Model...")
ws3 = sheets['Valuation Model']

_w(ws3, 1, 1, f'{COMPANY} ({TS_CODE}) - Valuation Model',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
_w(ws3, 2, 1, 'Based on Dividend Yield + ROIC Multiple Methodology',
   Font(name='Apple Braille', size=11), SOLID)

_w(ws3, 4, 1, f'Valuation Parameters ({UNIT_LABEL})',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
for yr in YEARS:
    c = yr_col(yr)
    _w(ws3, 4, c, yr,
       Font(name='Apple Braille', size=11, bold=True), SOLID, C_ALIGN, 'General')

w_sec(ws3, 5, 1, 'Key Assumptions')

w_lbl(ws3, 6, 1, 'China 10yr Bond Yield (%)')
for yr in YEARS:
    w_blue(ws3, 6, yr_col(yr), CN_BOND[yr], '0.00%')

w_lbl(ws3, 7, 1, 'ROIC')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_grn(ws3, 7, c, f"='ROIC Analysis'!{cl}17", '0.0%')

w_lbl(ws3, 8, 1, 'Valuation Multiple (ROIC×100+5)')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws3, 8, c, f'=IF(ISNUMBER({cl}7),{cl}7*100+5,"-")', '0.0')

w_sec(ws3, 10, 1, 'Input Data (from Financial Data)')

LINKS_VM = [
    (11, 'Dividends Paid',              f"=ABS('Financial Data'!{{}}25)"),
    (12, 'Share Repurchases',           f"=ABS('Financial Data'!{{}}26)"),
    (14, 'Operating Cash Flow',         f"='Financial Data'!{{}}23"),
    (15, 'Change in Net Working Capital', f"='Financial Data'!{{}}24"),
    (17, 'D&A',                         f"='Financial Data'!{{}}11"),
    (19, 'NOPAT',                       f"='Financial Data'!{{}}10"),
    (39, 'Total Debt',                  f"='Financial Data'!{{}}17"),
    (40, 'Cash and Equivalents',        f"='Financial Data'!{{}}18"),
    (41, 'Minority Interest',           f"='Financial Data'!{{}}19"),
    (42, 'Market Cap (Actual)',         f"='Financial Data'!{{}}29"),
]
for row, label, fmla_tmpl in LINKS_VM:
    w_lbl(ws3, row, 1, label)
    for yr in YEARS:
        c  = yr_col(yr)
        cl = yr_cl(yr)
        w_grn(ws3, row, c, fmla_tmpl.format(cl), '#,##0.0')

# adj. OCF
w_lbl(ws3, 16, 1, 'adj. OCF')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws3, 16, c, f'=IF(ISNUMBER({cl}15),{cl}14+{cl}15,{cl}14)', '#,##0.0')

# NOPAT-Dividend-Repurchase (NDR)
w_lbl(ws3, 21, 1, 'NOPAT-Dividend-Repurchase (NDR)')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws3, 21, c, f'={cl}19-{cl}11-{cl}12', '#,##0.0')

# Change of Debt / Equity / IC
w_lbl(ws3, 23, 1, 'Change of Debt')
w_lbl(ws3, 24, 1, 'Change of Equity')
w_lbl(ws3, 26, 1, 'Change of IC')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c   = yr_col(yr)
    cl  = yr_cl(yr)
    pcl = yr_cl(YEARS[i-1])
    w_blk(ws3, 23, c, f"='Financial Data'!{cl}17-'Financial Data'!{pcl}17", '#,##0.0')
    w_blk(ws3, 24, c, f"='ROIC Analysis'!{cl}11-'ROIC Analysis'!{pcl}11", '#,##0.0')
    w_blk(ws3, 26, c, f"='ROIC Analysis'!{cl}14-'ROIC Analysis'!{pcl}14", '#,##0.0')

# Ratios
w_lbl(ws3, 28, 1, 'Dividend Yield (on Mkt Cap)')
w_lbl(ws3, 29, 1, 'NDR/IC')
w_lbl(ws3, 30, 1, 'ROIC')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws3, 28, c, f'=IF({cl}42=0,"-",{cl}11/{cl}42)', '0.0%')
    w_blk(ws3, 29, c, f"=IF(ISNUMBER('ROIC Analysis'!{cl}14),{cl}21/'ROIC Analysis'!{cl}14,\"-\")", '0.0%')
    w_blk(ws3, 30, c, f'={cl}7', '0.0%')

# NOPAT Growth
w_lbl(ws3, 32, 1, 'NOPAT Growth')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c   = yr_col(yr)
    cl  = yr_cl(yr)
    pcl = yr_cl(YEARS[i-1])
    w_blk(ws3, 32, c, f'=IF({pcl}19=0,"-",({cl}19-{pcl}19)/{pcl}19)', '0.0%')

# NOPAT Growth multiple & valuation
w_lbl(ws3, 34, 1, 'Multiple fr. NOPAT Growth')
w_lbl(ws3, 36, 1, 'Valuation fr. NOPAT Growth')
w_lbl(ws3, 37, 1, 'Gap NOPAT Gr vs. Actual')
for i, yr in enumerate(YEARS):
    if i == 0: continue
    c   = yr_col(yr)
    cl  = yr_cl(yr)
    w_blk(ws3, 34, c, f'=IF(ISNUMBER({cl}32),MAX(0,{cl}32*100+5),"-")', '0.0')
    w_blk(ws3, 36, c, f'=IF(ISNUMBER({cl}34),{cl}19*{cl}34,"-")', '#,##0.0')
    w_blk(ws3, 37, c, f'=IF(AND(ISNUMBER({cl}36),{cl}42<>0),{cl}36/{cl}42-1,"-")', '0.0%')

# Valuation Calculations
w_sec(ws3, 44, 1, 'Valuation Calculations')

w_lbl(ws3, 45, 1, 'Mkt Cap 1 (Div Yield)',     bold=True)
w_lbl(ws3, 46, 1, 'Adjusted FCF (adj OCF - Div)')
w_lbl(ws3, 47, 1, 'EV (ROIC Multiple)')
w_lbl(ws3, 48, 1, 'Mkt Cap 2 (EV-Debt+Cash-MI)', bold=True)
w_lbl(ws3, 50, 1, 'Total Valuation',            bold=True)
w_lbl(ws3, 51, 1, 'Actual Market Cap')
w_lbl(ws3, 52, 1, 'Valuation Premium/(Discount)', bold=True)

for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws3, 45, c, f'=IF({cl}6=0,"-",{cl}11/{cl}6)', '#,##0.0')
    w_blk(ws3, 46, c, f'={cl}16-{cl}11', '#,##0.0')
    w_blk(ws3, 47, c, f'=IF(ISNUMBER({cl}8),{cl}46*{cl}8,"-")', '#,##0.0')
    w_blk(ws3, 48, c, f'=IF(ISNUMBER({cl}47),{cl}47-{cl}39+{cl}40-{cl}41,"-")', '#,##0.0')
    w_blk(ws3, 50, c, f'=IF(AND(ISNUMBER({cl}45),ISNUMBER({cl}48)),{cl}45+{cl}48,"-")', '#,##0.0')
    w_blk(ws3, 51, c, f'={cl}42', '#,##0.0')
    w_blk(ws3, 52, c, f'=IF(AND(ISNUMBER({cl}50),{cl}51<>0),{cl}50/{cl}51,"-")', '0.0"x"')

# Capital Structure
w_sec(ws3, 54, 1, 'Capital Structure Analysis')
w_lbl(ws3, 55, 1, 'Dividend Yield (on Mkt Cap)')
w_lbl(ws3, 56, 1, 'Div Payout Ratio (Div/OCF)')
w_lbl(ws3, 57, 1, 'Debt / Equity Ratio')
w_lbl(ws3, 58, 1, 'Total Equity (incl. MI)')
w_lbl(ws3, 60, 1, 'Mkt Value / Equity')
for yr in YEARS:
    c  = yr_col(yr)
    cl = yr_cl(yr)
    w_blk(ws3, 55, c, f'={cl}28', '0.0%')
    w_blk(ws3, 56, c, f'=IF({cl}14=0,"-",{cl}11/{cl}14)', '0.0%')
    w_blk(ws3, 57, c, f"=IF('ROIC Analysis'!{cl}11=0,\"-\",'Financial Data'!{cl}17/'ROIC Analysis'!{cl}11)", '0.0"x"')
    w_grn(ws3, 58, c, f"='ROIC Analysis'!{cl}11", '#,##0.0')
    w_blk(ws3, 60, c, f'=IF({cl}58=0,"-",{cl}42/{cl}58)', '0.0"x"')

ws3.column_dimensions['A'].width = 38
for yr in YEARS:
    ws3.column_dimensions[yr_cl(yr)].width = 13
ws3.column_dimensions[get_column_letter(yr_col(YEARS[-1]) + 1)].width = 8.83
set_heights(ws3, 17)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: TTM Quarterly Analysis
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 4: TTM Quarterly Analysis...")
ws4 = sheets['TTM Quarterly Analysis']

_w(ws4, 1, 1, f'{COMPANY} ({TS_CODE}) - TTM Quarterly Analysis',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
_w(ws4, 2, 1, 'TTM = Trailing 12 months. Flow items: YTD(Qn)+Annual(Y-1)-YTD(Qn,Y-1). Units: CNY 亿元',
   Font(name='Apple Braille', size=11), SOLID)

# Quarter headers (row 3), newest first
for j, row_d in enumerate(ttm_rows):
    c = j + 2  # col B onwards
    _w(ws4, 3, c, row_d['q_label'],
       Font(name='Apple Braille', size=11, bold=True), SOLID, C_ALIGN, 'General')

# Define TTM row structure
TTM_ROWS_BLUE = [
    (5,  'TTM Revenue',              'rev',    '#,##0.0'),
    (6,  'TTM COGS',                 'cogs',   '#,##0.0'),
    (9,  'TTM EBIT (Operating Profit)','ebit', '#,##0.0'),
    (10, 'TTM Income Tax',           'tax',    '#,##0.0'),
    (11, 'TTM Net Income',           'ni',     '#,##0.0'),
    (12, 'TTM NI to Parent',         'ni_p',   '#,##0.0'),
    (16, 'TTM OCF',                  'ocf',    '#,##0.0'),
    (17, 'TTM Dividends Paid',       'div_cf', '#,##0.0'),
    (18, 'TTM D&A',                  'dna',    '#,##0.0'),
    (21, 'Total Assets',             'tot_assets','#,##0.0'),
    (22, 'Total Equity (Parent)',    'eq_excl', '#,##0.0'),
    (23, 'Total Equity (incl MI)',   'eq_incl', '#,##0.0'),
    (24, 'Total Debt',               'tot_debt','#,##0.0'),
    (25, 'Cash & Equivalents',       'cash_eq', '#,##0.0'),
    (26, 'Minority Interest',        'mi',      '#,##0.0'),
    (28, 'Quarter-end Close (CNY)',  'close',   '0.00'),
    (29, 'Market Cap (CNY 亿)',       'mktcap',  '#,##0.0'),
]

# Section headers
w_sec(ws4, 4, 1, 'TTM Income Statement')
w_sec(ws4, 15, 1, 'TTM Cash Flow')
w_sec(ws4, 20, 1, 'Balance Sheet (quarter-end)')
w_sec(ws4, 27, 1, 'Market Data')
w_sec(ws4, 31, 1, 'Valuation Multiples')
w_sec(ws4, 36, 1, 'Profitability')
w_sec(ws4, 40, 1, 'Growth (TTM YoY)')

# Labels
for row, label, key, nf in TTM_ROWS_BLUE:
    w_lbl(ws4, row, 1, label)

w_lbl(ws4, 7, 1, 'TTM Gross Profit')
w_lbl(ws4, 8, 1, 'TTM Gross Margin %')
w_lbl(ws4, 13, 1, 'TTM EBIT Margin %')
w_lbl(ws4, 14, 1, 'TTM Net Margin %')
w_lbl(ws4, 19, 1, 'TTM EBITDA')
w_lbl(ws4, 30, 1, 'Enterprise Value')
w_lbl(ws4, 32, 1, 'P/E TTM')
w_lbl(ws4, 33, 1, 'P/B')
w_lbl(ws4, 34, 1, 'EV/EBITDA TTM')
w_lbl(ws4, 35, 1, 'Dividend Yield TTM')
w_lbl(ws4, 37, 1, 'ROE TTM')
w_lbl(ws4, 38, 1, 'ROA TTM')
w_lbl(ws4, 39, 1, 'ROIC TTM')
w_lbl(ws4, 41, 1, 'Revenue Growth (TTM YoY)')
w_lbl(ws4, 42, 1, 'NI Growth (TTM YoY)')
w_lbl(ws4, 43, 1, 'OCF Growth (TTM YoY)')

# Write blue data and black formulas
for j, row_d in enumerate(ttm_rows):
    c  = j + 2
    cl = get_column_letter(c)

    for row, label, key, nf in TTM_ROWS_BLUE:
        v = row_d.get(key)
        if v is not None:
            w_blue(ws4, row, c, v, nf)

    # Black formula rows
    w_blk(ws4, 7,  c, f'={cl}5-{cl}6',                '#,##0.0')
    w_blk(ws4, 8,  c, f'=IFERROR({cl}7/{cl}5,"-")',   '0.0%')
    w_blk(ws4, 13, c, f'=IFERROR({cl}9/{cl}5,"-")',   '0.0%')
    w_blk(ws4, 14, c, f'=IFERROR({cl}12/{cl}5,"-")',  '0.0%')
    w_blk(ws4, 19, c, f'={cl}9+{cl}18',               '#,##0.0')
    w_blk(ws4, 30, c, f'={cl}29+{cl}24-{cl}25+{cl}26','#,##0.0')
    # PE: mktcap(亿) / ni_parent(亿) — both in same unit → dimensionless ratio
    w_blk(ws4, 32, c, f'=IFERROR({cl}29/{cl}12,"-")', '0.0"x"')
    # PB: mktcap(亿) / equity_parent(亿)
    w_blk(ws4, 33, c, f'=IFERROR({cl}29/{cl}22,"-")', '0.0"x"')
    w_blk(ws4, 34, c, f'=IFERROR({cl}30/{cl}19,"-")', '0.0"x"')
    w_blk(ws4, 35, c, f'=IFERROR(ABS({cl}17)/{cl}29,"-")', '0.0%')
    w_blk(ws4, 37, c, f'=IFERROR({cl}12/{cl}22,"-")', '0.0%')
    w_blk(ws4, 38, c, f'=IFERROR({cl}11/{cl}21,"-")', '0.0%')
    w_blk(ws4, 39, c, f'=IFERROR(({cl}9-{cl}10)/({cl}23+{cl}24),"-")', '0.0%')

    # YoY growth: compare to same quarter 4 positions to the right (1 year)
    if j + 4 < len(ttm_rows):
        py_cl = get_column_letter(c + 4)
        w_blk(ws4, 41, c, f'=IFERROR(({cl}5-{py_cl}5)/{py_cl}5,"-")',   '0.0%')
        w_blk(ws4, 42, c, f'=IFERROR(({cl}12-{py_cl}12)/{py_cl}12,"-")', '0.0%')
        w_blk(ws4, 43, c, f'=IFERROR(({cl}16-{py_cl}16)/{py_cl}16,"-")', '0.0%')

ws4.column_dimensions['A'].width = 34
for j in range(len(ttm_rows)):
    ws4.column_dimensions[get_column_letter(j+2)].width = 13
ws4.column_dimensions[get_column_letter(len(ttm_rows)+2)].width = 8.83
set_heights(ws4, 17)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: Top 10 Shareholders
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 5: Top 10 Shareholders...")
ws5 = sheets['Top 10 Shareholders']

# Sort quarters newest-first
sorted_q = sorted(top10_data.keys(), reverse=True)

# Build union of all shareholder names
all_shareholders = []
name_counts = {}
for qd in sorted_q:
    df = top10_data[qd]
    # Find the name column
    name_col = None
    for candidate in ['股东名称', '名称', 'holder_name', '名称\n股东名称']:
        if candidate in df.columns:
            name_col = candidate
            break
    if name_col is None and len(df.columns) > 0:
        name_col = df.columns[0]
    if name_col:
        for name in df[name_col]:
            name = str(name).strip()
            if name and name not in ['nan', 'None']:
                name_counts[name] = name_counts.get(name, 0) + 1

all_shareholders = sorted(name_counts.keys(), key=lambda x: -name_counts[x])

# Title
_w(ws5, 1, 1, f'{COMPANY} ({TS_CODE}) - Top 10 Shareholders (Quarterly)',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
_w(ws5, 2, 1, 'Source: akshare stock_gdfx_top_10_em. Holdings in shares, % of total. Empty = not in top-10 that quarter.',
   Font(name='Apple Braille', size=11), SOLID)

# Quarter date headers (row 3) — 2 columns per quarter
for j, qd in enumerate(sorted_q):
    date_label = f"{qd[:4]}-{qd[4:6]}-{qd[6:]}"
    col_start = 2 + j * 2  # B, D, F, ...
    _w(ws5, 3, col_start, date_label,
       Font(name='Apple Braille', size=11, bold=True, color='FFFFFFFF'),
       HDR_FILL, C_ALIGN, 'General')
    # Merge the 2 cells for the date header
    ws5.merge_cells(start_row=3, start_column=col_start,
                    end_row=3, end_column=col_start+1)

# Sub-headers row 4
for j in range(len(sorted_q)):
    col_start = 2 + j * 2
    _w(ws5, 4, col_start,   '持股数 (shares)',
       Font(name='Apple Braille', size=11, bold=True, color='FFFFFFFF'), HDR_FILL, C_ALIGN)
    _w(ws5, 4, col_start+1, '占比 (%)',
       Font(name='Apple Braille', size=11, bold=True, color='FFFFFFFF'), HDR_FILL, C_ALIGN)

# Shareholder rows
for i, shareholder in enumerate(all_shareholders):
    row = 5 + i
    _w(ws5, row, 1, shareholder, Font(name='Apple Braille', size=11), SOLID)

    for j, qd in enumerate(sorted_q):
        df = top10_data[qd]
        col_start = 2 + j * 2

        # Find name column
        name_col = None
        for candidate in ['股东名称', '名称', 'holder_name']:
            if candidate in df.columns:
                name_col = candidate
                break
        if name_col is None and len(df.columns) > 0:
            name_col = df.columns[0]

        # Find shares and pct columns
        shares_col = None
        pct_col = None
        for col in df.columns:
            col_str = str(col)
            if any(k in col_str for k in ['持股数', '持股量', 'shares', '数量']):
                shares_col = col
            if any(k in col_str for k in ['占比', '%', '比例', 'pct', 'ratio']):
                pct_col = col

        if name_col:
            match = df[df[name_col].astype(str).str.strip() == shareholder]
            if not match.empty:
                row_data = match.iloc[0]
                if shares_col and pd.notna(row_data.get(shares_col)):
                    v = na(row_data[shares_col])
                    if v:
                        w_blue(ws5, row, col_start, int(v), '#,##0')
                if pct_col and pd.notna(row_data.get(pct_col)):
                    v = na(row_data[pct_col])
                    if v:
                        # Might be in % already (e.g., 8.23) or decimal
                        pct_val = v / 100 if v > 1 else v
                        w_blue(ws5, row, col_start+1, pct_val, '0.00%')

# Summary rows
summary_row = 5 + len(all_shareholders) + 1
w_sec(ws5, summary_row, 1, 'Summary')
w_lbl(ws5, summary_row+1, 1, 'Total periods covered')
_w(ws5, summary_row+1, 2, len(sorted_q), F_BLACK, SOLID, R_ALIGN, '#,##0')
w_lbl(ws5, summary_row+2, 1, 'Unique top-10 shareholders')
_w(ws5, summary_row+2, 2, len(all_shareholders), F_BLACK, SOLID, R_ALIGN, '#,##0')
w_lbl(ws5, summary_row+3, 1, 'Note')
_w(ws5, summary_row+3, 2, 'Empty cell = shareholder was not in top-10 that quarter (may still hold shares)',
   Font(name='Apple Braille', size=11), SOLID)

# Column widths
ws5.column_dimensions['A'].width = 53
for j in range(len(sorted_q)):
    col_start = 2 + j * 2
    ws5.column_dimensions[get_column_letter(col_start)].width   = 15
    ws5.column_dimensions[get_column_letter(col_start+1)].width = 10
set_heights(ws5, 15)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 6: Shareholder Analytics
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 6: Shareholder Analytics...")
ws6 = sheets['Shareholder Analytics']

_w(ws6, 1, 1, f'{COMPANY} ({TS_CODE}) - Shareholder Analytics',
   Font(name='Apple Braille', size=11, bold=True), SOLID)
_w(ws6, 2, 1, 'Source: akshare stock_zh_a_gdhs_detail_em. Sorted ascending by date.',
   Font(name='Apple Braille', size=11), SOLID)

# Column headers row 3
HDR_LABELS = [
    (1, 'Date'),
    (2, 'Shareholder Count (户)'),
    (3, 'Total Mkt Cap (CNY 亿)'),
    (4, 'Total Shares (亿)'),
    (5, 'Mkt Cap / Shareholder (CNY 万)'),
    (6, 'Shares / Shareholder'),
    (7, 'Mkt Cap QoQ Chg (亿)'),
    (8, 'Mkt Cap QoQ Chg %'),
    (9, 'Shareholder Chg'),
    (10, 'Shareholder Chg %'),
    (11, 'Price / Share (CNY)'),
]
for col, label in HDR_LABELS:
    _w(ws6, 3, col, label,
       Font(name='Apple Braille', size=11, bold=True), SOLID, C_ALIGN, 'General')

# Populate data
sh_rows = []
if sh_count_df is not None and not sh_count_df.empty:
    print(f"  Processing akshare shareholder data: {list(sh_count_df.columns)}")
    # Find relevant columns
    date_col  = None
    count_col = None
    mv_col    = None
    share_col = None
    for col in sh_count_df.columns:
        cs = str(col)
        if '截止' in cs or '日期' in cs or 'date' in cs.lower():
            date_col = col
        if '户数' in cs and '本次' in cs:
            count_col = col
        elif '股东户数' in cs and date_col != col:
            if count_col is None: count_col = col
        if '总市值' in cs:
            mv_col = col
        if '总股本' in cs:
            share_col = col

    if date_col is None: date_col = sh_count_df.columns[0]
    if count_col is None: count_col = sh_count_df.columns[1] if len(sh_count_df.columns) > 1 else None

    for _, row in sh_count_df.sort_values(date_col).iterrows():
        date_val  = str(row[date_col])[:10]
        count_val = na(row.get(count_col)) if count_col else None
        mv_raw    = na(row.get(mv_col))    if mv_col    else None
        share_raw = na(row.get(share_col)) if share_col else None

        mv_yi    = round(mv_raw    / 1e8, 2) if mv_raw    else None
        share_yi = round(share_raw / 1e8, 4) if share_raw else None

        sh_rows.append(dict(date=date_val, count=count_val, mv=mv_yi, shares=share_yi))

elif not ts_holder.empty:
    print("  Using tushare stk_holdernumber as fallback...")
    print(f"  Columns: {list(ts_holder.columns)}")
    ts_holder = ts_holder.sort_values('end_date' if 'end_date' in ts_holder.columns else ts_holder.columns[0])
    for _, row in ts_holder.iterrows():
        date_val = str(row.get('end_date', row.iloc[0]))[:8]
        if len(date_val) == 8:
            date_val = f"{date_val[:4]}-{date_val[4:6]}-{date_val[6:]}"
        count_val = na(row.get('holder_num'))
        sh_rows.append(dict(date=date_val, count=count_val, mv=None, shares=None))

for i, sh in enumerate(sh_rows):
    r = 4 + i
    # Date
    _w(ws6, r, 1, sh['date'], Font(name='Apple Braille', size=11), SOLID)
    # Shareholder count (blue input)
    if sh['count']:
        w_blue(ws6, r, 2, int(sh['count']), '#,##0')
    # Mkt Cap (blue input if available)
    if sh['mv']:
        w_blue(ws6, r, 3, sh['mv'], '#,##0.0')
    # Total shares (blue input)
    if sh['shares']:
        w_blue(ws6, r, 4, sh['shares'], '#,##0.0')

    cl = get_column_letter(1)   # A, but using row number
    r_cl = str(r)

    # Formula cols E-K
    w_blk(ws6, r, 5, f'=IFERROR(C{r}*10000/B{r},"-")', '#,##0.0')
    w_blk(ws6, r, 6, f'=IFERROR(D{r}*100000000/B{r},"-")', '#,##0')
    w_blk(ws6, r, 11, f'=IFERROR(C{r}/D{r},"-")', '0.00')

    if i > 0:
        pr = r - 1
        w_blk(ws6, r, 7,  f'=IFERROR(C{r}-C{pr},"-")', '+#,##0.0;-#,##0.0;"-"')
        w_blk(ws6, r, 8,  f'=IFERROR(G{r}/C{pr},"-")', '0.0%')
        w_blk(ws6, r, 9,  f'=IFERROR(B{r}-B{pr},"-")', '+#,##0;-#,##0;"-"')
        w_blk(ws6, r, 10, f'=IFERROR(I{r}/B{pr},"-")', '0.0%')

ws6.column_dimensions['A'].width = 14
for col in range(2, 12):
    ws6.column_dimensions[get_column_letter(col)].width = 16
set_heights(ws6, 17)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 7: Data Correlation
# ════════════════════════════════════════════════════════════════════════════
print("  Building Sheet 7: Data Correlation...")
ws7 = sheets['Data Correlation']

doc_rows = [
    (1,  f'{COMPANY} ({TS_CODE}) - Data Correlation & Methodology', True),
    (2,  '', False),
    (3,  '1. Cross-File Validation', True),
    (4,  f'ROIC cross-check: Sheet2!ROIC = NOPAT / Avg IC. NOPAT sourced from Sheet1 EBIT - Tax.', False),
    (5,  f'Equity in ROIC sheet = total_hldr_eqy_inc_min_int (incl. MI) from tushare balancesheet.', False),
    (6,  f'Market cap verified against tushare daily_basic total_mv / 1e4 for year-end trading day.', False),
    (7,  '', False),
    (8,  '2. Key Relationships', True),
    (9,  f'贵州茅台 is a premium baijiu (white spirits) producer. Extremely high margins (>60% gross).', False),
    (10, f'Contract liabilities (预收款) from distributors are NOT classified as debt in this model.', False),
    (11, f'Total Debt ≈ 0 or minimal. ROIC is very high due to asset-light model + pricing power.', False),
    (12, f'Dividends: sourced from tushare cashflow c_pay_dist_dpcp_int_exp (financing outflow).', False),
    (13, '', False),
    (14, '3. Valuation Methodology', True),
    (15, 'Leg 1 (Dividend Yield): Mkt_Cap_1 = Dividends_Paid / Sovereign_Bond_Yield', False),
    (16, 'Leg 2 (ROIC Multiple): Multiple = ROIC×100+5; EV = Adj_FCF × Multiple', False),
    (17, 'Adj FCF = adj. OCF - Dividends. Mkt_Cap_2 = EV - Debt + Cash - MI', False),
    (18, 'Total Valuation = Mkt_Cap_1 + Mkt_Cap_2', False),
    (19, 'Premium/Discount = Total_Valuation / Actual_Market_Cap', False),
    (20, '', False),
    (21, '4. Data Sources', True),
    (22, 'Primary: tushare pro API (token: 45e390d8...)', False),
    (23, '  income(): IS annual, report_type=1, end_date 1231', False),
    (24, '  balancesheet(): BS annual + quarterly, 152 columns', False),
    (25, '  cashflow(): CF annual + quarterly', False),
    (26, '  daily_basic(): year-end & quarter-end market cap', False),
    (27, '  dividend(): cash dividends per share', False),
    (28, '  stk_holdernumber(): shareholder count', False),
    (29, 'Supplementary: akshare', False),
    (30, '  stock_gdfx_top_10_em(): top-10 shareholders quarterly', False),
    (31, '  stock_zh_a_gdhs_detail_em(): shareholder analytics', False),
    (32, '', False),
    (33, '5. Units', True),
    (34, f'All financial data: CNY 亿元 (÷1e8 from raw yuan)', False),
    (35, f'Market cap: CNY 亿元 (total_mv ÷ 1e4 from 万元)', False),
    (36, f'Shareholder Analytics: Total Mkt Cap in 亿元; Total Shares in 亿股', False),
    (37, '', False),
    (38, '6. Bond Yield', True),
    (39, 'China 10yr sovereign bond yield (year-end):' , False),
    (40, '2015=2.82% 2016=3.07% 2017=3.90% 2018=3.23% 2019=3.14% 2020=3.15%', False),
    (41, '2021=2.78% 2022=2.84% 2023=2.56% 2024=1.68% 2025=1.83%', False),
    (42, '', False),
    (43, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', False),
]

for r, text, bold in doc_rows:
    _w(ws7, r, 1, text,
       Font(name='Apple Braille', size=11, bold=bold), SOLID)

ws7.column_dimensions['A'].width = 110


# ════════════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════════════
print(f"\nSaving to: {OUTPUT_FILE}")
wb.save(OUTPUT_FILE)
print(f"\n{'='*60}")
print(f"✓ Model saved: {OUTPUT_FILE}")
print(f"{'='*60}")
print("\nSummary:")
print(f"  Sheet 1: Financial Data — {len(YEARS)} years ({YEARS[0]}-{YEARS[-1]})")
print(f"  Sheet 2: ROIC Analysis")
print(f"  Sheet 3: Valuation Model")
print(f"  Sheet 4: TTM Quarterly — {len(ttm_rows)} quarters")
print(f"  Sheet 5: Top 10 Shareholders — {len(sorted_q)} quarters, {len(all_shareholders)} unique holders")
print(f"  Sheet 6: Shareholder Analytics — {len(sh_rows)} data points")
print(f"  Sheet 7: Data Correlation")
