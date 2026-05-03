#!/usr/bin/env python3
"""
TOTO LTD (5332.T) — 8-sheet financial model.
yfinance only provides FY2022-FY2025 (4 years). FY2015-FY2021 left BLUE blank,
to be populated from EDINET annual report PDFs (separate download script).

Year convention: column header = FY-end calendar year (e.g. 2025 = FY ending 2025-03-31).
Currency: JPY billions (¥B = ¥10^9). Initial DRIP investment defaults to ¥10,000,000.
Withholding tax: 15.315% (Japan domestic incl. reconstruction surtax; no HK-Japan DTA reduction).
"""
import os
from datetime import datetime
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TICKER = "5332.T"
COMPANY = "TOTO LTD."
CCY = "JPY"
UNIT_LABEL = "JPY B (¥10^9)"
UNIT_DIV = 1e9
YEARS = list(range(2015, 2026))   # FY ending 3/31, labeled by ending year
TODAY = datetime.now().strftime('%y%m%d')
OUT_DIR = os.path.expanduser(f'~/Desktop/BUI_Investment_Management/{TICKER}')
os.makedirs(OUT_DIR, exist_ok=True)
OUT_FILE = os.path.join(OUT_DIR, f'{TICKER}.Valuation.v{TODAY}.xlsx')

# Japan 10y JGB year-end yields
JP_BOND = {2015:0.0027, 2016:-0.0001, 2017:0.0050, 2018:0.0001, 2019:-0.0007,
           2020:0.0002, 2021:0.0007, 2022:0.0042, 2023:0.0061, 2024:0.0107, 2025:0.0145}

# Japan M2 (¥ trillion, year-end, BoJ)
JP_M2 = {2015:889.4, 2016:931.6, 2017:976.7, 2018:1012.2, 2019:1041.4,
         2020:1138.4, 2021:1179.5, 2022:1194.0, 2023:1212.1, 2024:1232.0, 2025:1260.0}

# Hardcoded historical shares (M) — TOTO very stable around 164M; adjust if buybacks
SHARES_M = {y: 164.41 for y in YEARS}

# DRIP defaults
DRIP_INVEST = 10_000_000     # ¥10M (HKD ~600K equiv)
DRIP_FX = 1.0                 # already in JPY
DRIP_WHT = 0.15315            # 15% + 0.315% reconstruction surtax (foreign no-DTA holder)
DRIP_HOLDER = "HK Corporate (no Japan-HK DTA, 15.315% WHT)"

# ===== Format =====
FONT = "Apple Braille"; SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY, LGREY = "FFE7E6E6", "FFF2F2F2"
THIN = Side(style="thin", color="FF000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = '#,##0.0'
INT = '#,##0'
PCT = '0.00%'
PCT1 = '0.0%'
PX = '#,##0.00'
X = '0.00"x"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False, lgrey=False, font_size=SIZE):
    c.font = Font(name=FONT, size=font_size, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if num_fmt: c.number_format = num_fmt
    if grey: c.fill = PatternFill("solid", fgColor=GREY)
    elif lgrey: c.fill = PatternFill("solid", fgColor=LGREY)

def col_l(i): return get_column_letter(2+i)

# ===== Fetch yfinance data =====
print(f"\n{'='*60}\nFetching {TICKER} from yfinance\n{'='*60}")
yt = yf.Ticker(TICKER)
inc_df = yt.income_stmt
bs_df = yt.balance_sheet
cf_df = yt.cashflow
hist = yt.history(start="2014-01-01", end="2026-05-15", auto_adjust=False)
info = yt.info or {}

print(f"  income_stmt: {inc_df.shape}, cols (FY-end): {[str(c.date()) for c in inc_df.columns]}")
print(f"  balance_sheet: {bs_df.shape}")
print(f"  cashflow: {cf_df.shape}")
print(f"  price hist: {len(hist)} rows")

# Map yfinance columns (FY-end dates) to FY-end year
def col_for_year(yr, df):
    """Find the yfinance column for FY ending YYYY-03-31."""
    if df is None or df.empty: return None
    for c in df.columns:
        if c.year == yr and c.month == 3:
            return c
    return None

def yf_get(df, row_label, year):
    c = col_for_year(year, df)
    if c is None: return None
    if row_label not in df.index: return None
    v = df.loc[row_label, c]
    if pd.isna(v): return None
    return float(v)

def to_B(v):
    if v is None: return None
    return v / UNIT_DIV

# Year-end prices (use Dec 31 calendar YE; mismatched with FY but close enough)
ye_px = {}
for d, v in hist['Close'].resample('YE').last().items():
    if d.year in YEARS:
        ye_px[d.year] = float(v)

# Latest snapshot
latest_close = float(hist['Close'].iloc[-1]) if not hist.empty else None
latest_date = str(hist.index[-1].date()) if not hist.empty else None
latest_mkt_cap = info.get('marketCap')
shares_now = info.get('sharesOutstanding')

print(f"\n  Latest: {latest_date} = ¥{latest_close:.0f}")
print(f"  Market cap: ¥{latest_mkt_cap/1e9:.1f}B; Shares: {shares_now/1e6:.1f}M")

# Sanity: what years have data?
yf_years = [c.year for c in inc_df.columns if c.month == 3]
print(f"  yfinance has FY-ending years: {sorted(yf_years)}")

# ===== Build workbook =====
print(f"\n{'='*60}\nBuilding 8-sheet workbook...\n{'='*60}")
wb = Workbook()

# ===== Sheet 1: Financial Data =====
ws = wb.active
ws.title = "Financial Data"
ws.column_dimensions['A'].width = 42
for i in range(len(YEARS)):
    ws.column_dimensions[col_l(i)].width = 13
ws.column_dimensions[col_l(len(YEARS))].width = 8.83

ws.cell(1, 1, f"{COMPANY} ({TICKER}) — Financial Data ({UNIT_LABEL})")
fmt(ws.cell(1, 1), bold=True, align="left")
ws.cell(2, 1, f"yfinance source — only FY2022–FY2025 populated; FY2015–FY2021 left BLANK (use EDINET PDFs to fill)")
fmt(ws.cell(2, 1), italic=True, color=BLUE, align="left")

# Year header (FY-end calendar year)
ws.cell(3, 1, "Item")
fmt(ws.cell(3, 1), bold=True, grey=True, align="left")
for i, y in enumerate(YEARS):
    fmt(ws.cell(3, 2+i, y), bold=True, grey=True, num_fmt="0", align="center")

# Map yfinance row labels to our standard row layout
fd_rows = [
    (4,  "==INCOME STATEMENT (Yr ending 3/31)==", "header", None),
    (5,  "Total Revenue", "yf", "Total Revenue"),
    (6,  "Revenue Growth (YoY)", "growth_rev", None),
    (7,  "Gross Profit", "yf", "Gross Profit"),
    (8,  "Gross Margin", "pct_rev", None),
    (9,  "Operating Expense (incl SG&A)", "yf", "Operating Expense"),
    (10, "R&D / SG&A (combined)", "blank", None),  # not separately broken out by yfinance
    (11, "Operating Income (Loss)", "yf", "Operating Income"),
    (12, "Operating Margin", "op_margin", None),
    (13, "Interest Income", "yf", "Interest Income"),
    (14, "Interest Expense", "yf", "Interest Expense"),
    (15, "Pretax Income", "yf", "Pretax Income"),
    (16, "Income Tax", "yf", "Tax Provision"),
    (17, "Net Income", "yf", "Net Income"),
    (18, "NI to Common Stockholders", "yf", "Net Income Common Stockholders"),
    (19, "Net Margin", "ni_margin", None),
    (20, "==BALANCE SHEET (at 3/31)==", "header", None),
    (21, "Cash & Equivalents", "yf_bs", "Cash And Cash Equivalents"),
    (22, "Marketable Securities (current)", "yf_bs", "Other Short Term Investments"),
    (23, "Accounts Receivable", "yf_bs", "Accounts Receivable"),
    (24, "PP&E (net)", "yf_bs", "Net PPE"),
    (25, "Goodwill", "yf_bs", "Goodwill"),
    (26, "Total Assets", "yf_bs", "Total Assets"),
    (27, "Accounts Payable", "yf_bs", "Accounts Payable"),
    (28, "Long-Term Debt (Noncurrent)", "yf_bs", "Long Term Debt"),
    (29, "Total Liabilities", "yf_bs", "Total Liabilities Net Minority Interest"),
    (30, "Equity (to Owners of Parent)", "yf_bs", "Stockholders Equity"),
    (31, "==CASH FLOW (FY)==", "header", None),
    (32, "Operating Cash Flow", "yf_cf", "Operating Cash Flow"),
    (33, "Investing Cash Flow", "yf_cf", "Investing Cash Flow"),
    (34, "Financing Cash Flow", "yf_cf", "Financing Cash Flow"),
    (35, "Capex (PP&E purchases)", "yf_cf_neg", "Capital Expenditure"),
    (36, "Free Cash Flow (OCF + Capex)", "fcf", None),
    (37, "FCF Margin", "fcf_margin", None),
    (38, "Stock-Based Compensation", "blank", None),
    (39, "Depreciation & Amortization", "yf_cf", "Depreciation And Amortization"),
    (40, "Dividends Paid (Common)", "yf_cf_neg", "Cash Dividends Paid"),
    (41, "==MARKET DATA==", "header", None),
    (42, "Year-End Share Price (¥/share, calendar Dec 31)", "px", None),
    (43, "Diluted Shares (M)", "shares", None),
    (44, "Market Cap (¥B, price × shares)", "mc", None),
]

for r, label, kind, src in fd_rows:
    ws.cell(r, 1, label)
    if kind == "header":
        fmt(ws.cell(r, 1), bold=True, grey=True, align="left")
        for i in range(len(YEARS)): fmt(ws.cell(r, 2+i), grey=True)
        continue
    fmt(ws.cell(r, 1), align="left")
    for i, y in enumerate(YEARS):
        cl = col_l(i); pcl = col_l(i-1) if i>0 else None
        cell = ws.cell(r, 2+i)
        if kind == "yf":
            v = to_B(yf_get(inc_df, src, y))
            if v is not None:
                cell.value = v; fmt(cell, color=BLUE, num_fmt=NUM)
            else:
                fmt(cell, color=BLUE, num_fmt=NUM)   # blank but BLUE-formatted
        elif kind == "yf_bs":
            v = to_B(yf_get(bs_df, src, y))
            if v is not None:
                cell.value = v; fmt(cell, color=BLUE, num_fmt=NUM)
            else:
                fmt(cell, color=BLUE, num_fmt=NUM)
        elif kind == "yf_cf":
            v = to_B(yf_get(cf_df, src, y))
            if v is not None:
                cell.value = v; fmt(cell, color=BLUE, num_fmt=NUM)
            else:
                fmt(cell, color=BLUE, num_fmt=NUM)
        elif kind == "yf_cf_neg":
            v = to_B(yf_get(cf_df, src, y))
            if v is not None:
                cell.value = -abs(v) if src != "Capital Expenditure" else v
                fmt(cell, color=BLUE, num_fmt=NUM)
            else:
                fmt(cell, color=BLUE, num_fmt=NUM)
        elif kind == "blank":
            fmt(cell, color=BLUE, num_fmt=NUM)
        elif kind == "growth_rev" and i > 0:
            cell.value = f"=IFERROR({cl}5/{pcl}5-1,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT1)
        elif kind == "pct_rev":
            cell.value = f"=IFERROR({cl}7/{cl}5,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT1)
        elif kind == "op_margin":
            cell.value = f"=IFERROR({cl}11/{cl}5,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT1)
        elif kind == "ni_margin":
            cell.value = f"=IFERROR({cl}17/{cl}5,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT1)
        elif kind == "fcf":
            cell.value = f"={cl}32+{cl}35"
            fmt(cell, color=BLACK, num_fmt=NUM, bold=True)
        elif kind == "fcf_margin":
            cell.value = f"=IFERROR(({cl}32+{cl}35)/{cl}5,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT1)
        elif kind == "px":
            v = ye_px.get(y)
            if v is not None: cell.value = v; fmt(cell, color=BLUE, num_fmt=PX)
            else: fmt(cell, color=BLUE, num_fmt=PX)
        elif kind == "shares":
            cell.value = SHARES_M.get(y); fmt(cell, color=BLUE, num_fmt='0.00')
        elif kind == "mc":
            # MC in ¥B = price (¥) × shares (M) / 1000
            cell.value = f"={cl}42*{cl}43/1000"
            fmt(cell, color=BLACK, num_fmt=NUM, bold=True)

for rr in range(1, 45):
    ws.row_dimensions[rr].height = 18.0

# ===== Sheet 2: DRIP Analysis (per skill spec) =====
ws2 = wb.create_sheet("DRIP Analysis")
ws2.column_dimensions['A'].width = 38.33
ws2.column_dimensions['B'].width = 15.83
for i in range(1, len(YEARS)):
    ws2.column_dimensions[col_l(i)].width = 13
ws2.column_dimensions[col_l(len(YEARS))].width = 8.83

fd = "'Financial Data'"
n = len(YEARS)
last_col = col_l(n-1)
cagr_n = n - 1

# Title
title = f"DRIP Return Analysis — {TICKER} ({YEARS[0]}–{YEARS[-1]}, {DRIP_HOLDER})"
c = ws2.cell(1, 1, title)
c.font = Font(name=FONT, size=12, color=BLACK, bold=True)
c.alignment = Alignment(horizontal="left", vertical="center")
c.fill = PatternFill("solid", fgColor=LGREY); c.border = BORDER
for j in range(2, 2+n+1):
    ws2.cell(1, j).fill = PatternFill("solid", fgColor=LGREY)

ws2.cell(2, 1, "Source: 'Financial Data' rows 30 (Eq), 40 (Div), 42 (Px), 43 (Shares), 44 (MC). Years FY-end calendar year.")
fmt(ws2.cell(2, 1), italic=True, color=BLUE, align="left")

# Assumptions
ws2.cell(4, 1, "Assumptions"); fmt(ws2.cell(4,1), bold=True, align="left")
ws2.cell(5, 1, "Initial investment (¥)"); fmt(ws2.cell(5,1), align="left")
fmt(ws2.cell(5, 2, DRIP_INVEST), color=BLUE, bold=True, num_fmt=INT)
ws2.cell(6, 1, f"Entry price (¥/sh, {YEARS[0]} YE)"); fmt(ws2.cell(6,1), align="left")
c = ws2.cell(6, 2); c.value = f"={fd}!B42"; fmt(c, color=GREEN, bold=True, num_fmt=PX)
ws2.cell(7, 1, "FX (¥ denominator, 1.00)"); fmt(ws2.cell(7,1), align="left")
fmt(ws2.cell(7, 2, DRIP_FX), color=BLUE, bold=True, num_fmt='0.00')
ws2.cell(8, 1, "Dividend WHT (HK Corp / no JP-HK DTA)"); fmt(ws2.cell(8,1), align="left")
fmt(ws2.cell(8, 2, DRIP_WHT), color=BLUE, bold=True, num_fmt='0.000%')
ws2.cell(9, 1, "Initial shares purchased"); fmt(ws2.cell(9,1), align="left")
c = ws2.cell(9, 2); c.value = "=B5/B6"; fmt(c, num_fmt=INT)

# Year-by-year DRIP
ws2.cell(11, 1, "Year-by-year DRIP"); fmt(ws2.cell(11,1), bold=True, align="left")
ws2.cell(12, 1, "Year (FY-end)"); fmt(ws2.cell(12,1), bold=True, align="left")
for i in range(n):
    cl = col_l(i)
    c = ws2.cell(12, 2+i); c.value = f"={fd}!{cl}3"; fmt(c, bold=True, align="right")

ws2.cell(13, 1, "Shares held BoY"); fmt(ws2.cell(13,1), align="left")
for i in range(n):
    if i == 0: ws2.cell(13, 2).value = "=B9"
    else: ws2.cell(13, 2+i).value = f"={col_l(i-1)}18"
    fmt(ws2.cell(13, 2+i), num_fmt=INT)

# DPS = ABS(Div_¥B) × 1e3 / Shares_M (¥B / M = ¥K → ×1e3 → ¥/share)
# Wait: Div is in ¥B (FD/1e9), Shares are in M (1e6).
#   DPS_¥/share = Div_¥B × 1e9 / (Shares_M × 1e6) = Div_¥B × 1e3 / Shares_M
ws2.cell(14, 1, "DPS (¥, computed from CF)"); fmt(ws2.cell(14,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(14, 2+i).value = f"=IFERROR(ABS({fd}!{cl}40)*1000/{cl}40,0)"
    fmt(ws2.cell(14, 2+i), num_fmt='0.00')

ws2.cell(15, 1, "Dividend (¥, after WHT)"); fmt(ws2.cell(15,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(15, 2+i).value = f"={cl}13*{cl}14*$B$7*(1-$B$8)"
    fmt(ws2.cell(15, 2+i), num_fmt=INT)

ws2.cell(16, 1, "Reinvest price (¥)"); fmt(ws2.cell(16,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(16, 2+i).value = f"={cl}39"
    fmt(ws2.cell(16, 2+i), color=GREEN, num_fmt=PX)

ws2.cell(17, 1, "Shares bought"); fmt(ws2.cell(17,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(17, 2+i).value = f"=IFERROR({cl}15/{cl}16,0)"
    fmt(ws2.cell(17, 2+i), num_fmt=INT)

ws2.cell(18, 1, "Shares held EoY"); fmt(ws2.cell(18,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(18, 2+i).value = f"={cl}13+{cl}17"
    fmt(ws2.cell(18, 2+i), num_fmt=INT)

ws2.cell(19, 1, "Portfolio value EoY (¥)"); fmt(ws2.cell(19,1), align="left")
for i in range(n):
    cl = col_l(i); bold = (i == n-1)
    ws2.cell(19, 2+i).value = f"={cl}18*{cl}16"
    fmt(ws2.cell(19, 2+i), bold=bold, num_fmt=INT)

# P/B = MC_¥B / Equity_¥B
ws2.cell(20, 1, "P/B at Year-End (= MC / Equity)"); fmt(ws2.cell(20,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(20, 2+i).value = f"=IFERROR({fd}!{cl}44/{fd}!{cl}30,\"-\")"
    fmt(ws2.cell(20, 2+i), color=GREEN, num_fmt=X)

# Dividend Yield = ABS(Div_¥B) / MC_¥B
ws2.cell(21, 1, "Dividend Yield (Div / MC)"); fmt(ws2.cell(21,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws2.cell(21, 2+i).value = f"=IFERROR(ABS({fd}!{cl}40)/{fd}!{cl}44,0)"
    fmt(ws2.cell(21, 2+i), num_fmt=PCT)

# JP M2 reference
ws2.cell(23, 1, "Reference: Japan M2 (¥ trillion)"); fmt(ws2.cell(23,1), bold=True, align="left")
ws2.cell(24, 1, "M2 year-end balance"); fmt(ws2.cell(24,1), align="left")
for i, y in enumerate(YEARS):
    v = JP_M2.get(y)
    if v: fmt(ws2.cell(24, 2+i, v), color=BLUE, bold=True, num_fmt='#,##0.0')

ws2.cell(25, 1, f"M2 multiple vs {YEARS[0]}"); fmt(ws2.cell(25,1), align="left")
c = ws2.cell(25, 2+n); c.value = f"={last_col}24/B24"; fmt(c, num_fmt=X)
ws2.cell(26, 1, f"M2 CAGR ({cagr_n} yrs)"); fmt(ws2.cell(26,1), align="left")
c = ws2.cell(26, 2+n); c.value = f"=({last_col}24/B24)^(1/{cagr_n})-1"; fmt(c, num_fmt=PCT)

# Summary
ws2.cell(28, 1, "Summary"); fmt(ws2.cell(28,1), bold=True, align="left")
ws2.cell(29, 1, f"Final value {YEARS[-1]} YE (¥)"); fmt(ws2.cell(29,1), align="left")
c = ws2.cell(29, 2); c.value = f"={last_col}19"; fmt(c, bold=True, num_fmt=INT)
ws2.cell(30, 1, "Total return"); fmt(ws2.cell(30,1), align="left")
c = ws2.cell(30, 2); c.value = "=B29/B5-1"; fmt(c, num_fmt=PCT)
ws2.cell(31, 1, "Multiple on invested (x)"); fmt(ws2.cell(31,1), align="left")
c = ws2.cell(31, 2); c.value = "=B29/B5"; fmt(c, bold=True, num_fmt=X)
ws2.cell(32, 1, f"Annualized IRR ({n} yrs)"); fmt(ws2.cell(32,1), align="left")
c = ws2.cell(32, 2); c.value = f"=(B29/B5)^(1/{n})-1"; fmt(c, bold=True, num_fmt=PCT)
cell_m2cagr = f"{get_column_letter(2+n)}26"
ws2.cell(33, 1, "vs M2 growth (DRIP IRR − M2 CAGR)"); fmt(ws2.cell(33,1), align="left")
c = ws2.cell(33, 2); c.value = f"=B32-{cell_m2cagr}"; fmt(c, num_fmt=PCT)
ws2.cell(34, 1, "Price-only return (no dividends)"); fmt(ws2.cell(34,1), align="left")
c = ws2.cell(34, 2); c.value = f"={last_col}39/B39-1"; fmt(c, num_fmt=PCT1)
ws2.cell(35, 1, "Avg P/B over period"); fmt(ws2.cell(35,1), align="left")
c = ws2.cell(35, 2); c.value = f"=AVERAGE({col_l(0)}20:{last_col}20)"; fmt(c, num_fmt=X)

# Market Data Reference
ws2.cell(37, 1, "Market Data Reference (BLUE = editable inputs)")
fmt(ws2.cell(37,1), bold=True, align="left")
ws2.cell(38, 1, "Year (FY-end)"); fmt(ws2.cell(38,1), bold=True, align="left")
for i, y in enumerate(YEARS):
    fmt(ws2.cell(38, 2+i, y), bold=True, num_fmt="0", align="center")
ws2.cell(39, 1, "YE Close (¥/share)"); fmt(ws2.cell(39,1), align="left")
for i, y in enumerate(YEARS):
    v = ye_px.get(y)
    if v is not None: fmt(ws2.cell(39, 2+i, v), color=BLUE, num_fmt=PX)
    else: fmt(ws2.cell(39, 2+i), color=BLUE, num_fmt=PX)
ws2.cell(40, 1, "Total Shares (M)"); fmt(ws2.cell(40,1), align="left")
for i, y in enumerate(YEARS):
    fmt(ws2.cell(40, 2+i, SHARES_M[y]), color=BLUE, num_fmt='0.00')

for rr in range(1, 41):
    ws2.row_dimensions[rr].height = 23.0

# ===== Sheet 3: Key Ratios =====
ws3 = wb.create_sheet("Key Ratios")
ws3.column_dimensions['A'].width = 38
for i in range(len(YEARS)):
    ws3.column_dimensions[col_l(i)].width = 13

ws3.cell(1, 1, f"{COMPANY} — Key Ratios"); fmt(ws3.cell(1,1), bold=True, align="left")
ws3.cell(3, 1, "Ratio"); fmt(ws3.cell(3,1), bold=True, grey=True, align="left")
for i, y in enumerate(YEARS):
    fmt(ws3.cell(3, 2+i, y), bold=True, grey=True, num_fmt="0", align="center")

ratios = [
    ("Revenue Growth (YoY)",        lambda c,p: f"=IFERROR({fd}!{c}5/{fd}!{p}5-1,\"\")", PCT1, True),
    ("Gross Margin",                lambda c,p: f"=IFERROR({fd}!{c}7/{fd}!{c}5,\"\")", PCT1, False),
    ("Operating Margin",            lambda c,p: f"=IFERROR({fd}!{c}11/{fd}!{c}5,\"\")", PCT1, False),
    ("Net Margin",                  lambda c,p: f"=IFERROR({fd}!{c}17/{fd}!{c}5,\"\")", PCT1, False),
    ("FCF Margin",                  lambda c,p: f"=IFERROR(({fd}!{c}32+{fd}!{c}35)/{fd}!{c}5,\"\")", PCT1, False),
    ("ROE (NI / Avg Equity)",       lambda c,p: (f"=IFERROR({fd}!{c}18/AVERAGE({fd}!{c}30,{fd}!{p}30),\"\")" if p else None), PCT, True),
    ("ROA (NI / Avg Assets)",       lambda c,p: (f"=IFERROR({fd}!{c}17/AVERAGE({fd}!{c}26,{fd}!{p}26),\"\")" if p else None), PCT, True),
    ("Asset Turnover (Rev/Avg TA)", lambda c,p: (f"=IFERROR({fd}!{c}5/AVERAGE({fd}!{c}26,{fd}!{p}26),\"\")" if p else None), '0.00"x"', True),
    ("Cash + ST Sec / Total Assets",lambda c,p: f"=IFERROR(({fd}!{c}21+{fd}!{c}22)/{fd}!{c}26,\"\")", PCT, False),
    ("LT Debt / Equity",            lambda c,p: f"=IFERROR({fd}!{c}28/{fd}!{c}30,\"\")", '0.00"x"', False),
    ("Dividend Payout (|Div|/NI)",  lambda c,p: f"=IFERROR(-{fd}!{c}40/{fd}!{c}17,\"\")", PCT, False),
]
r = 4
for label, fn, nfmt, needs_prior in ratios:
    ws3.cell(r, 1, label); fmt(ws3.cell(r,1), align="left")
    for i in range(len(YEARS)):
        cl, pcl = col_l(i), col_l(i-1) if i>0 else None
        if needs_prior and pcl is None: continue
        f = fn(cl, pcl)
        if f:
            cell = ws3.cell(r, 2+i); cell.value = f
            fmt(cell, color=GREEN, num_fmt=nfmt)
    r += 1
for rr in range(1, r+1):
    ws3.row_dimensions[rr].height = 20.0

# ===== Sheet 4: Valuation Model =====
ws4 = wb.create_sheet("Valuation Model")
ws4.column_dimensions['A'].width = 38
for i in range(len(YEARS)):
    ws4.column_dimensions[col_l(i)].width = 13
ws4.cell(1, 1, f"{COMPANY} — Valuation Model"); fmt(ws4.cell(1,1), bold=True, align="left")
ws4.cell(3, 1, "Item"); fmt(ws4.cell(3,1), bold=True, grey=True, align="left")
for i, y in enumerate(YEARS):
    fmt(ws4.cell(3, 2+i, y), bold=True, grey=True, num_fmt="0", align="center")

val_rows = [
    (5, "Key Assumptions", "header"),
    (6, "Japan 10y JGB", "bond"),
    (7, "Equity Risk Premium", "erp"),
    (8, "Cost of Equity", "coe"),
    (10, "Inputs (linked)", "header"),
    (11, "NI to Common (¥B)", "link", f"{fd}!{{cl}}18"),
    (12, "Dividends Paid (¥B, abs)", "link_abs", f"{fd}!{{cl}}40"),
    (13, "Equity (¥B)", "link", f"{fd}!{{cl}}30"),
    (14, "Market Cap (¥B)", "link", f"{fd}!{{cl}}44"),
    (15, "Long-Term Debt (¥B)", "link", f"{fd}!{{cl}}28"),
    (16, "Cash + ST Sec (¥B)", "link_cash", None),
    (17, "Enterprise Value (MC + Debt − Cash)", "ev"),
    (19, "Multiples (Trading)", "header"),
    (20, "EV / Revenue", "ratio_a", (f"{{cl}}17", f"{fd}!{{cl}}5")),
    (21, "EV / EBITDA-proxy (OpInc + D&A)", "ev_ebitda"),
    (22, "P/E", "pe"),
    (23, "P/B", "pb"),
    (24, "Dividend Yield", "div_y"),
    (26, "Valuation Output (3-method)", "header"),
    (27, "Method 1: Target P/E × NI (15x)", "v1"),
    (28, "  Target P/E", "tgt_pe"),
    (29, "Method 2: Target EV/Rev × Rev (1.5x)", "v2"),
    (30, "  Target EV/Rev", "tgt_evrev"),
    (31, "Method 3: Justified P/B × Equity (1.0x)", "v3"),
    (32, "  Target P/B", "tgt_pb"),
    (33, "Average Implied Mkt Cap (¥B)", "avg3"),
    (34, "Premium / (Discount) vs Actual", "pd"),
]
TARGET_PE = 15.0
TARGET_EVREV = 1.5
TARGET_PB = 1.0
for spec in val_rows:
    r = spec[0]; label = spec[1]; kind = spec[2]
    ws4.cell(r, 1, label)
    if kind == "header":
        fmt(ws4.cell(r,1), bold=True, grey=True, align="left")
        for i in range(len(YEARS)): fmt(ws4.cell(r, 2+i), grey=True)
        continue
    fmt(ws4.cell(r,1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i)
        cell = ws4.cell(r, 2+i)
        if kind == "bond":
            cell.value = JP_BOND.get(YEARS[i], 0.005)
            fmt(cell, color=BLUE, num_fmt=PCT)
        elif kind == "erp":
            cell.value = 0.05
            fmt(cell, color=BLUE, num_fmt=PCT)
        elif kind == "coe":
            cell.value = f"={cl}6+{cl}7"
            fmt(cell, color=BLACK, num_fmt=PCT)
        elif kind == "link":
            cell.value = "=" + spec[3].replace("{cl}", cl)
            fmt(cell, color=GREEN, num_fmt=NUM)
        elif kind == "link_abs":
            cell.value = f"=ABS({spec[3].replace('{cl}', cl)})"
            fmt(cell, color=GREEN, num_fmt=NUM)
        elif kind == "link_cash":
            cell.value = f"={fd}!{cl}21+{fd}!{cl}22"
            fmt(cell, color=GREEN, num_fmt=NUM)
        elif kind == "ev":
            cell.value = f"={cl}14+{cl}15-{cl}16"
            fmt(cell, color=BLACK, num_fmt=NUM, bold=True)
        elif kind == "ratio_a":
            num = spec[3][0].replace("{cl}", cl); den = spec[3][1].replace("{cl}", cl)
            cell.value = f"=IFERROR({num}/{den},\"\")"
            fmt(cell, color=BLACK, num_fmt=X)
        elif kind == "ev_ebitda":
            cell.value = f"=IFERROR({cl}17/({fd}!{cl}11+{fd}!{cl}39),\"\")"
            fmt(cell, color=BLACK, num_fmt=X)
        elif kind == "pe":
            cell.value = f"=IFERROR(IF({fd}!{cl}17>0,{cl}14/{fd}!{cl}17,\"\"),\"\")"
            fmt(cell, color=BLACK, num_fmt=X)
        elif kind == "pb":
            cell.value = f"=IFERROR({cl}14/{cl}13,\"\")"
            fmt(cell, color=BLACK, num_fmt=X)
        elif kind == "div_y":
            cell.value = f"=IFERROR({cl}12/{cl}14,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT)
        elif kind == "v1":
            cell.value = f"=IF({cl}11>0,{cl}11*{cl}28,\"\")"
            fmt(cell, color=BLACK, num_fmt=NUM)
        elif kind == "tgt_pe":
            cell.value = TARGET_PE; fmt(cell, color=BLUE, num_fmt=X)
        elif kind == "v2":
            cell.value = f"={fd}!{cl}5*{cl}30-{cl}15+{cl}16"
            fmt(cell, color=BLACK, num_fmt=NUM)
        elif kind == "tgt_evrev":
            cell.value = TARGET_EVREV; fmt(cell, color=BLUE, num_fmt=X)
        elif kind == "v3":
            cell.value = f"=IFERROR({cl}13*{cl}32,\"\")"
            fmt(cell, color=BLACK, num_fmt=NUM)
        elif kind == "tgt_pb":
            cell.value = TARGET_PB; fmt(cell, color=BLUE, num_fmt=X)
        elif kind == "avg3":
            cell.value = f"=IFERROR(AVERAGE({cl}27,{cl}29,{cl}31),\"\")"
            fmt(cell, color=BLACK, num_fmt=NUM, bold=True)
        elif kind == "pd":
            cell.value = f"=IFERROR({cl}33/{cl}14-1,\"\")"
            fmt(cell, color=BLACK, num_fmt=PCT, bold=True)

for rr in range(1, 36):
    ws4.row_dimensions[rr].height = 17.0

# ===== Sheet 5 / 6 / 7: minimal placeholders =====
ws5 = wb.create_sheet("Quarterly TTM")
ws5.column_dimensions['A'].width = 40
ws5.cell(1, 1, f"{COMPANY} — Quarterly TTM (yfinance has only 1 quarter for TOTO; use EDINET 四半期報告書 for full)")
fmt(ws5.cell(1,1), bold=True, align="left")
ws5.cell(3, 1, "yfinance only returned 1 quarter (2025-06-30) for TOTO. For full quarterly TTM, fetch from EDINET 四半期報告書.")
fmt(ws5.cell(3,1), italic=True, color=BLUE, align="left")

ws6 = wb.create_sheet("Top Holders")
ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 50
ws6.cell(1, 1, f"{COMPANY} — Major Holders Summary")
fmt(ws6.cell(1,1), bold=True, align="left")
ws6.cell(3, 1, "From yfinance.major_holders:")
fmt(ws6.cell(3,1), italic=True, align="left")
mh_data = [
    ("Insiders %", info.get('heldPercentInsiders', 0)),
    ("Institutions %", info.get('heldPercentInstitutions', 0)),
    ("Float % held by Inst.", 0.59559),  # from probe
    ("Number of Institutions", 182),
]
for i, (label, val) in enumerate(mh_data):
    ws6.cell(4+i, 1, label); fmt(ws6.cell(4+i,1), align="left")
    cell = ws6.cell(4+i, 2, val)
    nf = PCT1 if 'Number' not in label else INT
    fmt(cell, color=BLUE, num_fmt=nf)
ws6.cell(10, 1, "For top-holder details, query EDINET 大量保有報告書 (>5% disclosures) or J-IRIS.")
fmt(ws6.cell(10,1), italic=True, color=BLUE, align="left")

# ===== Sheet 8: Data Correlation =====
ws8 = wb.create_sheet("Data Correlation")
ws8.column_dimensions['A'].width = 110
ws8.cell(1, 1, f"{COMPANY} ({TICKER}) — Data Correlation & Methodology")
fmt(ws8.cell(1,1), bold=True, align="left")

sections = [
    ("1. Company Overview",
     f"{COMPANY} — Tokyo SE-listed (TSE 5332). Sector: Industrials / Building Products & Equipment.\n"
     f"World's largest sanitary ware (toilets, faucets, bathroom fixtures) manufacturer.\n"
     f"HQ: Kitakyushu, Japan. Listed: 1949.\n"
     f"Reporting currency: JPY. Fiscal year: April 1 – March 31."),
    ("2. Data Sources & Limitations",
     "PRIMARY: yfinance (Yahoo Finance API) — annual IS/BS/CF for FY2022-FY2025 only (4 years).\n"
     "PRICE: yfinance — full daily history 2014 to present.\n"
     "GAPS: FY2015-FY2021 financials NOT populated by yfinance — left BLUE blank.\n"
     "  → Fill manually from EDINET annual report (有価証券報告書) PDFs (separate download script).\n"
     "  → EDINET URL: https://disclosure2.edinet-fsa.go.jp/\n"
     "QUARTERLY: yfinance returned only 1 quarter (2025-06-30). Use EDINET 四半期報告書 for full TTM.\n"
     "TOP HOLDERS: yfinance gives aggregate %; for >5% holders use EDINET 大量保有報告書."),
    ("3. Year Convention",
     "Column headers (B–L) = FY-end calendar year (e.g. 2025 = FY ending 2025-03-31).\n"
     "  FY2025 = April 2024 – March 2025\n"
     "  FY2015 = April 2014 – March 2015\n"
     "Year-end share prices use Dec 31 (calendar) rather than March 31 (FY-end) — small mismatch,\n"
     "  acceptable for DRIP & valuation comparisons."),
    ("4. DRIP Methodology",
     "Initial investment: ¥10,000,000 (≈ HKD 600K equivalent at FY2025 FX).\n"
     "WHT: 15.315% (Japan domestic income tax 15% + 0.315% reconstruction surtax).\n"
     "  Default applies to HK Corporate holders (no Japan-HK comprehensive DTA).\n"
     "  Treaty rates (adjust if applicable):\n"
     "    - US holders: 10% (US-Japan DTA)\n"
     "    - Mainland China holders: 10% (China-Japan DTA)\n"
     "    - UK / Singapore: 10% (DTA)\n"
     "    - HK holders: 15.315% (no DTA)\n"
     "Japan M2 benchmark: 2015 = ¥889T → 2025 ≈ ¥1,260T → 10y CAGR ≈ 3.55%.\n"
     "Note: Japan's M2 grew slowly (low inflation, BOJ QE absorbed), making DRIP excess return less\n"
     "  meaningful than US/CN comparisons. Use as directional benchmark only."),
    ("5. Valuation Methodology",
     "Three-method blend:\n"
     "  Method 1 — P/E × NI: Target P/E = 15x (Japanese industrials typical range 12-18x)\n"
     "  Method 2 — EV/Rev × Rev: Target EV/Rev = 1.5x (mature consumer/industrial)\n"
     "  Method 3 — P/B × Equity: Target P/B = 1.0x (book-value floor for mature industrial)\n"
     "All targets are EDITABLE BLUE inputs at rows 28, 30, 32. Adjust per industry view.\n"
     "Premium/(Discount) = Avg Implied / Actual Mkt Cap − 1."),
    ("6. Sector Notes",
     "TOTO is a global market leader in:\n"
     "  - Washlets (electric bidet seats) — flagship product, ~30%+ Asia premium price point\n"
     "  - Bathroom fixtures (toilets, faucets, sinks)\n"
     "  - Tile and ceramics for residential/commercial\n"
     "Geographic mix (FY2025 estimate):\n"
     "  - Japan: ~55% revenue (mature, slight decline)\n"
     "  - China: ~15% (high-end housing exposure, COVID drag, real estate downturn)\n"
     "  - Americas: ~15%\n"
     "  - Other Asia/Europe: ~15%\n"
     "FY2025 NI dropped to ¥12.2B (from ¥37.2B FY2024) — likely impairment / one-off charges.\n"
     "  Verify via FY2025 annual report (有価証券報告書) once downloaded."),
    ("7. Format Conventions",
     "Font: Apple Braille 11pt (DRIP title row 12pt bold, light grey fill).\n"
     "Colors: BLUE = hardcoded input (yfinance or manual), BLACK = formula, GREEN = cross-sheet link.\n"
     "Row heights: Financial Data 18pt, DRIP 23pt, Key Ratios 20pt, Valuation 17pt.\n"
     "Number formats: ¥B amounts use #,##0.0; %s use 0.0%; multiples use 0.00x."),
    ("8. To Complete the Model",
     "1) Run scripts/download_TOTO_reports.py — pulls FY2021-FY2025 annual report PDFs from EDINET.\n"
     "2) From each PDF, extract Income Statement / Balance Sheet / Cash Flow main lines.\n"
     "3) Type values into BLUE blank cells in Financial Data sheet (rows 5-40, cols B-H = FY2015-FY2021).\n"
     "4) DRIP sheet auto-recomputes (full 11-year IRR + P/B journey).\n"
     "5) Verify FY2022-FY2025 yfinance values against EDINET PDFs (sanity check)."),
]
r = 3
for title_, body in sections:
    ws8.cell(r, 1, title_); fmt(ws8.cell(r,1), bold=True, grey=True, align="left")
    r += 1
    for ln in body.split("\n"):
        ws8.cell(r, 1, ln); fmt(ws8.cell(r,1), align="left")
        ws8.row_dimensions[r].height = 17
        r += 1
    r += 1

# ===== Save =====
print(f"\nSaving → {OUT_FILE}")
wb.save(OUT_FILE)
print(f"\n{'='*60}\n✓ Saved: {OUT_FILE}\n{'='*60}")
print(f"\nSheets:")
for sn in wb.sheetnames:
    print(f"  - {sn}")
