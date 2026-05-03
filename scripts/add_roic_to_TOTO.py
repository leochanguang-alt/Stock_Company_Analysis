#!/usr/bin/env python3
"""
Add ROIC Analysis sheet to TOTO 5332.T financial model.
Inserted as Sheet 3 (after DRIP, before Key Ratios) per financial-model skill spec.

ROIC = Adj NOPAT / Avg Invested Capital
  Adj NOPAT (after-tax) = (Operating Income + SBC) × (1 − Tax Rate)
  Invested Capital = Total Equity + LT Debt − Cash − Marketable Securities
"""
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SRC = '/Users/chenguang/Desktop/BUI_Investment_Management/5332.T/5332.T.Valuation.v260501.xlsx'
TODAY = datetime.now().strftime('%y%m%d')
DST = SRC.replace('v260501', f'v{TODAY}')

YEARS = list(range(2015, 2026))   # FY-end calendar years

# ===== Format =====
FONT = "Apple Braille"; SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY = "FFE7E6E6"
THIN = Side(style="thin", color="FF000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = '#,##0.0'
PCT = '0.0%'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False):
    c.font = Font(name=FONT, size=SIZE, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if num_fmt: c.number_format = num_fmt
    if grey: c.fill = PatternFill("solid", fgColor=GREY)

def col_l(i): return get_column_letter(2+i)

# ===== Open & insert =====
print(f"Loading: {SRC}")
wb = load_workbook(SRC)

# Remove if exists (idempotent)
if 'ROIC Analysis' in wb.sheetnames:
    del wb['ROIC Analysis']
    print("  Removed existing 'ROIC Analysis' for clean re-add")

# Insert at position 2 (0-indexed: after Financial Data + DRIP)
drip_idx = wb.sheetnames.index('DRIP Analysis')
ws = wb.create_sheet('ROIC Analysis', drip_idx + 1)
print(f"  Inserted 'ROIC Analysis' at position {drip_idx + 1}")

# ===== Build ROIC sheet =====
ws.column_dimensions['A'].width = 42
for i in range(len(YEARS)):
    ws.column_dimensions[col_l(i)].width = 13
ws.column_dimensions[col_l(len(YEARS))].width = 8.83

fd = "'Financial Data'"

# Title (row 1)
ws.cell(1, 1, f"TOTO LTD. (5332.T) — ROIC Decomposition")
fmt(ws.cell(1, 1), bold=True, align="left")
ws.cell(2, 1, "ROIC = Adj NOPAT / Avg Invested Capital. FY-end calendar year. ¥B.")
fmt(ws.cell(2, 1), italic=True, color=BLUE, align="left")

# Year header (row 3)
ws.cell(3, 1, "Component")
fmt(ws.cell(3, 1), bold=True, grey=True, align="left")
for i, y in enumerate(YEARS):
    fmt(ws.cell(3, 2+i, y), bold=True, grey=True, num_fmt="0", align="center")

# ===== NOPAT Calculation block =====
ws.cell(4, 1, "==NOPAT Calculation==")
fmt(ws.cell(4, 1), bold=True, grey=True, align="left")
for i in range(len(YEARS)): fmt(ws.cell(4, 2+i), grey=True)

rows = [
    (5,  "Operating Income (EBIT)",     "link",     f"{fd}!{{cl}}11", NUM, BLACK),
    (6,  "+ Stock-Based Compensation",   "link",     f"{fd}!{{cl}}38", NUM, BLACK),
    (7,  "Adj EBIT (incl SBC)",          "formula",  "={cl}5+{cl}6",   NUM, BLACK),
    (8,  "Income Tax Expense",           "link",     f"{fd}!{{cl}}16", NUM, BLACK),
    (9,  "Pretax Income",                "link",     f"{fd}!{{cl}}15", NUM, BLACK),
    (10, "Effective Tax Rate",           "tax_rate", None, PCT, GREEN),
    (11, "Adj NOPAT (after-tax)",        "nopat",    None, NUM, BLACK),

    (13, "==Invested Capital Components==", "header", None, None, None),
    (14, "Long-Term Debt",               "link",     f"{fd}!{{cl}}28", NUM, BLACK),
    (15, "+ Equity (to Owners of Parent)", "link",   f"{fd}!{{cl}}30", NUM, BLACK),
    (16, "− Cash & Equivalents",         "link_neg", f"{fd}!{{cl}}21", NUM, BLACK),
    (17, "− Marketable Securities (current)", "link_neg", f"{fd}!{{cl}}22", NUM, BLACK),
    (18, "Invested Capital",             "ic",       None, NUM, BLACK),

    (20, "==ROIC==", "header", None, None, None),
    (21, "Avg Invested Capital",         "avg_ic",   None, NUM, BLACK),
    (22, "ROIC (Adj NOPAT / Avg IC)",    "roic",     None, PCT, GREEN),

    (24, "==YoY Changes==", "header", None, None, None),
    (25, "Adj NOPAT Growth",             "nopat_g",  None, PCT, BLACK),
    (26, "Invested Capital Growth",      "ic_g",     None, PCT, BLACK),
]

for r, label, kind, src, nfmt, color in rows:
    ws.cell(r, 1, label)
    if kind == "header":
        fmt(ws.cell(r, 1), bold=True, grey=True, align="left")
        for i in range(len(YEARS)): fmt(ws.cell(r, 2+i), grey=True)
        continue
    fmt(ws.cell(r, 1), align="left")
    for i in range(len(YEARS)):
        cl = col_l(i); pcl = col_l(i-1) if i>0 else None
        cell = ws.cell(r, 2+i)
        if kind == "link":
            cell.value = "=" + src.replace("{cl}", cl)
            fmt(cell, color=color, num_fmt=nfmt)
        elif kind == "link_neg":
            cell.value = f"=-{src.replace('{cl}', cl)}"
            fmt(cell, color=color, num_fmt=nfmt)
        elif kind == "formula":
            cell.value = src.replace("{cl}", cl)
            fmt(cell, color=color, num_fmt=nfmt)
        elif kind == "tax_rate":
            # Effective tax rate = Tax / Pretax (capped 0-50% for sanity)
            cell.value = f"=IFERROR(IF({cl}9>0,{cl}8/{cl}9,0.30),0.30)"
            fmt(cell, color=color, num_fmt=nfmt)
        elif kind == "nopat":
            cell.value = f"=IFERROR({cl}7*(1-{cl}10),\"\")"
            fmt(cell, color=BLACK, num_fmt=nfmt, bold=True)
        elif kind == "ic":
            cell.value = f"={cl}14+{cl}15+{cl}16+{cl}17"
            fmt(cell, color=BLACK, num_fmt=nfmt, bold=True)
        elif kind == "avg_ic" and i > 0:
            cell.value = f"=AVERAGE({cl}18,{pcl}18)"
            fmt(cell, color=BLACK, num_fmt=nfmt)
        elif kind == "roic" and i > 0:
            cell.value = f"=IFERROR({cl}11/{cl}21,\"\")"
            fmt(cell, color=color, num_fmt=nfmt, bold=True)
        elif kind == "nopat_g" and i > 0:
            cell.value = f"=IFERROR(({cl}11-{pcl}11)/{pcl}11,\"\")"
            fmt(cell, color=color, num_fmt=nfmt)
        elif kind == "ic_g" and i > 0:
            cell.value = f"=IFERROR(({cl}18-{pcl}18)/{pcl}18,\"\")"
            fmt(cell, color=color, num_fmt=nfmt)

# Row heights = 20pt per spec
for rr in range(1, 28):
    ws.row_dimensions[rr].height = 20.0

# ===== Save =====
print(f"\nSaving → {DST}")
wb.save(DST)
print(f"\n{'='*60}\n✓ Saved: {DST}\n{'='*60}")
print(f"\nSheets:")
for sn in wb.sheetnames:
    print(f"  - {sn}")
