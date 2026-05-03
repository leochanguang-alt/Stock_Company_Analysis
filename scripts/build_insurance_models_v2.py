#!/usr/bin/env python3
"""
Build 8-sheet insurance valuation models for 601319.SH (PICC), 601318.SH (Ping An),
601601.SH (CPIC). Bespoke "Method C" template with EV/NBV roll-forward and
Combined Operating Ratio detail. A+H weighted market cap.

Output: ~/Desktop/BUI_Investment_Management/<TICKER>/<TICKER>.Valuation.v<YYMMDD>.xlsx
"""
import os, time
from datetime import datetime
import tushare as ts
import akshare as ak
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

TS_TOKEN = "45e390d8bf059895336edd8073d294a2f137495fee6cd659acfc6f43"
YEARS = list(range(2015, 2026))   # 2015-2025, 11 years
TODAY = datetime.now().strftime("%y%m%d")
OUT_BASE = os.path.expanduser("~/Desktop/BUI_Investment_Management")

# ===== Format constants =====
FONT = "Apple Braille"
SIZE = 11
BLUE   = "FF0000FF"   # input
BLACK  = "FF000000"   # formula
GREEN  = "FF008000"   # cross-sheet link
GREY   = "FFE7E6E6"   # section header bg
THIN   = Side(style="thin", color="FF000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.00;(#,##0.00);"–"'
PCT_FMT = '0.0%;(0.0%);"–"'
INT_FMT = '#,##0;(#,##0);"–"'

# Chinese 10-yr govt bond yields (%) used for cost of equity / dividend yield benchmark
CN_BOND = {2015:3.05, 2016:3.06, 2017:3.91, 2018:3.31, 2019:3.14,
           2020:3.18, 2021:2.78, 2022:2.84, 2023:2.56, 2024:1.68, 2025:1.85}

# ============================================================
# COMPANY CONFIG — EV/NBV/COR data manually compiled from annual reports
# (CNY 亿). These are USER-EDITABLE BLUE input cells; verify against 年报.
# ============================================================
CONFIG = {
    "601319.SH": {
        "name": "中国人保 (PICC Group)",
        "h_code5": "01339",
        "h_yf": "1339.HK",
        "biz_type": "P&C-dominant",
        # PICC Life subsidiary EV (人保寿险, ~80% of group EV)
        "ev_group": {2015:1180, 2016:1413, 2017:1708, 2018:1769, 2019:1962,
                     2020:2152, 2021:2289, 2022:2453, 2023:2536, 2024:2732, 2025:2900},
        # NBV — group level (Life + Health subsidiary)
        "nbv": {2015:65, 2016:108, 2017:130, 2018:34, 2019:30,
                2020:25, 2021:32, 2022:32, 2023:65, 2024:106, 2025:118},
        # PICC P&C Combined Operating Ratio (人保财险, %)
        "cor": {2015:96.5, 2016:98.1, 2017:97.0, 2018:98.5, 2019:98.9,
                2020:98.9, 2021:99.6, 2022:97.6, 2023:97.8, 2024:98.0, 2025:97.5},
        "loss_ratio":   {2015:62.5, 2016:64.0, 2017:62.0, 2018:62.5, 2019:63.0,
                         2020:65.0, 2021:71.0, 2022:71.5, 2023:70.0, 2024:70.5, 2025:70.0},
        "expense_ratio":{2015:34.0, 2016:34.1, 2017:35.0, 2018:36.0, 2019:35.9,
                         2020:33.9, 2021:28.6, 2022:26.1, 2023:27.8, 2024:27.5, 2025:27.5},
        "target_pe": 7.0,
        "target_pev": 0.8,
        "drip_entry": 2018,    # PICC Group A-share IPO 2018-11-16 → first full year 2018
    },
    "601318.SH": {
        "name": "中国平安 (Ping An)",
        "h_code5": "02318",
        "h_yf": "2318.HK",
        "biz_type": "Conglomerate (Life+P&C+Bank+Tech)",
        # Group EV (CNY 亿) — Ping An Group annual report
        "ev_group": {2015:5541, 2016:6377, 2017:8251, 2018:10025, 2019:12005,
                     2020:13281, 2021:13955, 2022:14238, 2023:13901, 2024:14221, 2025:14800},
        # Life NBV (CNY 亿) — Ping An Life & Health
        "nbv": {2015:308, 2016:508, 2017:674, 2018:723, 2019:759,
                2020:496, 2021:379, 2022:288, 2023:311, 2024:285, 2025:310},
        # Ping An P&C COR (%)
        "cor": {2015:95.6, 2016:95.9, 2017:96.2, 2018:96.0, 2019:96.4,
                2020:99.1, 2021:98.0, 2022:100.3, 2023:100.7, 2024:98.3, 2025:97.8},
        "loss_ratio":   {2015:58.0, 2016:58.6, 2017:60.1, 2018:59.7, 2019:60.9,
                         2020:65.7, 2021:70.0, 2022:71.5, 2023:70.0, 2024:70.0, 2025:70.0},
        "expense_ratio":{2015:37.6, 2016:37.3, 2017:36.1, 2018:36.3, 2019:35.5,
                         2020:33.4, 2021:28.0, 2022:28.8, 2023:30.7, 2024:28.3, 2025:27.8},
        "target_pe": 8.0,
        "target_pev": 0.7,
        "drip_entry": 2015,    # Listed since 2007; full 11-year DRIP
    },
    "601601.SH": {
        "name": "中国太保 (China Pacific / CPIC)",
        "h_code5": "02601",
        "h_yf": "2601.HK",
        "biz_type": "Life+P&C balanced",
        # Group EV (CNY 亿)
        "ev_group": {2015:2210, 2016:2461, 2017:2861, 2018:3361, 2019:3959,
                     2020:4478, 2021:4983, 2022:5200, 2023:5295, 2024:5628, 2025:5900},
        # CPIC Life NBV (CNY 亿)
        "nbv": {2015:121, 2016:191, 2017:267, 2018:271, 2019:246,
                2020:178, 2021:134, 2022:92, 2023:110, 2024:132, 2025:145},
        # CPIC P&C (太平洋财险) COR (%)
        "cor": {2015:99.8, 2016:99.2, 2017:98.8, 2018:98.4, 2019:98.3,
                2020:99.0, 2021:99.0, 2022:97.3, 2023:97.7, 2024:98.6, 2025:98.0},
        "loss_ratio":   {2015:60.0, 2016:61.5, 2017:61.2, 2018:62.1, 2019:60.2,
                         2020:62.7, 2021:69.6, 2022:69.6, 2023:69.1, 2024:70.0, 2025:69.5},
        "expense_ratio":{2015:39.8, 2016:37.7, 2017:37.6, 2018:36.3, 2019:38.1,
                         2020:36.3, 2021:29.4, 2022:27.7, 2023:28.6, 2024:28.6, 2025:28.5},
        "target_pe": 8.0,
        "target_pev": 0.5,
        "drip_entry": 2015,    # Listed since 2007; full 11-year DRIP
    },
}

# DRIP defaults — assume HK corporate holder via 沪股通
DRIP_INVEST = 1_000_000   # CNY 1,000,000
DRIP_FX     = 1.0
DRIP_WHT    = 0.10        # 10% PRC withholding tax for HK-incorporated holder via Stock Connect
DRIP_HOLDER = "HK Corporate (via 沪股通)"

# Light grey for DRIP title row (per skill spec)
LGREY = "FFF2F2F2"

# China M2 (CNY 万亿, year-end)
CN_M2 = {2015:139.23, 2016:155.01, 2017:167.68, 2018:182.67, 2019:198.65,
         2020:218.68, 2021:238.29, 2022:266.43, 2023:292.27, 2024:313.53, 2025:340.29}

# ============================================================
# Helpers
# ============================================================
def fmt(c, fill_color=None, font_color=BLACK, bold=False, italic=False,
        num_fmt=None, align="right", fill_grey=False):
    c.font = Font(name=FONT, size=SIZE, color=font_color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    c.border = BORDER
    if num_fmt: c.number_format = num_fmt
    if fill_grey: c.fill = PatternFill("solid", fgColor=GREY)
    if fill_color: c.fill = PatternFill("solid", fgColor=fill_color)

def to_yi(v):
    """元 → 亿元."""
    if v is None or pd.isna(v): return None
    return float(v) / 1e8

def safe_get(df, year, field):
    """tushare returns latest first; pick by end_date YYYY-12-31."""
    if df is None or df.empty: return None
    sub = df[df['end_date'] == f"{year}1231"]
    if sub.empty: return None
    v = sub.iloc[0].get(field)
    return None if pd.isna(v) else v

# ============================================================
# Data fetcher
# ============================================================
def fetch(ts_code, cfg):
    print(f"\n{'='*60}\nFetching {ts_code} ({cfg['name']})\n{'='*60}")
    pro = ts.pro_api(TS_TOKEN)

    # NOTE: tushare end_date filters by ANNOUNCEMENT date, not period end.
    # FY2025 annual was published 2026-03-27. To capture it, end_date must be ≥ 2026-04-30.
    print("  [1/5] Income statement (annual + Q1 2026)...")
    inc = pro.income(ts_code=ts_code, start_date='20141231', end_date='20260601', report_type='1')
    time.sleep(0.4)
    print("  [2/5] Balance sheet (annual + Q1 2026)...")
    bs = pro.balancesheet(ts_code=ts_code, start_date='20141231', end_date='20260601', report_type='1')
    time.sleep(0.4)
    print("  [3/5] Cash flow (annual + Q1 2026)...")
    cf = pro.cashflow(ts_code=ts_code, start_date='20141231', end_date='20260601', report_type='1')
    time.sleep(0.4)
    # Q1 2026 quarterly data (period='20260331')
    inc_q1 = pro.income(ts_code=ts_code, period='20260331', report_type='1')
    time.sleep(0.3)
    bs_q1 = pro.balancesheet(ts_code=ts_code, period='20260331', report_type='1')
    time.sleep(0.3)
    cf_q1 = pro.cashflow(ts_code=ts_code, period='20260331', report_type='1')
    time.sleep(0.3)

    print("  [4/5] Daily basic (A-share total_share, total_mv) + latest Q1 2026...")
    db = pro.daily_basic(ts_code=ts_code, start_date='20151231', end_date='20260430',
                         fields='ts_code,trade_date,close,total_share,total_mv,pe,pb')
    time.sleep(0.4)
    # YE A-share price/shares
    a_data = {}
    for y in YEARS:
        sub = db[db['trade_date'].astype(str).str.startswith(str(y))]
        if not sub.empty:
            row = sub.sort_values('trade_date').iloc[-1]
            a_data[y] = {
                'a_close': float(row['close']),
                'total_share_yi': float(row['total_share']) / 1e4,  # 万股 → 亿股
                'a_mktcap_yi': float(row['total_mv']) / 1e4,        # 万元 → 亿元
            }

    print("  [5/5] H-share price (yfinance)...")
    h_data = {}
    try:
        tk = yf.Ticker(cfg['h_yf'])
        hist = tk.history(start="2015-01-01", end="2026-04-30", auto_adjust=False)
        ye = hist['Close'].resample('YE').last()
        for ts_, v in ye.items():
            y = ts_.year
            if y in YEARS: h_data[y] = float(v)
        print(f"    yfinance: {len(h_data)} years")
    except Exception as e:
        print(f"    yfinance ERR: {e}")

    # Top 10 shareholders (latest)
    print("  [+] Top 10 shareholders...")
    top10 = None
    try:
        # ts_code → 'sh601318'
        code = ts_code.split('.')[0]
        market = 'sh' if ts_code.endswith('.SH') else 'sz'
        top10 = ak.stock_gdfx_top_10_em(symbol=f"{market}{code}", date="20251231")
    except Exception as e:
        print(f"    top10 ERR: {e}")

    # Try to get H-share count (from listing tables, fallback to manual)
    # We'll compute total H shares as: (total_share - A_total_share). For Chinese insurers,
    # A vs H split is published. Fall back: use a config H-share count.
    # Actually H-share total is fixed for these names; embed manually:
    H_SHARES_YI = {
        "601319.SH": 68.88,    # 中国人保 H 股 ≈ 68.88 亿股
        "601318.SH": 74.48,    # 中国平安 H 股 ≈ 74.48 亿股
        "601601.SH": 27.71,    # 中国太保 H 股 ≈ 27.71 亿股
    }

    # Latest 2026 trading day (for "current" market data row)
    latest = {}
    db26 = db[db['trade_date'].astype(str).str.startswith('2026')]
    if not db26.empty:
        rrow = db26.sort_values('trade_date').iloc[-1]
        latest = {
            'date': str(rrow['trade_date']),
            'close': float(rrow['close']),
            'total_share_yi': float(rrow['total_share']) / 1e4,
            'a_mktcap_yi': float(rrow['total_mv']) / 1e4,
            'pe': float(rrow['pe']) if pd.notna(rrow['pe']) else None,
            'pb': float(rrow['pb']) if pd.notna(rrow['pb']) else None,
        }

    return {
        'inc': inc, 'bs': bs, 'cf': cf, 'db': db,
        'inc_q1': inc_q1, 'bs_q1': bs_q1, 'cf_q1': cf_q1,
        'a_data': a_data, 'h_data': h_data, 'top10': top10,
        'h_shares_yi': H_SHARES_YI[ts_code],
        'latest': latest,
    }

# ============================================================
# Workbook builder
# ============================================================
def build_workbook(ts_code, cfg, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Data"

    inc, bs, cf = data['inc'], data['bs'], data['cf']
    a_data, h_data = data['a_data'], data['h_data']

    # ----- Sheet 1: Financial Data (with Market Data + Q1 2026 sections) -----
    build_sheet1_financial(ws, cfg, inc, bs, cf, a_data, data['inc_q1'], data['bs_q1'], data['latest'])

    # ----- Sheet 2: DRIP Analysis (NEW) -----
    ws_drip = wb.create_sheet("DRIP Analysis")
    build_sheet_drip(ws_drip, ts_code, cfg)

    # Sheet 3 — Key Ratios (was Sheet 2)
    ws2 = wb.create_sheet("Key Ratios")
    build_sheet2_ratios(ws2, cfg)

    # Sheet 3 — EV & NBV
    ws3 = wb.create_sheet("EV & NBV Analysis")
    build_sheet3_ev_nbv(ws3, cfg)

    # Sheet 4 — Combined Operating Ratio
    ws4 = wb.create_sheet("COR Analysis")
    build_sheet4_cor(ws4, cfg)

    # Sheet 5 — Valuation Model
    ws5 = wb.create_sheet("Valuation Model")
    build_sheet5_valuation(ws5, cfg)

    # Sheet 6 — A+H Market Cap
    ws6 = wb.create_sheet("A+H Market Cap")
    build_sheet6_mktcap(ws6, cfg, a_data, h_data, data['h_shares_yi'])

    # Sheet 7 — Top 10 Shareholders
    ws7 = wb.create_sheet("Top 10 Shareholders")
    build_sheet7_holders(ws7, cfg, data['top10'])

    # Sheet 8 — Data Correlation
    ws8 = wb.create_sheet("Data Correlation")
    build_sheet8_correlation(ws8, ts_code, cfg)

    return wb

# ===== Sheet 1: Financial Data =====
def build_sheet1_financial(ws, cfg, inc, bs, cf, a_data=None, inc_q1=None, bs_q1=None, latest=None):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+len(YEARS))].width = 8.83

    # Title
    ws.cell(1,1, f"{cfg['name']} — Financial Data (CNY 亿元)")
    fmt(ws.cell(1,1), bold=True, align="left")

    # Header row
    ws.cell(3,1, "Item")
    fmt(ws.cell(3,1), bold=True, fill_grey=True, align="left")
    for i,y in enumerate(YEARS):
        c = ws.cell(3,2+i, y)
        fmt(c, bold=True, fill_grey=True, num_fmt="0", align="center")

    rows = [
        ("==INCOME STATEMENT==", "header"),
        ("Gross Premium Written 原保费收入", "prem_income"),
        ("Earned Premium 已赚保费", "prem_earned"),
        ("Investment Income 投资收益", "invest_income"),
        ("Other Operating Income 其他业务收入", "_other_op_inc"),
        ("Total Operating Revenue 营业收入", "total_revenue"),
        ("Claims & Benefits 赔付支出", "compens_payout"),
        ("Reserve Provision 提取保险责任准备金", "reser_insur_liab"),
        ("Acquisition + Admin Costs 手续费佣金+管理费", "_acq_admin"),
        ("Other Operating Costs", "_other_op_cost"),
        ("Operating Profit 营业利润", "operate_profit"),
        ("Total Profit 利润总额", "total_profit"),
        ("Income Tax 所得税", "income_tax"),
        ("Net Income 净利润", "n_income"),
        ("NI to Parent 归母净利润", "n_income_attr_p"),
        ("Minority Interest 少数股东损益", "minority_gain"),
        ("==BALANCE SHEET==", "header"),
        ("Cash & Equivalents 货币资金", "money_cap"),
        ("Investment Assets 投资资产", "_invest_assets"),
        ("Reinsurance Receivables 应收分保账款", "reinsur_receiv"),
        ("Premium Receivables 应收保费", "premium_receiv"),
        ("Other Assets 其他资产", "_other_assets"),
        ("Total Assets 总资产", "total_assets"),
        ("Insurance Reserves 保险责任准备金合计", "_insurance_reserves"),
        ("Policyholder Investments 保户投资款", "ph_invest"),
        ("Other Liabilities 其他负债", "_other_liab"),
        ("Total Liabilities 总负债", "total_liab"),
        ("Equity to Parent 归母权益", "total_hldr_eqy_exc_min_int"),
        ("Minority Interest 少数股东权益", "minority_int"),
        ("Total Equity 合计权益", "total_hldr_eqy_inc_min_int"),
        ("==CASH FLOW==", "header"),
        ("Net Cash from Operating", "n_cashflow_act"),
        ("Net Cash from Investing", "n_cashflow_inv_act"),
        ("Net Cash from Financing", "n_cash_flows_fnc_act"),
        ("Capex 资本性支出", "c_pay_acq_const_fiolta"),
    ]

    r = 4
    for label, field in rows:
        ws.cell(r,1, label)
        if label.startswith("=="):
            fmt(ws.cell(r,1), bold=True, fill_grey=True, align="left")
            for i in range(len(YEARS)):
                fmt(ws.cell(r,2+i), fill_grey=True)
        else:
            fmt(ws.cell(r,1), align="left")
            for i, y in enumerate(YEARS):
                v = None
                if field == "_other_op_inc":
                    # total_revenue - prem_earned - invest_income
                    tot = safe_get(inc, y, 'total_revenue')
                    pe = safe_get(inc, y, 'prem_earned')
                    ii = safe_get(inc, y, 'invest_income')
                    if tot is not None:
                        v = tot - (pe or 0) - (ii or 0)
                elif field == "_acq_admin":
                    fc = safe_get(inc, y, 'comm_exp') or 0
                    ad = safe_get(inc, y, 'admin_exp') or 0
                    v = fc + ad if (fc or ad) else None
                elif field == "_other_op_cost":
                    # Plug from total_cogs minus key components
                    tc = safe_get(inc, y, 'total_cogs')
                    cp = safe_get(inc, y, 'compens_payout') or 0
                    rs = safe_get(inc, y, 'reser_insur_liab') or 0
                    cm = safe_get(inc, y, 'comm_exp') or 0
                    ad = safe_get(inc, y, 'admin_exp') or 0
                    if tc is not None:
                        v = tc - cp - rs - cm - ad
                elif field == "_invest_assets":
                    parts = ['fa_avail_for_sale','htm_invest','lt_eqt_invest',
                             'invest_real_estate','trad_asset','debt_invest','oth_debt_invest']
                    s = 0; got = False
                    for p in parts:
                        x = safe_get(bs, y, p)
                        if x is not None: s += x; got = True
                    v = s if got else None
                elif field == "_other_assets":
                    ta = safe_get(bs, y, 'total_assets') or 0
                    mc = safe_get(bs, y, 'money_cap') or 0
                    inv_parts = ['fa_avail_for_sale','htm_invest','lt_eqt_invest',
                                 'invest_real_estate','trad_asset','debt_invest','oth_debt_invest']
                    inv = sum((safe_get(bs, y, p) or 0) for p in inv_parts)
                    rr = safe_get(bs, y, 'reinsur_receiv') or 0
                    pr = safe_get(bs, y, 'premium_receiv') or 0
                    if ta:
                        v = ta - mc - inv - rr - pr
                elif field == "_insurance_reserves":
                    parts = ['reser_lins_liab','reser_lthins_liab','reser_une_prem','reser_outstd_claims']
                    s = 0; got = False
                    for p in parts:
                        x = safe_get(bs, y, p)
                        if x is not None: s += x; got = True
                    v = s if got else None
                elif field == "_other_liab":
                    tl = safe_get(bs, y, 'total_liab') or 0
                    parts = ['reser_lins_liab','reser_lthins_liab','reser_une_prem','reser_outstd_claims','ph_invest']
                    res = sum((safe_get(bs, y, p) or 0) for p in parts)
                    if tl:
                        v = tl - res
                elif field in ('n_cashflow_act','n_cashflow_inv_act','n_cash_flows_fnc_act','c_pay_acq_const_fiolta'):
                    v = safe_get(cf, y, field)
                else:
                    src = bs if field in [c for c in bs.columns] else inc
                    v = safe_get(src, y, field)
                c = ws.cell(r, 2+i, to_yi(v) if v is not None else None)
                fmt(c, font_color=BLUE, num_fmt=NUM_FMT)
        r += 1

    # ===== Market Data section (rows 39+) =====
    md_start = max(r, 39)
    # Section header
    ws.cell(md_start, 1, "==MARKET DATA==")
    fmt(ws.cell(md_start,1), bold=True, fill_grey=True, align="left")
    for i in range(len(YEARS)):
        fmt(ws.cell(md_start, 2+i), fill_grey=True)

    md_rows = [
        (md_start+1, "A-share YE Close (CNY)", "px"),
        (md_start+2, "Total Shares Outstanding (亿股)", "shares"),
        (md_start+3, "Market Cap (CNY 亿)", "mc"),
        (md_start+4, "Dividends Paid (CNY 亿, neg)", "div"),
        (md_start+5, "Book Value Per Share (CNY)", "bvps"),
        (md_start+6, "P/B at Year-End", "pb"),
        (md_start+7, "P/E at Year-End", "pe"),
        (md_start+8, "Dividend Yield", "dy"),
    ]
    for rr_, label, kind in md_rows:
        ws.cell(rr_, 1, label); fmt(ws.cell(rr_,1), align="left")
        for i, y in enumerate(YEARS):
            cl = get_column_letter(2+i)
            cell = ws.cell(rr_, 2+i)
            if kind == "px":
                v = (a_data or {}).get(y, {}).get('a_close')
                if v is not None:
                    cell.value = v; fmt(cell, font_color=BLUE, num_fmt='#,##0.00')
            elif kind == "shares":
                v = (a_data or {}).get(y, {}).get('total_share_yi')
                if v is not None:
                    cell.value = v; fmt(cell, font_color=BLUE, num_fmt='#,##0.00')
            elif kind == "mc":
                v = (a_data or {}).get(y, {}).get('a_mktcap_yi')
                if v is not None:
                    cell.value = v; fmt(cell, font_color=BLUE, num_fmt=NUM_FMT)
            elif kind == "div":
                # Dividends paid: c_pay_dist_dpcp_int_exp from cashflow
                vv = safe_get(cf, y, 'c_pay_dist_dpcp_int_exp')
                if vv is not None:
                    cell.value = -abs(to_yi(vv))
                    fmt(cell, font_color=BLUE, num_fmt=NUM_FMT)
            elif kind == "bvps":
                # = Equity to Parent / shares
                cell.value = f"=IFERROR({cl}30/{cl}{md_start+2},\"-\")"
                fmt(cell, font_color=BLACK, num_fmt='#,##0.00')
            elif kind == "pb":
                # = Mkt Cap / Equity to Parent
                cell.value = f"=IFERROR({cl}{md_start+3}/{cl}30,\"-\")"
                fmt(cell, font_color=BLACK, num_fmt='0.00"x"')
            elif kind == "pe":
                cell.value = f"=IFERROR({cl}{md_start+3}/{cl}18,\"-\")"
                fmt(cell, font_color=BLACK, num_fmt='0.00"x"')
            elif kind == "dy":
                cell.value = f"=IFERROR(ABS({cl}{md_start+4})/{cl}{md_start+3},\"-\")"
                fmt(cell, font_color=BLACK, num_fmt='0.00%')

    # ===== Q1 2026 LATEST section =====
    q1_start = md_start + 10
    ws.cell(q1_start, 1, "==Q1 2026 LATEST (披露 2026-04-30)==")
    fmt(ws.cell(q1_start,1), bold=True, fill_grey=True, align="left")
    for i in range(len(YEARS)):
        fmt(ws.cell(q1_start, 2+i), fill_grey=True)

    # Q1 2026 fields
    def q1_get(df, field):
        if df is None or df.empty: return None
        sub = df[df['end_date'] == '20260331']
        if sub.empty: return None
        v = sub.iloc[0].get(field)
        return None if pd.isna(v) else v

    q1_items = [
        (q1_start+1, "Q1 Operating Revenue 营业收入",   to_yi(q1_get(inc_q1, 'total_revenue'))),
        (q1_start+2, "Q1 Operating Profit 营业利润",     to_yi(q1_get(inc_q1, 'operate_profit'))),
        (q1_start+3, "Q1 Net Income 净利润",              to_yi(q1_get(inc_q1, 'n_income'))),
        (q1_start+4, "Q1 NI to Parent 归母净利润",        to_yi(q1_get(inc_q1, 'n_income_attr_p'))),
        (q1_start+5, "Q1 Total Assets 总资产 (期末)",     to_yi(q1_get(bs_q1, 'total_assets'))),
        (q1_start+6, "Q1 Equity to Parent 归母权益 (期末)", to_yi(q1_get(bs_q1, 'total_hldr_eqy_exc_min_int'))),
    ]
    for rr_, label, val in q1_items:
        ws.cell(rr_, 1, label); fmt(ws.cell(rr_,1), align="left")
        cell = ws.cell(rr_, 2, val)
        fmt(cell, font_color=BLUE, num_fmt=NUM_FMT)
        # Annualized run-rate (×4) in column 3
        ws.cell(rr_, 3); cell2 = ws.cell(rr_, 3)
        if val is not None:
            cell2.value = val * 4
            fmt(cell2, font_color=BLACK, num_fmt=NUM_FMT, italic=True)
    # Add column header for Q1 column
    ws.cell(q1_start, 2, "Q1 2026")
    fmt(ws.cell(q1_start, 2), bold=True, fill_grey=True, align="center")
    ws.cell(q1_start, 3, "× 4 Annualised")
    fmt(ws.cell(q1_start, 3), bold=True, fill_grey=True, align="center", italic=True)

    # Latest market snapshot
    if latest:
        ws.cell(q1_start+8, 1, f"Latest Trading Day: {latest.get('date','')}")
        fmt(ws.cell(q1_start+8, 1), bold=True, italic=True, align="left")
        for j, (label, key, nf) in enumerate([
            ("Close (CNY)", 'close', '#,##0.00'),
            ("Total Shares (亿)", 'total_share_yi', '#,##0.00'),
            ("Market Cap (亿)", 'a_mktcap_yi', NUM_FMT),
            ("PE", 'pe', '0.00"x"'),
            ("PB", 'pb', '0.00"x"'),
        ]):
            ws.cell(q1_start+9+j, 1, label); fmt(ws.cell(q1_start+9+j,1), align="left")
            v = latest.get(key)
            if v is not None:
                cell = ws.cell(q1_start+9+j, 2, v); fmt(cell, font_color=BLUE, num_fmt=nf)

    for rr in range(1, q1_start+15):
        ws.row_dimensions[rr].height = 20

# ===== Sheet 2: Key Ratios =====
def build_sheet2_ratios(ws, cfg):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} — Key Ratios"); fmt(ws.cell(1,1), bold=True, align="left")

    ws.cell(3,1, "Ratio"); fmt(ws.cell(3,1), bold=True, fill_grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(3,2+i, y), bold=True, fill_grey=True, num_fmt="0", align="center")

    fd = "'Financial Data'"
    rows = [
        ("Premium Growth (YoY)", "growth", 5),  # row 5 of FD = prem_income
        ("Investment Yield (Inv Inc / Avg Inv Assets)", "inv_yield", None),
        ("Operating Profit Margin", "op_margin", None),
        ("Net Profit Margin", "ni_margin", None),
        ("Effective Tax Rate", "etr", None),
        ("ROE (NI to Parent / Avg Equity to Parent)", "roe", None),
        ("ROA (NI / Avg Total Assets)", "roa", None),
        ("Equity / Assets (Leverage)", "lev", None),
        ("Reserve / Total Liabilities", "rsv_ratio", None),
        ("Dividend Payout Ratio", "payout", None),
    ]
    r = 4
    # Map row offsets in Financial Data sheet:
    # FD row 4 = "==IS==" header. Then 5=prem_income, 6=earned_prem, 7=invest_income,
    # ... NI to Parent at row 18 (5+13), total_revenue at row 9, operate_profit at 14
    # Let me recount: rows list starts at 4 with "==IS=="
    # 4: ==IS==
    # 5: prem_income, 6: prem_earned, 7: invest_income, 8: other_op_inc, 9: total_revenue
    # 10: compens_payout, 11: reser_insur, 12: acq_admin, 13: other_op_cost, 14: operate_profit
    # 15: total_profit, 16: income_tax, 17: n_income, 18: n_income_attr_p, 19: minority_gain
    # 20: ==BS==
    # 21: money_cap, 22: invest_assets, 23: reinsur_receiv, 24: premium_receiv, 25: other_assets
    # 26: total_assets, 27: insurance_reserves, 28: ph_invest, 29: other_liab, 30: total_liab
    # 31: equity_parent, 32: minority_int, 33: total_equity
    R = {'prem_income':5,'prem_earned':6,'invest_income':7,'total_revenue':9,
         'compens_payout':10,'reser':11,'operate_profit':14,'total_profit':15,
         'income_tax':16,'n_income':17,'ni_p':18,'mi':19,
         'money_cap':21,'invest_assets':22,'total_assets':26,'reserves':27,
         'total_liab':30,'eq_p':31,'mi_eq':32,'tot_eq':33}

    for label, key, _ in rows:
        ws.cell(r,1, label); fmt(ws.cell(r,1), align="left")
        for i,y in enumerate(YEARS):
            col = get_column_letter(2+i); pcol = get_column_letter(1+i)
            f = None
            if key == "growth" and i > 0:
                f = f"=IFERROR({fd}!{col}{R['prem_income']}/{fd}!{pcol}{R['prem_income']}-1,\"\")"
            elif key == "inv_yield" and i > 0:
                f = f"=IFERROR({fd}!{col}{R['invest_income']}/AVERAGE({fd}!{col}{R['invest_assets']},{fd}!{pcol}{R['invest_assets']}),\"\")"
            elif key == "op_margin":
                f = f"=IFERROR({fd}!{col}{R['operate_profit']}/{fd}!{col}{R['total_revenue']},\"\")"
            elif key == "ni_margin":
                f = f"=IFERROR({fd}!{col}{R['n_income']}/{fd}!{col}{R['total_revenue']},\"\")"
            elif key == "etr":
                f = f"=IFERROR({fd}!{col}{R['income_tax']}/{fd}!{col}{R['total_profit']},\"\")"
            elif key == "roe" and i > 0:
                f = f"=IFERROR({fd}!{col}{R['ni_p']}/AVERAGE({fd}!{col}{R['eq_p']},{fd}!{pcol}{R['eq_p']}),\"\")"
            elif key == "roa" and i > 0:
                f = f"=IFERROR({fd}!{col}{R['n_income']}/AVERAGE({fd}!{col}{R['total_assets']},{fd}!{pcol}{R['total_assets']}),\"\")"
            elif key == "lev":
                f = f"=IFERROR({fd}!{col}{R['tot_eq']}/{fd}!{col}{R['total_assets']},\"\")"
            elif key == "rsv_ratio":
                f = f"=IFERROR({fd}!{col}{R['reserves']}/{fd}!{col}{R['total_liab']},\"\")"
            if f:
                c = ws.cell(r, 2+i)
                c.value = f
                fmt(c, font_color=GREEN, num_fmt=PCT_FMT)
        r += 1
    for rr in range(1, r+1):
        ws.row_dimensions[rr].height = 20

# ===== Sheet 3: EV & NBV =====
def build_sheet3_ev_nbv(ws, cfg):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} — EV & NBV Analysis (CNY 亿元)"); fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(2,1, "All EV/NBV figures are MANUALLY input from annual reports — verify before use.")
    fmt(ws.cell(2,1), italic=True, align="left", font_color=BLUE)

    ws.cell(4,1, "Item"); fmt(ws.cell(4,1), bold=True, fill_grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(4,2+i, y), bold=True, fill_grey=True, num_fmt="0", align="center")

    rows = [
        ("Group Embedded Value (EV)", "ev"),
        ("EV Growth (YoY)", "ev_growth"),
        ("New Business Value (NBV)", "nbv"),
        ("NBV Growth (YoY)", "nbv_growth"),
        ("NBV / EV (NB contribution)", "nbv_ev"),
        ("Implied Value of In-Force (VIF) ≈ EV - Net Asset", "vif"),
    ]
    r = 5
    for label, key in rows:
        ws.cell(r,1, label); fmt(ws.cell(r,1), align="left")
        for i,y in enumerate(YEARS):
            col = get_column_letter(2+i); pcol = get_column_letter(1+i)
            if key == "ev":
                v = cfg['ev_group'].get(y)
                c = ws.cell(r,2+i, v); fmt(c, font_color=BLUE, num_fmt=NUM_FMT)
            elif key == "nbv":
                v = cfg['nbv'].get(y)
                c = ws.cell(r,2+i, v); fmt(c, font_color=BLUE, num_fmt=NUM_FMT)
            elif key == "ev_growth" and i > 0:
                f = f"=IFERROR({col}5/{pcol}5-1,\"\")"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=PCT_FMT)
            elif key == "nbv_growth" and i > 0:
                f = f"=IFERROR({col}7/{pcol}7-1,\"\")"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=PCT_FMT)
            elif key == "nbv_ev":
                f = f"=IFERROR({col}7/{col}5,\"\")"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=PCT_FMT)
            elif key == "vif":
                # = EV - Equity to Parent (from Financial Data sheet row 31)
                f = f"=IFERROR({col}5-'Financial Data'!{col}31,\"\")"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=GREEN, num_fmt=NUM_FMT)
        r += 1
    for rr in range(1, r+2):
        ws.row_dimensions[rr].height = 20

# ===== Sheet 4: Combined Operating Ratio =====
def build_sheet4_cor(ws, cfg):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} — P&C Combined Operating Ratio")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(2,1, "P&C subsidiary metrics (人保财险 / 平安产险 / 太平洋产险). Verify against 年报.")
    fmt(ws.cell(2,1), italic=True, align="left", font_color=BLUE)

    ws.cell(4,1, "Ratio (%)"); fmt(ws.cell(4,1), bold=True, fill_grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(4,2+i, y), bold=True, fill_grey=True, num_fmt="0", align="center")

    rows_cor = [
        ("Loss Ratio 综合赔付率", "loss_ratio"),
        ("Expense Ratio 综合费用率", "expense_ratio"),
        ("Combined Operating Ratio (COR) 综合成本率", "cor"),
        ("Underwriting Margin (1 - COR)", "uw_margin"),
    ]
    r = 5
    for label, key in rows_cor:
        ws.cell(r,1, label); fmt(ws.cell(r,1), align="left")
        for i,y in enumerate(YEARS):
            col = get_column_letter(2+i)
            if key in ("loss_ratio","expense_ratio","cor"):
                v = cfg[key].get(y)
                c = ws.cell(r,2+i, v/100 if v else None)
                fmt(c, font_color=BLUE, num_fmt=PCT_FMT)
            elif key == "uw_margin":
                f = f"=1-{col}7"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=PCT_FMT)
        r += 1
    for rr in range(1, r+2):
        ws.row_dimensions[rr].height = 20

# ===== Sheet 5: Valuation Model (3-method) =====
def build_sheet5_valuation(ws, cfg):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} — Valuation Model (3-method avg)")
    fmt(ws.cell(1,1), bold=True, align="left")

    ws.cell(3,1, "Item"); fmt(ws.cell(3,1), bold=True, fill_grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(3,2+i, y), bold=True, fill_grey=True, num_fmt="0", align="center")

    fd = "'Financial Data'"
    ev_sheet = "'EV & NBV Analysis'"
    mc_sheet = "'A+H Market Cap'"

    # Inputs at top
    rows = [
        ("==INPUTS==", "header"),
        ("Bond Yield (10y CGB) %", "bond"),
        ("COE Premium %", "coe"),
        ("Cost of Equity %", "cost_eq"),
        ("==METHOD 1: DIVIDEND YIELD==", "header"),
        ("Dividends Paid (亿)", "div_paid"),
        ("Implied Mkt Cap = Div / (Bond+1%)", "v1"),
        ("==METHOD 2: P/EV (寿险核心估值)==", "header"),
        ("Group EV (亿)", "ev_link"),
        ("Target P/EV multiple", "tgt_pev"),
        ("Implied Mkt Cap = EV × P/EV", "v2"),
        ("==METHOD 3: TARGET P/E==", "header"),
        ("NI to Parent (亿)", "ni_link"),
        ("Target P/E", "tgt_pe"),
        ("Implied Mkt Cap = NI × P/E", "v3"),
        ("==SUMMARY==", "header"),
        ("Average Implied Mkt Cap", "avg_v"),
        ("Actual A+H Mkt Cap", "actual_mc"),
        ("Premium / (Discount)", "premdisc"),
    ]
    r = 4
    for label, key in rows:
        ws.cell(r,1, label)
        if label.startswith("=="):
            fmt(ws.cell(r,1), bold=True, fill_grey=True, align="left")
            for i in range(len(YEARS)):
                fmt(ws.cell(r,2+i), fill_grey=True)
        else:
            fmt(ws.cell(r,1), align="left")
            for i,y in enumerate(YEARS):
                col = get_column_letter(2+i)
                if key == "bond":
                    c = ws.cell(r,2+i, CN_BOND.get(y, 3.0)/100); fmt(c, font_color=BLUE, num_fmt=PCT_FMT)
                elif key == "coe":
                    c = ws.cell(r,2+i, 0.05); fmt(c, font_color=BLUE, num_fmt=PCT_FMT)
                elif key == "cost_eq":
                    f = f"={col}5+{col}6"
                    c = ws.cell(r,2+i); c.value=f; fmt(c, font_color=BLACK, num_fmt=PCT_FMT)
                elif key == "div_paid":
                    # rough: from CF - assume in cf row, but we don't have it on sheet 1.
                    # Use 30% of NI to Parent as default placeholder (BLUE so editable)
                    c = ws.cell(r,2+i)
                    c.value = f"=0.3*{fd}!{col}18"
                    fmt(c, font_color=BLUE, num_fmt=NUM_FMT)
                elif key == "v1":
                    f = f"=IFERROR({col}9/({col}5+0.01),\"\")"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT)
                elif key == "ev_link":
                    f = f"={ev_sheet}!{col}5"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=GREEN, num_fmt=NUM_FMT)
                elif key == "tgt_pev":
                    c = ws.cell(r,2+i, cfg['target_pev']); fmt(c, font_color=BLUE, num_fmt='0.00"x"')
                elif key == "v2":
                    f = f"={col}12*{col}13"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT)
                elif key == "ni_link":
                    f = f"={fd}!{col}18"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=GREEN, num_fmt=NUM_FMT)
                elif key == "tgt_pe":
                    c = ws.cell(r,2+i, cfg['target_pe']); fmt(c, font_color=BLUE, num_fmt='0.0"x"')
                elif key == "v3":
                    f = f"={col}16*{col}17"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT)
                elif key == "avg_v":
                    f = f"=AVERAGE({col}10,{col}14,{col}18)"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT, bold=True)
                elif key == "actual_mc":
                    f = f"={mc_sheet}!{col}9"   # row 9 of mktcap sheet = total
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=GREEN, num_fmt=NUM_FMT)
                elif key == "premdisc":
                    f = f"=IFERROR({col}21/{col}20-1,\"\")"
                    c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=PCT_FMT, bold=True)
        r += 1
    for rr in range(1, r+2):
        ws.row_dimensions[rr].height = 17

# ===== Sheet 6: A+H Market Cap =====
def build_sheet6_mktcap(ws, cfg, a_data, h_data, h_shares_yi):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} — A+H Weighted Market Cap (CNY 亿)")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(2,1, f"H-share count assumed constant at {h_shares_yi:.2f} 亿股 (verify post-buybacks)")
    fmt(ws.cell(2,1), italic=True, align="left", font_color=BLUE)

    ws.cell(3,1, "Item"); fmt(ws.cell(3,1), bold=True, fill_grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(3,2+i, y), bold=True, fill_grey=True, num_fmt="0", align="center")

    rows = [
        ("Total Shares 总股本 (亿股)", "tot_sh"),
        ("H-share Count (亿股)", "h_sh"),
        ("A-share Count = Total - H (亿股)", "a_sh"),
        ("A-share YE Price (CNY)", "a_px"),
        ("H-share YE Price (HKD)", "h_px"),
        ("HKD/CNY FX (avg)", "fx"),
        ("A-share Mkt Cap (亿 CNY)", "a_mc"),
        ("H-share Mkt Cap (亿 CNY equiv)", "h_mc"),
        ("Total A+H Mkt Cap (亿 CNY)", "tot_mc"),
    ]
    # HKD/CNY rough averages
    FX = {2015:0.81,2016:0.86,2017:0.87,2018:0.84,2019:0.88,
          2020:0.89,2021:0.83,2022:0.86,2023:0.90,2024:0.92,2025:0.92}
    r = 4
    for label, key in rows:
        ws.cell(r,1, label); fmt(ws.cell(r,1), align="left")
        for i,y in enumerate(YEARS):
            col = get_column_letter(2+i)
            if key == "tot_sh":
                v = a_data.get(y,{}).get('total_share_yi')
                c = ws.cell(r,2+i, v); fmt(c, font_color=BLUE, num_fmt=NUM_FMT)
            elif key == "h_sh":
                c = ws.cell(r,2+i, h_shares_yi); fmt(c, font_color=BLUE, num_fmt=NUM_FMT)
            elif key == "a_sh":
                f = f"={col}4-{col}5"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT)
            elif key == "a_px":
                v = a_data.get(y,{}).get('a_close')
                c = ws.cell(r,2+i, v); fmt(c, font_color=BLUE, num_fmt='0.00')
            elif key == "h_px":
                v = h_data.get(y)
                c = ws.cell(r,2+i, v); fmt(c, font_color=BLUE, num_fmt='0.00')
            elif key == "fx":
                c = ws.cell(r,2+i, FX.get(y, 0.90)); fmt(c, font_color=BLUE, num_fmt='0.0000')
            elif key == "a_mc":
                f = f"={col}6*{col}7"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT)
            elif key == "h_mc":
                f = f"={col}5*{col}8*{col}9"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT)
            elif key == "tot_mc":
                f = f"={col}10+{col}11"
                c = ws.cell(r,2+i); c.value = f; fmt(c, font_color=BLACK, num_fmt=NUM_FMT, bold=True)
        r += 1
    for rr in range(1, r+2):
        ws.row_dimensions[rr].height = 20

# ===== Sheet 7: Top 10 Shareholders =====
def build_sheet7_holders(ws, cfg, top10):
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    ws.cell(1,1, f"{cfg['name']} — Top 10 Shareholders (latest)")
    fmt(ws.cell(1,1), bold=True, align="left")

    headers = ["#", "Shareholder", "Shares Held", "% of Total", "Type"]
    for i,h in enumerate(headers):
        c = ws.cell(3,1+i, h); fmt(c, bold=True, fill_grey=True, align="center" if i==0 else "left")

    if top10 is not None and not top10.empty:
        # Try common column names
        cols = list(top10.columns)
        name_col = next((c for c in cols if '股东' in c or '名称' in c or 'holder' in c.lower()), cols[0])
        share_col = next((c for c in cols if '持股数' in c or '数量' in c or '股份' in c), None)
        pct_col = next((c for c in cols if '比例' in c or '%' in c), None)
        for i, row in top10.head(10).iterrows():
            r = 4 + i
            ws.cell(r,1, i+1); fmt(ws.cell(r,1), align="center")
            ws.cell(r,2, str(row[name_col])[:60]); fmt(ws.cell(r,2), align="left")
            if share_col:
                v = row[share_col]
                try: v = float(str(v).replace(',','').replace('万股','').replace('亿股',''))
                except: v = None
                ws.cell(r,3, v); fmt(ws.cell(r,3), num_fmt=INT_FMT)
            if pct_col:
                v = row[pct_col]
                try: v = float(str(v).replace('%',''))/100
                except: v = None
                ws.cell(r,4, v); fmt(ws.cell(r,4), num_fmt=PCT_FMT)
            fmt(ws.cell(r,5))
    else:
        ws.cell(4,1, "Data not retrieved — query 东财/同花顺 manually")
        fmt(ws.cell(4,1), italic=True, align="left")

    for rr in range(1, 16):
        ws.row_dimensions[rr].height = 15

# ===== Sheet 8: Data Correlation =====
def build_sheet8_correlation(ws, ts_code, cfg):
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 110

    ws.cell(1,1, f"{cfg['name']} ({ts_code}) — Data Correlation & Methodology")
    fmt(ws.cell(1,1), bold=True, align="left")

    sections = [
        ("1. Company Overview",
         f"{cfg['name']} — {cfg['biz_type']}. Dual-listed: A-share {ts_code} on SSE, "
         f"H-share {cfg['h_code5']}.HK ({cfg['h_yf']}). 11-year history 2015-2025 in CNY 亿元."),
        ("2. Data Sources",
         "• tushare pro API: I/S, B/S, CF (annual report_type=1) — A-share filer.\n"
         "• yfinance: H-share year-end closing price (split/div-adjusted).\n"
         "• tushare daily_basic: A-share total_share, total_mv → A-side market cap.\n"
         "• akshare stock_gdfx_top_10_em: top-10 shareholders snapshot.\n"
         "• Manual input (BLUE cells): EV, NBV, P&C COR — sourced from annual reports."),
        ("3. EV & NBV Caveats",
         "• EV = Embedded Value, the regulator-approved actuarial valuation of in-force life book + adjusted net asset.\n"
         "• NBV = New Business Value, reflects PV of future profits from current-year new policies.\n"
         "• Pre-2023: Reported under C-ROSS Phase II (China Risk Oriented Solvency System).\n"
         "• 2023+: Some insurers restated EV/NBV under IFRS 17 / new accounting standards — "
         "discontinuity vs. prior years; CHECK 年报 reconciliation tables.\n"
         "• Group EV here = whole-group consolidated EV (Life + Health + P&C book value).\n"
         "• For PICC: P&C dominates revenue, so EV figure is materially smaller than P&C book."),
        ("4. Combined Operating Ratio (COR) Caveats",
         "• COR = Loss Ratio + Expense Ratio. <100% = underwriting profit; >100% = underwriting loss.\n"
         "• Numbers shown are P&C subsidiary level (not group consolidated):\n"
         "  - 601319.SH → 人保财险 (PICC P&C, separately listed as 2328.HK)\n"
         "  - 601318.SH → 平安产险 (Ping An P&C, sub of group)\n"
         "  - 601601.SH → 太平洋产险 (CPIC P&C, sub of group)\n"
         "• Auto VAT reform (2020) caused expense ratio drop ~5pp across industry — methodology break.\n"
         "• 2024-2025: Catastrophe losses (typhoons, floods, EV claims inflation) pushed COR higher."),
        ("5. Valuation Methodology",
         "Three-method average (weights: 1/3 each):\n"
         "Method 1 — Dividend Yield: Implied MC = Div / (Bond + 1%). Conservative floor.\n"
         "Method 2 — P/EV: Implied MC = Group EV × Target P/EV multiple.\n"
         "  • Target P/EV defaults: PICC 0.8x, Ping An 0.7x, CPIC 0.5x (mainland avg 0.4-1.0x).\n"
         "Method 3 — Target P/E: Implied MC = NI × Target P/E.\n"
         "  • Target P/E defaults: PICC 7x, Ping An 8x, CPIC 8x.\n"
         "Premium/(Discount) = Avg Implied / Actual A+H Mkt Cap − 1."),
        ("6. A+H Market Cap Methodology",
         "• A-share: A_close (CNY) × A_share_count\n"
         "• H-share: H_close (HKD) × H_share_count × HKD/CNY FX\n"
         "• Both blended at YE FX (rough avg, refine with PBoC central parity if needed).\n"
         "• Total = A_MC + H_MC, in CNY 亿元.\n"
         "• H-share counts (亿股, fixed): PICC 68.88 / Ping An 74.48 / CPIC 27.71.\n"
         "  Verify against latest 年报 if there have been buybacks or H→A conversions."),
        ("7. Insurance Reserve Methodology",
         "• Insurance Reserves row aggregates: 寿险责任准备金 + 长期健康险责任准备金 + 未到期责任准备金 + 未决赔款.\n"
         "• Policyholder Investment Funds (保户投资款) shown separately — 'investment-linked' liabilities.\n"
         "• These two together typically = 80-90% of total liabilities for life-heavy insurers."),
        ("8. Known Limitations",
         "• EV/NBV are MANUAL input, sourced from annual reports — verify each cell.\n"
         "• 2025 figures (EV/NBV/COR) are PRELIMINARY estimates pending FY25 disclosure (~April 2026).\n"
         "• Tushare insurance fields use generic schema — some 'investment income' subtotals may "
         "differ from 年报 line items by reclassification.\n"
         "• PICC Group consolidates 人保财险 (1339.HK separately listed); intra-group eliminations "
         "not separately broken out.\n"
         "• Ping An Group includes Ping An Bank (000001.SZ) — banking NII not separated here.\n"
         "• Dividend payout in Valuation sheet defaulted to 30% of NI (BLUE editable) — replace with "
         "actual cash dividend from 利润分配表 for precision."),
    ]
    r = 3
    for title, body in sections:
        ws.cell(r,1, "")
        ws.cell(r,2, title); fmt(ws.cell(r,2), bold=True, fill_grey=True, align="left")
        r += 1
        for line in body.split("\n"):
            ws.cell(r,2, line); fmt(ws.cell(r,2), align="left")
            ws.row_dimensions[r].height = 17
            r += 1
        r += 1

# ============================================================
# Sheet 2: DRIP Analysis (NEW)
# ============================================================
def build_sheet_drip(ws, ts_code, cfg):
    """DRIP roll-forward for HK corporate holder via 沪股通. WHT 10%, FX 1.0, init CNY 1M.
    Years from cfg['drip_entry'] to 2025. Includes per-year P/B row.
    """
    entry = cfg['drip_entry']
    drip_years = [y for y in YEARS if y >= entry]
    n_years = len(drip_years)

    # Column widths per skill spec
    ws.column_dimensions['A'].width = 38.33
    ws.column_dimensions['B'].width = 15.83
    for i, y in enumerate(drip_years):
        ws.column_dimensions[get_column_letter(2+i)].width = 13
    ws.column_dimensions[get_column_letter(2+n_years)].width = 8.83

    fd = "'Financial Data'"
    # Find FD-sheet rows for: Equity to Parent (row 30), Dividends (md_start+4), Px (md_start+1),
    # Shares (md_start+2), MktCap (md_start+3), PB (md_start+6).
    # md_start = 39 always since rows are fixed in build_sheet1_financial.
    FD_EQ_PARENT = 30      # Equity to Parent row in FD
    FD_PX        = 40      # = md_start+1
    FD_SHARES    = 41      # = md_start+2
    FD_MC        = 42      # = md_start+3
    FD_DIV       = 43      # = md_start+4
    FD_BVPS      = 44      # = md_start+5
    FD_PB        = 45      # = md_start+6
    FD_NI_P      = 18      # NI to Parent

    # Year-column offset on FD: drip_years[0] maps to FD col B + (entry-2015)
    year_offset = entry - YEARS[0]   # for entry=2018, offset=3 → FD col E

    def fd_col(i):
        """Column letter on FD sheet for the i-th drip year."""
        return get_column_letter(2 + year_offset + i)

    def col(i):
        """Column letter on this DRIP sheet for the i-th drip year."""
        return get_column_letter(2 + i)

    # Title (12pt bold, light grey)
    title = f"DRIP Return Analysis — {ts_code} ({drip_years[0]}–{drip_years[-1]}, {DRIP_HOLDER})"
    c = ws.cell(1,1, title)
    c.font = Font(name=FONT, size=12, color=BLACK, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = PatternFill("solid", fgColor=LGREY)
    c.border = BORDER
    # Span fill across columns
    for j in range(2, 2+n_years+1):
        cc = ws.cell(1, j)
        cc.fill = PatternFill("solid", fgColor=LGREY)
        cc.font = Font(name=FONT, size=12, bold=True)

    ws.cell(2,1, f"Source: 'Financial Data' rows 18 (NI), 30 (Equity), {FD_DIV} (Div), {FD_PX} (Price), {FD_SHARES} (Shares)")
    fmt(ws.cell(2,1), italic=True, font_color=BLUE, align="left")

    # ---- Assumptions ----
    ws.cell(4,1, "Assumptions"); fmt(ws.cell(4,1), bold=True, align="left")
    ws.cell(5,1, "Initial investment (CNY)"); fmt(ws.cell(5,1), align="left")
    fmt(ws.cell(5,2, DRIP_INVEST), font_color=BLUE, bold=True, num_fmt=INT_FMT)
    ws.cell(6,1, f"Entry price (CNY/sh, {entry} YE)"); fmt(ws.cell(6,1), align="left")
    c = ws.cell(6,2); c.value = f"={fd}!{fd_col(0)}{FD_PX}"
    fmt(c, font_color=GREEN, bold=True, num_fmt='#,##0.00')
    ws.cell(7,1, "FX (CNY denominator)"); fmt(ws.cell(7,1), align="left")
    fmt(ws.cell(7,2, DRIP_FX), font_color=BLUE, bold=True, num_fmt='0.00')
    ws.cell(8,1, "Dividend WHT (HK Corp via 沪股通)"); fmt(ws.cell(8,1), align="left")
    fmt(ws.cell(8,2, DRIP_WHT), font_color=BLUE, bold=True, num_fmt='0.00%')
    ws.cell(9,1, "Initial shares purchased"); fmt(ws.cell(9,1), align="left")
    c = ws.cell(9,2); c.value = "=B5/B6"; fmt(c, num_fmt=INT_FMT)

    # ---- Year-by-year DRIP ----
    ws.cell(11,1, "Year-by-year DRIP"); fmt(ws.cell(11,1), bold=True, align="left")

    # Row 12: Year header
    ws.cell(12,1, "Year"); fmt(ws.cell(12,1), bold=True, align="left")
    for i, y in enumerate(drip_years):
        c = ws.cell(12, 2+i); c.value = f"={fd}!{fd_col(i)}3"
        fmt(c, bold=True, align="right")

    # Row 13: Shares BoY
    ws.cell(13,1, "Shares held BoY"); fmt(ws.cell(13,1), align="left")
    for i in range(n_years):
        if i == 0:
            ws.cell(13,2).value = "=B9"
        else:
            ws.cell(13, 2+i).value = f"={col(i-1)}18"
        fmt(ws.cell(13, 2+i), num_fmt=INT_FMT)

    # Row 14: DPS (CNY/share)
    ws.cell(14,1, "DPS (CNY, from CF)"); fmt(ws.cell(14,1), align="left")
    for i in range(n_years):
        ws.cell(14, 2+i).value = f"=IFERROR(ABS({fd}!{fd_col(i)}{FD_DIV})/{fd}!{fd_col(i)}{FD_SHARES},0)"
        fmt(ws.cell(14, 2+i), num_fmt='0.0000')

    # Row 15: After-tax dividend
    ws.cell(15,1, "Dividend (CNY, after WHT)"); fmt(ws.cell(15,1), align="left")
    for i in range(n_years):
        ws.cell(15, 2+i).value = f"={col(i)}13*{col(i)}14*$B$7*(1-$B$8)"
        fmt(ws.cell(15, 2+i), num_fmt=INT_FMT)

    # Row 16: Reinvest price
    ws.cell(16,1, "Reinvest price (CNY)"); fmt(ws.cell(16,1), align="left")
    for i in range(n_years):
        ws.cell(16, 2+i).value = f"={fd}!{fd_col(i)}{FD_PX}"
        fmt(ws.cell(16, 2+i), font_color=GREEN, num_fmt='#,##0.00')

    # Row 17: Shares bought
    ws.cell(17,1, "Shares bought"); fmt(ws.cell(17,1), align="left")
    for i in range(n_years):
        ws.cell(17, 2+i).value = f"=IFERROR({col(i)}15/{col(i)}16,0)"
        fmt(ws.cell(17, 2+i), num_fmt=INT_FMT)

    # Row 18: Shares EoY
    ws.cell(18,1, "Shares held EoY"); fmt(ws.cell(18,1), align="left")
    for i in range(n_years):
        ws.cell(18, 2+i).value = f"={col(i)}13+{col(i)}17"
        fmt(ws.cell(18, 2+i), num_fmt=INT_FMT)

    # Row 19: Portfolio value EoY
    ws.cell(19,1, "Portfolio value EoY (CNY)"); fmt(ws.cell(19,1), align="left")
    for i in range(n_years):
        bold = (i == n_years-1)
        ws.cell(19, 2+i).value = f"={col(i)}18*{col(i)}16"
        fmt(ws.cell(19, 2+i), bold=bold, num_fmt=INT_FMT)

    # Row 20: P/B at YE  (NEW per user request)
    ws.cell(20,1, "P/B at Year-End (= MktCap / Equity_to_Parent)")
    fmt(ws.cell(20,1), align="left")
    for i in range(n_years):
        ws.cell(20, 2+i).value = f"=IFERROR({fd}!{fd_col(i)}{FD_MC}/{fd}!{fd_col(i)}{FD_EQ_PARENT},\"-\")"
        fmt(ws.cell(20, 2+i), font_color=GREEN, num_fmt='0.00"x"')

    # Row 21: Dividend Yield
    ws.cell(21,1, "Dividend Yield (Div / MktCap)")
    fmt(ws.cell(21,1), align="left")
    for i in range(n_years):
        ws.cell(21, 2+i).value = f"=IFERROR(ABS({fd}!{fd_col(i)}{FD_DIV})/{fd}!{fd_col(i)}{FD_MC},0)"
        fmt(ws.cell(21, 2+i), num_fmt='0.00%')

    # ---- M2 reference ----
    ws.cell(23,1, "Reference: China M2 (CNY 万亿)"); fmt(ws.cell(23,1), bold=True, align="left")
    ws.cell(24,1, "M2 year-end balance"); fmt(ws.cell(24,1), align="left")
    for i, y in enumerate(drip_years):
        v = CN_M2.get(y)
        if v: fmt(ws.cell(24, 2+i, v), font_color=BLUE, bold=True, num_fmt='#,##0.00')

    last_col = col(n_years-1)
    cagr_n = n_years - 1
    ws.cell(25,1, f"M2 multiple vs {drip_years[0]}"); fmt(ws.cell(25,1), align="left")
    c = ws.cell(25, 2+n_years)
    c.value = f"={last_col}24/B24"; fmt(c, num_fmt='0.00"x"')
    ws.cell(26,1, f"M2 CAGR ({cagr_n} yrs)"); fmt(ws.cell(26,1), align="left")
    c = ws.cell(26, 2+n_years)
    c.value = f"=({last_col}24/B24)^(1/{cagr_n})-1"; fmt(c, num_fmt='0.00%')

    # ---- Summary ----
    ws.cell(28,1, "Summary"); fmt(ws.cell(28,1), bold=True, align="left")
    ws.cell(29,1, f"Final value {drip_years[-1]} YE (CNY)"); fmt(ws.cell(29,1), align="left")
    c = ws.cell(29,2); c.value = f"={last_col}19"; fmt(c, bold=True, num_fmt=INT_FMT)
    ws.cell(30,1, "Total return"); fmt(ws.cell(30,1), align="left")
    c = ws.cell(30,2); c.value = "=B29/B5-1"; fmt(c, num_fmt='0.00%')
    ws.cell(31,1, "Multiple on invested (x)"); fmt(ws.cell(31,1), align="left")
    c = ws.cell(31,2); c.value = "=B29/B5"; fmt(c, bold=True, num_fmt='0.00"x"')
    ws.cell(32,1, f"Annualized IRR ({n_years} yrs)"); fmt(ws.cell(32,1), align="left")
    c = ws.cell(32,2); c.value = f"=(B29/B5)^(1/{n_years})-1"; fmt(c, bold=True, num_fmt='0.00%')
    cell_m2cagr = f"{get_column_letter(2+n_years)}26"
    ws.cell(33,1, "vs M2 growth (DRIP IRR − M2 CAGR)"); fmt(ws.cell(33,1), align="left")
    c = ws.cell(33,2); c.value = f"=B32-{cell_m2cagr}"; fmt(c, num_fmt='0.00%')
    ws.cell(34,1, "Price-only return (no dividends)"); fmt(ws.cell(34,1), align="left")
    c = ws.cell(34,2); c.value = f"={fd}!{fd_col(n_years-1)}{FD_PX}/{fd}!{fd_col(0)}{FD_PX}-1"
    fmt(c, num_fmt='0.0%')
    ws.cell(35,1, "Avg P/B over period"); fmt(ws.cell(35,1), align="left")
    pb_range = f"{col(0)}20:{last_col}20"
    c = ws.cell(35,2); c.value = f"=IFERROR(AVERAGE({pb_range}),\"-\")"
    fmt(c, num_fmt='0.00"x"')

    # Row heights = 23pt per spec
    for rr in range(1, 36):
        ws.row_dimensions[rr].height = 23.0


# ============================================================
# Main
# ============================================================
def main():
    for ts_code, cfg in CONFIG.items():
        try:
            data = fetch(ts_code, cfg)
            print("Building workbook...")
            wb = build_workbook(ts_code, cfg, data)
            outdir = os.path.join(OUT_BASE, ts_code)
            os.makedirs(outdir, exist_ok=True)
            outpath = os.path.join(outdir, f"{ts_code}.Valuation.v{TODAY}.xlsx")
            wb.save(outpath)
            print(f"\n{'='*60}\n✓ Saved: {outpath}\n{'='*60}\n")
        except Exception as e:
            import traceback
            print(f"\n!! ERROR for {ts_code}: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    main()
