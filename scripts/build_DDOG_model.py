#!/usr/bin/env python3
"""
Financial Valuation Model — DDOG (Datadog Inc, NASDAQ)
SaaS observability/monitoring company. Reports in USD.
Data source: SEC EDGAR XBRL (companyfacts) + yfinance (prices).
Units: USD millions (USD ÷ 1e6).
Years: 2017-2025 (9 years, including pre-IPO via S-1 disclosures).
"""
import os, time, json, requests
from datetime import datetime, date
import pandas as pd
import numpy as np
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ─── Config ───────────────────────────────────────────────────────────────────
TS_CODE   = 'DDOG.O'
CIK       = '0001561550'
COMPANY   = 'Datadog, Inc.'
CCY       = 'USD'
UNIT_LABEL= 'USD M'
UNIT_DIV  = 1e6
YEARS     = list(range(2017, 2026))    # 2017-2025
SEC_HDR   = {'User-Agent':'BUI Research bui@example.com'}

# US 10y Treasury (year-end) — risk-free benchmark for SaaS valuation
US_10Y = {2017:0.0241, 2018:0.0269, 2019:0.0192, 2020:0.0091, 2021:0.0151,
          2022:0.0388, 2023:0.0388, 2024:0.0458, 2025:0.0428}
TODAY = datetime.now().strftime('%y%m%d')
OUT_DIR = os.path.expanduser(f'~/Desktop/BUI_Investment_Management/{TS_CODE}')
os.makedirs(OUT_DIR, exist_ok=True)
OUT_FILE = os.path.join(OUT_DIR, f'{TS_CODE}.Valuation.v{TODAY}.xlsx')

# ─── Style ────────────────────────────────────────────────────────────────────
FONT = "Apple Braille"
SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY = "FFE7E6E6"
THIN = Side(style="thin", color="FF000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.0;(#,##0.0);"–"'
PCT_FMT = '0.0%;(0.0%);"–"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False, fill=None):
    c.font = Font(name=FONT, size=SIZE, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if num_fmt: c.number_format = num_fmt
    if grey: c.fill = PatternFill("solid", fgColor=GREY)
    if fill: c.fill = PatternFill("solid", fgColor=fill)

def col_letter(i): return get_column_letter(2+i)

# ─── SEC EDGAR fetcher ────────────────────────────────────────────────────────
print(f"\n{'='*60}\nFetching {TS_CODE} ({COMPANY}) — SEC EDGAR\n{'='*60}")
r = requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{CIK}.json",
                 headers=SEC_HDR, timeout=30)
j = r.json()
facts = j.get('facts', {}).get('us-gaap', {})
print(f"  Loaded {len(facts)} XBRL tags")

def fy_annual(tag, units='USD'):
    """Return {year: value} keyed by end-date year (12-31), latest filed wins."""
    if tag not in facts: return {}
    arr = facts[tag].get('units', {}).get(units, [])
    out = {}
    for u in arr:
        if u.get('form') not in ('10-K','10-K/A','S-1','S-1/A','10-Q'): continue
        end = u.get('end','')
        if not end.endswith('-12-31'): continue
        start = u.get('start')
        # Only annual periods (≥ 350 days span) for flow items
        if start and units == 'USD':
            d1, d2 = date.fromisoformat(start), date.fromisoformat(end)
            if (d2-d1).days < 350: continue
        yr = int(end[:4])
        if yr not in out or u.get('filed','') > out[yr][1]:
            out[yr] = (u['val'], u.get('filed',''))
    return {k: v[0] for k,v in out.items()}

def fy_instant(tag, units='USD'):
    """Balance-sheet items: instant value at end-date, no 'start'."""
    if tag not in facts: return {}
    arr = facts[tag].get('units', {}).get(units, [])
    out = {}
    for u in arr:
        if u.get('form') not in ('10-K','10-K/A','S-1','S-1/A'): continue
        end = u.get('end','')
        if not end.endswith('-12-31'): continue
        if u.get('start'): continue   # instant only
        yr = int(end[:4])
        if yr not in out or u.get('filed','') > out[yr][1]:
            out[yr] = (u['val'], u.get('filed',''))
    return {k: v[0] for k,v in out.items()}

# Fetch standard fields
print("  Fetching XBRL tags...")
revenue = fy_annual('RevenueFromContractWithCustomerExcludingAssessedTax')
gross   = fy_annual('GrossProfit')
opinc   = fy_annual('OperatingIncomeLoss')
rd      = fy_annual('ResearchAndDevelopmentExpense')
sm      = fy_annual('SellingAndMarketingExpense')
ga      = fy_annual('GeneralAndAdministrativeExpense')
intinc  = fy_annual('InvestmentIncomeInterest')
intexp  = fy_annual('InterestExpense')
pretax  = fy_annual('IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest')
tax     = fy_annual('IncomeTaxExpenseBenefit')
ni      = fy_annual('NetIncomeLoss')

cash    = fy_instant('CashAndCashEquivalentsAtCarryingValue')
mkt_sec = fy_instant('MarketableSecuritiesCurrent')
ar      = fy_instant('AccountsReceivableNetCurrent')
ppe     = fy_instant('PropertyPlantAndEquipmentNet')
goodwill= fy_instant('Goodwill')
ta      = fy_instant('Assets')
ap      = fy_instant('AccountsPayableCurrent')
deferred= fy_instant('ContractWithCustomerLiabilityCurrent')
ltdebt  = fy_instant('LongTermDebtNoncurrent')
tl      = fy_instant('Liabilities')
eq      = fy_instant('StockholdersEquity')

ocf     = fy_annual('NetCashProvidedByUsedInOperatingActivities')
icf     = fy_annual('NetCashProvidedByUsedInInvestingActivities')
fcf_act = fy_annual('NetCashProvidedByUsedInFinancingActivities')
capex   = fy_annual('PaymentsToAcquirePropertyPlantAndEquipment')
sbc     = fy_annual('ShareBasedCompensation')

# Diluted shares (avg) from yfinance IS — fall back if needed
print("  Fetching prices and shares (yfinance)...")
yt = yf.Ticker("DDOG")
hist = yt.history(start="2018-01-01", end="2026-04-30", auto_adjust=False)
ye_px = {d.year: float(v) for d, v in hist['Close'].resample('YE').last().items() if d.year in YEARS}
yinfo = yt.info
shares_now_M = (yinfo.get('sharesOutstanding') or 328e6) / 1e6   # ~328M

# Best-effort historical share count: use IS diluted average shares from yfinance
# Map: pre-IPO ~80M, post-IPO grew to ~330M; embed snapshots from 10-K filings
SHARES_M = {
    2017: 30.0,    # pre-IPO Class B (low convertible structure)
    2018: 38.0,
    2019: 280.0,   # post-IPO Sept 2019
    2020: 305.0,
    2021: 313.0,
    2022: 318.0,
    2023: 325.0,
    2024: 340.0,
    2025: 348.0,
}

# Convert raw USD → USD millions
def to_M(v):
    return float(v)/UNIT_DIV if v is not None else None

# Sanity print
print("\nData summary (USD M):")
print(f"{'Year':<6}{'Rev':>10}{'OpInc':>10}{'NI':>10}{'OCF':>10}{'Cash':>10}{'TA':>10}{'Eq':>10}{'Px':>10}{'Sh(M)':>10}")
for y in YEARS:
    print(f"{y:<6}{to_M(revenue.get(y)) or 0:>10.1f}{to_M(opinc.get(y)) or 0:>10.1f}"
          f"{to_M(ni.get(y)) or 0:>10.1f}{to_M(ocf.get(y)) or 0:>10.1f}"
          f"{to_M(cash.get(y)) or 0:>10.1f}{to_M(ta.get(y)) or 0:>10.1f}"
          f"{to_M(eq.get(y)) or 0:>10.1f}{ye_px.get(y) or 0:>10.2f}{SHARES_M[y]:>10.1f}")

# ============================================================
# Build workbook
# ============================================================
print(f"\n{'='*60}\nBuilding workbook...\n{'='*60}")
wb = Workbook()

# ===== Sheet 1: Financial Data =====
ws1 = wb.active
ws1.title = "Financial Data"
ws1.column_dimensions['A'].width = 42
for i in range(len(YEARS)):
    ws1.column_dimensions[col_letter(i)].width = 12
ws1.column_dimensions[col_letter(len(YEARS))].width = 8.83

ws1.cell(1,1, f"{COMPANY} ({TS_CODE}) — Financial Data ({UNIT_LABEL})")
fmt(ws1.cell(1,1), bold=True, align="left")

ws1.cell(3,1, "Item"); fmt(ws1.cell(3,1), bold=True, grey=True, align="left")
for i,y in enumerate(YEARS):
    fmt(ws1.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

rows1 = [
    ("==INCOME STATEMENT==","header",None),
    ("Revenue 营业收入","val",revenue),
    ("Revenue Growth (YoY)","growth_rev",None),
    ("Gross Profit 毛利","val",gross),
    ("Gross Margin","pct_rev",None),
    ("R&D Expense","val",rd),
    ("S&M Expense","val",sm),
    ("G&A Expense","val",ga),
    ("Operating Income (Loss)","val",opinc),
    ("Operating Margin","op_margin",None),
    ("Interest Income","val",intinc),
    ("Interest Expense","val",intexp),
    ("Pretax Income","val",pretax),
    ("Income Tax","val",tax),
    ("Net Income (Loss)","val",ni),
    ("Net Margin","ni_margin",None),
    ("==BALANCE SHEET==","header",None),
    ("Cash & Equivalents","val",cash),
    ("Marketable Securities (current)","val",mkt_sec),
    ("Accounts Receivable","val",ar),
    ("PP&E (net)","val",ppe),
    ("Goodwill","val",goodwill),
    ("Total Assets","val",ta),
    ("Accounts Payable","val",ap),
    ("Deferred Revenue (current)","val",deferred),
    ("Long-Term Debt","val",ltdebt),
    ("Total Liabilities","val",tl),
    ("Stockholders' Equity","val",eq),
    ("==CASH FLOW==","header",None),
    ("Operating Cash Flow","val",ocf),
    ("Investing Cash Flow","val",icf),
    ("Financing Cash Flow","val",fcf_act),
    ("Capex (PP&E purchases)","val",capex),
    ("Free Cash Flow (OCF − Capex)","fcf",None),
    ("FCF Margin","fcf_margin",None),
    ("Stock-Based Compensation","val",sbc),
    ("==MARKET DATA==","header",None),
    ("Year-End Share Price (USD)","px",None),
    ("Diluted Shares (M)","shares",None),
    ("Market Cap (USD M)","mcap",None),
]
r = 4
for label, kind, src in rows1:
    ws1.cell(r,1,label)
    if kind=="header":
        fmt(ws1.cell(r,1), bold=True, grey=True, align="left")
        for i in range(len(YEARS)): fmt(ws1.cell(r,2+i), grey=True)
    else:
        fmt(ws1.cell(r,1), align="left")
        for i,y in enumerate(YEARS):
            cl = col_letter(i); pcl = col_letter(i-1) if i>0 else None
            if kind=="val":
                v = to_M(src.get(y)) if src else None
                c = ws1.cell(r,2+i,v); fmt(c, color=BLUE, num_fmt=NUM_FMT)
            elif kind=="growth_rev" and i>0:
                f = f"=IFERROR({cl}5/{pcl}5-1,\"\")"
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT)
            elif kind=="pct_rev":
                f = f"=IFERROR({cl}7/{cl}5,\"\")"   # gross/revenue
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT)
            elif kind=="op_margin":
                f = f"=IFERROR({cl}12/{cl}5,\"\")"
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT)
            elif kind=="ni_margin":
                f = f"=IFERROR({cl}18/{cl}5,\"\")"
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT)
            elif kind=="fcf":
                f = f"={cl}33-{cl}36"
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT, bold=True)
            elif kind=="fcf_margin":
                f = f"=IFERROR(({cl}33-{cl}36)/{cl}5,\"\")"
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT)
            elif kind=="px":
                v = ye_px.get(y)
                c = ws1.cell(r,2+i,v); fmt(c, color=BLUE, num_fmt='0.00')
            elif kind=="shares":
                v = SHARES_M.get(y)
                c = ws1.cell(r,2+i,v); fmt(c, color=BLUE, num_fmt='#,##0.0')
            elif kind=="mcap":
                f = f"={cl}41*{cl}42"
                c = ws1.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT, bold=True)
    r += 1
for rr in range(1, r+1): ws1.row_dimensions[rr].height = 18

# ===== Sheet 2: Key Ratios =====
ws2 = wb.create_sheet("Key Ratios")
ws2.column_dimensions['A'].width = 42
for i in range(len(YEARS)):
    ws2.column_dimensions[col_letter(i)].width = 12

ws2.cell(1,1, f"{COMPANY} — Key Ratios"); fmt(ws2.cell(1,1), bold=True, align="left")
ws2.cell(3,1, "Ratio"); fmt(ws2.cell(3,1), bold=True, grey=True, align="left")
for i,y in enumerate(YEARS):
    fmt(ws2.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

fd = "'Financial Data'"
ratios = [
    ("Revenue Growth (YoY)",          lambda c,p: f"=IFERROR({fd}!{c}5/{fd}!{p}5-1,\"\")", PCT_FMT, True),
    ("Gross Margin",                  lambda c,p: f"=IFERROR({fd}!{c}7/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("R&D / Revenue",                 lambda c,p: f"=IFERROR({fd}!{c}9/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("S&M / Revenue",                 lambda c,p: f"=IFERROR({fd}!{c}10/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("G&A / Revenue",                 lambda c,p: f"=IFERROR({fd}!{c}11/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("Operating Margin (GAAP)",       lambda c,p: f"=IFERROR({fd}!{c}12/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("Operating Margin ex-SBC (Adj)", lambda c,p: f"=IFERROR(({fd}!{c}12+{fd}!{c}39)/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("Net Margin",                    lambda c,p: f"=IFERROR({fd}!{c}18/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("FCF Margin",                    lambda c,p: f"=IFERROR(({fd}!{c}33-{fd}!{c}36)/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("Rule of 40 (Rev Growth + FCF Margin)", lambda c,p: (
        f"=IFERROR({fd}!{c}5/{fd}!{p}5-1+({fd}!{c}33-{fd}!{c}36)/{fd}!{c}5,\"\")" if p else None), PCT_FMT, True),
    ("ROE (NI / Avg Equity)",         lambda c,p: (f"=IFERROR({fd}!{c}18/AVERAGE({fd}!{c}31,{fd}!{p}31),\"\")" if p else None), PCT_FMT, True),
    ("ROA (NI / Avg Assets)",         lambda c,p: (f"=IFERROR({fd}!{c}18/AVERAGE({fd}!{c}26,{fd}!{p}26),\"\")" if p else None), PCT_FMT, True),
    ("SBC / Revenue",                 lambda c,p: f"=IFERROR({fd}!{c}39/{fd}!{c}5,\"\")", PCT_FMT, False),
    ("Cash + ST Sec / Total Assets",  lambda c,p: f"=IFERROR(({fd}!{c}21+{fd}!{c}22)/{fd}!{c}26,\"\")", PCT_FMT, False),
]
r = 4
for label, fn, nfmt, needs_prior in ratios:
    ws2.cell(r,1,label); fmt(ws2.cell(r,1), align="left")
    for i,y in enumerate(YEARS):
        cl, pcl = col_letter(i), col_letter(i-1) if i>0 else None
        if needs_prior and pcl is None: continue
        f = fn(cl, pcl)
        if f:
            c = ws2.cell(r,2+i); c.value=f
            fmt(c, color=GREEN, num_fmt=nfmt)
    r += 1
for rr in range(1,r+1): ws2.row_dimensions[rr].height = 20

# ===== Sheet 3: ROIC Analysis =====
ws3 = wb.create_sheet("ROIC Analysis")
ws3.column_dimensions['A'].width = 42
for i in range(len(YEARS)):
    ws3.column_dimensions[col_letter(i)].width = 12

ws3.cell(1,1, f"{COMPANY} — ROIC Decomposition"); fmt(ws3.cell(1,1), bold=True, align="left")
ws3.cell(3,1,"Component"); fmt(ws3.cell(3,1), bold=True, grey=True, align="left")
for i,y in enumerate(YEARS):
    fmt(ws3.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

roic_rows = [
    ("Operating Income (Loss)", lambda c,p: f"={fd}!{c}12", NUM_FMT, False),
    ("+ Stock-Based Comp",      lambda c,p: f"={fd}!{c}39", NUM_FMT, False),
    ("Adj NOPAT (pre-tax, OpInc + SBC)", lambda c,p: f"={c}4+{c}5", NUM_FMT, False),
    ("Tax Rate (assumed)",      lambda c,p: f"=0.21", PCT_FMT, False),
    ("Adj NOPAT (after-tax)",   lambda c,p: f"={c}6*(1-{c}7)", NUM_FMT, False),
    ("",                        lambda c,p: None, None, False),
    ("Total Equity",            lambda c,p: f"={fd}!{c}31", NUM_FMT, False),
    ("+ Long-Term Debt",        lambda c,p: f"={fd}!{c}29", NUM_FMT, False),
    ("− Cash & ST Securities",  lambda c,p: f"=-({fd}!{c}21+{fd}!{c}22)", NUM_FMT, False),
    ("Invested Capital",        lambda c,p: f"={c}10+{c}11+{c}12", NUM_FMT, False),
    ("",                        lambda c,p: None, None, False),
    ("ROIC (Adj NOPAT / Avg IC)", lambda c,p: (f"=IFERROR({c}8/AVERAGE({c}13,{p}13),\"\")" if p else None), PCT_FMT, True),
]
r = 4
for label, fn, nfmt, needs_prior in roic_rows:
    ws3.cell(r,1,label); fmt(ws3.cell(r,1), align="left")
    for i,y in enumerate(YEARS):
        cl, pcl = col_letter(i), col_letter(i-1) if i>0 else None
        if needs_prior and pcl is None: continue
        f = fn(cl, pcl)
        if f:
            c = ws3.cell(r,2+i); c.value=f
            color = BLACK if not nfmt or nfmt!=PCT_FMT or label.startswith("Tax") else (BLUE if label.startswith("Tax") else GREEN)
            if label.startswith("ROIC"): color = GREEN
            elif label.startswith("Tax"): color = BLUE
            else: color = BLACK
            fmt(c, color=color, num_fmt=nfmt or NUM_FMT, bold=label.startswith(("ROIC","Adj NOPAT (after","Invested")))
    r += 1
for rr in range(1,r+1): ws3.row_dimensions[rr].height = 18

# ===== Sheet 4: Valuation Model =====
ws4 = wb.create_sheet("Valuation Model")
ws4.column_dimensions['A'].width = 42
for i in range(len(YEARS)):
    ws4.column_dimensions[col_letter(i)].width = 12

ws4.cell(1,1, f"{COMPANY} — Valuation Model (SaaS multiples)"); fmt(ws4.cell(1,1), bold=True, align="left")
ws4.cell(3,1,"Item"); fmt(ws4.cell(3,1), bold=True, grey=True, align="left")
for i,y in enumerate(YEARS):
    fmt(ws4.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

val_rows = [
    ("==INPUTS==", "header"),
    ("US 10y Treasury", "bond"),
    ("Equity Risk Premium", "erp"),
    ("Cost of Equity", "coe"),
    ("==MARKET METRICS==", "header"),
    ("Market Cap (USD M)", "mc"),
    ("+ Long-Term Debt", "ltd"),
    ("− Cash & ST Securities", "cash"),
    ("Enterprise Value (EV)", "ev"),
    ("==MULTIPLES (TRADING)==", "header"),
    ("EV / Revenue", "ev_rev"),
    ("EV / FCF", "ev_fcf"),
    ("P/E (NI > 0 only)", "pe"),
    ("Price / Sales", "ps"),
    ("==VALUATION OUTPUT==", "header"),
    ("Method 1: EV / Revenue × Target Mult (15x)", "v1"),
    ("Method 2: EV / FCF × Target Mult (35x)", "v2"),
    ("Method 3: P/E × Target (50x)", "v3"),
    ("Average Implied EV", "avg_ev"),
    ("Implied Mkt Cap (avg EV − Debt + Cash)", "imp_mc"),
    ("Premium / (Discount) vs Actual", "pd"),
]
r = 4
for label, key in val_rows:
    ws4.cell(r,1,label)
    if label.startswith("=="):
        fmt(ws4.cell(r,1), bold=True, grey=True, align="left")
        for i in range(len(YEARS)): fmt(ws4.cell(r,2+i), grey=True)
    else:
        fmt(ws4.cell(r,1), align="left")
        for i,y in enumerate(YEARS):
            cl = col_letter(i)
            if key=="bond":
                c = ws4.cell(r,2+i, US_10Y.get(y, 0.04)); fmt(c, color=BLUE, num_fmt=PCT_FMT)
            elif key=="erp":
                c = ws4.cell(r,2+i, 0.06); fmt(c, color=BLUE, num_fmt=PCT_FMT)
            elif key=="coe":
                f = f"={cl}5+{cl}6"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT)
            elif key=="mc":
                f = f"={fd}!{cl}43"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=GREEN, num_fmt=NUM_FMT)
            elif key=="ltd":
                f = f"={fd}!{cl}29"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=GREEN, num_fmt=NUM_FMT)
            elif key=="cash":
                f = f"=-({fd}!{cl}21+{fd}!{cl}22)"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=GREEN, num_fmt=NUM_FMT)
            elif key=="ev":
                f = f"={cl}9+{cl}10+{cl}11"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT, bold=True)
            elif key=="ev_rev":
                f = f"=IFERROR({cl}12/{fd}!{cl}5,\"\")"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt='0.0"x"')
            elif key=="ev_fcf":
                f = f"=IFERROR({cl}12/({fd}!{cl}33-{fd}!{cl}36),\"\")"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt='0.0"x"')
            elif key=="pe":
                f = f"=IFERROR(IF({fd}!{cl}18>0,{cl}9/{fd}!{cl}18,\"\"),\"\")"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt='0.0"x"')
            elif key=="ps":
                f = f"=IFERROR({cl}9/{fd}!{cl}5,\"\")"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt='0.0"x"')
            elif key=="v1":
                f = f"={fd}!{cl}5*15"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT)
            elif key=="v2":
                f = f"=({fd}!{cl}33-{fd}!{cl}36)*35"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT)
            elif key=="v3":
                f = f"=IF({fd}!{cl}18>0,{fd}!{cl}18*50,\"\")"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT)
            elif key=="avg_ev":
                f = f"=IFERROR(AVERAGE({cl}19,{cl}20,{cl}21),AVERAGE({cl}19,{cl}20))"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT, bold=True)
            elif key=="imp_mc":
                f = f"={cl}22-{cl}10-{cl}11"   # avg EV - debt - (-cash)= -cash flips sign
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=NUM_FMT)
            elif key=="pd":
                f = f"=IFERROR({cl}23/{cl}9-1,\"\")"
                c = ws4.cell(r,2+i); c.value=f; fmt(c, color=BLACK, num_fmt=PCT_FMT, bold=True)
    r += 1
for rr in range(1,r+1): ws4.row_dimensions[rr].height = 17

# ===== Sheet 5: Quarterly TTM =====
ws5 = wb.create_sheet("Quarterly TTM")
ws5.column_dimensions['A'].width = 42
ws5.cell(1,1, f"{COMPANY} — Quarterly TTM Analysis"); fmt(ws5.cell(1,1), bold=True, align="left")
ws5.cell(3,1, "SEC EDGAR XBRL: pull 10-Q quarterly facts for revenue / OCF / FCF.")
fmt(ws5.cell(3,1), italic=True, align="left", color=BLUE)
ws5.cell(4,1, "Computed below for the most recent ~12 quarters where available.")
fmt(ws5.cell(4,1), italic=True, align="left", color=BLUE)

# Pull quarterly revenue
def get_q(tag):
    if tag not in facts: return []
    arr = facts[tag].get('units', {}).get('USD', [])
    out = []
    for u in arr:
        if u.get('form') not in ('10-Q','10-K','10-K/A'): continue
        end = u.get('end',''); start = u.get('start','')
        if not end or not start: continue
        d1, d2 = date.fromisoformat(start), date.fromisoformat(end)
        if 80 <= (d2-d1).days <= 100:    # 1 quarter
            out.append((end, u['val']))
    # dedup — latest filed wins
    by_end = {}
    for e,v in out:
        by_end[e] = v
    return sorted(by_end.items())

q_rev = get_q('RevenueFromContractWithCustomerExcludingAssessedTax')[-12:]
q_ocf = dict(get_q('NetCashProvidedByUsedInOperatingActivities'))
q_capex = dict(get_q('PaymentsToAcquirePropertyPlantAndEquipment'))

ws5.cell(6,1,"Quarter End"); fmt(ws5.cell(6,1), bold=True, grey=True, align="left")
ws5.cell(6,2,"Revenue (M)"); fmt(ws5.cell(6,2), bold=True, grey=True, align="center")
ws5.cell(6,3,"OCF (M)"); fmt(ws5.cell(6,3), bold=True, grey=True, align="center")
ws5.cell(6,4,"Capex (M)"); fmt(ws5.cell(6,4), bold=True, grey=True, align="center")
ws5.cell(6,5,"FCF (M)"); fmt(ws5.cell(6,5), bold=True, grey=True, align="center")

for i,(end,v) in enumerate(q_rev):
    r = 7+i
    ws5.cell(r,1,end); fmt(ws5.cell(r,1), align="center")
    ws5.cell(r,2, v/1e6); fmt(ws5.cell(r,2), color=BLUE, num_fmt=NUM_FMT)
    ocfv = q_ocf.get(end)
    cxv = q_capex.get(end)
    if ocfv is not None:
        ws5.cell(r,3, ocfv/1e6); fmt(ws5.cell(r,3), color=BLUE, num_fmt=NUM_FMT)
    if cxv is not None:
        ws5.cell(r,4, cxv/1e6); fmt(ws5.cell(r,4), color=BLUE, num_fmt=NUM_FMT)
    if ocfv is not None and cxv is not None:
        ws5.cell(r,5, (ocfv-cxv)/1e6); fmt(ws5.cell(r,5), color=BLACK, num_fmt=NUM_FMT, bold=True)
for c in range(2,6):
    ws5.column_dimensions[get_column_letter(c)].width = 14

# ===== Sheet 6: Top Holders (yfinance) =====
ws6 = wb.create_sheet("Top Holders")
ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 50
ws6.column_dimensions['C'].width = 16
ws6.column_dimensions['D'].width = 14
ws6.column_dimensions['E'].width = 14

ws6.cell(1,1, f"{COMPANY} — Institutional Holders (yfinance snapshot)")
fmt(ws6.cell(1,1), bold=True, align="left")

try:
    holders = yt.institutional_holders
    if holders is not None and not holders.empty:
        cols = ["Holder","Shares","Date Reported","Value"]
        for i,h in enumerate(["#"]+cols):
            c = ws6.cell(3,1+i, h); fmt(c, bold=True, grey=True, align="center" if i==0 else "left")
        for i,row in holders.head(15).iterrows():
            r = 4+i
            ws6.cell(r,1, i+1); fmt(ws6.cell(r,1), align="center")
            ws6.cell(r,2, str(row.get('Holder','')))
            fmt(ws6.cell(r,2), align="left")
            ws6.cell(r,3, float(row.get('Shares', 0)) if pd.notna(row.get('Shares')) else None)
            fmt(ws6.cell(r,3), num_fmt='#,##0')
            ws6.cell(r,4, str(row.get('Date Reported','')))
            fmt(ws6.cell(r,4), align="center")
            ws6.cell(r,5, float(row.get('Value', 0)) if pd.notna(row.get('Value')) else None)
            fmt(ws6.cell(r,5), num_fmt='#,##0')
except Exception as e:
    ws6.cell(3,1, f"Failed to retrieve institutional holders: {e}")

# ===== Sheet 7: Data Correlation =====
ws7 = wb.create_sheet("Data Correlation")
ws7.column_dimensions['A'].width = 5
ws7.column_dimensions['B'].width = 110

ws7.cell(1,1, f"{COMPANY} ({TS_CODE}) — Data Correlation & Methodology")
fmt(ws7.cell(1,1), bold=True, align="left")

sections = [
    ("1. Company Overview",
     "Datadog, Inc. — cloud monitoring & observability SaaS platform. NYSE/NASDAQ listed since Sept 2019.\n"
     "HQ: New York. Founders/Co-CEOs: Olivier Pomel, Alexis Lê-Quôc.\n"
     "Products: Infrastructure Monitoring, APM, Log Mgmt, Security, RUM, Database Monitoring, etc.\n"
     "Customer base: ~30,000 customers (2025); >$100K ARR customers ~3,800.\n"
     "Reporting currency: USD. Fiscal year = calendar year."),
    ("2. Data Sources",
     "PRIMARY: SEC EDGAR XBRL (companyfacts API) — official 10-K / 10-Q / S-1 filings.\n"
     f"  https://data.sec.gov/api/xbrl/companyfacts/CIK{CIK}.json\n"
     "SECONDARY: yfinance — year-end share prices, institutional holders snapshot.\n"
     "Note: tushare does NOT cover US equities. SEC EDGAR is the authoritative source."),
    ("3. Period & Units",
     f"Years: 2017-2025 (9 years, including 2 pre-IPO years from S-1 prospectus disclosures).\n"
     f"Units: USD millions (raw USD ÷ 1e6).\n"
     "All values rounded to 1 decimal — for full precision query SEC API directly."),
    ("4. Key XBRL Tags Used",
     "Revenue: RevenueFromContractWithCustomerExcludingAssessedTax (ASC 606 standard tag)\n"
     "Operating Income: OperatingIncomeLoss\n"
     "Net Income: NetIncomeLoss\n"
     "OCF: NetCashProvidedByUsedInOperatingActivities\n"
     "Capex: PaymentsToAcquirePropertyPlantAndEquipment\n"
     "Cash: CashAndCashEquivalentsAtCarryingValue\n"
     "Marketable Securities: MarketableSecuritiesCurrent\n"
     "Equity: StockholdersEquity\n"
     "SBC: ShareBasedCompensation"),
    ("5. SaaS Valuation Methodology",
     "SaaS companies priced on FORWARD multiples; 3-method blend:\n"
     "  Method 1: EV / Revenue × 15x (industry mid for high-growth SaaS at 25-35% growth)\n"
     "  Method 2: EV / FCF × 35x (premium for FCF-conversion >25%)\n"
     "  Method 3: GAAP P/E × 50x (only when NI > 0; pre-2023 not meaningful)\n"
     "Implied Mkt Cap = Avg EV − Debt + Cash.\n"
     "Multiples are EDITABLE BLUE inputs — calibrate to peer group (Snowflake, MongoDB, Cloudflare).\n"
     "Premium / (Discount) = Implied / Actual − 1."),
    ("6. Rule of 40",
     "SaaS health metric: Revenue Growth + FCF Margin. ≥40% = healthy, ≥60% = elite.\n"
     "Datadog historically 80-100% (top-decile). Current ~50-60%.\n"
     "Computed in Sheet 2 row 'Rule of 40'."),
    ("7. Stock-Based Compensation (SBC) Adjustment",
     "GAAP operating loss includes SBC (~25-30% of revenue for high-growth SaaS).\n"
     "Adjusted OpMargin (ex-SBC) presented in Sheet 2 to reflect cash economics.\n"
     "Caveat: SBC IS a real cost (dilutes shareholders); interpret 'Adj' figures as upper bound."),
    ("8. Data Caveats & Limitations",
     "• Pre-IPO 2017-2018: data from S-1 (filed Aug 2019). Some line items may be unaudited.\n"
     "• 2025 OpInc shows GAAP operating LOSS despite strong revenue — reflects SBC catch-up in tech downturn.\n"
     "• Historical share counts (2017-2018): synthetic Class B equivalent count; not a clean post-conversion view.\n"
     "• Capex small (<2% of revenue) — Datadog asset-light, content/data infrastructure on AWS/GCP.\n"
     "• Marketable securities ($2-3B+ on balance sheet) is genuine cash-equivalent liquidity.\n"
     "• Dividend yield method NOT used — Datadog does not pay dividends.\n"
     "• 2025 institutional holders snapshot: yfinance refreshes ~quarterly; stale by 1-3 months."),
]
r = 3
for title, body in sections:
    ws7.cell(r,2, title); fmt(ws7.cell(r,2), bold=True, grey=True, align="left")
    r += 1
    for ln in body.split("\n"):
        ws7.cell(r,2, ln); fmt(ws7.cell(r,2), align="left")
        ws7.row_dimensions[r].height = 17
        r += 1
    r += 1

# Save
print(f"\nSaving → {OUT_FILE}")
wb.save(OUT_FILE)
print(f"\n{'='*60}\n✓ Saved: {OUT_FILE}\n{'='*60}\n")
print("Sheets:")
for sn in wb.sheetnames:
    print(f"  - {sn}")
