#!/usr/bin/env python3
"""
Add unified-format DRIP Analysis sheet to 00883.HK 中海油 file.
Renames existing "DRIP Simulation" → "DRIP (Granular Semi-Annual)" to preserve.
Adds new "DRIP Analysis" sheet matching bank-financial-model skill spec.
"""
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SRC = '/Users/chenguang/Desktop/BUI_Investment_Management/BUI-公司研究/00883.HK.中海油/00883.HK.Valuation.v260413.xlsx'
TODAY = datetime.now().strftime('%y%m%d')
DST = SRC.replace('v260413', f'v{TODAY}')

YEARS = list(range(2015, 2026))   # 2015–2025 = 11 years

# ========== Format ==========
FONT = "Apple Braille"
SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY    = "FFE7E6E6"
LGREY   = "FFF2F2F2"
THIN    = Side(style="thin", color="FF000000")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.0;(#,##0.0);"–"'
INT_FMT = '#,##0;(#,##0);"–"'
PCT_FMT = '0.00%;(0.00%);"–"'
PCT1_FMT = '0.0%;(0.0%);"–"'
PX_FMT  = '#,##0.00'
X_FMT   = '0.00"x"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False, lgrey=False, font_size=SIZE):
    c.font = Font(name=FONT, size=font_size, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if num_fmt: c.number_format = num_fmt
    if grey: c.fill = PatternFill("solid", fgColor=GREY)
    elif lgrey: c.fill = PatternFill("solid", fgColor=LGREY)

def col_l(i): return get_column_letter(2+i)

# ========== Inputs (calibrated) ==========

# CNOOC year-end HKD close prices (yfinance 0883.HK, auto_adjust=False)
YE_PX = {2015:8.07, 2016:9.70, 2017:11.22, 2018:12.10, 2019:12.96,
         2020:7.22, 2021:8.03, 2022:9.98, 2023:13.00, 2024:19.10, 2025:21.28}

# Total shares outstanding (millions). CNOOC very stable; ~44,647 M from 2014-present.
SHARES_M = {y: 44647.0 for y in YEARS}

# HK M2 (HKD trillion, year-end, HKMA reported)
HK_M2 = {2015:12.13, 2016:12.85, 2017:13.66, 2018:13.96, 2019:14.93,
         2020:16.85, 2021:16.98, 2022:16.93, 2023:17.00, 2024:17.20, 2025:17.50}

# China M2 (CNY 万亿, for cross-reference comment)
CN_M2 = {2015:139.23, 2025:340.29}

# DRIP defaults
DRIP_INVEST = 1_000_000     # HKD 1M
DRIP_FX     = 1.10          # HKD per CNY (used to convert FD CNY div → HKD)
DRIP_WHT    = 0.0           # HK Corp direct holding HK-incorporated red chip = 0%

# ========== Open workbook ==========
print(f"Loading: {SRC}")
wb = load_workbook(SRC)

# Rename old DRIP sheet
if 'DRIP Simulation' in wb.sheetnames:
    wb['DRIP Simulation'].title = 'DRIP (Granular Semi-Annual)'
    print("  Renamed: 'DRIP Simulation' → 'DRIP (Granular Semi-Annual)'")

# Insert new DRIP Analysis sheet after Financial Data
fd_idx = wb.sheetnames.index('Financial Data')
ws = wb.create_sheet('DRIP Analysis', fd_idx + 1)
print(f"  Inserted: 'DRIP Analysis' at position {fd_idx + 1}")

# ========== Build DRIP Analysis sheet ==========

n = len(YEARS)
ws.column_dimensions['A'].width = 38.33
ws.column_dimensions['B'].width = 15.83
for i in range(1, n):
    ws.column_dimensions[get_column_letter(2+i)].width = 13
ws.column_dimensions[col_l(n)].width = 8.83

fd = "'Financial Data'"

# Title (row 1, 12pt bold, light grey)
title = f"DRIP Return Analysis — 0883.HK ({YEARS[0]}–{YEARS[-1]}, HK Corporate direct holding)"
c = ws.cell(1, 1, title)
c.font = Font(name=FONT, size=12, color=BLACK, bold=True)
c.alignment = Alignment(horizontal="left", vertical="center")
c.fill = PatternFill("solid", fgColor=LGREY)
for j in range(2, 2+n+1):
    cc = ws.cell(1, j)
    cc.fill = PatternFill("solid", fgColor=LGREY)

ws.cell(2, 1, "Source: 'Financial Data' rows 25 (Div CNY M), 16 (Eq CNY M), 29 (MC HKD M); YE prices/shares hardcoded below")
fmt(ws.cell(2,1), italic=True, color=BLUE, align="left")

# ========== Assumptions (rows 4–9) ==========
ws.cell(4, 1, "Assumptions")
fmt(ws.cell(4,1), bold=True, align="left")
ws.cell(5, 1, "Initial investment (HKD)")
fmt(ws.cell(5,1), align="left")
fmt(ws.cell(5,2, DRIP_INVEST), color=BLUE, bold=True, num_fmt=INT_FMT)

ws.cell(6, 1, f"Entry price (HKD/sh, {YEARS[0]} YE)")
fmt(ws.cell(6,1), align="left")
c = ws.cell(6, 2); c.value = "=B39"   # link to Market Data row 39, col B (first year HKD price)
fmt(c, color=GREEN, bold=True, num_fmt=PX_FMT)

ws.cell(7, 1, "FX (HKD per CNY, for div conversion)")
fmt(ws.cell(7,1), align="left")
fmt(ws.cell(7,2, DRIP_FX), color=BLUE, bold=True, num_fmt='0.0000')

ws.cell(8, 1, "Dividend WHT (HK Corp direct, red chip → 0%)")
fmt(ws.cell(8,1), align="left")
fmt(ws.cell(8,2, DRIP_WHT), color=BLUE, bold=True, num_fmt=PCT_FMT)

ws.cell(9, 1, "Initial shares purchased")
fmt(ws.cell(9,1), align="left")
c = ws.cell(9, 2); c.value = "=B5/B6"
fmt(c, num_fmt=INT_FMT)

# ========== Year-by-year DRIP (rows 11–21) ==========
ws.cell(11, 1, "Year-by-year DRIP")
fmt(ws.cell(11,1), bold=True, align="left")

# Row 12: Year header
ws.cell(12, 1, "Year")
fmt(ws.cell(12,1), bold=True, align="left")
for i, y in enumerate(YEARS):
    cl = col_l(i)
    c = ws.cell(12, 2+i); c.value = f"={fd}!{cl}3"
    fmt(c, bold=True, align="right")

# Row 13: Shares held BoY
ws.cell(13, 1, "Shares held BoY")
fmt(ws.cell(13,1), align="left")
for i in range(n):
    if i == 0:
        ws.cell(13, 2).value = "=B9"
    else:
        ws.cell(13, 2+i).value = f"={col_l(i-1)}18"
    fmt(ws.cell(13, 2+i), num_fmt=INT_FMT)

# Row 14: DPS (HKD/share)
# DPS_HKD = ABS(FD!Div_CNY) * FX / Shares
ws.cell(14, 1, "DPS (HKD, computed from CF)")
fmt(ws.cell(14,1), align="left")
for i in range(n):
    cl = col_l(i)
    # FD row 25 = Dividends Paid (CNY M, negative); B40 = total shares (M)
    ws.cell(14, 2+i).value = f"=IFERROR(ABS({fd}!{cl}25)*$B$7/{cl}40,0)"
    fmt(ws.cell(14, 2+i), num_fmt='0.0000')

# Row 15: After-tax dividend (HKD)
ws.cell(15, 1, "Dividend (HKD, after WHT)")
fmt(ws.cell(15,1), align="left")
for i in range(n):
    cl = col_l(i)
    # B13 * B14 * (1-WHT). Note FX is already baked into DPS (row 14), so no FX again.
    ws.cell(15, 2+i).value = f"={cl}13*{cl}14*(1-$B$8)"
    fmt(ws.cell(15, 2+i), num_fmt=INT_FMT)

# Row 16: Reinvest price (HKD) — linked from Market Data row 39
ws.cell(16, 1, "Reinvest price (HKD)")
fmt(ws.cell(16,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws.cell(16, 2+i).value = f"={cl}39"
    fmt(ws.cell(16, 2+i), color=GREEN, num_fmt=PX_FMT)

# Row 17: Shares bought
ws.cell(17, 1, "Shares bought")
fmt(ws.cell(17,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws.cell(17, 2+i).value = f"=IFERROR({cl}15/{cl}16,0)"
    fmt(ws.cell(17, 2+i), num_fmt=INT_FMT)

# Row 18: Shares held EoY
ws.cell(18, 1, "Shares held EoY")
fmt(ws.cell(18,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws.cell(18, 2+i).value = f"={cl}13+{cl}17"
    fmt(ws.cell(18, 2+i), num_fmt=INT_FMT)

# Row 19: Portfolio value EoY (HKD)
ws.cell(19, 1, "Portfolio value EoY (HKD)")
fmt(ws.cell(19,1), align="left")
for i in range(n):
    cl = col_l(i)
    bold = (i == n-1)
    ws.cell(19, 2+i).value = f"={cl}18*{cl}16"
    fmt(ws.cell(19, 2+i), bold=bold, num_fmt=INT_FMT)

# Row 20: P/B at YE (proper, FX-adjusted)
# P/B = MC_HKD (FD!29) / (Eq_CNY (FD!16) * FX_HKD_per_CNY)
ws.cell(20, 1, "P/B at Year-End (HKD-adjusted, MC / (Eq × FX))")
fmt(ws.cell(20,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws.cell(20, 2+i).value = f"=IFERROR({fd}!{cl}29/({fd}!{cl}16*$B$7),\"-\")"
    fmt(ws.cell(20, 2+i), color=GREEN, num_fmt=X_FMT)

# Row 21: Dividend yield
# Yield = ABS(FD!Div_CNY) * FX / FD!MC_HKD
ws.cell(21, 1, "Dividend Yield (Div_HKD / MC_HKD)")
fmt(ws.cell(21,1), align="left")
for i in range(n):
    cl = col_l(i)
    ws.cell(21, 2+i).value = f"=IFERROR(ABS({fd}!{cl}25)*$B$7/{fd}!{cl}29,0)"
    fmt(ws.cell(21, 2+i), num_fmt=PCT_FMT)

# ========== HK M2 reference (rows 23–26) ==========
ws.cell(23, 1, "Reference: HK M2 (HKD 万亿)")
fmt(ws.cell(23,1), bold=True, align="left")
ws.cell(24, 1, "M2 year-end balance")
fmt(ws.cell(24,1), align="left")
for i, y in enumerate(YEARS):
    v = HK_M2.get(y)
    if v: fmt(ws.cell(24, 2+i, v), color=BLUE, bold=True, num_fmt='#,##0.00')

last_col = col_l(n-1)
cagr_n = n - 1
ws.cell(25, 1, f"M2 multiple vs {YEARS[0]}")
fmt(ws.cell(25,1), align="left")
c = ws.cell(25, 2+n)
c.value = f"={last_col}24/B24"
fmt(c, num_fmt=X_FMT)

ws.cell(26, 1, f"M2 CAGR ({cagr_n} yrs)")
fmt(ws.cell(26,1), align="left")
c = ws.cell(26, 2+n)
c.value = f"=({last_col}24/B24)^(1/{cagr_n})-1"
fmt(c, num_fmt=PCT_FMT)

# ========== Summary (rows 28–35) ==========
ws.cell(28, 1, "Summary")
fmt(ws.cell(28,1), bold=True, align="left")

ws.cell(29, 1, f"Final value {YEARS[-1]} YE (HKD)")
fmt(ws.cell(29,1), align="left")
c = ws.cell(29, 2); c.value = f"={last_col}19"
fmt(c, bold=True, num_fmt=INT_FMT)

ws.cell(30, 1, "Total return")
fmt(ws.cell(30,1), align="left")
c = ws.cell(30, 2); c.value = "=B29/B5-1"
fmt(c, num_fmt=PCT_FMT)

ws.cell(31, 1, "Multiple on invested (x)")
fmt(ws.cell(31,1), align="left")
c = ws.cell(31, 2); c.value = "=B29/B5"
fmt(c, bold=True, num_fmt=X_FMT)

ws.cell(32, 1, f"Annualized IRR ({n} yrs)")
fmt(ws.cell(32,1), align="left")
c = ws.cell(32, 2); c.value = f"=(B29/B5)^(1/{n})-1"
fmt(c, bold=True, num_fmt=PCT_FMT)

cell_m2cagr = f"{get_column_letter(2+n)}26"
ws.cell(33, 1, "vs HK M2 growth (DRIP IRR − M2 CAGR)")
fmt(ws.cell(33,1), align="left")
c = ws.cell(33, 2); c.value = f"=B32-{cell_m2cagr}"
fmt(c, num_fmt=PCT_FMT)

ws.cell(34, 1, "Price-only return (no dividends)")
fmt(ws.cell(34,1), align="left")
c = ws.cell(34, 2); c.value = f"={last_col}39/B39-1"
fmt(c, num_fmt=PCT1_FMT)

ws.cell(35, 1, "Avg P/B over period")
fmt(ws.cell(35,1), align="left")
pb_range = f"{col_l(0)}20:{last_col}20"
c = ws.cell(35, 2); c.value = f"=AVERAGE({pb_range})"
fmt(c, num_fmt=X_FMT)

# ========== Market Data reference (rows 37–40) ==========
ws.cell(37, 1, "Market Data Reference (BLUE = editable inputs)")
fmt(ws.cell(37,1), bold=True, align="left")

ws.cell(38, 1, "Year")
fmt(ws.cell(38,1), bold=True, grey=True, align="left")
for i, y in enumerate(YEARS):
    fmt(ws.cell(38, 2+i, y), bold=True, grey=True, num_fmt="0", align="center")

ws.cell(39, 1, "YE Close (HKD/share)")
fmt(ws.cell(39,1), align="left")
for i, y in enumerate(YEARS):
    fmt(ws.cell(39, 2+i, YE_PX[y]), color=BLUE, num_fmt=PX_FMT)

ws.cell(40, 1, "Total Shares (M)")
fmt(ws.cell(40,1), align="left")
for i, y in enumerate(YEARS):
    fmt(ws.cell(40, 2+i, SHARES_M[y]), color=BLUE, num_fmt=INT_FMT)

# ========== Notes (rows 42+) ==========
notes_start = 42
ws.cell(notes_start, 1, "Notes")
fmt(ws.cell(notes_start,1), bold=True, align="left")

notes = [
    "1. CNOOC Limited (0883.HK) is HK-incorporated red chip → no HK withholding on dividends to HK Corp holders.",
    "   This DRIP uses 0% WHT. Compare with the granular 'DRIP (Granular Semi-Annual)' sheet which used 10%.",
    "   Mainland investors via 港股通 face 20% (individual) or 10% (corporate) WHT — adjust B8 if applicable.",
    "2. CNOOC reports financials in CNY but pays dividends in HKD per share. Cash flow row 25 (FD!) is",
    "   the CNY-equivalent of HKD dividends declared. We convert back via FX = 1.10 (HKD per CNY) at row B7.",
    "3. Year-end share prices (row 39) and total shares (row 40) are USER-EDITABLE BLUE inputs.",
    "   YE prices sourced from yfinance 0883.HK (auto_adjust=False). Total shares ~44,647M (CNOOC very stable).",
    "4. P/B at row 20 is FX-adjusted: MC_HKD / (Eq_CNY × FX_HKD_per_CNY). The existing FD row 47 P/B was",
    "   uncorrected (HKD/CNY ratio, ~10% overstated). Use row 20 here for proper P/B comparison.",
    "5. HK M2 benchmark (row 24) is the appropriate macro benchmark for HKD-denominated investment.",
    f"   HK M2 10y CAGR ({YEARS[0]}–{YEARS[-1]}) ≈ 3.73%, much lower than China M2 (~9.35%) — HK is a",
    "   monetary 'tail end' (HKD pegged to USD via Linked Exchange Rate System).",
    "6. Dividend Yield row 21 uses HKD-converted dividend over HKD market cap — direct comparable to",
    "   the HKEX yield published numbers.",
]
for i, line in enumerate(notes):
    ws.cell(notes_start+1+i, 1, line)
    fmt(ws.cell(notes_start+1+i, 1), align="left")
    ws.row_dimensions[notes_start+1+i].height = 17

# Row heights for top section (per skill spec, 23pt for DRIP)
for rr in range(1, notes_start+1):
    ws.row_dimensions[rr].height = 23.0

# ========== Save ==========
print(f"\nSaving to: {DST}")
wb.save(DST)
print(f"\n{'='*60}\n✓ Saved: {DST}\n{'='*60}")
print(f"\nSheets in updated workbook:")
for sn in wb.sheetnames:
    print(f"  - {sn}")
