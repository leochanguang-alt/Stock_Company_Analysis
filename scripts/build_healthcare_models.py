#!/usr/bin/env python3
"""
Build 7-sheet financial valuation models for 4 healthcare companies:
  ILMN — Illumina (US, USD M)
  PHG  — Royal Philips (NL, EUR M, IFRS)
  JNJ  — Johnson & Johnson (US, USD M)
  TMO  — Thermo Fisher Scientific (US, USD M)

Data: SEC EDGAR XBRL (companyfacts) + yfinance (prices, holders).
Years: 2015-2025 (11 years). PHG in EUR (Plan A — no FX conversion).
"""
import os, requests
from datetime import datetime, date
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

YEARS = list(range(2015, 2026))
TODAY = datetime.now().strftime('%y%m%d')
OUT_BASE = os.path.expanduser('~/Desktop/BUI_Investment_Management')
SEC_HDR = {'User-Agent':'BUI Research bui@example.com'}

# ============================================================
# Companies
# ============================================================
CONFIG = {
    "ILMN": dict(
        cik="0001110803", name="Illumina, Inc.",
        ccy="USD", unit_label="USD M", unit_div=1e6,
        ns="us-gaap",
        sector="Diagnostics / Sequencing",
        target_pe=25.0, target_evrev=4.0,
        # Historical share counts (M); ILMN ~145-160M throughout
        shares={2015:145, 2016:145, 2017:145, 2018:145, 2019:148,
                2020:147, 2021:155, 2022:158, 2023:159, 2024:160, 2025:152},
    ),
    "PHG":  dict(
        cik="0000313216", name="Koninklijke Philips N.V.",
        ccy="EUR", unit_label="EUR M", unit_div=1e6,
        ns="ifrs-full",
        sector="HealthTech / Industrial Healthcare",
        target_pe=18.0, target_evrev=2.0,
        # Philips ~870M-960M shares
        shares={2015:920, 2016:920, 2017:920, 2018:920, 2019:910,
                2020:900, 2021:880, 2022:870, 2023:910, 2024:935, 2025:951},
    ),
    "JNJ":  dict(
        cik="0000200406", name="Johnson & Johnson",
        ccy="USD", unit_label="USD M", unit_div=1e6,
        ns="us-gaap",
        sector="Mega-Pharma",
        target_pe=18.0, target_evrev=4.0,
        # JNJ ~2,400-2,800M shares
        shares={2015:2750, 2016:2710, 2017:2680, 2018:2670, 2019:2630,
                2020:2630, 2021:2630, 2022:2620, 2023:2410, 2024:2410, 2025:2407},
    ),
    "TMO":  dict(
        cik="0000097745", name="Thermo Fisher Scientific Inc.",
        ccy="USD", unit_label="USD M", unit_div=1e6,
        ns="us-gaap",
        sector="Life Sciences Tools / Diagnostics",
        target_pe=25.0, target_evrev=5.0,
        shares={2015:399, 2016:393, 2017:391, 2018:401, 2019:402,
                2020:393, 2021:393, 2022:388, 2023:382, 2024:382, 2025:372},
    ),
}

# US 10y Treasury (year-end)
US_10Y = {2015:0.0227, 2016:0.0244, 2017:0.0241, 2018:0.0269, 2019:0.0192,
          2020:0.0091, 2021:0.0151, 2022:0.0388, 2023:0.0388, 2024:0.0458, 2025:0.0428}

# German Bund 10y (proxy for EUR risk-free, year-end)
DE_BUND = {2015:0.0063, 2016:0.0021, 2017:0.0042, 2018:0.0024, 2019:-0.0019,
           2020:-0.0058, 2021:-0.0018, 2022:0.0257, 2023:0.0202, 2024:0.0231, 2025:0.0245}

# ============================================================
# Format
# ============================================================
FONT = "Apple Braille"
SIZE = 11
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
GREY    = "FFE7E6E6"
THIN    = Side(style="thin", color="FF000000")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.0;(#,##0.0);"–"'
INT_FMT = '#,##0;(#,##0);"–"'
PCT_FMT = '0.0%;(0.0%);"–"'
PX_FMT  = '#,##0.00'
X_FMT   = '0.00"x"'

def fmt(c, color=BLACK, bold=False, italic=False, num_fmt=None,
        align="right", grey=False):
    c.font = Font(name=FONT, size=SIZE, color=color, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if num_fmt: c.number_format = num_fmt
    if grey: c.fill = PatternFill("solid", fgColor=GREY)

def col_l(i): return get_column_letter(2+i)

# ============================================================
# SEC EDGAR fetcher (handles fiscal year-end window + IFRS)
# ============================================================
def make_facts_loader(cik):
    r = requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json",
                     headers=SEC_HDR, timeout=30)
    return r.json().get('facts', {})

def smart_annual(facts, ns, tags, units):
    """Try multiple tags; accept fiscal year-ends in Dec 26 - Jan 4 window. Annual flow only."""
    if isinstance(tags, str): tags = [tags]
    out_combined = {}
    for tag in tags:
        if tag not in facts.get(ns, {}): continue
        arr = facts[ns][tag].get('units', {}).get(units, [])
        for u in arr:
            if u.get('form') not in ('10-K','10-K/A','20-F','20-F/A'): continue
            end = u.get('end','')
            try: ed = date.fromisoformat(end)
            except: continue
            if not ((ed.month == 12 and ed.day >= 26) or (ed.month == 1 and ed.day <= 4)):
                continue
            fy_year = ed.year if ed.month == 12 else ed.year - 1
            start = u.get('start')
            if start:
                try:
                    d1 = date.fromisoformat(start)
                    if (ed - d1).days < 350: continue
                except: continue
            if fy_year not in out_combined or u.get('filed','') > out_combined[fy_year][1]:
                out_combined[fy_year] = (u['val'], u.get('filed',''))
    return {k: v[0] for k, v in out_combined.items()}

def smart_instant(facts, ns, tags, units):
    """Balance-sheet items at fiscal year-end."""
    if isinstance(tags, str): tags = [tags]
    out = {}
    for tag in tags:
        if tag not in facts.get(ns, {}): continue
        arr = facts[ns][tag].get('units', {}).get(units, [])
        for u in arr:
            if u.get('form') not in ('10-K','10-K/A','20-F','20-F/A'): continue
            end = u.get('end','')
            try: ed = date.fromisoformat(end)
            except: continue
            if not ((ed.month == 12 and ed.day >= 26) or (ed.month == 1 and ed.day <= 4)):
                continue
            if u.get('start'): continue
            fy_year = ed.year if ed.month == 12 else ed.year - 1
            if fy_year not in out or u.get('filed','') > out[fy_year][1]:
                out[fy_year] = (u['val'], u.get('filed',''))
    return {k: v[0] for k, v in out.items()}

# ============================================================
# Per-company tag maps
# ============================================================
def load_data(ticker, cfg):
    print(f"\n{'='*60}\n{ticker} {cfg['name']}\n{'='*60}")
    facts = make_facts_loader(cfg['cik'])
    print(f"  Loaded {sum(len(facts.get(ns,{})) for ns in facts)} XBRL tags across {list(facts.keys())} namespaces")

    ns = cfg['ns']
    ccy = cfg['ccy']

    if ns == "us-gaap":
        rev   = smart_annual(facts, ns, ['Revenues','RevenueFromContractWithCustomerExcludingAssessedTax'], ccy)
        rev2  = smart_annual(facts, ns, ['RevenueFromContractWithCustomerExcludingAssessedTax'], ccy)
        for y, v in rev2.items():
            if y not in rev: rev[y] = v
        gross = smart_annual(facts, ns, 'GrossProfit', ccy)
        opinc = smart_annual(facts, ns, 'OperatingIncomeLoss', ccy)
        rd    = smart_annual(facts, ns, 'ResearchAndDevelopmentExpense', ccy)
        sga   = smart_annual(facts, ns, 'SellingGeneralAndAdministrativeExpense', ccy)
        intinc= smart_annual(facts, ns, ['InvestmentIncomeInterest','InterestIncomeOperating'], ccy)
        intexp= smart_annual(facts, ns, 'InterestExpense', ccy)
        pretax= smart_annual(facts, ns,
                ['IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest',
                 'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments'],
                ccy)
        tax   = smart_annual(facts, ns, 'IncomeTaxExpenseBenefit', ccy)
        ni    = smart_annual(facts, ns, 'NetIncomeLoss', ccy)
        ni_p  = smart_annual(facts, ns,
                ['NetIncomeLossAvailableToCommonStockholdersBasic','NetIncomeLoss'], ccy)
        # Fallback: derive OpInc from GrossProfit - SGA - R&D (for JNJ post-2014 etc.)
        for y in YEARS:
            if y not in opinc and y in gross and y in sga:
                opinc[y] = gross[y] - sga[y] - (rd.get(y, 0) or 0)
        # If still missing, fallback to Pretax (less precise; includes non-op)
        for y in YEARS:
            if y not in opinc and y in pretax:
                opinc[y] = pretax[y]
        cash    = smart_instant(facts, ns, 'CashAndCashEquivalentsAtCarryingValue', ccy)
        mkt_sec = smart_instant(facts, ns, ['MarketableSecuritiesCurrent','ShortTermInvestments'], ccy)
        ar      = smart_instant(facts, ns, 'AccountsReceivableNetCurrent', ccy)
        ppe     = smart_instant(facts, ns, 'PropertyPlantAndEquipmentNet', ccy)
        goodwill= smart_instant(facts, ns, 'Goodwill', ccy)
        ta      = smart_instant(facts, ns, 'Assets', ccy)
        ap      = smart_instant(facts, ns, 'AccountsPayableCurrent', ccy)
        ltdebt  = smart_instant(facts, ns,
                  ['LongTermDebtNoncurrent','LongTermDebt'], ccy)
        tl      = smart_instant(facts, ns, 'Liabilities', ccy)
        eq      = smart_instant(facts, ns,
                  ['StockholdersEquity',
                   'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest'], ccy)
        ocf     = smart_annual(facts, ns, 'NetCashProvidedByUsedInOperatingActivities', ccy)
        icf     = smart_annual(facts, ns, 'NetCashProvidedByUsedInInvestingActivities', ccy)
        ffin    = smart_annual(facts, ns, 'NetCashProvidedByUsedInFinancingActivities', ccy)
        capex   = smart_annual(facts, ns,
                  ['PaymentsToAcquirePropertyPlantAndEquipment',
                   'PaymentsToAcquireProductiveAssets'], ccy)
        sbc     = smart_annual(facts, ns, 'ShareBasedCompensation', ccy)
        divs    = smart_annual(facts, ns,
                  ['PaymentsOfDividendsCommonStock',
                   'PaymentsOfDividends',
                   'PaymentsOfOrdinaryDividends'], ccy)
        da      = smart_annual(facts, ns,
                  ['DepreciationDepletionAndAmortization','DepreciationAndAmortization'], ccy)

    else:   # ifrs-full (PHG)
        rev   = smart_annual(facts, ns, 'Revenue', ccy)
        gross = smart_annual(facts, ns, 'GrossProfit', ccy)
        opinc = smart_annual(facts, ns,
                ['ProfitLossFromOperatingActivities','OperatingIncomeLoss'], ccy)
        rd    = smart_annual(facts, ns, 'ResearchAndDevelopmentExpense', ccy)
        sga   = smart_annual(facts, ns,
                ['DistributionCostsAndAdministrativeExpenses','AdministrativeExpense'], ccy)
        intinc= smart_annual(facts, ns, 'FinanceIncome', ccy)
        intexp= smart_annual(facts, ns, ['FinanceCosts','InterestExpense'], ccy)
        pretax= smart_annual(facts, ns,
                ['ProfitLossBeforeTax','ProfitLossFromContinuingOperationsBeforeTax'], ccy)
        tax   = smart_annual(facts, ns, 'IncomeTaxExpenseContinuingOperations', ccy)
        ni    = smart_annual(facts, ns, 'ProfitLoss', ccy)
        ni_p  = smart_annual(facts, ns,
                ['ProfitLossAttributableToOwnersOfParent','ProfitLoss'], ccy)
        cash    = smart_instant(facts, ns, 'CashAndCashEquivalents', ccy)
        mkt_sec = smart_instant(facts, ns, 'OtherFinancialAssetsCurrent', ccy)
        ar      = smart_instant(facts, ns, 'TradeAndOtherCurrentReceivables', ccy)
        ppe     = smart_instant(facts, ns, 'PropertyPlantAndEquipment', ccy)
        goodwill= smart_instant(facts, ns, 'Goodwill', ccy)
        ta      = smart_instant(facts, ns, 'Assets', ccy)
        ap      = smart_instant(facts, ns, 'TradeAndOtherCurrentPayables', ccy)
        ltdebt  = smart_instant(facts, ns,
                  ['NoncurrentBorrowings','LongtermBorrowings','BorrowingsNoncurrent'], ccy)
        tl      = smart_instant(facts, ns, 'Liabilities', ccy)
        eq      = smart_instant(facts, ns,
                  ['EquityAttributableToOwnersOfParent','Equity'], ccy)
        ocf     = smart_annual(facts, ns,
                  ['CashFlowsFromUsedInOperatingActivities',
                   'CashFlowsFromUsedInOperatingActivitiesContinuingOperations'], ccy)
        icf     = smart_annual(facts, ns,
                  ['CashFlowsFromUsedInInvestingActivities',
                   'CashFlowsFromUsedInInvestingActivitiesContinuingOperations'], ccy)
        ffin    = smart_annual(facts, ns,
                  ['CashFlowsFromUsedInFinancingActivities',
                   'CashFlowsFromUsedInFinancingActivitiesContinuingOperations'], ccy)
        capex   = smart_annual(facts, ns,
                  ['PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities'], ccy)
        sbc     = {}    # IFRS often consolidates SBC into operating expenses
        divs    = smart_annual(facts, ns, 'DividendsPaidClassifiedAsFinancingActivities', ccy)
        da      = smart_annual(facts, ns,
                  ['DepreciationAndAmortisationExpense',
                   'DepreciationAmortisationAndImpairmentLossReversalOfImpairmentLossRecognisedInProfitOrLoss'], ccy)

    # YE prices
    print("  Fetching prices (yfinance)...")
    yt = yf.Ticker(ticker)
    hist = yt.history(start="2014-01-01", end="2026-05-15", auto_adjust=False)
    ye_px = {d.year: float(v) for d, v in hist['Close'].resample('YE').last().items() if d.year in YEARS}
    info = {}
    try: info = yt.info
    except: pass
    holders = None
    try: holders = yt.institutional_holders
    except: pass

    return dict(
        rev=rev, gross=gross, opinc=opinc, rd=rd, sga=sga,
        intinc=intinc, intexp=intexp, pretax=pretax, tax=tax, ni=ni, ni_p=ni_p,
        cash=cash, mkt_sec=mkt_sec, ar=ar, ppe=ppe, goodwill=goodwill,
        ta=ta, ap=ap, ltdebt=ltdebt, tl=tl, eq=eq,
        ocf=ocf, icf=icf, ffin=ffin, capex=capex, sbc=sbc, divs=divs, da=da,
        ye_px=ye_px, info=info, holders=holders,
    )

def to_M(v, div):
    if v is None or pd.isna(v): return None
    return float(v) / div

# ============================================================
# Workbook builder
# ============================================================
def build(ticker, cfg, data):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Financial Data"
    s1_financial(ws1, ticker, cfg, data)
    ws2 = wb.create_sheet("Key Ratios"); s2_ratios(ws2, ticker, cfg)
    ws3 = wb.create_sheet("ROIC Analysis"); s3_roic(ws3, ticker, cfg)
    ws4 = wb.create_sheet("Valuation Model"); s4_valuation(ws4, ticker, cfg)
    ws5 = wb.create_sheet("Quarterly TTM"); s5_quarterly(ws5, ticker, cfg, data)
    ws6 = wb.create_sheet("Top Holders"); s6_holders(ws6, ticker, cfg, data['holders'])
    ws7 = wb.create_sheet("Data Correlation"); s7_correlation(ws7, ticker, cfg)
    return wb

# ===== Sheet 1: Financial Data =====
def s1_financial(ws, ticker, cfg, d):
    ud = cfg['unit_div']; ul = cfg['unit_label']; ccy = cfg['ccy']
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 12
    ws.column_dimensions[col_l(len(YEARS))].width = 8.83

    ws.cell(1,1, f"{cfg['name']} ({ticker}) — Financial Data ({ul})")
    fmt(ws.cell(1,1), bold=True, align="left")
    src_note = "SEC EDGAR XBRL " + ("(IFRS, " + ccy + ")" if cfg['ns']=='ifrs-full' else "(US-GAAP, " + ccy + ")")
    ws.cell(2,1, src_note); fmt(ws.cell(2,1), italic=True, color=BLUE, align="left")

    ws.cell(3,1, "Item"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")

    rows = [
        (4,  "==INCOME STATEMENT==", "header", None),
        (5,  "Revenue 营业收入", "data", d['rev']),
        (6,  "Revenue Growth (YoY)", "growth_rev", None),
        (7,  "Gross Profit 毛利", "data", d['gross']),
        (8,  "Gross Margin", "pct_rev", None),
        (9,  "R&D Expense", "data", d['rd']),
        (10, "SG&A Expense", "data", d['sga']),
        (11, "Operating Income (Loss)", "data", d['opinc']),
        (12, "Operating Margin", "op_margin", None),
        (13, "Interest Income", "data", d['intinc']),
        (14, "Interest Expense", "data", d['intexp']),
        (15, "Pretax Income", "data", d['pretax']),
        (16, "Income Tax", "data", d['tax']),
        (17, "Net Income (Loss)", "data", d['ni']),
        (18, "NI to Owners of Parent / Common", "data", d['ni_p']),
        (19, "Net Margin", "ni_margin", None),
        (20, "==BALANCE SHEET==", "header", None),
        (21, "Cash & Equivalents", "data", d['cash']),
        (22, "Marketable Securities (current)", "data", d['mkt_sec']),
        (23, "Accounts Receivable", "data", d['ar']),
        (24, "PP&E (net)", "data", d['ppe']),
        (25, "Goodwill", "data", d['goodwill']),
        (26, "Total Assets", "data", d['ta']),
        (27, "Accounts Payable", "data", d['ap']),
        (28, "Long-Term Debt (Noncurrent)", "data", d['ltdebt']),
        (29, "Total Liabilities", "data", d['tl']),
        (30, "Equity (to Owners of Parent)", "data", d['eq']),
        (31, "==CASH FLOW==", "header", None),
        (32, "Operating Cash Flow", "data", d['ocf']),
        (33, "Investing Cash Flow", "data", d['icf']),
        (34, "Financing Cash Flow", "data", d['ffin']),
        (35, "Capex (PP&E purchases)", "data_neg", d['capex']),
        (36, "Free Cash Flow (OCF − Capex)", "fcf", None),
        (37, "FCF Margin", "fcf_margin", None),
        (38, "Stock-Based Compensation", "data", d['sbc']),
        (39, "Depreciation & Amortization", "data", d['da']),
        (40, "Dividends Paid (Common)", "data_neg", d['divs']),
        (41, "==MARKET DATA==", "header", None),
        (42, "Year-End Share Price (local ccy)", "px", None),
        (43, "Diluted Shares (M)", "shares", None),
        (44, "Market Cap (price × shares)", "mc", None),
    ]
    for r, label, kind, src in rows:
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, grey=True, align="left")
            for i in range(len(YEARS)): fmt(ws.cell(r,2+i), grey=True)
            continue
        fmt(ws.cell(r,1), align="left")
        for i, y in enumerate(YEARS):
            cl = col_l(i); pcl = col_l(i-1) if i>0 else None
            cell = ws.cell(r, 2+i)
            if kind == "data":
                v = to_M(src.get(y), ud) if src else None
                if v is not None: cell.value = v; fmt(cell, color=BLUE, num_fmt=NUM_FMT)
                else: fmt(cell, color=BLUE, num_fmt=NUM_FMT)
            elif kind == "data_neg":
                v = to_M(src.get(y), ud) if src else None
                if v is not None:
                    cell.value = -abs(v)
                    fmt(cell, color=BLUE, num_fmt=NUM_FMT)
            elif kind == "growth_rev" and i > 0:
                cell.value = f"=IFERROR({cl}5/{pcl}5-1,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "pct_rev":
                cell.value = f"=IFERROR({cl}7/{cl}5,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "op_margin":
                cell.value = f"=IFERROR({cl}11/{cl}5,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "ni_margin":
                cell.value = f"=IFERROR({cl}17/{cl}5,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "fcf":
                cell.value = f"={cl}32+{cl}35"   # capex already negative
                fmt(cell, color=BLACK, num_fmt=NUM_FMT, bold=True)
            elif kind == "fcf_margin":
                cell.value = f"=IFERROR(({cl}32+{cl}35)/{cl}5,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "px":
                v = d['ye_px'].get(y)
                if v is not None: cell.value = v; fmt(cell, color=BLUE, num_fmt=PX_FMT)
                else: fmt(cell, color=BLUE, num_fmt=PX_FMT)
            elif kind == "shares":
                v = cfg['shares'].get(y)
                if v is not None: cell.value = v; fmt(cell, color=BLUE, num_fmt='#,##0.0')
            elif kind == "mc":
                cell.value = f"={cl}42*{cl}43"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT, bold=True)
    for rr in range(1, 45):
        ws.row_dimensions[rr].height = 18.0

# ===== Sheet 2: Key Ratios =====
def s2_ratios(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 38
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 12

    ws.cell(1,1, f"{cfg['name']} — Key Ratios"); fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1, "Ratio"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i, y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")
    fd = "'Financial Data'"
    ratios = [
        ("Revenue Growth (YoY)",        lambda c,p: f"=IFERROR({fd}!{c}5/{fd}!{p}5-1,\"\")", PCT_FMT, True),
        ("Gross Margin",                lambda c,p: f"=IFERROR({fd}!{c}7/{fd}!{c}5,\"\")", PCT_FMT, False),
        ("R&D / Revenue",               lambda c,p: f"=IFERROR({fd}!{c}9/{fd}!{c}5,\"\")", PCT_FMT, False),
        ("SG&A / Revenue",              lambda c,p: f"=IFERROR({fd}!{c}10/{fd}!{c}5,\"\")", PCT_FMT, False),
        ("Operating Margin (GAAP)",     lambda c,p: f"=IFERROR({fd}!{c}11/{fd}!{c}5,\"\")", PCT_FMT, False),
        ("Net Margin",                  lambda c,p: f"=IFERROR({fd}!{c}17/{fd}!{c}5,\"\")", PCT_FMT, False),
        ("FCF Margin",                  lambda c,p: f"=IFERROR(({fd}!{c}32+{fd}!{c}35)/{fd}!{c}5,\"\")", PCT_FMT, False),
        ("ROE (NI to Parent / Avg Equity)", lambda c,p: (f"=IFERROR({fd}!{c}18/AVERAGE({fd}!{c}30,{fd}!{p}30),\"\")" if p else None), PCT_FMT, True),
        ("ROA (NI / Avg Assets)",       lambda c,p: (f"=IFERROR({fd}!{c}17/AVERAGE({fd}!{c}26,{fd}!{p}26),\"\")" if p else None), PCT_FMT, True),
        ("Asset Turnover (Rev/Avg TA)", lambda c,p: (f"=IFERROR({fd}!{c}5/AVERAGE({fd}!{c}26,{fd}!{p}26),\"\")" if p else None), '0.00"x"', True),
        ("Cash + ST Sec / Total Assets", lambda c,p: f"=IFERROR(({fd}!{c}21+{fd}!{c}22)/{fd}!{c}26,\"\")", PCT_FMT, False),
        ("LT Debt / Equity",            lambda c,p: f"=IFERROR({fd}!{c}28/{fd}!{c}30,\"\")", '0.00"x"', False),
        ("Goodwill / Total Assets",     lambda c,p: f"=IFERROR({fd}!{c}25/{fd}!{c}26,\"\")", PCT_FMT, False),
        ("Dividend Payout (|Div|/NI)",  lambda c,p: f"=IFERROR(-{fd}!{c}40/{fd}!{c}17,\"\")", PCT_FMT, False),
    ]
    r = 4
    for label, fn, nfmt, needs_prior in ratios:
        ws.cell(r,1,label); fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl, pcl = col_l(i), col_l(i-1) if i>0 else None
            if needs_prior and pcl is None: continue
            f = fn(cl, pcl)
            if f:
                cell = ws.cell(r, 2+i); cell.value = f
                fmt(cell, color=GREEN, num_fmt=nfmt)
        r += 1
    for rr in range(1, r+1):
        ws.row_dimensions[rr].height = 20.0

# ===== Sheet 3: ROIC Analysis =====
def s3_roic(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 12
    ws.cell(1,1, f"{cfg['name']} — ROIC Decomposition")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1,"Component"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")
    fd = "'Financial Data'"
    rows = [
        ("Operating Income (Loss)",  lambda c,p: f"={fd}!{c}11", NUM_FMT, False, BLACK),
        ("+ Stock-Based Comp",       lambda c,p: f"={fd}!{c}38", NUM_FMT, False, BLACK),
        ("Adj NOPAT (pre-tax)",      lambda c,p: f"={c}4+{c}5", NUM_FMT, False, BLACK),
        ("Tax Rate (assumed)",       lambda c,p: "=0.21", PCT_FMT, False, BLUE),
        ("Adj NOPAT (after-tax)",    lambda c,p: f"={c}6*(1-{c}7)", NUM_FMT, False, BLACK),
        ("",                         lambda c,p: None, None, False, BLACK),
        ("Total Equity",             lambda c,p: f"={fd}!{c}30", NUM_FMT, False, BLACK),
        ("+ Long-Term Debt",         lambda c,p: f"={fd}!{c}28", NUM_FMT, False, BLACK),
        ("− Cash & ST Securities",   lambda c,p: f"=-({fd}!{c}21+{fd}!{c}22)", NUM_FMT, False, BLACK),
        ("Invested Capital",         lambda c,p: f"={c}10+{c}11+{c}12", NUM_FMT, False, BLACK),
        ("",                         lambda c,p: None, None, False, BLACK),
        ("ROIC (Adj NOPAT / Avg IC)", lambda c,p: (f"=IFERROR({c}8/AVERAGE({c}13,{p}13),\"\")" if p else None), PCT_FMT, True, GREEN),
    ]
    r = 4
    for label, fn, nfmt, needs_prior, color in rows:
        ws.cell(r,1,label); fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl, pcl = col_l(i), col_l(i-1) if i>0 else None
            if needs_prior and pcl is None: continue
            f = fn(cl, pcl)
            if f:
                cell = ws.cell(r, 2+i); cell.value = f
                fmt(cell, color=color, num_fmt=nfmt or NUM_FMT, bold=label.startswith(("ROIC","Adj NOPAT (after","Invested")))
        r += 1
    for rr in range(1, r+1):
        ws.row_dimensions[rr].height = 20.0

# ===== Sheet 4: Valuation Model =====
def s4_valuation(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 42
    for i in range(len(YEARS)):
        ws.column_dimensions[col_l(i)].width = 12
    ws.cell(1,1, f"{cfg['name']} — Valuation Model"); fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1,"Item"); fmt(ws.cell(3,1), bold=True, grey=True, align="left")
    for i,y in enumerate(YEARS):
        fmt(ws.cell(3,2+i,y), bold=True, grey=True, num_fmt="0", align="center")
    fd = "'Financial Data'"
    bond_table = US_10Y if cfg['ccy'] == 'USD' else DE_BUND
    bond_label = "US 10y Treasury" if cfg['ccy'] == 'USD' else "German Bund 10y"
    val_rows = [
        (5, "Key Assumptions", "header"),
        (6, bond_label, "bond"),
        (7, "Equity Risk Premium", "erp"),
        (8, "Cost of Equity", "coe"),
        (10, "Inputs (linked)", "header"),
        (11, "NI to Parent / Common", "link", f"{fd}!{{cl}}18"),
        (12, "Dividends Paid (abs)", "link_abs", f"{fd}!{{cl}}40"),
        (13, "Equity (to Parent)", "link", f"{fd}!{{cl}}30"),
        (14, "Market Cap", "link", f"{fd}!{{cl}}44"),
        (15, "Long-Term Debt", "link", f"{fd}!{{cl}}28"),
        (16, "Cash + ST Sec", "link_cash", None),
        (17, "Enterprise Value (MC + Debt − Cash)", "ev"),
        (19, "Multiples (Trading)", "header"),
        (20, "EV / Revenue", "ratio_a", (f"{{cl}}17", f"{fd}!{{cl}}5")),
        (21, "EV / EBITDA-proxy (OpInc + D&A + SBC)", "ev_ebitda"),
        (22, "P/E (NI > 0 only)", "pe"),
        (23, "P/B (Mkt Cap / Equity)", "pb"),
        (24, "Dividend Yield", "div_y"),
        (26, "Valuation Output (3-method)", "header"),
        (27, "Method 1: Target P/E × NI", "v1"),
        (28, f"  Target P/E ({cfg['target_pe']}x)", "tgt_pe"),
        (29, "Method 2: Target EV/Rev × Rev → MC", "v2"),
        (30, f"  Target EV/Rev ({cfg['target_evrev']}x)", "tgt_evrev"),
        (31, "Method 3: Justified P/B (Gordon)", "v3"),
        (32, "Average Implied Mkt Cap", "avg3"),
        (33, "Premium / (Discount) vs Actual", "pd"),
    ]
    for spec in val_rows:
        r = spec[0]; label = spec[1]; kind = spec[2]
        ws.cell(r,1, label)
        if kind == "header":
            fmt(ws.cell(r,1), bold=True, grey=True, align="left")
            for i in range(len(YEARS)): fmt(ws.cell(r,2+i), grey=True)
            continue
        fmt(ws.cell(r,1), align="left")
        for i in range(len(YEARS)):
            cl = col_l(i)
            cell = ws.cell(r, 2+i)
            if kind == "bond":
                cell.value = bond_table.get(YEARS[i], 0.03)
                fmt(cell, color=BLUE, num_fmt=PCT_FMT)
            elif kind == "erp":
                cell.value = 0.05
                fmt(cell, color=BLUE, num_fmt=PCT_FMT)
            elif kind == "coe":
                cell.value = f"={cl}6+{cl}7"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "link":
                cell.value = "=" + spec[3].replace("{cl}", cl)
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "link_abs":
                cell.value = f"=ABS({spec[3].replace('{cl}', cl)})"
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "link_cash":
                cell.value = f"={fd}!{cl}21+{fd}!{cl}22"
                fmt(cell, color=GREEN, num_fmt=NUM_FMT)
            elif kind == "ev":
                cell.value = f"={cl}14+{cl}15-{cl}16"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT, bold=True)
            elif kind == "ratio_a":
                num = spec[3][0].replace("{cl}", cl)
                den = spec[3][1].replace("{cl}", cl)
                cell.value = f"=IFERROR({num}/{den},\"\")"
                fmt(cell, color=BLACK, num_fmt=X_FMT)
            elif kind == "ev_ebitda":
                cell.value = f"=IFERROR({cl}17/({fd}!{cl}11+{fd}!{cl}38+{fd}!{cl}39),\"\")"
                fmt(cell, color=BLACK, num_fmt=X_FMT)
            elif kind == "pe":
                cell.value = f"=IFERROR(IF({fd}!{cl}17>0,{cl}14/{fd}!{cl}17,\"\"),\"\")"
                fmt(cell, color=BLACK, num_fmt=X_FMT)
            elif kind == "pb":
                cell.value = f"=IFERROR({cl}14/{cl}13,\"\")"
                fmt(cell, color=BLACK, num_fmt=X_FMT)
            elif kind == "div_y":
                cell.value = f"=IFERROR({cl}12/{cl}14,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT)
            elif kind == "v1":
                cell.value = f"=IF({cl}11>0,{cl}11*{cl}28,\"\")"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "tgt_pe":
                cell.value = cfg['target_pe']
                fmt(cell, color=BLUE, num_fmt=X_FMT)
            elif kind == "v2":
                cell.value = f"={fd}!{cl}5*{cl}30-{cl}15+{cl}16"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "tgt_evrev":
                cell.value = cfg['target_evrev']
                fmt(cell, color=BLUE, num_fmt=X_FMT)
            elif kind == "v3":
                # Justified P/B = (ROE-g)/(r-g) × Equity, simplistic
                # Use ROE = NI/Equity; g = ROE × (1 - Payout); r = bond + 5%
                cell.value = (f"=IFERROR(IF({cl}13>0,"
                              f"({fd}!{cl}18/{cl}13-({fd}!{cl}18/{cl}13)*0.7)/"
                              f"(({cl}6+0.05)-({fd}!{cl}18/{cl}13)*0.7)*{cl}13,\"\"),\"\")")
                fmt(cell, color=BLACK, num_fmt=NUM_FMT)
            elif kind == "avg3":
                cell.value = f"=IFERROR(AVERAGE({cl}27,{cl}29,{cl}31),\"\")"
                fmt(cell, color=BLACK, num_fmt=NUM_FMT, bold=True)
            elif kind == "pd":
                cell.value = f"=IFERROR({cl}32/{cl}14-1,\"\")"
                fmt(cell, color=BLACK, num_fmt=PCT_FMT, bold=True)
    for rr in range(1, 35):
        ws.row_dimensions[rr].height = 17.0

# ===== Sheet 5: Quarterly TTM =====
def s5_quarterly(ws, ticker, cfg, data):
    ws.column_dimensions['A'].width = 42
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.cell(1,1, f"{cfg['name']} — Quarterly TTM Snapshot")
    fmt(ws.cell(1,1), bold=True, align="left")
    ws.cell(3,1, "Quarterly data from SEC 10-Q (US-GAAP) or 6-K (foreign filer).")
    fmt(ws.cell(3,1), italic=True, color=BLUE, align="left")
    ws.cell(4,1, "PHG (foreign private issuer) typically files semi-annually only on SEC.")
    fmt(ws.cell(4,1), italic=True, color=BLUE, align="left")
    # Pull quarterly revenue
    ns = cfg['ns']; ccy = cfg['ccy']
    facts = make_facts_loader(cfg['cik'])
    rev_tags = (['Revenues','RevenueFromContractWithCustomerExcludingAssessedTax']
                if ns == 'us-gaap' else ['Revenue'])
    def get_q(tag_list):
        out = {}
        for tag in tag_list:
            if tag not in facts.get(ns, {}): continue
            arr = facts[ns][tag].get('units', {}).get(ccy, [])
            for u in arr:
                if u.get('form') not in ('10-Q','6-K','10-K','10-K/A','20-F','20-F/A'): continue
                end = u.get('end',''); start = u.get('start','')
                if not end or not start: continue
                try:
                    d1 = date.fromisoformat(start); d2 = date.fromisoformat(end)
                except: continue
                # Quarter = 80-100 days
                if 80 <= (d2-d1).days <= 100:
                    out[end] = u['val']
        return sorted(out.items())[-12:]
    q_rev = get_q(rev_tags)
    q_ni  = get_q(['NetIncomeLoss'] if ns=='us-gaap' else ['ProfitLoss'])
    q_ni_d = dict(q_ni)
    ws.cell(6,1,"Quarter End"); fmt(ws.cell(6,1), bold=True, grey=True, align="left")
    ws.cell(6,2,f"Revenue ({cfg['unit_label']})"); fmt(ws.cell(6,2), bold=True, grey=True, align="center")
    ws.cell(6,3,f"Net Income ({cfg['unit_label']})"); fmt(ws.cell(6,3), bold=True, grey=True, align="center")
    for i,(end,v) in enumerate(q_rev):
        r = 7+i
        ws.cell(r,1,end); fmt(ws.cell(r,1), align="center")
        ws.cell(r,2, v/cfg['unit_div']); fmt(ws.cell(r,2), color=BLUE, num_fmt=NUM_FMT)
        if end in q_ni_d:
            ws.cell(r,3, q_ni_d[end]/cfg['unit_div']); fmt(ws.cell(r,3), color=BLUE, num_fmt=NUM_FMT)

# ===== Sheet 6: Top Holders =====
def s6_holders(ws, ticker, cfg, holders):
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.cell(1,1, f"{cfg['name']} — Top Institutional Holders (yfinance snapshot)")
    fmt(ws.cell(1,1), bold=True, align="left")
    headers = ["#","Holder","Shares","Date Reported","Value"]
    for i, h in enumerate(headers):
        c = ws.cell(3,1+i,h)
        fmt(c, bold=True, color="FFFFFFFF", align="center" if i==0 else "left")
        c.fill = PatternFill("solid", fgColor="FF305496")
    if holders is not None and not holders.empty:
        for i, (_, row) in enumerate(holders.head(15).iterrows()):
            r = 4 + i
            ws.cell(r,1, i+1); fmt(ws.cell(r,1), align="center")
            ws.cell(r,2, str(row.get('Holder',''))[:60]); fmt(ws.cell(r,2), align="left")
            try:
                sh = float(row.get('Shares', 0)) if pd.notna(row.get('Shares')) else None
                if sh is not None:
                    ws.cell(r,3, sh); fmt(ws.cell(r,3), color=BLUE, num_fmt=INT_FMT)
            except: pass
            ws.cell(r,4, str(row.get('Date Reported','')))
            fmt(ws.cell(r,4), align="center")
            try:
                v = float(row.get('Value', 0)) if pd.notna(row.get('Value')) else None
                if v is not None:
                    ws.cell(r,5, v); fmt(ws.cell(r,5), color=BLUE, num_fmt=INT_FMT)
            except: pass
    else:
        ws.cell(4,1, "Top holders data not retrieved — use 13F filings on SEC EDGAR")
        fmt(ws.cell(4,1), italic=True, align="left")
    for rr in range(1, 20):
        ws.row_dimensions[rr].height = 15.0

# ===== Sheet 7: Data Correlation =====
def s7_correlation(ws, ticker, cfg):
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 110
    ws.cell(1,1, f"{cfg['name']} ({ticker}) — Data Correlation & Methodology")
    fmt(ws.cell(1,1), bold=True, align="left")
    sections = [
        ("1. Company Overview",
         f"{cfg['name']} — {cfg['sector']}.\n"
         f"Years modeled: 2015-2025 (11 years). Reporting currency: {cfg['ccy']} ({cfg['unit_label']}).\n"
         f"CIK: {cfg['cik']} | Filing standard: {'IFRS (20-F foreign filer)' if cfg['ns']=='ifrs-full' else 'US GAAP (10-K)'}\n"),
        ("2. Data Sources",
         f"PRIMARY: SEC EDGAR XBRL companyfacts API\n"
         f"  https://data.sec.gov/api/xbrl/companyfacts/CIK{cfg['cik']}.json\n"
         f"  Namespace: {cfg['ns']} | Units: {cfg['ccy']}\n"
         "SECONDARY: yfinance — year-end share prices, institutional holders snapshot.\n"
         f"yfinance only provides 4-5 years of financial history, INSUFFICIENT for 11-year template.\n"
         "SEC EDGAR is the authoritative source for full historical financials."),
        ("3. Currency & Unit Convention",
         f"All financials in {cfg['unit_label']} (raw {cfg['ccy']} ÷ 1e6).\n"
         f"Shares in millions. Market Cap = YE price × shares (in {cfg['ccy']} M).\n"
         f"{'IMPORTANT: PHG reports in EUR. NO conversion to USD performed.' if cfg['ccy']=='EUR' else ''}\n"
         f"{'For cross-comparison with USD healthcare peers, apply EUR/USD FX (~1.07-1.10).' if cfg['ccy']=='EUR' else ''}"),
        ("4. Fiscal Year-End Quirks",
         "ILMN, JNJ: Fiscal year ends Sunday-nearest-Dec-31 (e.g. Dec 28 / Dec 29 / Dec 31).\n"
         "  Our model accepts year-ends Dec 26 - Jan 4, mapping to the relevant FY.\n"
         "TMO: Calendar year-end Dec 31.\n"
         "PHG: Calendar year-end Dec 31, but reports semi-annually (no quarterly 10-Q on SEC).\n"
         "  → Quarterly TTM sheet may be sparse for PHG."),
        ("5. Key XBRL Tag Mapping",
         "US GAAP (ILMN, JNJ, TMO):\n"
         "  Revenue: Revenues OR RevenueFromContractWithCustomerExcludingAssessedTax\n"
         "    (TMO needs both: pre-2018 used 'Revenues', post-ASC 606 used the longer name)\n"
         "  Operating Income: OperatingIncomeLoss\n"
         "  NI: NetIncomeLoss\n"
         "  OCF: NetCashProvidedByUsedInOperatingActivities\n"
         "  Capex: PaymentsToAcquirePropertyPlantAndEquipment\n"
         "  Equity: StockholdersEquity\n"
         "IFRS (PHG):\n"
         "  Revenue: Revenue\n"
         "  Operating Income: ProfitLossFromOperatingActivities\n"
         "  NI: ProfitLoss / ProfitLossAttributableToOwnersOfParent\n"
         "  OCF: CashFlowsFromUsedInOperatingActivities\n"
         "  Capex: PurchaseOfPropertyPlantAndEquipmentClassifiedAsInvestingActivities\n"
         "  Equity: EquityAttributableToOwnersOfParent\n"
         "  LT Debt: NoncurrentBorrowings"),
        ("6. Valuation Methodology",
         "Three-method blend:\n"
         f"  Method 1 — P/E × NI (Target P/E = {cfg['target_pe']}x for {cfg['sector']})\n"
         f"  Method 2 — EV/Revenue × Revenue (Target EV/Rev = {cfg['target_evrev']}x); convert EV → MC by removing debt and adding cash\n"
         "  Method 3 — Justified P/B (Gordon): (ROE − g) / (r − g) × Equity, with payout assumed 70%\n"
         "All targets are EDITABLE BLUE inputs — calibrate to peer group."),
        ("7. Sector-Specific Notes",
         {
            'ILMN': "Illumina: NGS sequencing platform. Major hits: GRAIL acquisition (2021)\n"
                    "    led to FTC/EU divestiture order, ~$4.4B writedown 2022. Operating losses 2022-2024.\n"
                    "    2025 turnaround under new CEO. Watch: Element / Ultima / MGI competition.",
            'PHG':  "Royal Philips: Sleep & Respiratory Care recall (2021-2024) → ~€1.2B charges,\n"
                    "    DOJ settlement May 2024 (~€1.1B). Re-focused on diagnostic imaging + personal health.\n"
                    "    Connected Care growing.",
            'JNJ':  "J&J: Talc litigation (~$10B+ reserved), 2023 Kenvue spin-off (consumer health → KVUE).\n"
                    "    Post-spin, JNJ = pharma + medtech only. 2024 NI dipped on talc + Stelara LOE.",
            'TMO':  "Thermo Fisher: Acquisitive growth (PPD 2021 ~$17.4B, Olink 2024). Largest life-sci tools\n"
                    "    company globally. COVID-tail (~$5B in 2022 → $0 by 2024) explains recent decline.\n"
                    "    Now stabilizing; bioprocess & analytical instruments core."
         }.get(ticker, "")),
        ("8. Caveats",
         "• Share counts (row 43) hardcoded from 10-K cover pages; mid-year buybacks not captured.\n"
         "• PHG: NO USD conversion applied (Plan A). To compare to USD peers, multiply by EUR/USD ~1.08.\n"
         "• 2025 figures reflect FY25 reports (typically published Jan-Mar 2026).\n"
         "• Target P/E and EV/Rev are CYCLICAL benchmarks; for ILMN's loss years, P/E method returns blank.\n"
         "• ILMN/PHG: substantial impairment charges create one-off NI distortions — clean for valuation."),
    ]
    r = 3
    for title, body in sections:
        ws.cell(r,2, title); fmt(ws.cell(r,2), bold=True, grey=True, align="left")
        r += 1
        for ln in body.split("\n"):
            ws.cell(r,2, ln); fmt(ws.cell(r,2), align="left")
            ws.row_dimensions[r].height = 17
            r += 1
        r += 1


def main():
    for ticker, cfg in CONFIG.items():
        try:
            data = load_data(ticker, cfg)
            print(f"  Building workbook...")
            wb = build(ticker, cfg, data)
            outdir = os.path.join(OUT_BASE, ticker)
            os.makedirs(outdir, exist_ok=True)
            outpath = os.path.join(outdir, f"{ticker}.Valuation.v{TODAY}.xlsx")
            wb.save(outpath)
            print(f"  ✓ Saved: {outpath}")
        except Exception as e:
            import traceback
            print(f"\n!! ERR {ticker}: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    main()
